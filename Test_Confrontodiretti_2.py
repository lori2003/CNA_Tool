import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import os
import re
import sys
from copy import copy

def extract_year_from_filename(file_path):
    name = os.path.basename(file_path)
    match = re.search(r"(?:19|20)\d{2}", name)
    if not match:
        return None
    return int(match.group(0))

def last_non_empty_row(ws, col_idx, start_row=4):
    for row in range(ws.max_row, start_row - 1, -1):
        if ws.cell(row=row, column=col_idx).value not in (None, ""):
            return row
    return start_row - 1

def normalize_name(value):
    if value is None:
        return None
    text = "".join(str(value).split()).upper()
    return text if text else None

def rename_province(value):
    if value is None:
        return None
    raw = str(value).strip()
    key = normalize_name(raw)
    rename_map = {
        normalize_name("VALLE D'AOS."): "VALLE D'AOSTA",
        normalize_name("VERBANO-C.-O."): "VERBANO C.O.",
        normalize_name("VARESE"): "LOMBARDIA N.OVEST",
        normalize_name("LECCO"): "LOMBARDIA N.OVEST",
        normalize_name("LARIO BRIANZA"): "LOMBARDIA N.OVEST",
        normalize_name("MILANO"): "MILANO*",
    }
    return rename_map.get(key, raw)

def to_float(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if not txt:
            return 0.0
        txt = txt.replace(".", "").replace(",", ".")
        try:
            return float(txt)
        except ValueError:
            return 0.0
    return 0.0

def get_header_value(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return ws.cell(row=row, column=col).value

def set_cell_value_safe(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    else:
        cell.value = value

def get_writable_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell

def apply_second_phase(ws, source_paths):
    if not source_paths:
        return
    if len(source_paths) != 2:
        print("Seconda fase: seleziona esattamente 2 file.")
        return

    sorted_sources = sorted(
        source_paths,
        key=lambda p: (
            extract_year_from_filename(p) if extract_year_from_filename(p) is not None else 9999,
            os.path.basename(p).lower()
        )
    )

    # Usa le colonne vuote gia' predisposte dall'utente: 2 colonne per file.
    base_col = 2  # primo file: B-C, secondo file: D-E
    base_last_row = last_non_empty_row(ws, 1, 4)

    # Mappa provincia (colonna A) -> riga, per allineamento sulla stessa riga
    a_name_to_row = {}
    for r in range(4, base_last_row + 1):
        a_norm = normalize_name(ws.cell(row=r, column=1).value)
        if a_norm and a_norm not in a_name_to_row:
            a_name_to_row[a_norm] = r

    # Pulisce l'area B:E da riga 4 in poi, così niente residui
    for r in range(4, base_last_row + 1):
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    fill_match = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    fill_miss = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    for idx, source_path in enumerate(sorted_sources):
        source_wb = openpyxl.load_workbook(source_path, data_only=True)
        source_ws = source_wb["Foglio2"] if "Foglio2" in source_wb.sheetnames else source_wb.active

        year_val = extract_year_from_filename(source_path)
        year_label = str(year_val) if year_val is not None else "Anno"

        start_col = base_col + (idx * 2)
        # Intestazioni anno sul blocco da 2 colonne
        for c in range(start_col, start_col + 2):
            ws.cell(row=3, column=c).value = year_label

        last_src_row = max(last_non_empty_row(source_ws, 1, 4), last_non_empty_row(source_ws, 2, 4))
        # Aggrega per riga target allineata alla colonna A
        sum_by_target_row = {}
        for r in range(4, last_src_row + 1):
            src_name = rename_province(source_ws.cell(row=r, column=1).value)
            src_norm = normalize_name(src_name)
            if not src_norm:
                continue
            target_row = a_name_to_row.get(src_norm)
            if target_row is None:
                continue
            sum_by_target_row[target_row] = sum_by_target_row.get(target_row, 0.0) + to_float(
                source_ws.cell(row=r, column=2).value
            )

        # Scrive provincia + valore sulla stessa riga della provincia in A
        for target_row, num_value in sum_by_target_row.items():
            ws.cell(row=target_row, column=start_col).value = ws.cell(row=target_row, column=1).value
            ws.cell(row=target_row, column=start_col + 1).value = num_value

        # Colora A e la colonna nome del file se provincia combacia
        for r in range(4, base_last_row + 1):
            a_name = normalize_name(ws.cell(row=r, column=1).value)
            b_name = normalize_name(ws.cell(row=r, column=start_col).value)

            if a_name and a_name == b_name:
                target_fill = fill_match
            elif a_name or b_name:
                target_fill = fill_miss
            else:
                continue

            ws.cell(row=r, column=1).fill = target_fill
            ws.cell(row=r, column=start_col).fill = target_fill

def main():
    # Configurazione finestra di dialogo
    root = tk.Tk()
    root.withdraw()  # Nasconde la finestra principale

    skip_first_phase = messagebox.askyesno(
        "Workflow",
        "Vuoi saltare la prima fase ed eseguire solo la seconda parte?"
    )

    if skip_first_phase:
        print("Seleziona il file Excel elaborato...")
        elaborato_path = filedialog.askopenfilename(
            title="Seleziona file Excel elaborato",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if not elaborato_path:
            print("Nessun file selezionato. Operazione annullata.")
            return

        print("Seleziona uno o piu file Excel anni...")
        source_paths = filedialog.askopenfilenames(
            title="Seleziona file Excel anni",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if not source_paths:
            print("Nessun file anni selezionato. Operazione annullata.")
            return

        wb = openpyxl.load_workbook(elaborato_path)
        if "Foglio2" not in wb.sheetnames:
            print("Errore: Il foglio 'Foglio2' non esiste nel file elaborato.")
            return
        ws = wb["Foglio2"]
        apply_second_phase(ws, source_paths)

        base_name, ext = os.path.splitext(elaborato_path)
        new_file_path = f"{base_name}_SecondaFase{ext}"
        wb.save(new_file_path)
        print(f"Seconda fase completata. File salvato come: {new_file_path}")
        return

    print("Seleziona il file Excel dalla finestra di dialogo...")
    
    # Apre la finestra di dialogo per selezionare il file
    file_path = filedialog.askopenfilename(
        title="Seleziona il file Excel",
        filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
    )

    if not file_path:
        print("Nessun file selezionato. Operazione annullata.")
        return

    print(f"Elaborazione del file: {file_path}")

    try:
        # Carica il workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Verifica esistenza Foglio2
        if "Foglio2" not in wb.sheetnames:
            print("Errore: Il foglio 'Foglio2' non esiste nel file.")
            # Se vuoi, potresti chiedere all'utente o prendere il foglio attivo
            # ws = wb.active
            return
        
        ws = wb["Foglio2"]
        
        # Funzione helper per normalizzare (rimuove tutti gli spazi e converte in minuscolo)
        def normalizza(valore):
            if valore is None:
                return None
            return str(valore).replace(" ", "").lower()
            
        # 1. Raccogli i valori della Colonna A in un set per confronto veloce
        valori_colonna_A = set()
        # Itera su tutte le celle della colonna A che hanno dati
        for cell in ws['A']:
            if cell.value is not None:
                valori_colonna_A.add(normalizza(cell.value))
        
        # Colori ARGB (FF = opaco). Con 6 cifre openpyxl usa alpha 00 (trasparente).
        fill_verde = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # verde chiaro
        fill_rosso = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # rosso chiaro
        font_verde = Font(color="FF008000")
        font_rosso = Font(color="FFFF0000")
        no_fill = PatternFill(fill_type=None)
        
        # Contatore per la colonna 20 (T)
        contatore = 1
        
        # 2. Itera su tutte le celle della Colonna B
        # ws['B'] restituisce le celle della colonna B presenti nel range utilizzato
        for cell in ws['B']:
            original_val_b = cell.value
            
            if original_val_b is not None:
                valore_b_norm = normalizza(original_val_b)
                
                # Condizione: Identici a quelli in colonna A (normalizzati)
                if valore_b_norm in valori_colonna_A:
                    # TROVATO: Azione 1: Colora la cella di verde chiaro
                    cell.fill = fill_verde
                    
                    # TROVATO: Azione 2: Numerazione in colonna 20 (T)
                    if cell.row >= 4:
                        # Scrive il numero nella colonna 20 (T) della stessa riga
                        target_cell = ws.cell(row=cell.row, column=20)
                        
                        if isinstance(target_cell, MergedCell):
                             # Se la cella e' unita, trova la cella principale (top-left) e scrivi li'
                            found = False
                            for merged_range in ws.merged_cells.ranges:
                                if target_cell.coordinate in merged_range:
                                    ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = contatore
                                    found = True
                                    break
                            if not found:
                                # Fallback se non trovato nel range (strano ma possibile)
                                target_cell.value = contatore
                        else:
                            # Cella normale
                            target_cell.value = contatore

                        contatore += 1
                else:
                    # NON TROVATO: Azione 3: Colora la cella di rosso chiaro
                    cell.fill = fill_rosso
        
        # 3. Gestione Formattazione (Colonne 3-19)
        # Colonne da formattare come numeri: 3-19 (C-S)
        # Colonne da colorare (Verde/Rosso) con formule Excel: 5, 6, 10, 11, 15, 16
        colonne_da_colorare = [5, 6, 10, 11, 15, 16]

        max_row = ws.max_row
        if max_row >= 4:
            for row in range(4, max_row + 1):
                for col_idx in range(3, 20): # range(3, 20) include 3 fino a 19
                    cell = ws.cell(row=row, column=col_idx)
                    valore = cell.value
                    
                    # 3A. CONVERSIONE E FORMATTAZIONE
                    val_num = None
                    
                    if isinstance(valore, (int, float)):
                         val_num = float(valore)
                    elif isinstance(valore, str) and valore.strip():
                        try:
                            # Tenta conversione robusta (gestisce 1.000,00)
                            # Rimuove punti migliaia e sostituisce virgola con punto
                            clean_val = valore.replace('.', '').replace(',', '.')
                            val_num = float(clean_val)
                            # Aggiorna il valore nella cella con il numero convertito
                            cell.value = val_num
                        except ValueError:
                            pass # Mantieni come testo se non è convertibile

                    # Imposta formato cella a Numero
                    cell.number_format = '#,##0.00'
                    
                    # Evita colori statici nelle colonne che useranno formule di formattazione condizionale
                    if col_idx in colonne_da_colorare:
                        cell.fill = no_fill

            # 3B. Colorazione via formule Excel (formattazione condizionale)
            for col_idx in colonne_da_colorare:
                col_letter = get_column_letter(col_idx)
                range_addr = f"{col_letter}4:{col_letter}{max_row}"

                # Verde se valore > 0
                ws.conditional_formatting.add(
                    range_addr,
                    FormulaRule(formula=[f"{col_letter}4>0"], font=font_verde)
                )
                # Rosso se valore < 0
                ws.conditional_formatting.add(
                    range_addr,
                    FormulaRule(formula=[f"{col_letter}4<0"], font=font_rosso)
                )

        # 4. Import da 2 file Excel esterni, ordinati per anno nel nome file
        print("Seleziona i 2 file Excel sorgente da cui importare i dati...")
        source_paths = filedialog.askopenfilenames(
            title="Seleziona 2 file sorgente",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )

        if len(source_paths) != 2:
            print("Errore: devi selezionare esattamente 2 file sorgente.")
            return

        sorted_sources = sorted(
            source_paths,
            key=lambda p: (
                extract_year_from_filename(p) if extract_year_from_filename(p) is not None else 9999,
                os.path.basename(p).lower()
            )
        )

        # Pulisce area destinazione (U:AB) per evitare residui da elaborazioni precedenti
        target_clear_max = max(ws.max_row, 4)
        for col_idx in range(21, 29):
            ws.cell(row=3, column=col_idx).value = None
            for row in range(4, target_clear_max + 1):
                ws.cell(row=row, column=col_idx).value = None

        for file_idx, source_path in enumerate(sorted_sources):
            source_year = extract_year_from_filename(source_path)
            source_year_label = str(source_year) if source_year is not None else os.path.basename(source_path)

            source_wb = openpyxl.load_workbook(source_path, data_only=True)
            if "Foglio2" in source_wb.sheetnames:
                source_ws = source_wb["Foglio2"]
            else:
                source_ws = source_wb.active

            dest_col_start = 21 + (file_idx * 4)
            ws.cell(row=3, column=dest_col_start, value=source_year_label)
            ws.cell(row=3, column=dest_col_start + 1, value="concomitanti")
            ws.cell(row=3, column=dest_col_start + 2, value="deleghe")
            ws.cell(row=3, column=dest_col_start + 3, value="revoche")

            source_last_row = last_non_empty_row(source_ws, 1, 4)
            if source_last_row < 4:
                continue

            # Aggrega per nome provincia rinominato (somma valori se duplicato)
            aggregated = {}
            for row in range(4, source_last_row + 1):
                province_name = rename_province(source_ws.cell(row=row, column=1).value)
                norm_name = normalize_name(province_name)
                if not norm_name:
                    continue
                if norm_name not in aggregated:
                    aggregated[norm_name] = {
                        "name": province_name,
                        "c15": 0.0,
                        "c41": 0.0,
                        "c28": 0.0,
                    }
                aggregated[norm_name]["c15"] += to_float(source_ws.cell(row=row, column=15).value)
                aggregated[norm_name]["c41"] += to_float(source_ws.cell(row=row, column=41).value)
                aggregated[norm_name]["c28"] += to_float(source_ws.cell(row=row, column=28).value)

            # Scrive dati aggregati in U:AB
            out_row = 4
            for item in sorted(aggregated.values(), key=lambda x: normalize_name(x["name"]) or ""):
                ws.cell(row=out_row, column=dest_col_start, value=item["name"])
                ws.cell(row=out_row, column=dest_col_start + 1, value=item["c15"])
                ws.cell(row=out_row, column=dest_col_start + 2, value=item["c41"])
                ws.cell(row=out_row, column=dest_col_start + 3, value=item["c28"])
                out_row += 1

        # 5. Match nomi (B vs U=21) e spostamento valori tra colonne
        # Prima svuota completamente le colonne target richieste
        colonne_da_svuotare = [3, 4, 8, 9, 13, 14, 18, 19]
        max_row_reconcile = max(ws.max_row, 4)
        for row in range(4, max_row_reconcile + 1):
            for col_idx in colonne_da_svuotare:
                ws.cell(row=row, column=col_idx).value = None

        # Indicizza i nomi in colonna B per match veloce
        b_name_to_row = {}
        for row in range(4, max_row_reconcile + 1):
            n = normalize_name(ws.cell(row=row, column=2).value)
            if n and n not in b_name_to_row:
                b_name_to_row[n] = row

        # Mappature richieste per i due blocchi anno
        block_mappings = [
            (21, {22: 3, 23: 8, 24: 13}),  # primo file (anno piu' vecchio)
            (25, {26: 4, 27: 9, 28: 14}),  # secondo file
        ]
        fill_giallo = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")  # giallo chiaro visibile
        fill_rosso_chiaro = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

        # Per ogni riga sorgente in U:AB, confronta nome (colonna 21/25) con colonna B
        for row in range(4, max_row_reconcile + 1):
            for name_col, source_to_dest in block_mappings:
                source_name_norm = normalize_name(ws.cell(row=row, column=name_col).value)
                matched_row = b_name_to_row.get(source_name_norm)

                for source_col, dest_col in source_to_dest.items():
                    source_cell = ws.cell(row=row, column=source_col)
                    src_val = source_cell.value

                    if src_val in (None, ""):
                        continue

                    if matched_row is not None:
                        dest_cell = ws.cell(row=matched_row, column=dest_col)
                        # Se ci sono piu' righe sorgente con stesso nome, somma nella stessa riga destinazione
                        dest_cell.value = to_float(dest_cell.value) + to_float(src_val)
                        dest_cell.fill = fill_giallo
                        # Elimina valore dalla cella origine dopo lo spostamento
                        source_cell.value = None
                        source_cell.fill = no_fill
                    else:
                        # Nessun match: evidenzia la cella origine in rosso chiaro
                        source_cell.fill = fill_rosso_chiaro

        # 6. Formule su colonne 5,6,10,11,15,16 per righe con testo in colonna B
        final_max_row = max(ws.max_row, 4)
        for row in range(4, final_max_row + 1):
            b_val = ws.cell(row=row, column=2).value
            if b_val is None or str(b_val).strip() == "":
                continue

            # Equivalenti A1 delle formule richieste in stile RC
            ws.cell(row=row, column=5).value = f"=D{row}-C{row}"
            ws.cell(row=row, column=6).value = f"=E{row}/D{row}/100%"
            ws.cell(row=row, column=10).value = f"=I{row}-H{row}"
            ws.cell(row=row, column=11).value = f"=J{row}/I{row}/100%"
            ws.cell(row=row, column=15).value = f"=N{row}-M{row}"
            ws.cell(row=row, column=16).value = f"=O{row}/N{row}/100%"

            # Formato percentuale per le colonne rapporto
            ws.cell(row=row, column=6).number_format = "0.00%"
            ws.cell(row=row, column=11).number_format = "0.00%"
            ws.cell(row=row, column=16).number_format = "0.00%"

        # Mantiene la colorazione positivo/negativo sulle colonne formula
        for col_idx in colonne_da_colorare:
            col_letter = get_column_letter(col_idx)
            range_addr = f"{col_letter}4:{col_letter}{final_max_row}"
            ws.conditional_formatting.add(
                range_addr,
                FormulaRule(formula=[f"{col_letter}4>0"], font=font_verde)
            )
            ws.conditional_formatting.add(
                range_addr,
                FormulaRule(formula=[f"{col_letter}4<0"], font=font_rosso)
            )

        # 7. Grassetto + bordo piu' spesso su righe con specifici numeri in colonna 20
        target_numbers = {
            6, 9, 15, 21, 32, 37, 44, 49, 59, 65, 68, 78, 85, 92, 102, 113, 116, 119, 121, 130, 131
        }
        medium_side = Side(style="medium", color="FF000000")
        group_ranges = [(2, 6), (8, 11), (13, 16), (18, 19)]

        for row in range(4, final_max_row + 1):
            idx_val = ws.cell(row=row, column=20).value
            try:
                idx_num = int(float(idx_val))
            except (TypeError, ValueError):
                continue

            if idx_num not in target_numbers:
                continue

            # Applica stile solo nei gruppi richiesti tra colonna 2 e 19
            for start_col, end_col in group_ranges:
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value in (None, ""):
                        continue

                    f = copy(cell.font)
                    f.bold = True
                    cell.font = f

                    # Bordo "a gruppo": top/bottom su tutte, left solo inizio, right solo fine
                    left = medium_side if col == start_col else cell.border.left
                    right = medium_side if col == end_col else cell.border.right
                    cell.border = Border(
                        left=left,
                        right=right,
                        top=medium_side,
                        bottom=medium_side,
                    )

        # 8. Trasforma ex col. 18 e 19 in due gruppi da 2 colonne: (18-19) e (20-21)
        # Salva le intestazioni esistenti (riga 1/2) delle vecchie colonne 18 e 19
        old_header_18 = get_header_value(ws, 1, 18)
        if old_header_18 in (None, ""):
            old_header_18 = get_header_value(ws, 2, 18)
        old_header_19 = get_header_value(ws, 1, 19)
        if old_header_19 in (None, ""):
            old_header_19 = get_header_value(ws, 2, 19)

        # Inserisce 2 colonne prima della vecchia 19 per ottenere spazio per i 2 gruppi
        ws.insert_cols(19, 2)

        # Pulisce eventuali merge che toccano l'area header target
        to_unmerge = []
        for mr in ws.merged_cells.ranges:
            if mr.max_row >= 1 and mr.min_row <= 3 and mr.max_col >= 18 and mr.min_col <= 21:
                to_unmerge.append(str(mr))
        for rng in to_unmerge:
            try:
                ws.unmerge_cells(rng)
            except KeyError:
                # Alcuni file hanno merge corrotti/parziali: ignora e continua
                pass

        # Ricrea intestazioni unite su righe 1-2
        ws.merge_cells(start_row=1, start_column=18, end_row=2, end_column=19)
        set_cell_value_safe(ws, 1, 18, old_header_18)
        ws.merge_cells(start_row=1, start_column=20, end_row=2, end_column=21)
        set_cell_value_safe(ws, 1, 20, old_header_19)

        # Intestazioni riga 3 richieste
        set_cell_value_safe(ws, 3, 18, "C+D-R")
        set_cell_value_safe(ws, 3, 19, "Tasso")
        set_cell_value_safe(ws, 3, 20, "D-R")
        set_cell_value_safe(ws, 3, 21, "Tasso.")

        # Formato richiesto per colonne 18-21 (header)
        yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        medium_side = Side(style="medium", color="FF000000")
        medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
        thin_side = Side(style="thin", color="FF000000")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        # Larghezza colonne 18-21 = 9
        for col in range(18, 22):
            ws.column_dimensions[get_column_letter(col)].width = 9

        # Grassetto, giallo e bordo spesso su area header (righe 1-3, colonne 18-21)
        for row in range(1, 4):
            for col in range(18, 22):
                c = get_writable_cell(ws, row, col)
                c.font = bold_font
                c.fill = yellow_fill
                c.border = medium_border
                c.alignment = center_align

        # Bordi sottili su tutte le celle delle colonne 19 e 20
        max_row_all = max(ws.max_row, 3)
        for row in range(1, max_row_all + 1):
            for col in (19, 20):
                c = get_writable_cell(ws, row, col)
                c.border = thin_border
                c.alignment = center_align

        # Seconda parte: import comparativo dal file 2024
        print("Seleziona uno o piu file Excel anni per la seconda parte...")
        source_paths = filedialog.askopenfilenames(
            title="Seleziona file Excel anni",
            filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xltx;*.xltm")]
        )
        if source_paths:
            apply_second_phase(ws, source_paths)

        # Salvataggio
        base_name, ext = os.path.splitext(file_path)
        new_file_path = f"{base_name}_Elaborato{ext}"
        
        wb.save(new_file_path)
        print(f"Operazione completata con successo!")
        print(f"File salvato come: {new_file_path}")
        
        # Apre la cartella del file generato (opzionale, utile per l'utente)
        try:
           os.startfile(os.path.dirname(new_file_path))
        except:
           pass

    except Exception as e:
        print(f"Si è verificato un errore durante l'elaborazione: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
