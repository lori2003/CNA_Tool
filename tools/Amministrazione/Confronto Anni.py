from __future__ import annotations

import os
import json
import shutil
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# =========================
# CONFIGURAZIONE PERSISTENTE
# =========================
CONFIG_FILE = Path(__file__).parent / "confronto_anni_config.json"

def load_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r") as f:
                return json.load(f)
        except:
            pass
    return {"output_filename": "CONFRONTO_DATI_RINNOVO"}

def save_config(data: dict):
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(data, f, indent=4)
        return True
    except:
        return False

# =========================
# TOOLBOX CONFIGURATION
# =========================
TOOL = {
    'id': 'confronto_anni',
    'name': 'Confronto Anni',
    'region': 'Amministrazione',
    'email_reminder': "File da mandare all'amministrazione — questo script produce/modifica 2 file/fogli.",
    'description': '#### 📌 1. FINALITÀ DEL TOOL\n'
                   'Automatizza il **Confronto Dati Rinnovo** tra due anni, generando un prospetto riepilogativo con calcoli delle differenze, percentuali ed evidenziazione delle criticità.\n'
                   '\n'
                   '#### 🚀 2. COME UTILIZZARLO\n'
                   '1. **Input:** Carica il file Excel elaborato precedentemente da "Elaborazione SindTabe".\n'
                   '2. **Configura:** Verifica ed eventualmente modifica l\'Anno (B3) e il Titolo (B1) del confronto.\n'
                   '3. **Esegui:** Il tool storicizza i dati, inserisce i nuovi valori e calcola le variazioni in automatico.\n'
                   '4. **Scarica:** Ottieni il file pronto per l\'analisi con le formule attive.\n'
                   '\n'
                   '#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n'
                   '* **Storicizzazione:** Sposta i dati della Colonna B (Anno Corrente) nella Colonna C (Anno Precedente).\n'
                   '* **Aggiornamento:** Copia i valori della Colonna O del file di input nella Colonna B del template.\n'
                   '* **Formule Dinamiche:** Mantiene le formule per Differenza (Col D) e Percentuale (Col E), forzando il ricalcolo.\n'
                    '* **Evidenziazione Critica:** Se la colonna F contiene il codice "98765" e la differenza è negativa, il testo della cella in Colonna A diventa rosso.\n'
                   '* **Pulizia:** Rimuove i codici di controllo dalla Colonna F prima del salvataggio.\n'
                   '\n'
                   '#### 📂 4. RISULTATO FINALE\n'
                   'File Excel (.xlsx) formattato, contenente il confronto aggiornato con i nuovi dati, formule attive e righe critiche evidenziate.',
    'inputs': [
        {'key': 'format_warning', 'label': '📌 **Attenzione:** Caricare solo il file **TABULATO-ECONOMICO** già elaborato dal codice Python **"Elaborazione SindTabe"** così che i dati saranno correttamente aggiornati.', 'type': 'warning'},
        {'key': 'file_xlsx_input', 'label': 'Carica File Elaborato (Dati)', 'type': 'file_single', 'required': False}
    ],
    'params': [
        {'key': 'template_status',
         'label': '',
         'type': 'dynamic_info',
         'function': 'get_template_status_minimal',
         'section': 'Configurazione'},
        {'key': 'excel_template_name',
         'label': '📄 Configurazione Template Excel',
         'type': 'file_path_info',
         'default': 'C:/Users/simoncellil/Desktop/toolbox/tools/Amministrazione/FileProspetti_Formattati/Confronti/CONFRONTO_DATI_RINNOVO_2026_2025_OK.xlsx',
         'section': 'Configurazione',
         'help': 'ℹ️ Path del file template. Usa 📂 per cercarlo nel PC o 🔍 per aprire la cartella attuale.'},
        {'key': 'year_title_config',
         'label': 'Anno e Titolo',
         'type': 'dynamic_info',
         'function': 'render_year_and_title_config',
         'section': 'Configurazione'},
        {'key': 'output_config_ui',
         'label': 'Configurazione Output',
         'type': 'dynamic_info',
         'function': 'render_output_config_ui',
         'section': 'Impostazioni Output'}
    ]
}

# =========================
# HELPERS DINAMICI (UI)
# =========================
def get_template_status_minimal(values: Dict[str, Any]) -> str:
    template_val = values.get("excel_template_name")
    if not template_val:
        # Fallback default
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break
    
    if not template_val:
        return "❌ Nessun template"
    
    template_path = Path(template_val)
    if not template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / template_val
    
    if template_path.exists():
        st.error("⚠️ **ATTENZIONE:** Il template contiene **2 FOGLI**! Assicurati che vengano elaborati correttamente **ENTRAMBI**. Per la compilazione userai i dati del file **'effettivo_dopo_totale'** del mese di dicembre di quell'anno e copierai i valori o manualmente o in automatico.")
        st.info("💡 **Nota Annuale:** Ogni anno ricordati di aggiornare la **Colonna B** del template con i dati dell'**anno passato** prima di caricare i nuovi tramite codice.")
        return "✅ **Template trovato**"
    else:
        return "❌ **Template non trovato**"

def render_output_config_ui(values: Dict[str, Any]) -> str:
    """Renderizza l'input box per il nome del file di output e il pulsante di salvataggio."""
    
    # 1. Carica configurazione
    config = load_config()
    current_stored_name = config.get("output_filename", "CONFRONTO_DATI_RINNOVO")
    
    st.markdown("Definisci il nome del file che verrà generato (senza estensione .xlsx).")
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        new_name = st.text_input("Nome File Output", value=current_stored_name, key="output_filename_input")
        
    with col2:
        st.write("") # Spacer
        st.write("") 
        if st.button("💾 Salva Default"):
            if new_name.strip():
                if save_config({"output_filename": new_name.strip()}):
                    st.success("Salvato!")
                else:
                    st.error("Errore salvataggio")
            else:
                st.warning("Nome non valido")

    return "" # Ritorna vuoto perché renderizza direttamente

def render_year_and_title_config(values: Dict[str, Any]) -> str:
    """Renderizza i campi per modificare anno e titolo del confronto."""
    
    st.markdown("<span style='font-size: 20px;'><b>📅 Anno e Titolo</b></span>", unsafe_allow_html=True)
    st.markdown("_Questi valori verranno inseriti nel file di output generato._")

    # --- SEZIONE SIMULAZIONE ---
    with st.expander("🔍 Verifica logica dinamica (Test anni futuri)"):
        st.info("Cambiando l'anno qui sotto, i campi 'Anno' e 'Titolo' si aggiorneranno immediatamente per mostrare come funzionerà il tool in futuro.")
        sim_year = st.number_input("Simula Anno Corrente", value=datetime.now().year, step=1, key="sim_year_val")
        
    # Calcolo valori dinamici
    effective_year = sim_year
    prev_year = effective_year - 1
    dynamic_year = f"01/01/{effective_year}"
    dynamic_title = f"ASSOCIATI AL {dynamic_year} CONFRONTATI COL {prev_year}"
    
    # Sincronizzazione Session State: se i valori calcolati sono diversi da quelli in memoria 
    # e l'utente non li ha appena modificati a mano, li aggiorniamo.
    # In Streamlit, per forzare l'aggiornamento di un widget con 'key', dobbiamo scrivere in session_state.
    if "last_sim_year" not in st.session_state or st.session_state.last_sim_year != sim_year:
        st.session_state["year_input"] = dynamic_year
        st.session_state["title_input"] = dynamic_title
        st.session_state["last_sim_year"] = sim_year

    # --- RECUPERO VALORI DAL TEMPLATE (Per contesto) ---
    template_val = values.get("excel_template_name")
    if not template_val:
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break
    
    current_year_in_template = ""
    current_title_in_template = ""
    if template_val:
        template_path = Path(template_val)
        if not template_path.is_absolute():
            project_root = Path(__file__).resolve().parent.parent.parent
            template_path = project_root / template_val
        if template_path.exists():
            try:
                wb = load_workbook(template_path, data_only=True)
                ws = wb.active
                current_year_in_template = str(ws['B3'].value or "").strip()
                current_title_in_template = str(ws['B1'].value or "").strip()
                wb.close()
            except: pass

    # --- RENDER CAMPI ---
    st.text_input(
        "Anno (cella B3)",
        key="year_input",
        help=f"Valore calcolato: {dynamic_year}. Nel template fisico c'è: '{current_year_in_template}'"
    )
    
    st.text_area(
        "Titolo (cella B1 - celle unite)",
        key="title_input",
        height=80,
        help=f"Valore calcolato: {dynamic_title}. Nel template fisico c'è: '{current_title_in_template}'"
    )
    
    return ""

# =========================
# RUNNER
# =========================
def run(file_xlsx_input: Path, excel_template_name: str, out_dir: Path, **kwargs) -> List[Path]:
    # Recupera nome file output dalla config attuale (o usa input se passato, ma qui usiamo la config persistente come 'preferenza')
    # Tuttavia, l'utente potrebbe aver cambiato il testo nell'UI 'output_filename_input' senza salvare.
    # Streamlit rerun cycle: quando clicca 'Esegui' nel toolbox, i valori dei widget "custom" come l'input text 
    # potrebbero non essere passati automaticamente a 'run' se non sono definiti come 'params' standard del framework.
    # MA: Il framework toolbox probabilmente passa 'values' con tutti i widget?
    # Se il framework non passa i valori dei widget interni alle dynamic_info, dobbiamo fare affidamento sulla config salvata
    # OPPURE leggere session_state se possibile.
    # Per sicurezza, leggiamo session_state se presente, altrimenti config.
    
    output_name = st.session_state.get("output_filename_input")
    if not output_name:
        config = load_config()
        output_name = config.get("output_filename", "CONFRONTO_DATI_RINNOVO")
    
    # Assicuriamoci che non abbia estensione
    if output_name.lower().endswith(".xlsx"):
        output_name = output_name[:-5]

    # 1. Risoluzione Template Path
    template_path = Path(excel_template_name)
    if not template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / excel_template_name
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template non trovato al percorso: {template_path}")

    # 2. Caricamento Template
    wb = load_workbook(template_path)
    ws = wb.active

    # --- LOGICA DI SPOSTAMENTO (B -> C) ---
    # Spostare B3 -> C3 (anno vecchio)
    val_b3 = ws["B3"].value
    ws["C3"].value = val_b3
    
    # Centratura B3 e C3
    center_alignment = Alignment(horizontal="center", vertical="center")
    ws["B3"].alignment = center_alignment
    ws["C3"].alignment = center_alignment

    # Spostare Colonna B (Riga 5+) -> Colonna C (Riga 5+)
    # Prima svuotiamo C da riga 5 in poi
    max_row = ws.max_row
    for r in range(5, max_row + 200): # +200 per sicurezza
        ws.cell(row=r, column=3).value = None # Col C = 3

    # Copia B -> C
    for r in range(5, max_row + 1):
        val_b = ws.cell(row=r, column=2).value # Col B = 2
        
        # Scrivi in C solo se c'è valore (o anche None per sovrascrivere)
        ws.cell(row=r, column=3).value = val_b
        
        # Opzionale: pulire B? 
        # "spostare quelli che trovi in colonna B... nella C"
        # Solitamente spostare implica togliere dall'origine, ma dato che poi ci incolliamo sopra,
        # la pulizia è implicita nella sovrascrittura. Ma meglio pulire per le righe dove non avremo nuovi dati.
        ws.cell(row=r, column=2).value = None

    # --- LOGICA DI INSERIMENTO (Input Col O -> Template Col B) ---
    # Leggiamo input (data_only=True per i valori, importante!)
    try:
        wb_in = load_workbook(file_xlsx_input, data_only=True)
        ws_in = wb_in.active
    except Exception as e:
        raise RuntimeError(f"Errore nella lettura del file Excel di input: {e}")

    # Copiare Colonna O (15) da riga 7 in poi -> Colonna B (2) da riga 5 in poi
    # Iteriamo sul file di input
    max_row_in = ws_in.max_row
    
    # Indici
    row_in_start = 7
    col_in_idx = 15 # O
    
    row_out_start = 5
    col_out_idx = 2 # B
    
    # Quante righe copiare? Fino all'ultima dell'input
    # Usiamo un ciclo while o for
    current_row_out = row_out_start
    
    for r_in in range(row_in_start, max_row_in + 1):
        val_in = ws_in.cell(row=r_in, column=col_in_idx).value
        
        # Filtriamo solo valori numerici come richiesto? 
        # "copiare... tutti i valori numerici"
        # Se è vuoto o testo header? Copiamo tutto ciò che troviamo nella colonna O?
        # Il prompt dice "copiare ... tutti i valori numerici".
        # Se c'è testo, lo ignoriamo? O lo copiamo? 
        # Solitamente nei tabulati ci sono 0 o numeri. 
        # Copiamo il valore grezzo.
        
        ws.cell(row=current_row_out, column=col_out_idx).value = val_in
        current_row_out += 1

    # --- EVIDENZIAZIONE RIGHE (Logica Condizionale) ---
    # Colore carattere: Rosso (#FF0000)
    red_font = Font(color="FF0000", bold=True)
    target_code = "98765"
    
    # Iteriamo sulle righe popolate
    # Ricalcoliamo max_row perché abbiamo aggiunto dati
    current_max_row = ws.max_row
    
    for r in range(5, current_max_row + 1):
        # Leggi valori necessari
        val_b = ws.cell(row=r, column=2).value or 0
        val_c = ws.cell(row=r, column=3).value or 0
        val_f = str(ws.cell(row=r, column=6).value).strip() if ws.cell(row=r, column=6).value else ""
        
        # Calcola Differenza (D = B - C) in memoria
        try:
            diff = float(val_b) - float(val_c)
        except:
            diff = 0
            
        # Condizione: F == "98765" AND D < 0 (valore negativo)
        if val_f == target_code and diff < 0:
            # Applica colore carattere rosso solo alla colonna A (1)
            ws.cell(row=r, column=1).font = red_font
    
    # --- PULIZIA COLONNA F ---
    # Svuotiamo la colonna F (6)
    for r in range(5, current_max_row + 1):
        ws.cell(row=r, column=6).value = None

    # --- LOGICA SOMMA PIEMONTE NORD ---
    # Cerca "PIEMONTE NORD" in colonna A e somma le 3 righe sottostanti in colonna B
    for r in range(5, current_max_row + 1):
        cell_a_val = str(ws.cell(row=r, column=1).value or "").strip()
        if cell_a_val == "PIEMONTE NORD":
            somma_pn = 0
            for i in range(1, 4): # Righe +1, +2, +3
                row_val = ws.cell(row=r + i, column=2).value
                try:
                    somma_pn += float(row_val or 0)
                except (ValueError, TypeError):
                    pass
            # Scrivi il risultato fisso in B
            ws.cell(row=r, column=2).value = somma_pn
            break # Trovata la prima occorrenza, fermati

    # --- AGGIORNAMENTO ANNO E TITOLO ---
    # Recupera i valori dall'UI
    new_year = st.session_state.get("year_input", "01/01/2026")
    new_title = st.session_state.get("title_input", "ASSICURATI AL 01/01/2026 CONFRONTATI COL 2025")
    
    # Aggiorna B3 con il nuovo anno
    ws["B3"].value = new_year.strip()
    
    # Aggiorna B1 con il nuovo titolo
    ws["B1"].value = new_title.strip()
    
    # --- FORZA RICALCOLO FORMULE ---
    # Excel ricalcolerà tutte le formule quando il file viene aperto
    wb.calculation.fullCalcOnLoad = True
    
    # --- SALVATAGGIO ---
    out_filename = f"{output_name}.xlsx"
    out_path = out_dir / out_filename
    
    wb.save(out_path)
    
    # --- CONTROLLO VALORI A ZERO ---
    rows_with_zeros = []
    # Controlliamo da riga 5 fino all'ultima riga popolata (current_max_row)
    for r in range(5, current_max_row + 1):
        val_b = ws.cell(row=r, column=2).value
        val_c = ws.cell(row=r, column=3).value
        
        # Consideriamo zero se il valore è effettivamente 0 (non None o vuoto)
        # Se vuoi segnalare anche celle vuote come 0, rimuovi il controllo is not None
        if val_b == 0 or val_c == 0:
            rows_with_zeros.append(r)
            
    if rows_with_zeros:
        st.warning(f"⚠️ **Attenzione:** Trovati valori uguali a 0 nelle colonne B o C alle righe: {', '.join(map(str, rows_with_zeros))}")
    else:
        st.info("✅ **Check Dati:** Nessun valore pari a 0 trovato nelle colonne B e C.")

    st.success(f"File generato con successo: {out_filename}")
    
    return [out_path]

# =========================
# GUIDA DINAMICA
# =========================
def get_guide(values: Dict[str, Any]) -> str:
    return """
### 📘 Guida e Logica del Tool

Questo strumento automatizza la creazione del **Confronto Dati Rinnovo**, gestendo storicizzazione e calcoli complessi secondo gli standard CNA.

#### 📌 1. FINALITÀ DEL TOOL
Il tool risolve la necessità di confrontare i dati dell'anno in corso con quelli dell'anno precedente, automatizzando il travaso dei dati e l'evidenziazione delle anomalie senza errori manuali di copia-incolla.

#### 🚀 2. COME UTILIZZARLO
1. **Dati:** Carica il file Excel prodotto dal tool "Elaborazione SindTabe".
2. **Setup:** Modifica se necessario l'Anno (cella B3) e il Titolo (B1) tramite i campi forniti.
3. **Elaborazione:** Clicca sul pulsante di esecuzione; il sistema sposterà i dati attuali nella colonna storica e caricherà i nuovi.
4. **Verifica:** Controlla le righe evidenziate in rosso, che indicano diminuzioni critiche nei codici monitorati.

#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)
*   **Architettura:** Utilizza `openpyxl` per manipolare direttamente il template mantenendo stili e formule Excel originali.
*   **Storicizzazione:** Esegue lo shift dei valori: Colonna B → Colonna C e B3 → C3.
*   **Analisi Condizionale:** Esegue un calcolo in memoria `(B - C)`. Se il codice in Colonna F è `98765` e la differenza è negativa, cambia il colore del carattere in rosso (grassetto) nella cella di Colonna A.
*   **Ricalcolo:** Imposta i flag di Excel per forzare il ricalcolo delle formule (Colonne D/E) all'apertura del file.

#### 📂 4. RISULTATO FINALE
Viene prodotto un file `.xlsx` pronto per l'apertura, con i dati storicizzati, i nuovi dati caricati, la Colonna F svuotata dai codici di controllo e le evidenziazioni grafiche per le variazioni negative d'interesse.
"""
