from __future__ import annotations

import os
import re
import shutil
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================
# TOOLBOX CONFIGURATION
# =========================
# =========================
# TOOLBOX CONFIGURATION
# =========================
TOOL = {'id': 'elaborazione_sindtabe',
 'name': 'Elaborazione SindTabe',
 'region': 'Amministrazione',
 'email_reminder': True,
 'description': '#### 📌 1. FINALITÀ DEL TOOL\n'
                "Automatizza l'aggiornamento massivo del **Tabulato Economico** CNA Amministrazione, gestendo il "
                'passaggio dei dati da file Excel grezzi al prospetto contabile ufficiale.\n'
                '\n'
                '#### 🚀 2. COME UTILIZZARLO\n'
                '1. **Dati:** Carica il file Excel contenente le sedi e gli importi da caricare.\n'
                '2. **Template:** Il tool utilizza automaticamente il file `TABULATO_ECONOMICO_AMMINISTRAZIONE.xlsx` '
                'come base.\n'
                "3. **Review:** Controlla l'anteprima dei dati caricati per verificare che le colonne siano allineate "
                'correttamente.\n'
                '\n'
                '#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n'
                "* **Normalizzazione Sedi:** Applica la regola dello 'Zero iniziale' per i codici a 3 cifre e "
                'normalizza le sedi provinciali (XX00) mantenendo intatti i codici regionali (99XX).\n'
                '* **Matching & Sum:** Utilizza la Colonna A come chiave univoca per sommare gli importi dal file di '
                'input alle singole celle del Tabulato (Colonne C-P).\n'
                '* **Formule Dinamiche:** Inserisce automaticamente formule `=SUM()` nei righi totali regionali e '
                'nazionali (9999) basandosi sulla configurazione fornita.\n'
                '\n'
                '#### 📂 4. RISULTATO FINALE\n'
                "Tabulato Economico aggiornato, con colonne formattate (larghezza 13) e pronto per l'editing "
                'immediato.',
 'inputs': [{'key': 'format_warning', 'label': '📌 **Atteso:** file **SindTabe** in formato **Excel (.xlsx)**, non CSV. Se converti manualmente, assicurati che gli zeri iniziali non vengano rimossi!', 'type': 'warning'},
            {'key': 'file_xlsx_input', 'label': 'Carica File Excel (Dati)', 'type': 'file_single', 'required': False}],
 'params': [{'key': 'excel_preview',
             'label': 'Anteprima e Controllo Dati Excel',
             'type': 'dynamic_info',
             'function': 'get_excel_preview',
             'section': 'Anteprima File Caricato'},
            {'key': 'template_status',
             'label': '',
             'type': 'dynamic_info',
             'function': 'get_template_status_minimal',
             'section': 'Riferimento Template'},
            {'key': 'excel_template_name',
             'label': '📄 Configurazione Template Excel',
             'type': 'file_path_info',
             'default': 'C:/Users/simoncellil/Desktop/toolbox/tools/Amministrazione/FileProspetti_Formattati/SindTabe/TABULATO_ECONOMICO_AMMINISTRAZIONE.xlsx',
             'section': 'Riferimento Template',
             'help': 'ℹ️ Path del file template. Usa 📂 per cercarlo nel PC o 🔍 per aprire la cartella attuale.'},
            {'key': 'template_header_preview',
             'label': 'Anteprima Intestazione Template',
             'type': 'dynamic_info',
             'function': 'get_template_header_preview',
             'section': 'Riferimento Template'},
            {'key': 'mapping_sedi',
             'label': 'Mappatura Sedi Speciali',
             'type': 'textarea',
             'default': '9901=9901  # Servizi al Territorio 1\n'
                        '3802=3802  # AVEZZANO\n'
                        '2202=2202  # Vibo Valentia\n'
                        '2203=2203  # CROTONE\n'
                        '1301=1301  # IMOLA\n'
                        '3201=3201  # RIMINI\n'
                        '7006=7006  # CIVITAVECCHIA\n'
                        '4927=4927  # LODI\n'
                        '0690=0690  # FERMO\n'
                        '0901=0901  # BAT\n'
                        '1701=1701  # CARBONIA IGLESIS\n'
                        '7390=7390  # OLBIA\n'
                        '3001=3001  # PRATO\n'
                        '5290=5290  # VERBANIA C. OSSOLA\n'
                        '8901=8901  # BIELLA\n'
                        '3290=3200  # FORLI CESENA\n'
                        '4901=2400  # MONZA -> LARIO BRIANZA\n'
                        '4995=2400  # DESIO -> LARIO BRIANZA\n'
                        '4909=2400  # SEREGNO -> LARIO BRIANZA\n'
                        '9999=9999  # NAZIONALE\n'
                        '8700=2400  # LOMBARDIA N.OVEST',
             'section': 'Configurazione',
             'help': 'Codici sede che non seguono la regola dello zero finale.'},
            {'key': 'regional_totals_config',
             'label': 'Configurazione Codici Totali (Regionali/Nazionali)',
             'type': 'textarea',
             'default': '9999, 9920, 9919, 9918, 9917, 9916, 9915, 9914, 9913, 9912, 9911, 9910, 9909, 9908, 9907, '
                        '9906, 9905, 9904, 9903, 9902, 9935, 9937',
             'section': 'Configurazione',
             'help': 'Elenco dei codici sede in cui il tool deve inserire automaticamente le formule di somma.'}]}

# =========================
# HELPERS DINAMICI (UI)
# =========================
def detect_sep(file_obj) -> str:
    """Rileva automaticamente il separatore analizzando la prima riga."""
    try:
        file_obj.seek(0)
        first_line = file_obj.readline().decode('latin1')
        file_obj.seek(0)
        if ';' in first_line: return ';'
        if '\t' in first_line: return '\t'
        # Se ci sono molte virgole e pochi o zero punti e virgola, è virgola
        return ','
    except:
        return ','

def get_input_file_reminder(values: Dict[str, Any]) -> str:
    """Mostra un promemoria sul formato file atteso."""
    return "📌 **Atteso:** file **SindTabe** in formato **Excel (.xlsx)**, non CSV. Se converti manualmente, assicurati che gli zeri iniziali non vengano rimossi!"

def get_excel_preview(values: Dict[str, Any]) -> str:
    """Genera un'anteprima dell'Excel caricato per verifica utente."""
    # Cerchiamo l'input caricato nello session_state di Streamlit
    excel_key = None
    for k in st.session_state.keys():
        if k.endswith("file_xlsx_input") and k.startswith("up_"):
            excel_key = k
            break
            
    if not excel_key or st.session_state[excel_key] is None:
        return "ℹ️ **In attesa del caricamento:** Carica un file Excel per vedere l'anteprima e il controllo qualità."
    
    try:
        file_obj = st.session_state[excel_key]
        # Lettura per ottenere info complete
        file_obj.seek(0)
        # Leggiamo tutto per avere il conteggio righe preciso (specifica engine per evitare errori)
        df_info = pd.read_excel(file_obj, dtype=str, engine='openpyxl')
        num_cols = len(df_info.columns)
        num_rows = len(df_info)
        
        status_color = "🟢" if num_cols >= 3 else "🔴"
        
        msg = f"### 📊 Controllo Qualità Excel\n"
        msg += f"- **Righe totali rilevate:** `{num_rows}`\n"
        msg += f"- **Colonne rilevate:** {status_color} `{num_cols}`\n"
        
        if num_cols < 3:
            msg += f"\n⚠️ **ATTENZIONE:** Il file sembra avere poche colonne. Verifica che i dati inizino correttamente (Sede in 1ª colonna, Importi dalla 3ª in poi).\n"
        
        st.markdown(msg)
        st.markdown("**Dati trovati (Elenco completo):**")
        st.dataframe(df_info, use_container_width=True)
        return "" 
    except Exception as e:
        return f"❌ **Errore durante l'anteprima:** {e}"

def get_warning_message(values: Dict[str, Any]) -> str:
    # Cerchiamo di capire se il template esiste
    template_val = values.get("excel_template_name")
    
    # Fallback se non ancora nel dict values (per via dell'ordine di rendering o prima esecuzione)
    if not template_val:
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break
                
    if not template_val:
        return "❌ Nessun template configurato."
        
    template_path = Path(template_val)
    if not template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / template_val
    
    # Messaggio di avviso per controllare l'anno
    return "⚠️ **Promemoria:** Ricorda di controllare sempre l'**anno** nell'intestazione 'Titolo' prima di procedere!"

def get_template_status_minimal(values: Dict[str, Any]) -> str:
    """Mostra lo stato del template in modo minimal, per il titolo della sezione."""
    template_val = values.get("excel_template_name")
    
    if not template_val:
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break
    
    if not template_val:
        return "❌ Nessun template"
    
    template_path = Path(template_val)
    if not template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / template_val
    
    if template_path.exists():
        return "✅ Template trovato"
    else:
        return "❌ Template non trovato"

def get_template_header_preview(values: Dict[str, Any]) -> str:
    """Mostra le intestazioni del template in campi editabili con possibilità di salvataggio."""
    template_val = values.get("excel_template_name")
    
    # Fallback se non ancora nel dict values
    if not template_val:
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break
    
    if not template_val:
        return "❌ Nessun template configurato."
    
    template_path = Path(template_val)
    if not template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / template_val
    
    if not template_path.exists():
        return f"❌ Template non trovato: `{template_path.name}`"
    
    try:
        from openpyxl import load_workbook
        
        # Titolo (ancora più piccolo)
        st.markdown("#### 📋 Intestazione Template")
        # Promemoria anno
        st.warning("⚠️ Ricorda di controllare sempre l'**anno** nell'intestazione 'Titolo' prima di procedere!")
        
        # Legge le intestazioni dal file
        wb = load_workbook(template_path, data_only=False)
        ws = wb.active
        
        # Leggiamo le righe di intestazione (righe 1 e 5)
        header_rows_to_read = [1, 5]
        header_data = {}
        
        for row_num in header_rows_to_read:
            cells_values = []
            for col in range(1, 17):  # Colonne A-P
                val = ws.cell(row=row_num, column=col).value
                cells_values.append(str(val).strip() if val else "")
            
            # Verifica se la riga ha contenuto significativo
            if any(cells_values):
                if row_num == 1:
                    # Riga 1 (titolo): rimuovi celle vuote e unisci con " - "
                    non_empty = [c for c in cells_values if c]
                    header_data[row_num] = " - ".join(non_empty)
                else:
                    # Altre righe (mesi): rimuovi celle vuote consecutive e usa " | "
                    non_empty = [c for c in cells_values if c]
                    header_data[row_num] = " | ".join(non_empty)
        
        wb.close()
        
        if not header_data:
            return "⚠️ Nessuna intestazione rilevata nel template."
        
        # Crea campi editabili per ogni riga di intestazione
        edited_headers = {}
        for row_num, row_content in header_data.items():
            label = "Titolo" if row_num == 1 else "Colonne"
            edited_headers[row_num] = st.text_area(
                label,
                value=row_content,
                key=f"header_edit_{row_num}",
                height=68
            )
        
        # Pulsante per salvare le modifiche
        if st.button("💾 Salva Modifiche Intestazione", key="save_header_btn"):
            try:
                from openpyxl.cell.cell import MergedCell
                wb_write = load_workbook(template_path)
                ws_write = wb_write.active
                
                for row_num, edited_content in edited_headers.items():
                    # Splitta per | e scrivi nelle celle
                    cells = [c.strip() for c in edited_content.split("|")]
                    for col_idx, cell_value in enumerate(cells[:16], start=1):
                        cell = ws_write.cell(row=row_num, column=col_idx)
                        # Scrivi solo se NON è una cella unita secondaria
                        if not isinstance(cell, MergedCell):
                            cell.value = cell_value if cell_value else None
                
                wb_write.save(template_path)
                wb_write.close()
                st.success("✅ Intestazioni salvate con successo!")
            except PermissionError:
                st.error("❌ Impossibile salvare: il file è aperto in un altro programma.")
            except Exception as e:
                st.error(f"❌ Errore durante il salvataggio: {e}")
        
        return ""  # Ritorna stringa vuota perché abbiamo già renderizzato
        
    except Exception as e:
        return f"❌ Errore lettura template: {e}"


def get_guide(values: Dict[str, Any]) -> str:
    return """
### 📘 Guida e Logica del Tool

Il tool automatizza l'aggiornamento del **Tabulato Economico** partendo da un CSV di input.

#### 1. Logica di Normalizzazione Sedi (CSV)
Ogni riga del CSV viene elaborata per far corrispondere il codice sede a quello del tabulato excel:
*   **Formattazione:** Se un codice ha 3 cifre (es: `100`), viene aggiunto uno `0` iniziale (`0100`).
*   **Mappature Speciali:** Vengono applicate le conversioni definite (es: Monza `4901` -> Lario Brianza `2400`).
*   **Regola Provinciale:** Per tutti gli altri codici, il 3° e 4° carattere diventano `00` (es: `3456` -> `3400`).
*   **Codici Regionali:** I codici che iniziano con `99` (Nazionali/Regionali) non vengono modificati per consentire il caricamento dei totali.

#### 2. Processo di Aggiornamento
1.  **Caricamento Template:** Il tool cerca il file `TABULATO_ECONOMICO_AMMINISTRAZIONE.xlsx` nella cartella specifica.
2.  **Reset Dati:** Dalla **riga 7**, le colonne da **C a P** (importi) vengono azzerate in tutto il foglio.
3.  **Lettura CSV:** Il file CSV viene letto riga per riga con rilevamento automatico del separatore.
4.  **Matching:** Se la sede normalizzata del CSV è presente nella colonna A dell'Excel, gli importi del CSV (colonne C-P) vengono **sommati** alle celle corrispondenti dell'Excel.
5.  **Risultato:** Viene prodotto un file excel finale pronto per il download.
"""

# =========================
# LOGICA DI NORMALIZZAZIONE
# =========================
# Costanti di default (verranno sovrascritte se passate dal runner)
REGIONAL_CODES_DEFAULT = {
    "9999", "9920", "9919", "9918", "9917", "9916", "9915", "9914", 
    "9913", "9912", "9911", "9910", "9909", "9908", "9907", "9906", 
    "9905", "9904", "9903", "9902", "9935"
}

def parse_mapping_config(text: str) -> Dict[str, str]:
    mapping = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
        if "=" in line:
            k, v = line.split("=", 1)
            k_s, v_s = k.strip().zfill(4), v.strip().zfill(4)
            if k_s and v_s:
                mapping[k_s] = v_s
    return mapping

def normalize_sede(code_raw: Any, special_map: Dict[str, str], regional_set: set) -> str:
    # Pulizia
    s = str(code_raw).strip()
    # Se il codice è un numero letto da Excel (es 1300.0), lo puliamo
    if s.endswith('.0'):
        s = s[:-2]
    digits = re.sub(r"\D+", "", s)
    if not digits:
        return ""
    
    # Padding a 4 cifre
    code = digits.zfill(4)[-4:]
    
    # 1. Check in mappatura speciale
    if code in special_map:
        return special_map[code]
    
    # 2. Check codici regionali (Solo quelli protetti in configurazione)
    if code in regional_set:
        return code
    
    # 3. Regola di default: azzera byte 3 e 4
    return code[:2] + "00"

# =========================
# RUNNER
# =========================
def run(file_xlsx_input: Path, excel_template_name: str, mapping_sedi: str, regional_totals_config: str, out_dir: Path) -> List[Path]:
    # 1. Risoluzione Template Path
    template_path = Path(excel_template_name)
    if not template_path.is_absolute():
        # Calcola la root del progetto partendo dalla posizione di questo script (tools/Amministrazione/)
        project_root = Path(__file__).resolve().parent.parent.parent
        template_path = project_root / excel_template_name
    
    if not template_path.exists():
        raise FileNotFoundError(f"Template non trovato al percorso: {template_path}")

    # 2. Caricamento mappatura e codici regionali
    special_map = parse_mapping_config(mapping_sedi)
    
    # Parsing codici dai textbox
    master_list_from_ui = {s.strip() for s in regional_totals_config.replace("\n", ",").split(",") if s.strip()}
    
    # Il set per la PROTEZIONE (Punto 3) - Escludiamo 9937 come richiesto espressamente
    protection_set = {c for c in master_list_from_ui if c != "9937"}
    
    # Il set per le FORMULE (Punto 1) - Include tutto
    formula_set = master_list_from_ui
    
    # 3. Caricamento Excel Destinazione
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    ws = wb.active 
    
    # 4. Analisi e Azzeramento area dati (Scansioniamo fino alla riga 2000 per sicurezza)
    excel_sede_map: Dict[str, int] = {}
    
    # Usiamo un range ampio per coprire tutto il tabulato
    for r in range(7, 2001):
        v = ws.cell(row=r, column=1).value
        
        # Azzeriamo sempre le colonne degli importi (C..P) nell'area dati
        for c in range(3, 17): 
            ws.cell(row=r, column=c).value = None 

        if v is not None:
            # Puliamo il codice in colonna A dell'Excel 
            s_v = str(v).strip()
            if s_v.endswith('.0'): s_v = s_v[:-2]
            s_code = re.sub(r"\D+", "", s_v).zfill(4)[-4:]
            if s_code:
                excel_sede_map[s_code] = r
                # Se è una riga sede valida, impostiamo a 0.0 per iniziare le somme
                for c in range(3, 17):
                    ws.cell(row=r, column=c).value = 0.0

    # 5. Lettura File Excel di Input (usiamo openpyxl per essere certi di leggere TUTTO)
    try:
        wb_in = load_workbook(file_xlsx_input, data_only=True)
        ws_in = wb_in.active
    except Exception as e:
        raise RuntimeError(f"Errore nella lettura del file Excel di input: {e}")

    match_count = 0
    row_count = 0
    
    # Iteriamo su tutte le righe del file di input (saltando intestazioni se necessario)
    # Per sicurezza leggiamo tutto e filtriamo le righe vuote
    for r_in in range(1, ws_in.max_row + 1):
        raw_sede = ws_in.cell(row=r_in, column=1).value
        if raw_sede is None:
            continue
            
        row_count += 1
        norm_s = normalize_sede(raw_sede, special_map, protection_set)
        
        if norm_s in excel_sede_map:
            match_count += 1
            target_row = excel_sede_map[norm_s]
            
            # Somma colonne C..P (Excel in: 3..16 -> Excel out: 3..16)
            for c_idx in range(3, 17): 
                val_raw = ws_in.cell(row=r_in, column=c_idx).value
                if val_raw is None:
                    continue
                    
                try:
                    # Pulizia numerica migliorata
                    if isinstance(val_raw, (int, float)):
                        val_float = float(val_raw)
                    else:
                        val_str = str(val_raw).strip()
                        val_str = val_str.replace("€", "").replace(" ", "")
                        if "." in val_str and "," in val_str:
                            val_str = val_str.replace(".", "")
                        val_str = val_str.replace(",", ".")
                        val_float = float(val_str)
                except:
                    val_float = 0.0
                
                if val_float != 0:
                    current_val = ws.cell(row=target_row, column=c_idx).value or 0.0
                    ws.cell(row=target_row, column=c_idx).value = current_val + val_float
                    
                    # Formattazione Colonna P: Punto per migliaia, niente virgola/decimali
                    if c_idx == 16:
                        ws.cell(row=target_row, column=c_idx).number_format = '#,##0'

    # 6. Inserimento Formule per i Totali Regionali e Nazionale
    cols_to_sum = range(3, 17) # C..P
    
    # Lista ufficiale per il totale nazionale 9999 (Solo quelli definiti nella dashboard escludendo 9999 e 9936)
    MASTER_TOTAL_LIST = {c for c in formula_set if c not in {"9999", "9936"}}
    
    # Mappa le righe dei codici regionali (Usa la lista completa definita dall'utente)
    reg_rows_found = {c: r for c, r in excel_sede_map.items() if c in formula_set}
    
    # Per definire i blocchi di somma regionale, usiamo solo i codici in MASTER_TOTAL_LIST
    boundary_items = sorted([(r, c) for c, r in reg_rows_found.items() if c in MASTER_TOTAL_LIST])
    
    last_processed_row = 6 
    row_9936 = excel_sede_map.get("9936")
    
    for t_row, s_code in boundary_items:
        # Calcolo range di somma
        start_sum = last_processed_row + 1
        end_sum = t_row - 1
        
        if start_sum <= end_sum:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                formula = f"=SUM({col_let}{start_sum}:{col_let}{end_sum})"
                
                # Se 9936 è dentro questo range, lo sottraiamo per non contarlo due volte
                if row_9936 and start_sum <= row_9936 <= end_sum:
                    formula += f"-{col_let}{row_9936}"
                
                ws.cell(row=t_row, column=c_idx).value = formula
                # Formattazione Colonna P anche per i totali
                if c_idx == 16:
                    ws.cell(row=t_row, column=c_idx).number_format = '#,##0'
        
        last_processed_row = t_row

    # Logica NAZIONALE (9999): somma solo le righe della MASTER_TOTAL_LIST
    if "9999" in reg_rows_found:
        t_row_9999 = reg_rows_found["9999"]
        relevant_rows = [r for c, r in reg_rows_found.items() if c in MASTER_TOTAL_LIST]
        
        if relevant_rows:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                cells_to_add = [f"{col_let}{r}" for r in relevant_rows]
                ws.cell(row=t_row_9999, column=c_idx).value = f"=SUM({','.join(cells_to_add)})"
                # Formattazione Colonna P anche per il nazionale
                if c_idx == 16:
                    ws.cell(row=t_row_9999, column=c_idx).number_format = '#,##0'

    # --- RIGA 9936: Somma specifica di 9100 e 9000 ---
    if row_9936:
        row_9000 = excel_sede_map.get("9000")
        row_9100 = excel_sede_map.get("9100")
        if row_9000 and row_9100:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                ws.cell(row=row_9936, column=c_idx).value = f"={col_let}{row_9000}+{col_let}{row_9100}"

    # --- LARGHEZZA COLONNE ---
    # Impostiamo larghezza 13 solo per le colonne degli importi (da C (3) a P (16))
    for c_idx in range(3, 17): 
        col_let = get_column_letter(c_idx)
        ws.column_dimensions[col_let].width = 13

    # 7. Salvataggio Output con sblocco editing
    out_filename = f"AGGIORNATO_{Path(excel_template_name).name}"
    out_path = out_dir / out_filename
    
    # Impostiamo proprietà per evitare "Protected View" (abilitazione modifica)
    wb.properties.title = "Tabulato Elaborato SindTabe"
    wb.properties.creator = "Toolbox Amministrazione"
    
    wb.save(out_path)
    
    st.success(f"Fatto! Risultati pronti. (Processate {row_count} righe dal file di input con {match_count} corrispondenze trovate)")
    
    return [out_path]
