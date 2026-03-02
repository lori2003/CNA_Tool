from __future__ import annotations

import os
import io
import re
import sys
import json
import time
import datetime
import subprocess
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any

import streamlit as st

# ── Safe imports with auto-install ─────────────────────────────────────────
try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests"])
    import requests

try:
    from dotenv import load_dotenv
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "python-dotenv"])
    from dotenv import load_dotenv

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import load_workbook

# ── Load LocationIQ API key from .env ──────────────────────────────────────
_ENV_PATH = Path(__file__).parent / "API" / "x-Estrazione_Deleghe_estere" / ".env"
if _ENV_PATH.exists():
    load_dotenv(_ENV_PATH)
LOCATIONIQ_API_KEY = os.getenv("LOCATIONIQ_API_KEY", "")

# =========================
# TOOLBOX CONFIGURATION
# =========================
TOOL = {
    'id': 'estrazione_deleghe_estere',
    'name': 'Estrazione Deleghe Estere',
    'region': 'Amministrazione',
    'email_reminder': True,
    'description': '#### 📌 1. FINALITÀ DEL TOOL\n'
                   'Estrae i record relativi alle **deleghe estere** dal file **SindRinn** (TXT), '
                   'identifica automaticamente la **nazione estera** di ogni record, e produce un file '
                   'Excel di riepilogo completo con totali economici e anagrafica.\n'
                   '\n'
                   '#### 🚀 2. COME UTILIZZARLO\n'
                   '1. **File SindRinn (TXT):** Carica il file TXT dell\'anno attuale (tasto "Carica").\n'
                   '2. **File CODICI_NAZIONI (Excel):** Carica il file Excel con la tabella ufficiale '
                   'dei codici catastali esteri (colonna B = codice, colonna C = nazione).\n'
                   '3. **Clicca Esegui:** Il tool avvia l\'elaborazione. Per i record senza codice estero '
                   'parte il geocoding via **LocationIQ** (può durare alcuni minuti — vedrai un timer con il '
                   'tempo trascorso e stimato).\n'
                   '4. **Scarica il risultato:** Al termine, comparirà il riepilogo con i totali per nazione '
                   'e il pulsante per scaricare il file Excel.\n'
                   '\n'
                   '#### 🧠 3. LOGICA DI ELABORAZIONE\n'
                   '* **Filtro Estero:** Seleziona i record dove la posizione 316 o 317 = "1".\n'
                   '* **Parsing Economico:** Somma 12 campi economici (pos. 19→102), ciascuno di 7 caratteri, '
                   'dividendo per 100.\n'
                   '* **Classificazione Dinamica:** Ogni record viene classificato a runtime come:\n'
                   '  - **Z-code** (es. Z404) → lookup istantaneo dal file CODICI_NAZIONI\n'
                   '  - **NaN/vuoto** → geocoding LocationIQ per Indirizzo/Comune/Frazione\n'
                   '  - **Codice italiano** (es. A182) → geocoding LocationIQ\n'
                   '* **Geocoding LocationIQ:** Query progressive (indirizzo completo → solo comune → solo '
                   'frazione) con retry automatico in caso di errore.\n'
                   '* **Tutti i numeri** mostrati nella dashboard sono **dinamici**, calcolati sui dati effettivi del file.\n'
                   '\n'
                   '#### 📂 4. RISULTATO FINALE\n'
                   '* File Excel `Estrazione_Deleghe_Estere_ANNO.xlsx` con colonne: Totale, Indirizzo, Frazione, '
                   'Comune, Pr, Cap, Codice Catastale Nazione, **Nazione Estera**, Sede Gestione.\n'
                   '* Dashboard riepilogativa con totale per nazione e record non identificati.',
    'inputs': [
        {'key': 'format_warning', 'label': '📌 **Atteso:** file **SindRinn** in formato **TXT** del anno attuale.', 'type': 'warning'},
        {'key': 'file_txt_input', 'label': 'Carica File TXT (SindRinn)', 'type': 'file_single', 'required': True},
        {'key': 'codici_nazioni_reminder', 'label': '📌 **Promemoria:** file da controllare e ricaricare periodicamente, verificando sempre il corretto posizionamento delle colonne.', 'type': 'warning'},
        {'key': 'file_codici_nazioni', 'label': 'Carica File CODICI_NAZIONI (Excel) — ✅ Precaricato: CODICI_NAZIONI.xlsx', 'type': 'file_single', 'required': False}
    ],
    'params': [
        {
            'key': 'template_path',
            'label': '📄 Configurazione Template Excel',
            'type': 'file_path_info',
            'default': str(Path(__file__).parent / "FileProspetti_Formattati" / "Estero_Nazioni" / "AMM_RIPARTO_IMPORTI_PER NAZIONE_2025_OK.xlsx"),
            'help': 'Percorso del template Excel per il prospetto di riparto importi per nazione.'
        },
        {
            'key': 'titolo_b4',
            'label': 'Titolo prospetto (cella B4)',
            'type': 'text',
            'default': f"RIPARTO IMPORTI PER NAZIONI ANNO {datetime.date.today().year}",
            'help': 'Contenuto della cella B4 del prospetto generato. Modifica il testo se necessario prima di produrre il file.'
        },
        {
            'key': 'produce_estrazione',
            'label': 'Produci anche il file di estrazione dettagliato',
            'type': 'checkbox',
            'default': False,
            'help': (
                'Se spuntata, produce in aggiunta al prospetto per nazione anche il file '
                'Estrazione_Deleghe_Estere_ANNO.xlsx con il dettaglio completo di ogni record: '
                'Totale, Indirizzo, Frazione, Comune, Pr, Cap, Codice Catastale Nazione, '
                'Nazione Estera, Sede Gestione.'
            )
        },
        {
            'key': 'check_reminder',
            'section': 'Check',
            'label': 'Promemoria verifica',
            'type': 'dynamic_info',
            'function': '_check_reminder_text'
        }
    ]
}


# ─────────────────────────────────────────────────────────────────────────────
#  GEOCODING CACHE
# ─────────────────────────────────────────────────────────────────────────────
_CACHE_DIR = Path(__file__).parent / "API" / "x-Estrazione_Deleghe_estere"
_CACHE_FILE = _CACHE_DIR / "geocode_cache.json"

# ── Template path ──────────────────────────────────────────────────────────
_TEMPLATE_PATH = (
    Path(__file__).parent
    / "FileProspetti_Formattati"
    / "Estero_Nazioni"
    / "AMM_RIPARTO_IMPORTI_PER NAZIONE_2025_OK.xlsx"
)

_CODICI_NAZIONI_DEFAULT_PATH = (
    Path(__file__).parent
    / "FileProspetti_Formattati"
    / "Estero_Nazioni"
    / "CODICI_NAZIONI.xlsx"
)


# ── Valid nation codes (template column A) ─────────────────────────────────
_VALID_CODES_G1 = ['Z404', 'Z700', 'Z401', 'Z602', 'Z103', 'Z133']
_VALID_CODES_G2 = ['Z600', 'Z112', 'Z110', 'Z114', 'ZZZZ']
_ALL_VALID_CODES = _VALID_CODES_G1 + _VALID_CODES_G2

_CODE_TO_ROW: Dict[str, int] = {
    'Z404': 7,  'Z700': 8,  'Z401': 9,  'Z602': 10, 'Z103': 11, 'Z133': 12,
    '9999': 13,
    'Z600': 15, 'Z112': 16, 'Z110': 17, 'Z114': 18, 'ZZZZ': 19,
    '9998': 20,
    '9997': 22,
}


def _load_cache() -> Dict[str, str]:
    """Load geocoding cache from disk."""
    if _CACHE_FILE.exists():
        try:
            with open(_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_cache(cache: Dict[str, str]) -> None:
    """Persist geocoding cache to disk."""
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with open(_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def _clear_cache() -> None:
    """Delete geocoding cache file from disk."""
    if _CACHE_FILE.exists():
        try:
            _CACHE_FILE.unlink()
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
#  PROMEMORIA CHECK
# ─────────────────────────────────────────────────────────────────────────────

def _check_reminder_text(_values: dict) -> str:
    return (
        "📋 Verificare **i totali** delle deleghe e dell'importo sulla **riga generale**, "
        "con quelli del prospetto estero prodotto dal tool "
        "**\"Sedi CNA - Standard e Estero\"**."
    )\


# ─────────────────────────────────────────────────────────────────────────────
#  CODICE CATASTALE ENRICHMENT
# ─────────────────────────────────────────────────────────────────────────────

def _enrich_codice_catastale(df: pd.DataFrame) -> pd.DataFrame:
    """
    Per le righe dove 'Codice Catastale Nazione' NON è già uno Z-code,
    cerca un'altra riga con la stessa 'Nazione Estera' che abbia uno Z-code
    e ne copia il valore. Se non trovato, imposta 'ZZZZ'.
    """
    df = df.copy()
    mask_z = df['Codice Catastale Nazione'].astype(str).str.strip().str.startswith('Z')
    nazione_to_zcode: Dict[str, str] = {}
    for _, row in df[mask_z].iterrows():
        nazione = str(row['Nazione Estera']).strip()
        zcode = str(row['Codice Catastale Nazione']).strip()
        if nazione not in nazione_to_zcode:
            nazione_to_zcode[nazione] = zcode

    def _resolve(row):
        codice = str(row['Codice Catastale Nazione']).strip()
        if codice.startswith('Z'):
            return codice
        nazione = str(row['Nazione Estera']).strip()
        return nazione_to_zcode.get(nazione, 'ZZZZ')

    df['Codice Catastale Nazione'] = df.apply(_resolve, axis=1)
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  TEMPLATE FILLER
# ─────────────────────────────────────────────────────────────────────────────

def _fill_template(df: pd.DataFrame, out_dir: Path, year: int, tmpl_override: str = "", titolo_b4: str = "") -> Optional[Path]:
    """
    Carica il template Excel, lo compila con i dati aggregati per codice nazione
    e salva il risultato nella cartella di output.
    C = n. record, D = importo totale, E = 50% importo. Colonna A lasciata vuota.
    I codici non presenti nella lista valida confluiscono nel bucket 'ZZZZ'.
    """
    tpl = Path(tmpl_override) if tmpl_override else _TEMPLATE_PATH
    if not tpl.exists():
        st.warning(f"⚠️ Template non trovato: {tpl}")
        return None

    df = df.copy()
    df['_codice_norm'] = df['Codice Catastale Nazione'].astype(str).str.strip()
    df.loc[~df['_codice_norm'].isin(_ALL_VALID_CODES), '_codice_norm'] = 'ZZZZ'

    agg = (
        df.groupby('_codice_norm')
        .agg(count=('Totale', 'count'), totale=('Totale', 'sum'))
        .reset_index()
    )
    agg_dict: Dict[str, tuple] = {
        r['_codice_norm']: (int(r['count']), float(r['totale']))
        for _, r in agg.iterrows()
    }

    wb = load_workbook(tpl)
    ws = wb.active

    # Aggiorna cella B4 con il titolo (eventualmente modificato dall'utente)
    if titolo_b4:
        ws['B4'] = titolo_b4

    # Righe singole nazioni (colonna A lasciata intatta)
    for code in _ALL_VALID_CODES:
        row_num = _CODE_TO_ROW[code]
        count, totale = agg_dict.get(code, (0, 0.0))
        ws.cell(row=row_num, column=3, value=count)
        ws.cell(row=row_num, column=4, value=round(totale, 2))
        ws.cell(row=row_num, column=5, value=round(totale / 2, 2))

    # Totale parziale gruppo 1 (9999)
    g1_count = sum(agg_dict.get(c, (0, 0))[0] for c in _VALID_CODES_G1)
    g1_tot   = sum(agg_dict.get(c, (0, 0.0))[1] for c in _VALID_CODES_G1)
    r9999 = _CODE_TO_ROW['9999']
    ws.cell(row=r9999, column=3, value=g1_count)
    ws.cell(row=r9999, column=4, value=round(g1_tot, 2))
    ws.cell(row=r9999, column=5, value=round(g1_tot / 2, 2))

    # Totale parziale gruppo 2 (9998)
    g2_count = sum(agg_dict.get(c, (0, 0))[0] for c in _VALID_CODES_G2)
    g2_tot   = sum(agg_dict.get(c, (0, 0.0))[1] for c in _VALID_CODES_G2)
    r9998 = _CODE_TO_ROW['9998']
    ws.cell(row=r9998, column=3, value=g2_count)
    ws.cell(row=r9998, column=4, value=round(g2_tot, 2))
    ws.cell(row=r9998, column=5, value=round(g2_tot / 2, 2))

    # Totale generale (9997)
    tot_count = g1_count + g2_count
    tot_all   = g1_tot + g2_tot
    r9997 = _CODE_TO_ROW['9997']
    ws.cell(row=r9997, column=3, value=tot_count)
    ws.cell(row=r9997, column=4, value=round(tot_all, 2))
    ws.cell(row=r9997, column=5, value=round(tot_all / 2, 2))

    out_path = out_dir / f"AMM_RIPARTO_IMPORTI_PER_NAZIONE_{year}.xlsx"
    wb.save(out_path)
    return out_path


# ─────────────────────────────────────────────────────────────────────────────
#  TEMPLATE VIEWER
# ─────────────────────────────────────────────────────────────────────────────

def _render_template_section() -> None:
    """Mostra la sezione Template con i pulsanti Sfoglia e Apri."""
    st.markdown("---")
    st.markdown("### 📋 Template")
    exists = _TEMPLATE_PATH.exists()
    status = "✅" if exists else "❌ non trovato"
    col_info, col_b, col_o = st.columns([4, 1, 1])
    with col_info:
        st.markdown(f"{status} &nbsp; `{_TEMPLATE_PATH.name}`")
    with col_b:
        if st.button("📂 Sfoglia", key="btn_tmpl_browse", use_container_width=True):
            try:
                subprocess.Popen(f'explorer /select,"{_TEMPLATE_PATH}"', shell=True)
            except Exception as e:
                st.error(str(e))
    with col_o:
        if st.button("📄 Apri", key="btn_tmpl_open", use_container_width=True):
            try:
                os.startfile(str(_TEMPLATE_PATH))
            except Exception as e:
                st.error(str(e))
    st.markdown("---")


# ─────────────────────────────────────────────────────────────────────────────
#  LOCATIONIQ GEOCODING
# ─────────────────────────────────────────────────────────────────────────────

def _geocode_single_query(query: str, api_key: str, max_retries: int = 3) -> Optional[str]:
    """
    Geocode a single query via LocationIQ with retry + exponential backoff.
    Returns the country name in UPPERCASE or None.
    """
    url = "https://us1.locationiq.com/v1/search.php"
    params = {
        'key': api_key,
        'q': query,
        'format': 'json',
        'accept-language': 'it',
        'limit': 1
    }

    for attempt in range(max_retries):
        try:
            response = requests.get(url, params=params, timeout=10)

            if response.status_code == 200:
                data = response.json()
                if data and isinstance(data, list) and len(data) > 0:
                    display_name = data[0].get('display_name', '')
                    # The last part of display_name is typically the country
                    parts = [p.strip() for p in display_name.split(',')]
                    if parts:
                        country = parts[-1].upper()
                        return country
                return None

            elif response.status_code == 404:
                # No results found
                return None

            elif response.status_code == 429:
                # Rate limit hit - wait longer
                wait = 2 ** (attempt + 1)
                time.sleep(wait)

            elif response.status_code == 401:
                # Invalid API key
                return None

            else:
                # Other errors - retry
                wait = 2 ** (attempt + 1)
                time.sleep(wait)

        except requests.exceptions.Timeout:
            wait = 2 ** (attempt + 1)
            time.sleep(wait)
        except Exception:
            return None

    return None


def _geocode_country(indirizzo: str, comune: str, frazione: str, api_key: str, cache: Dict[str, str]) -> Optional[str]:
    """
    Try to resolve the country name from address fields using LocationIQ.
    Uses progressive query combinations (most specific → least specific).
    Checks cache first to avoid redundant API calls.
    Returns the country name in UPPERCASE or None if not found.
    """
    # Build query combinations from most to least specific
    queries = []

    # 1. Full: Indirizzo, Comune, Frazione
    parts_full = [p for p in [indirizzo, comune, frazione] if p]
    if parts_full:
        queries.append(", ".join(parts_full))

    # 2. Comune + Frazione
    parts_cf = [p for p in [comune, frazione] if p]
    if parts_cf and parts_cf != parts_full:
        queries.append(", ".join(parts_cf))

    # 3. Indirizzo + Comune (without Frazione)
    parts_ic = [p for p in [indirizzo, comune] if p]
    if parts_ic and len(parts_ic) == 2 and parts_ic != parts_full:
        queries.append(", ".join(parts_ic))

    # 4. Comune only
    if comune:
        queries.append(comune)

    # 5. Frazione only
    if frazione and frazione != comune:
        queries.append(frazione)

    for query in queries:
        # Check cache first
        cache_key = query.upper().strip()
        if cache_key in cache:
            return cache[cache_key]

        # Query LocationIQ
        country = _geocode_single_query(query, api_key)
        if country:
            cache[cache_key] = country  # save to cache
            return country

        # Rate limit: ~2 requests/sec for free tier
        time.sleep(0.5)

    return None


# ─────────────────────────────────────────────────────────────────────────────
#  RECORD PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_sindrinn_line(line: str) -> Optional[Dict[str, Any]]:
    """Estrae i dati da una singola riga del file SindRinn se soddisfa i criteri estero."""
    if len(line) < 321:
        return None

    # Filtro Estero: Pos 316 (index 315) o Pos 317 (index 316) == '1'
    if line[315] != '1' and line[316] != '1':
        return None

    data = {}

    # 1. Calcolo Totale (12 campi da 7 caratteri, partendo da colonna 19)
    totale_importo = 0.0
    for i in range(12):
        start = 18 + (i * 7)
        end = start + 7
        try:
            val_str = line[start:end].strip()
            if val_str:
                totale_importo += float(val_str) / 100.0
        except ValueError:
            pass

    data['Totale'] = totale_importo

    # 2. Estrazione Anagrafica (Posizioni 1-based)
    data['Indirizzo'] = line[160:212].strip()
    data['Frazione'] = line[212:247].strip()
    data['Comune'] = line[247:277].strip()
    data['Pr'] = line[277:280].strip()
    data['Cap'] = line[280:289].strip()
    data['Codice Catastale Nazione'] = line[311:315].strip()
    data['Sede Gestione'] = line[317:321].strip()

    return data


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(file_txt_input: Path, file_codici_nazioni: Path, out_dir: Path, template_path: str = "", titolo_b4: str = "", produce_estrazione: bool = False) -> List[Path]:
    """Funzione principale eseguita dalla Toolbox."""

    # ── Check API key ────────────────────────────────────────────────────
    if not LOCATIONIQ_API_KEY:
        st.error("⚠️ API key LocationIQ non trovata! Verifica il file .env in API/x-Estrazione_Deleghe_estere/")
        return []

    # ── Clear old cache on every execution ────────────────────────────────
    _clear_cache()

    results = []

    # ── 1. Read TXT ──────────────────────────────────────────────────────
    try:
        with open(file_txt_input, 'r', encoding='latin1') as f:
            lines = f.readlines()
    except Exception as e:
        st.error(f"Errore nella lettura del file TXT: {e}")
        return []

    for line in lines:
        parsed = parse_sindrinn_line(line)
        if parsed:
            results.append(parsed)

    if not results:
        st.warning("Nessun record trovato con i criteri specificati.")
        return []

    df = pd.DataFrame(results)
    tot_records = len(df)

    # ── 2. Load CODICI_NAZIONI lookup ────────────────────────────────────
    if file_codici_nazioni is None:
        if _CODICI_NAZIONI_DEFAULT_PATH.exists():
            file_codici_nazioni = _CODICI_NAZIONI_DEFAULT_PATH
            pass  # precaricato silenziosamente
        else:
            st.error(f"File CODICI_NAZIONI non trovato: {_CODICI_NAZIONI_DEFAULT_PATH}")
            return []
    try:
        df_nazioni = pd.read_excel(file_codici_nazioni)
        col_codice = df_nazioni.columns[1]   # Column B = codice_belfiore
        col_nome = df_nazioni.columns[2]     # Column C = denominazione_nazione
        lookup_nazioni = dict(zip(
            df_nazioni[col_codice].astype(str).str.strip(),
            df_nazioni[col_nome].astype(str).str.strip()
        ))
    except Exception as e:
        st.error(f"Errore nella lettura del file CODICI_NAZIONI: {e}")
        return []

    # ── 3. Classify records dynamically ──────────────────────────────────
    def _classify(codice):
        if pd.isna(codice) or str(codice).strip() == '':
            return 'nan'
        codice_str = str(codice).strip()
        if codice_str.startswith('Z'):
            return 'z_code'
        return 'italian'

    df['_tipo_record'] = df['Codice Catastale Nazione'].apply(_classify)

    n_zcode = (df['_tipo_record'] == 'z_code').sum()
    n_nan = (df['_tipo_record'] == 'nan').sum()
    n_italian = (df['_tipo_record'] == 'italian').sum()

    st.info(
        f"📊 Rilevati **{tot_records}** record totali:\n"
        f"- ✅ **{n_zcode}** con codice catastale estero (Z-code) → lookup da CODICI_NAZIONI\n"
        f"- ❓ **{n_nan}** senza codice catastale → geocoding LocationIQ\n"
        f"- 🏠 **{n_italian}** con codice catastale italiano → geocoding LocationIQ"
    )

    # ── 4. Resolve Z-code records (instant) ──────────────────────────────
    mask_z = df['_tipo_record'] == 'z_code'
    df.loc[mask_z, 'Nazione Estera'] = df.loc[mask_z, 'Codice Catastale Nazione'].apply(
        lambda c: lookup_nazioni.get(str(c).strip(), "NON IDENTIFICATO")
    )

    # ── 5. Geocode NaN + Italian records (by Comune dedup) ────────────────
    mask_geocode = df['_tipo_record'].isin(['nan', 'italian'])
    api_rows = df[mask_geocode].index.tolist()
    n_api_rows = len(api_rows)

    if n_api_rows > 0:
        # Build unique Comune lookup (dedup key)
        df_api = df.loc[api_rows].copy()
        df_api['_comune_key'] = df_api['Comune'].fillna('').str.strip().str.upper()
        unique_comuni = df_api['_comune_key'].unique().tolist()
        n_unique = len(unique_comuni)

        # Geocode each unique Comune once
        comune_to_country: Dict[str, Optional[str]] = {}
        progress_bar = st.progress(0, text="Geocoding comuni unici...")
        status_text = st.empty()
        resolved_count = 0
        failed_count = 0
        start_time = time.time()

        for i, comune in enumerate(unique_comuni):
            if not comune:
                comune_to_country[comune] = None
                failed_count += 1
            else:
                # Primary: Comune alone
                country = _geocode_single_query(comune, LOCATIONIQ_API_KEY)

                # Fallback 1: Indirizzo + Comune
                if country is None:
                    sample = df_api[df_api['_comune_key'] == comune].iloc[0]
                    indirizzo = str(sample['Indirizzo']).strip() if pd.notna(sample['Indirizzo']) else ''
                    frazione = str(sample['Frazione']).strip() if pd.notna(sample['Frazione']) else ''
                    if indirizzo:
                        country = _geocode_single_query(f"{indirizzo}, {comune}", LOCATIONIQ_API_KEY)

                    # Fallback 2: Comune + Frazione
                    if country is None and frazione:
                        country = _geocode_single_query(f"{comune}, {frazione}", LOCATIONIQ_API_KEY)

                    # Fallback 3: Indirizzo + Comune + Frazione
                    if country is None and indirizzo and frazione:
                        country = _geocode_single_query(f"{indirizzo}, {comune}, {frazione}", LOCATIONIQ_API_KEY)

                comune_to_country[comune] = country
                if country:
                    resolved_count += 1
                else:
                    failed_count += 1

            # Timer
            elapsed = time.time() - start_time
            done = i + 1
            remaining = n_unique - done
            avg = elapsed / done if done > 0 else 1
            eta = remaining * avg
            el_m, el_s = divmod(int(elapsed), 60)
            et_m, et_s = divmod(int(eta), 60)

            progress_bar.progress(done / n_unique, text=f"Geocoding {done}/{n_unique} comuni")
            status_text.markdown(
                f"⏱️ **Trascorso:** {el_m:02d}:{el_s:02d} · "
                f"**Rimanente:** ~{et_m:02d}:{et_s:02d} · "
                f"✅ {resolved_count} · ❌ {failed_count}"
            )

        progress_bar.empty()
        status_text.empty()

        # Map results back to all rows
        failed_count_records = 0
        for idx in api_rows:
            comune_key = str(df.at[idx, 'Comune']).strip().upper() if pd.notna(df.at[idx, 'Comune']) else ''
            country = comune_to_country.get(comune_key)
            if country:
                df.at[idx, 'Nazione Estera'] = country
            else:
                df.at[idx, 'Nazione Estera'] = "NON IDENTIFICATO"
                failed_count_records += 1

        total_elapsed = time.time() - start_time
        tot_m, tot_s = divmod(int(total_elapsed), 60)

        # Calcola anche gli Z-code non risolti per avere il conteggio esatto nel box finale
        zcode_unresolved = (df['_tipo_record'] == 'z_code') & (df['Nazione Estera'] == 'NON IDENTIFICATO')
        total_failed_records = failed_count_records + zcode_unresolved.sum()
        total_resolved_records = (df['Nazione Estera'] != 'NON IDENTIFICATO').sum()

        st.success(
            f"🌐 Geocoding completato in **{tot_m} min {tot_s} sec**:\n"
            f"- 🔍 **{tot_records}** record analizzati\n"
            f"- ✅ **{total_resolved_records}** risolti\n"
            f"- ❌ **{total_failed_records}** non identificati"
        )

    # ── 5c. Enrich Codice Catastale Nazione → Z-code unificato ───────────
    df = _enrich_codice_catastale(df)

    # ── 6. Build final output ────────────────────────────────────────────
    df.drop(columns=['_tipo_record'], inplace=True)

    cols_order = ['Totale', 'Indirizzo', 'Frazione', 'Comune', 'Pr', 'Cap',
                  'Codice Catastale Nazione', 'Nazione Estera', 'Sede Gestione']
    df = df[cols_order]

    current_year = datetime.date.today().year

    # ── 6b. Fill template with aggregated data per nation ─────────────
    tmpl_out = _fill_template(df, out_dir, current_year, template_path, titolo_b4)
    if tmpl_out:
        st.success(f"📊 Prospetto per nazione compilato: **{tmpl_out.name}**")

    # ── 6c. Estrazione dettagliata (solo se checkbox spuntata) ─────────
    out_path = None
    if produce_estrazione:
        out_filename = f"Estrazione_Deleghe_Estere_{current_year}.xlsx"
        out_path = out_dir / out_filename
        df.to_excel(out_path, index=False)
        st.info(f"📄 File di estrazione dettagliato prodotto: **{out_filename}**")

    # ── 7. Compute summary stats and persist in session_state ─────────
    n_empty = (df['Nazione Estera'] == 'NON IDENTIFICATO').sum()
    country_counts = (
        df['Nazione Estera']
        .value_counts()
        .reset_index()
    )
    country_counts.columns = ['Nazione', 'N. Record']

    # Store in session_state so the dashboard survives reruns/downloads
    st.session_state['deleghe_estere_dashboard'] = {
        'tot_records': tot_records,
        'n_empty': int(n_empty),
        'country_table': country_counts.to_dict('records'),
    }

    # ── 8. Display final dashboard ────────────────────────────────────
    _render_dashboard()

    out_files = []
    if tmpl_out:
        out_files.append(tmpl_out)
    if out_path:
        out_files.append(out_path)
    return out_files


def _render_dashboard() -> None:
    """Display the summary dashboard from session_state."""
    data = st.session_state.get('deleghe_estere_dashboard')
    if not data:
        return

    tot = data['tot_records']
    n_empty = data['n_empty']
    table = data['country_table']

    st.success(f"✅ Elaborazione completata! Estratti **{tot}** record.")

    # Country totals table
    st.markdown("### 🌍 Totale per Nazione")
    df_table = pd.DataFrame(table)
    st.dataframe(df_table, use_container_width=True, hide_index=True)
