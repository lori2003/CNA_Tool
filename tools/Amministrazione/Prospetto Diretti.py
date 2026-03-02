import streamlit as st
import pandas as pd
import os
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

TOOL = {
    'id': 'prospetto_diretti',
    'name': 'Prospetto Diretti',
    'region': 'Amministrazione',
    'email_reminder': True,
    'description': 'Genera il prospetto dei diretti aggregato per Regione, partendo da un file Excel con Sede, Provincia e Quota.',
    'version': '1.0.1',
    'author': 'Assistant',
    'inputs': [
        {
            'key': 'info_msg', 
            'label': '⚠️ **PROMEMORIA FORMATO FILE**\n\n'
                     'Il file deve avere:\n'
                     '- **Colonna A**: Sede CNA\n'
                     '- **Colonna B**: Sigla Provincia\n'
                     '- **Colonna C**: Quota\n\n'
                     '⚠️ **Attenzione**: se il file originale è in formato **.slk**, deve essere convertito in **.xlsx** prima del caricamento.\n\n'
                     '📅 **Nota Anno**: se fatto nel 2026, l\'anno dei diretti sarà per il 2025 e così via.',
            'type': 'info'
        },
        {
            'key': 'logic_note', 
            'label': 'ℹ️ **Nota Logiche Applicate**\n\n'
                     '- **Bolzano/Bozen** → **Bolzano** (pulizia automatica).\n'
                     '- **Monza, Varese, Lecco, Como** → Aggregati in **LOMBARDIA N.OVEST**.\n'
                     '- **Verona, Vicenza** → Aggregati in **VENETO NORD** (in coda al Veneto).',
            'type': 'info'
        },
        {'key': 'file_input', 'label': 'Carica File Excel (Dati)', 'type': 'file_single', 'accept': ['.xlsx'], 'required': True},
    ],
    'params': [

        {
            'key': 'header_preview',
            'type': 'dynamic_info',
            'function': 'get_template_header_preview',
            'section': 'Configurazione Template📝'
        }
    ]
}

def get_template_header_preview(values):
    """Permette di visualizzare e modificare l'intestazione del template (Cella A1)."""
    try:
        template_path = os.path.join("tools", "Amministrazione", "FileProspetti_Formattati", "Template-Diretti", "Template-Diretti.xlsx")
        
        if not os.path.exists(template_path):
            return f"❌ Template non trovato: `{template_path}`"

        # Leggi l'intestazione attuale
        wb = load_workbook(template_path)
        ws = wb.active
        current_header = str(ws['A1'].value) if ws['A1'].value else ""
        wb.close()

        # UI Editabile
        st.info("ℹ️ L'anno verrà aggiornato automaticamente durante l'esecuzione (Anno Corrente - 1).")
        new_header = st.text_input("Intestazione Report (Base)", value=current_header, key="custom_header_edit")

        if st.button("💾 Salva Intestazione nel Template", key="save_header_btn"):
            try:
                wb_write = load_workbook(template_path)
                ws_write = wb_write.active
                ws_write['A1'] = new_header
                wb_write.save(template_path)
                wb_write.close()
                st.success(f"✅ Intestazione aggiornata: **{new_header}**")
                # Ricarica per sicurezza visuale
                st.rerun()
            except Exception as e:
                st.error(f"Errore salvataggio: {e}")

        return "" # La UI è renderizzata direttamente

    except Exception as e:
        return f"Errore lettura header: {e}"

@st.cache_data(show_spinner="Elaborazione in corso...")
def process_excel_data(file_content, report_year):
    """Elabora i dati e ritorna il dataframe processato e le metriche."""
    try:
        # Leggi Excel
        df = pd.read_excel(file_content, header=None) 
            
        # Rinomina colonne per facilità (A=0, B=1, C=2)
        if len(df.columns) < 3:
            return None, "Il file deve avere almeno 3 colonne!", {}

        # Assumiamo che la prima riga possa essere un'intestazione o dati.
        first_cell = str(df.iloc[0, 0]).lower()
        if "sede" in first_cell:
            df = df.iloc[1:].reset_index(drop=True)
            
        # Rinomina colonne critiche
        df = df.rename(columns={0: 'Sede_Codice', 1: 'Provincia_Originale', 2: 'Quota'})
        
        # Conteggi Iniziali
        total_rows_input = len(df)
        
        # Filtro righe vuote o spurie (Sede mancante)
        dropped_rows_count = df['Sede_Codice'].isna().sum()
        df = df.dropna(subset=['Sede_Codice'])
        valid_rows_count = len(df)
        
        # Assicurati che Quota sia numerico
        df['Quota'] = pd.to_numeric(df['Quota'], errors='coerce').fillna(0)

        # 2. Estrazione Sigla Provincia dalla Sede (Byte 3 e 4 -> indici 2:4)
        df['Extracted_Prov'] = df['Sede_Codice'].astype(str).str[2:4].str.upper()

        # 3. Caricamento Lookup Regioni + Nome Completo Provincia
        lookup_path = os.path.join("tools", "Amministrazione", "FileProspetti_Formattati", "Elenco_Comuni_Regione", "Elenco-comuni-italiani.xlsx")
        
        if not os.path.exists(lookup_path):
             return None, f"File di lookup non trovato: {lookup_path}", {}
            
        # Leggi colonna O (Sigla auto, indice 14), K (Regione, indice 10), L (Nome Provincia, indice 11)
        df_lookup = pd.read_excel(lookup_path, usecols=[10, 11, 14])
        
        if len(df_lookup.columns) == 3:
             df_lookup.columns = ['Regione', 'Nome_Provincia', 'Sigla_Auto']
        else:
             return None, "Errore lettura colonne lookup.", {}

        # Pulizia Lookup
        df_lookup = df_lookup.dropna(subset=['Sigla_Auto'])
        # Mapping
        prov_to_region = df_lookup.drop_duplicates(subset=['Sigla_Auto']).set_index('Sigla_Auto')['Regione'].to_dict()
        prov_to_name = df_lookup.drop_duplicates(subset=['Sigla_Auto']).set_index('Sigla_Auto')['Nome_Provincia'].to_dict()

        # 4. Arricchimento DataFrame
        df['Regione'] = df['Extracted_Prov'].map(prov_to_region)
        df['Nome_Provincia'] = df['Extracted_Prov'].map(prov_to_name)
        
        # Gestione non trovati
        missing_mask = df['Nome_Provincia'].isna()
        missing_provs = []
        if missing_mask.any():
            missing_provs = df[missing_mask]['Extracted_Prov'].unique().tolist()
            df.loc[missing_mask, 'Regione'] = 'NON IDENTIFICATO'
            df.loc[missing_mask, 'Nome_Provincia'] = df.loc[missing_mask, 'Sede_Codice'] # Fallback

        # --- CUSTOM LOGIC: PULIZIA E RAGGRUPPAMENTI ---
        
        # 1. Pulizia Nomi (Rimuovi tutto dopo "/")
        # Es: "Bolzano/Bozen" -> "Bolzano"
        df['Nome_Provincia'] = df['Nome_Provincia'].astype(str).apply(lambda x: x.split('/')[0].strip()).str.upper()
        df['Regione'] = df['Regione'].astype(str).apply(lambda x: x.split('/')[0].strip()).str.upper()

        # 2. Aggregazione LOMBARDIA N.OVEST
        # Province da accorpare: Monza e della Brianza, Varese, Lecco, Como
        lombardia_novest_provs = ['MONZA E DELLA BRIANZA', 'VARESE', 'LECCO', 'COMO'] # Uppercase ora
        
        # Identifica le righe target
        mask_lomb_novest = df['Nome_Provincia'].isin(lombardia_novest_provs)
        
        # Se ci sono dati per queste province, li aggreghiamo
        if mask_lomb_novest.any():
            # Cambiamo i nomi provincia in "LOMBARDIA N.OVEST"
            # Manteniamo la Regione "LOMBARDIA" (o quella originale che dovrebbe essere Lombardia)
            df.loc[mask_lomb_novest, 'Nome_Provincia'] = 'LOMBARDIA N.OVEST'
            # Nota: La regione rimane quella mappata (Lombardia), perfetto.

        # 3. Marcatura VENETO NORD (Senza aggregare)
        # Province target: Verona, Vicenza -> Devono apparire in fondo al Veneto sotto "VENETO NORD"
        veneto_nord_provs = ['VERONA', 'VICENZA'] # Uppercase
        # Non modifichiamo il Nome_Provincia qui, lasciamo i dettagli.
        # La logica di "Raggruppamento Visivo" avverrà in fase di scrittura Excel e Ordinamento.

        # 5. Aggregazione Finale
        grouped = df.groupby(['Regione', 'Nome_Provincia']).agg(
            Associati=('Sede_Codice', 'count'),
            Importo=('Quota', 'sum')
        ).reset_index()

        # Ordinamento Personalizzato
        # Vogliamo che "VENETO NORD" sia in fondo alla lista del Veneto.
        # Creiamo una chiave di ordinamento temporanea.
        grouped['SortKey'] = grouped['Nome_Provincia']
        # Prefisso ZZZ per mandarlo in fondo
        grouped.loc[grouped['Nome_Provincia'].isin(veneto_nord_provs), 'SortKey'] = 'ZZZ_' + grouped['Nome_Provincia']
        
        grouped = grouped.sort_values(by=['Regione', 'SortKey'])
        grouped = grouped.drop(columns=['SortKey'])
        
        metrics = {
            'total_rows_input': total_rows_input,
            'valid_rows_count': valid_rows_count,
            'dropped_rows_count': dropped_rows_count,
            'missing_provs': missing_provs
        }
        
        return grouped, None, metrics

    except Exception as e:
        return None, str(e), {}

def get_ui_results():
    if "prospetto_diretti_dashboard_data" not in st.session_state:
        return
        
    val_data = st.session_state["prospetto_diretti_dashboard_data"]
    metrics = val_data.get("metrics", {})
    grouped = val_data.get("grouped")
    
    if grouped is None:
        return

    # Calcolo Anno Referenza (lo stesso usato in run)
    current_year = datetime.datetime.now().year
    report_year = current_year - 1
    
    st.write(f"### 🗓️ Generazione Prospetto Diretti: **{report_year}**")

    # Estrai metriche
    total_rows_input = metrics.get('total_rows_input', 0)
    valid_rows_count = metrics.get('valid_rows_count', 0)
    dropped_rows_count = metrics.get('dropped_rows_count', 0)
    
    # Totali calcolati
    calc_total_associati = grouped['Associati'].sum()
    calc_total_importo = grouped['Importo'].sum()

    # 📊 DASHBOARD DI CONTROLLO UNIFICATA
    # Header rimosso su richiesta utente (ridondante)
    # st.markdown("### 📊 Dashboard di Controllo Elaborazione")
    st.markdown("---")
    
    # Riga 1: Metriche Input
    c1, c2, c3 = st.columns(3)
    c1.metric("Record Totali Input", total_rows_input)
    c2.metric("Record Validi", valid_rows_count)
    c3.metric("Record Scartati (No Sede)", dropped_rows_count, delta_color="inverse")
    
    if dropped_rows_count > 0:
        st.warning(f"⚠️ Attenzione: sono stati scartati **{dropped_rows_count}** record perché privi di Codice Sede.")

    # Riga 2: Metriche Calcolate
    st.markdown("#### ✅ Verifica Totali Calcolati")
    c1_calc, c2_calc = st.columns(2)
    c1_calc.metric("Totale Associati Calcolato", calc_total_associati)
    c2_calc.metric("Totale Importo Calcolato", f"€ {calc_total_importo:,.2f}")
    
    st.markdown("---")
    
    # Tabelle di Riepilogo
    col_reg, col_prov = st.columns(2)
    
    with col_reg:
        st.markdown("##### 📍 Riepilogo per Regione")
        grouped_region = grouped.groupby('Regione')[['Associati', 'Importo']].sum().reset_index()
        st.dataframe(grouped_region, use_container_width=True, hide_index=True)
        
    with col_prov:
        st.markdown("##### 🏢 Riepilogo per Provincia (Sede)")
        st.dataframe(grouped, use_container_width=True, hide_index=True)

    st.markdown("---")

def run(file_input, **kwargs):
    if not file_input:
        st.warning("Per favore carica un file Excel.")
        return []

    try:
        current_year = datetime.datetime.now().year
        report_year = current_year - 1
        
        # Elaborazione Dati (Cached per performance e persistenza)
        grouped, error_msg, metrics = process_excel_data(file_input, report_year) # report_year impacting nothing but cache key differentiation
        
        if error_msg:
            st.error(f"Errore: {error_msg}")
            return []
            
        if grouped is None:
             return []

        # Salvataggio Dati in Session State per la Dashboard Persistente
        st.session_state["prospetto_diretti_dashboard_data"] = {
            "metrics": metrics,
            "grouped": grouped
        }
        
        # Estrai metriche per uso logica business (scrittura excel)
        # La dashboard visuale è demandata a get_ui_results() che verrà chiamata dall'app main
        
        # 6. Scrittura su Template
        template_path = os.path.join("tools", "Amministrazione", "FileProspetti_Formattati", "Template-Diretti", "Template-Diretti.xlsx")
        if not os.path.exists(template_path):
            st.error(f"❌ Template non trovato: `{template_path}`")
            return []

        wb = load_workbook(template_path)
        ws = wb.active
        
        # AGGIORNAMENTO AUTOMATICO ANNO
        header_text = ws['A1'].value
        ws['A1'].value = f"ASSOCIATI DIRETTI - {report_year}"

        current_row = 3
        grand_total_associati = 0 # Usato solo per controllo logico o se servisse valore statico
        grand_total_start_row = 3 # Prima riga dati per formula totale complessivo

        # Stili
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        alignment_right = Alignment(horizontal='right')

        regions = grouped['Regione'].unique()
        progress_bar = st.progress(0)
        total_regions = len(regions)
        
        # Liste per tracciare le celle dei Totali Regionali per il Grand Total
        regional_totals_refs_associati = []
        regional_totals_refs_importo = []

        for idx, region in enumerate(regions):
            region_data = grouped[grouped['Regione'] == region]
            
            # Layout STANDARD: Prima Dettagli, poi Totale sotto
            start_row = current_row
            
            # Divide rows into Standard and Veneto Nord
            veneto_nord_keys = ['VERONA', 'VICENZA']
            rows_standard = []
            rows_veneto_nord = []
            
            for _, row in region_data.iterrows():
                if region == "VENETO" and row['Nome_Provincia'] in veneto_nord_keys:
                    rows_veneto_nord.append(row)
                else:
                    rows_standard.append(row)

            # 1. Scrivi Dettagli Standard
            for row in rows_standard:
                c1 = ws.cell(row=current_row, column=1, value=row['Nome_Provincia'])
                c2 = ws.cell(row=current_row, column=2, value=row['Associati'])
                c3 = ws.cell(row=current_row, column=3, value=row['Importo'])
                
                c1.border = thin_border
                c2.border = thin_border
                c3.border = thin_border
                c3.number_format = '#,##0.00'
                
                current_row += 1
            
            # 1b. Scrivi Sottogruppo VENETO NORD (se presente)
            if rows_veneto_nord:
                # Intestazione Veneto Nord (con Somme)
                vn_header_row = current_row
                
                # Calcola range righe dettagli per formule
                vn_start = vn_header_row + 1
                vn_end = vn_header_row + len(rows_veneto_nord)
                
                # Scrivi Header
                c1_vn = ws.cell(row=vn_header_row, column=1, value="VENETO NORD")
                c2_vn = ws.cell(row=vn_header_row, column=2, value=f"=SUM(B{vn_start}:B{vn_end})")
                c3_vn = ws.cell(row=vn_header_row, column=3, value=f"=SUM(C{vn_start}:C{vn_end})")
                
                # Stile Header VN (simile a dettaglio ma grassetto?) o normale? Utente dice "come tutte".
                # Ma è un totale, quindi meglio distinguerlo leggermente? 
                # "NON avere il grassetto, trattala come tutte".
                # OK, allora NO grassetto.
                c1_vn.border = thin_border
                c2_vn.border = thin_border
                c3_vn.border = thin_border
                c3_vn.number_format = '#,##0.00'
                
                current_row += 1
                
                # Dettagli Indentati
                for row in rows_veneto_nord:
                    c1 = ws.cell(row=current_row, column=1, value=row['Nome_Provincia'])
                    c2 = ws.cell(row=current_row, column=2, value=row['Associati'])
                    c3 = ws.cell(row=current_row, column=3, value=row['Importo'])
                    
                    # Indentazione
                    c1.alignment = alignment_right
                    
                    c1.border = thin_border
                    c2.border = thin_border
                    c3.border = thin_border
                    c3.number_format = '#,##0.00'
                    
                    current_row += 1
            
            end_row = current_row - 1
            
            # 2. Scrivi Riga Totale Regione
            # NOTA: Ora abbiamo righe "Doppie" nei valori (Header VN + Dettagli VN).
            # La somma standard =SUM(start:end) includerebbe sia il subtotale VN che i dettagli VN! = DOPPIO CONTEGGIO.
            # Dobbiamo escludere la riga Header VN dal totale regione?
            # Oppure escludere i dettagli VN?
            # Se usiamo =SUM(), conterà tutto.
            # Dobbiamo usare la matematica: (Totale Colonna - Header VN) ? No.
            # Soluzione: SUBTOTALE(9, ...) ignora altri SUBTOTALE(9, ...).
            # Se mettiamo SUBTOTAL(9) nell'header VN?
            # Ma l'utente vuole =SUM().
            
            # Alternativa: Somma Standard (B...:B...) / 2 ?
            # Se Veneto Nord c'è:
            #   Standard Rows: contate 1 volta.
            #   Veneto Nord Header: Somma(Dettagli).
            #   Veneto Nord Dettagli: Dettagli.
            #   Totale = Standard + VN_Header + VN_Details = Standard + 2*VN_Details.
            #   Non è divisibile per 2 pulito.
            
            # Soluzione Formula Regione: =SUM(Standard) + SUM(Header VN).
            # Costruiamo la formula a pezzi.
            
            # Range Standard: start_row : start_row + len(rows_standard) - 1
            # Range VN Header: quel singolo riferimento.
            
            formulas_associati = []
            formulas_importo = []
            
            # Aggiungi range standard se esiste
            if rows_standard:
                std_end = start_row + len(rows_standard) - 1
                if std_end >= start_row:
                    formulas_associati.append(f"B{start_row}:B{std_end}")
                    formulas_importo.append(f"C{start_row}:C{std_end}")
            
            # Aggiungi riferimento header VN se esiste
            if rows_veneto_nord:
                # L'header VN è stato scritto a riga: start_row + len(rows_standard)
                vn_head_r = start_row + len(rows_standard)
                formulas_associati.append(f"B{vn_head_r}")
                formulas_importo.append(f"C{vn_head_r}")
                
            # Componi Formula Regionale
            # =SUM(Range1, Cell2)
            if not formulas_associati:
                f_ass = "0"
                f_imp = "0"
            else:
                f_ass = "=SUM(" + ",".join(formulas_associati) + ")"
                f_imp = "=SUM(" + ",".join(formulas_importo) + ")"
            
            c1_tot = ws.cell(row=current_row, column=1, value=f"{region}")
            c2_tot = ws.cell(row=current_row, column=2, value=f_ass)
            c3_tot = ws.cell(row=current_row, column=3, value=f_imp)
            
            # Aggiungi i riferimenti di QUESTO totale regionale alle liste per il Grand Total
            regional_totals_refs_associati.append(c2_tot.coordinate)
            regional_totals_refs_importo.append(c3_tot.coordinate)
            
            # Stile Totale
            for c in [c1_tot, c2_tot, c3_tot]:
                c.font = bold_font
                c.border = thick_border
            c3_tot.number_format = '#,##0.00'
            
            current_row += 1

            # Salta riga (Empty row separator)
            current_row += 1 
            progress_bar.progress((idx + 1) / total_regions)
            
        # Riga Totale Complessivo
        # Soluzione Richiesta: Somma ESPLICITA dei Totali Regionali.
        
        gt_ass_formula = "=SUM(" + ",".join(regional_totals_refs_associati) + ")" if regional_totals_refs_associati else "0"
        gt_imp_formula = "=SUM(" + ",".join(regional_totals_refs_importo) + ")" if regional_totals_refs_importo else "0"
        
        c1_grand = ws.cell(row=current_row, column=1, value="TOTALE COMPLESSIVO")
        c2_grand = ws.cell(row=current_row, column=2, value=gt_ass_formula)
        c3_grand = ws.cell(row=current_row, column=3, value=gt_imp_formula)

        for c in [c1_grand, c2_grand, c3_grand]:
            c.font = bold_font
            c.border = thick_border
            
        # Formattazione Numerica Totale Complessivo
        c3_grand.number_format = '#,##0.00'

        # Salva File
        output_dir = os.path.join("tools", "Amministrazione", "Output")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            
        output_filename = f"Prospetto Diretti {report_year}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        wb.save(output_path)
        progress_bar.empty()

        st.success(f"✅ Prospetto generato con successo!")
            
        return [output_path]

    except Exception as e:
        st.error(f"Errore critico durante l'elaborazione: {e}")
        st.exception(e)
        return []





