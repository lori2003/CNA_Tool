from __future__ import annotations

import os
import re
import datetime
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# =========================
# TOOLBOX CONFIGURATION
# =========================
TOOL = {'id': 'riepilogo_sindrinn_sedi_cna',
 'name': 'Sedi CNA - Standard e Estero',
 'region': 'Amministrazione',
 'email_reminder': "File da mandare all'amministrazione — questo script produce/modifica 2 file/fogli.",
 'description': '#### 📌 1. FINALITÀ DEL TOOL\n'
                'Automatizza il riepilogo del file **SindRinn** (TXT) producendo **due output distinti** (Standard ed '
                'Estero).\n'
                '\n'
                '#### 🚀 2. COME UTILIZZARLO\n'
                "1. **Dati:** Carica il file TXT (SindRinn) dell'anno attuale.\n"
                '2. **Template:** Il tool utilizza il file `TABULATO_ECONOMICO_AMMINISTRAZIONE.xlsx` specifico per '
                'SindRinn.\n'
                "3. **Verifica:** Genera simultaneamente il report completo e quello filtrato per l'estero.\n"
                '\n'
                '#### 🧠 3. LOGICA DI ELABORAZIONE\n'
                '* **Normalizzazione Sedi:** Applica mappature speciali e protegge i codici regionali.\n'
                '* **Parsing TXT:** Estrae il codice sede (pos 318, lung 4) e 12 campi economici (pos 19, lung 7 '
                'ciascuno).\n'
                "* **Calcolo:** Divide i valori per 100 per ottenere l'importo in euro.\n"
                '* **Aggregazione:** Somma i valori nelle colonne C-N, calcola il totale in colonna O e incrementa il '
                'conteggio record in colonna P.\n'
                '* **Formule Dinamiche:** Inserisce automaticamente formule `=SUM()` nei righi totali regionali e '
                'nazionali.',
 'inputs': [{'key': 'dual_output_info',
             'label': '🔄 **INFO:** Questo tool produrrà **2 file in output**: Standard ed Estero.',
             'type': 'info'},
            {'key': 'format_warning',
             'label': '📌 **Atteso:** file **SindRinn** in formato **TXT** del anno attuale.',
             'type': 'warning'},
            {'key': 'deleghe_reminder',
             'label': '⚠️ **Promemoria:** Verificare se il totale delle deleghe del SindTabe è uguale al numero dei '
                      'record del Sindrinn Txt. SE LA DIFFERENZA dai due è minima (±2) è passabile, se di più, '
                      "scrivere all'INPS.",
             'type': 'info'},
            {'key': 'file_txt_input', 'label': 'Carica File TXT (SindRinn)', 'type': 'file_single', 'required': True}],
 'params': [{'key': 'template_path',
             'label': '📄 Configurazione Template Excel',
             'type': 'file_path_info',
             'default': 'tools/Amministrazione/FileProspetti_Formattati/Riepilogo SindRinn Sedi '
                        'CNA/TABULATO_ECONOMICO_AMMINISTRAZIONE.xlsx',
             'section': 'Riferimento Template'},
            {'key': 'mapping_sedi',
             'label': 'Mappatura Sedi Speciali',
             'type': 'textarea',
             'default': '9901=9901  # Servizi al Territorio 1\n'
                        '3802=3802  # AVEZZANO\n'
                        '2202=2202  # Vibo Valentia\n'
                        '2203=2203  # CROTONE\n'
                        '1301=1301  # IMOLA\n'
                        '3201=3201  # RIMINI\n'
                        '7006=7006  # CIVITAVECCHIA\n'
                        '4927=4927  # LODI\n'
                        '0690=0690  # FERMO\n'
                        '0901=0901  # BAT\n'
                        '1701=1701  # CARBONIA IGLESIS\n'
                        '7390=7390  # OLBIA\n'
                        '3001=3001  # PRATO\n'
                        '5290=5290  # VERBANIA C. OSSOLA\n'
                        '8901=8901  # BIELLA\n'
                        '3290=3200  # FORLI CESENA\n'
                        '4901=2400  # MONZA -> LARIO BRIANZA\n'
                        '4995=2400  # DESIO -> LARIO BRIANZA\n'
                        '4909=2400  # SEREGNO -> LARIO BRIANZA\n'
                        '9999=9999  # NAZIONALE\n'
                        '8700=2400  # LOMBARDIA N.OVEST',
             'section': 'Configurazione'},
            {'key': 'regional_totals_config',
             'label': 'Configurazione Codici Totali (Regionali/Nazionali)',
             'type': 'textarea',
             'default': '9999, 9920, 9919, 9918, 9917, 9916, 9915, 9914, 9913, 9912, 9911, 9910, 9909, 9908, 9907, '
                        '9906, 9905, 9904, 9903, 9902, 9935, 9937',
             'section': 'Configurazione'}]}

# =========================
# LOGICA DI NORMALIZZAZIONE
# =========================
def parse_mapping_config(text: str) -> Dict[str, str]:
    mapping = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
        if "=" in line:
            k, v = line.split("=", 1)
            k_s, v_s = k.strip().zfill(4), v.strip().zfill(4)
            if k_s and v_s:
                mapping[k_s] = v_s
    return mapping

def normalize_sede(code_raw: Any, special_map: Dict[str, str], regional_set: set) -> str:
    s = str(code_raw).strip()
    if s.endswith('.0'):
        s = s[:-2]
    digits = re.sub(r"\D+", "", s)
    if not digits:
        return ""
    code = digits.zfill(4)[-4:]
    if code in special_map:
        return special_map[code]
    if code in regional_set:
        return code
    return code[:2] + "00"

# =========================
# RUNNER
# =========================
def process_data(file_txt_input: Path, template_path: Path, mapping_sedi: str, regional_totals_config: str, out_path: Path, is_estero: bool = False):
    # 2. Caricamento mappatura e codici regionali
    special_map = parse_mapping_config(mapping_sedi)
    master_list_from_ui = {s.strip() for s in regional_totals_config.replace("\n", ",").split(",") if s.strip()}
    protection_set = {c for c in master_list_from_ui if c != "9937"}
    formula_set = master_list_from_ui

    # 3. Caricamento Excel
    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        raise RuntimeError(f"Errore nel caricamento del template: {e}")
    
    # Modifica Titolo per Estero
    if is_estero:
        # Il titolo è solitamente in C1 (celle unite)
        # Verifichiamo se C1 esiste e cambiamo il valore mantenendo lo stile
        current_year = datetime.date.today().year
        cell_c1 = ws['C1']
        cell_c1.value = f"DISTRIBUZIONE DELEGHE ESTERO PER TERRITORIO - RINNOVO {current_year}"

    # 4. Analisi e Azzeramento (Colonne C-P)
    excel_sede_map: Dict[str, int] = {}
    for r in range(7, 3000): # Range ampio
        v = ws.cell(row=r, column=1).value
        # Azzeramento aree C-P (3-16)
        for c in range(3, 17):
            ws.cell(row=r, column=c).value = None

        if v is not None:
            s_v = str(v).strip()
            if s_v.endswith('.0'): s_v = s_v[:-2]
            s_code = re.sub(r"\D+", "", s_v).zfill(4)[-4:]
            if s_code:
                excel_sede_map[s_code] = r
                # Inizializziamo a 0 per somme
                for c in range(3, 17):
                    ws.cell(row=r, column=c).value = 0.0

    # 5. Lettura File TXT
    match_count = 0
    record_count = 0
    
    try:
        with open(file_txt_input, 'r', encoding='latin1') as f:
            lines = f.readlines()
            
        for line in lines:
            if len(line) < 321:
                continue
            
            # FILTRO ESTERO
            # Pos 316 (index 315) o Pos 317 (index 316) == '1'
            if is_estero:
                if line[315] != '1' and line[316] != '1':
                    continue
                
            record_count += 1
            raw_sede = line[317:321]
            norm_s = normalize_sede(raw_sede, special_map, protection_set)
            
            if norm_s in excel_sede_map:
                match_count += 1
                target_row = excel_sede_map[norm_s]
                
                sum_row_importi = 0.0
                for i in range(12):
                    start_idx = 18 + (i * 7)
                    end_idx = start_idx + 7
                    val_str = line[start_idx:end_idx].strip()
                    
                    try:
                        val_float = float(val_str) / 100.0
                    except:
                        val_float = 0.0
                        
                    current_cell_val = ws.cell(row=target_row, column=i+3).value or 0.0
                    ws.cell(row=target_row, column=i+3).value = current_cell_val + val_float
                    sum_row_importi += val_float
                
                current_tot_val = ws.cell(row=target_row, column=15).value or 0.0
                ws.cell(row=target_row, column=15).value = current_tot_val + sum_row_importi
                
                current_rec_count = ws.cell(row=target_row, column=16).value or 0.0
                ws.cell(row=target_row, column=16).value = current_rec_count + 1
                
    except Exception as e:
        raise RuntimeError(f"Errore durante l'elaborazione del file TXT: {e}")

    # 6. Inserimento Formule per i Totali
    cols_to_sum = range(3, 17) # C..P
    MASTER_TOTAL_LIST = {c for c in formula_set if c not in {"9999", "9936"}}
    reg_rows_found = {c: r for c, r in excel_sede_map.items() if c in formula_set}
    boundary_items = sorted([(r, c) for c, r in reg_rows_found.items() if c in MASTER_TOTAL_LIST])
    
    last_processed_row = 6 
    row_9936 = excel_sede_map.get("9936")
    
    for t_row, s_code in boundary_items:
        start_sum = last_processed_row + 1
        end_sum = t_row - 1
        if start_sum <= end_sum:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                formula = f"=SUM({col_let}{start_sum}:{col_let}{end_sum})"
                if row_9936 and start_sum <= row_9936 <= end_sum:
                    formula += f"-{col_let}{row_9936}"
                ws.cell(row=t_row, column=c_idx).value = formula
        last_processed_row = t_row

    if "9999" in reg_rows_found:
        t_row_9999 = reg_rows_found["9999"]
        relevant_rows = [r for c, r in reg_rows_found.items() if c in MASTER_TOTAL_LIST]
        if relevant_rows:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                cells_to_add = [f"{col_let}{r}" for r in relevant_rows]
                ws.cell(row=t_row_9999, column=c_idx).value = f"=SUM({','.join(cells_to_add)})"

    if row_9936:
        row_9000 = excel_sede_map.get("9000")
        row_9100 = excel_sede_map.get("9100")
        if row_9000 and row_9100:
            for c_idx in cols_to_sum:
                col_let = get_column_letter(c_idx)
                ws.cell(row=row_9936, column=c_idx).value = f"={col_let}{row_9000}+{col_let}{row_9100}"

    # Formattazione numerica
    for c_idx in range(3, 16):
        for r in (list(excel_sede_map.values()) + list(reg_rows_found.values())):
             ws.cell(row=r, column=c_idx).number_format = '#,##0.00'
    for r in (list(excel_sede_map.values()) + list(reg_rows_found.values())):
         ws.cell(row=r, column=16).number_format = '#,##0'

    wb.save(out_path)
    return record_count, match_count

def run(file_txt_input: Path, template_path: str, mapping_sedi: str, regional_totals_config: str, out_dir: Path) -> List[Path]:
    # 1. Risoluzione Template Path
    abs_template_path = Path(template_path)
    if not abs_template_path.is_absolute():
        project_root = Path(__file__).resolve().parent.parent.parent
        abs_template_path = project_root / template_path
    
    if not abs_template_path.exists():
        raise FileNotFoundError(f"Template non trovato: {abs_template_path}")

    # Output paths
    out_standard = out_dir / "TABULATO_ECONOMICO_DA_SINDRINN.xlsx"
    out_estero = out_dir / "DISTRIBUZIONE DELEGHE ESTERO PER TERRITORIO.xlsx"

    # Process Standard
    rec_std, match_std = process_data(file_txt_input, abs_template_path, mapping_sedi, regional_totals_config, out_standard, is_estero=False)
    
    # Process Estero
    rec_est, match_est = process_data(file_txt_input, abs_template_path, mapping_sedi, regional_totals_config, out_estero, is_estero=True)

    st.success(f"Elaborazione completata!\n"
               f"- **Standard**: {rec_std} record, {match_std} corrispondenze.\n"
               f"- **Estero**: {rec_est} record, {match_est} corrispondenze.")

    return [out_standard, out_estero]
