from __future__ import annotations

import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import Dict, List, Any

from core.toolkit import ctx
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ══════════════════════════════════════════════
# COSTANTI
# ══════════════════════════════════════════════
TEMPLATE_PATH = Path(
    r"C:/Users/simoncellil/Desktop/toolbox/tools/Amministrazione"
    r"/FileProspetti_Formattati/Risultato di Gestione/RISULTATO DI GESTIONE 2024_OK.xlsx"
)

GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
RED_FILL   = PatternFill(start_color="FF2222", end_color="FF2222", fill_type="solid")

# Indici colonne (1-based, openpyxl)
COL_A  = 1   # A
COL_B  = 2   # B
COL_C  = 3   # C
COL_D  = 4   # D
COL_E  = 5   # E
COL_F  = 6   # F
COL_G  = 7   # G
COL_H  = 8   # H
COL_N  = 14  # N  (nota "Convertita in ...")
COL_I  = 9   # I
COL_J  = 10  # J
COL_K  = 11  # K  ← Concomitanti
COL_L  = 12  # L  ← Deleghe
COL_M  = 13  # M  ← Revoche
COL_O  = 15  # O  (file 1 – Concomitanti)
COL_AB = 28  # AB (file 1 – Revoche)
COL_AO = 41  # AO (file 1 – Deleghe)

# Alias per il confronto normalizzato (chiavi già normalizzate con normalize_name)
ALIAS_MAP = {
    'VENETOOVEST':  'VENETONORD',       # VENETO OVEST → VENETO NORD
    'VALLEDAOS':    'VALLEDAOSTA',      # VALLE D'AOS → VALLE D'AOSTA
    'LARIOBRIANZA': 'LOMBARDIANOVEST',  # LARIO BRIANZA → LOMBARDIA N.OVEST
    'VARESE':       'LOMBARDIANOVEST',  # VARESE → LOMBARDIA N.OVEST
}

# ══════════════════════════════════════════════
# CONFIGURAZIONE TOOL
# ══════════════════════════════════════════════
TOOL = {
    'id': 'risultato_di_gestione',
    'name': 'Risultato di Gestione',
    'region': 'Amministrazione',
    'email_reminder': "File da mandare all'amministrazione",
    'description': (
        '#### 📌 1. FINALITÀ DEL TOOL\n'
        'Compila il prospetto **Risultato di Gestione** partendo da due sorgenti: '
        'il file *effettivo dopo totale* di dicembre e il *Tabulato Economico Sindrinn* dell\'anno corrente.\n'
        '\n'
        '#### 🚀 2. COME UTILIZZARLO\n'
        '1. **Input 1:** Carica il file **"effettivo dopo totale"** di **dicembre dell\'anno passato**.\n'
        '2. **Input 2:** Carica il **Tabulato Economico da Sindrinn** già elaborato con l\'anno corrente.\n'
        '3. **Esegui:** Il tool popola il template, confronta le sedi e evidenzia i match.\n'
        '4. **Scarica:** Ottieni il file compilato con evidenziazione verde/rossa.\n'
        '\n'
        '#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n'
        '* **Pulizia preventiva:** Svuota le colonne B, C, D, E, F del template.\n'
        '* **Effettivo (J-M):** Copia col A, O (Concomitanti), AO (Deleghe), AB (Revoche) dal file 1 (da riga 4) → template J-M (da riga 5, intestazioni in riga 4).\n'
        '* **Sindrinn (H-I):** Copia col B e P dal file 2 (da riga 7) → template H-I (da riga 5).\n'
        '* **Confronto:** Normalizza H e colonna A (maiuscolo, rimuove spazi/trattini). Se match → valore I in colonna B, H:I **verde**. Altrimenti H:I **rosso**.\n'
        '\n'
        '#### 📂 4. RISULTATO FINALE\n'
        'File Excel (.xlsx) con il prospetto compilato e l\'evidenziazione verde/rossa del confronto Sindrinn ↔ Effettivo.'
    ),
    'inputs': [
        {
            'key': 'warning_effettivo',
            'label': (
                '⚠️ **Attenzione:** Caricare SOLO il file **"effettivo dopo totale"** '
                'del mese di **DICEMBRE dell\'anno passato**.'
            ),
            'type': 'warning'
        },
        {
            'key': 'file_effettivo',
            'label': 'Carica "Effettivo Dopo Totale" (Dicembre Anno Passato)',
            'type': 'file_single',
            'required': True
        },
        {
            'key': 'warning_tabulato',
            'label': (
                '⚠️ **File Atteso:** "Tabulato Economico da Sindrinn, già elaborato con anno corrente"'
            ),
            'type': 'warning'
        },
        {
            'key': 'file_tabulato',
            'label': 'Carica Tabulato Economico (Sindrinn – Anno Corrente)',
            'type': 'file_single',
            'required': True
        },
    ],
    'params': [
        {
            'key': 'template_info',
            'label': '',
            'type': 'dynamic_info',
            'function': 'get_template_status',
            'section': 'Configurazione'
        },
        {
            'key': 'excel_template_name',
            'label': '📄 Template Excel',
            'type': 'file_path_info',
            'default': str(TEMPLATE_PATH),
            'section': 'Configurazione',
            'help': 'ℹ️ Path del file template. Usa 📂 per cercarlo nel PC o 🔍 per aprire la cartella attuale.'
        },
    ]
}

# ══════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════
def normalize_name(name) -> str:
    """Normalizza un nome comune/provincia per il confronto.

    Converte in maiuscolo, rimuove accenti, apostrofi, spazi e altri
    separatori per rendere il confronto robusto a varianti tipografiche.
    """
    if name is None:
        return ""
    s = str(name).upper().strip()
    # Rimuove accenti (À→A, È→E, ecc.)
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    # Rimuove apostrofi, spazi, trattini, underscore, slash, punti
    s = re.sub(r"[\s\-_/\\.,'\u2019\u2018`\"]+", '', s)
    return s


def canonical_name(name) -> str:
    """Normalizza e applica le regole di alias per il confronto."""
    n = normalize_name(name)
    if not n:
        return ''
    # MILANO* → MILANO (qualsiasi suffisso viene ignorato)
    if n.startswith('MILANO'):
        return 'MILANO'
    return ALIAS_MAP.get(n, n)


def is_aliased(name) -> bool:
    """Restituisce True se il nome viene trasformato dall'alias (o dalla regola MILANO*)."""
    return canonical_name(name) != normalize_name(name)


def get_template_status(values: Dict[str, Any]) -> str:
    """Mostra lo stato del template nella UI."""
    template_val = values.get("excel_template_name")
    if not template_val:
        for p in TOOL.get("params", []):
            if p.get("key") == "excel_template_name":
                template_val = p.get("default")
                break

    if not template_val:
        ctx.error("❌ Nessun template configurato.")
        return ""

    if Path(template_val).exists():
        return "✅ **Template trovato**"
    else:
        ctx.error(f"❌ Template non trovato: `{template_val}`")
        return ""


# ══════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════
def run(
    file_effettivo: Path,
    file_tabulato: Path,
    excel_template_name: str,
    out_dir: Path,
    **kwargs
) -> List[Path]:

    # ── 1. Risoluzione template ──────────────────────────────────────────────
    template_path = Path(excel_template_name) if excel_template_name else TEMPLATE_PATH
    if not template_path.exists():
        raise FileNotFoundError(f"Template non trovato: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    # ── 1b. Aggiornamento date/anno nel template ──────────────────────────────
    current_year = date.today().year
    ws["B3"] = date(current_year, 1, 1)
    c1_val = ws["C1"].value
    if c1_val:
        ws["C1"].value = re.sub(r'\b\d{4}\b', str(current_year - 1), str(c1_val))

    # ── 2. Pulizia preventiva colonne B, C, D, E, F (da riga 5) ────────────
    safe_max = ws.max_row + 500
    for r in range(5, safe_max):
        for col in (COL_B, COL_C, COL_D, COL_E, COL_F):
            ws.cell(row=r, column=col).value = None

    # ── 3. File 1 – "effettivo dopo totale" ─────────────────────────────────
    try:
        wb1 = load_workbook(file_effettivo, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Errore apertura file 'effettivo dopo totale': {exc}")

    ws1 = wb1.active

    # Intestazioni fisse in riga 4 del template (colonne K, L, M)
    ws.cell(row=4, column=COL_K).value = "Concomitanti"
    ws.cell(row=4, column=COL_L).value = "Deleghe"
    ws.cell(row=4, column=COL_M).value = "Revoche"

    # Raccoglie i dati dal file 1 (salta righe completamente vuote)
    data_f1 = []
    for r_in in range(4, ws1.max_row + 1):
        val_a  = ws1.cell(row=r_in, column=COL_A).value
        val_o  = ws1.cell(row=r_in, column=COL_O).value
        val_ao = ws1.cell(row=r_in, column=COL_AO).value
        val_ab = ws1.cell(row=r_in, column=COL_AB).value
        if all(v is None for v in (val_a, val_o, val_ao, val_ab)):
            continue
        data_f1.append((val_a, val_o, val_ao, val_ab))

    wb1.close()

    # ── 3b. Raggruppamento File 1: somma K,L,M per nome canonico ────────────────
    # Più voci con lo stesso canonical (es. VARESE + LARIO BRIANZA → LOMBARDIA
    # N.OVEST) vengono sommate in un unico gruppo; ogni riga sorgente è tracciata
    # per indice → nessun valore va perduto.
    j_groups: Dict[str, Any] = {}
    for idx, (val_a, val_o, val_ao, val_ab) in enumerate(data_f1):
        key = canonical_name(val_a)
        if not key:
            continue
        if key not in j_groups:
            j_groups[key] = {'k': 0.0, 'l': 0.0, 'm': 0.0,
                              'indices': [], 'raw': [], 'has_num': False}
        g = j_groups[key]
        for attr, v in (('k', val_o), ('l', val_ao), ('m', val_ab)):
            if v is not None:
                try:
                    g[attr] += float(v)
                    g['has_num'] = True
                except (TypeError, ValueError):
                    pass
        g['indices'].append(idx)
        g['raw'].append((val_a, val_o, val_ao, val_ab))

    ctx.info(
        f"**📂 File 1 — Effettivo Dopo Totale**  \n"
        f"Righe catturate: **{len(data_f1)}** — "
        f"raggruppate in **{len(j_groups)}** voci canoniche"
    )

    # ── 4. File 2 – Tabulato Economico Sindrinn ──────────────────────────────
    try:
        wb2 = load_workbook(file_tabulato, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Errore apertura file 'Tabulato Economico Sindrinn': {exc}")

    ws2 = wb2.active

    COL_P_FILE2 = 16
    data_f2 = []
    for r_in in range(7, ws2.max_row + 1):
        val_b = ws2.cell(row=r_in, column=2).value
        val_p = ws2.cell(row=r_in, column=COL_P_FILE2).value
        if not isinstance(val_b, str) or not val_b.strip():
            continue
        data_f2.append((val_b.strip(), val_p))

    wb2.close()

    with ctx.expander(f"🔍 Dati letti dal Tabulato Sindrinn — {len(data_f2)} righe"):
        for nome, valore in data_f2:
            ctx.write(f"• {nome}  →  {valore}")

    # ── 4b. Raggruppamento File 2: somma val P per nome canonico ─────────────
    h_groups: Dict[str, Any] = {}
    for idx, (val_b, val_p) in enumerate(data_f2):
        key = canonical_name(val_b)
        if not key:
            continue
        if key not in h_groups:
            h_groups[key] = {'p': 0.0, 'indices': [], 'raw': [], 'has_num': False}
        g = h_groups[key]
        if val_p is not None:
            try:
                g['p'] += float(val_p)
                g['has_num'] = True
            except (TypeError, ValueError):
                pass
        g['indices'].append(idx)
        g['raw'].append((val_b, val_p))

    ctx.info(
        f"**📂 File 2 — Tabulato Sindrinn**  \n"
        f"Righe catturate: **{len(data_f2)}** — "
        f"raggruppate in **{len(h_groups)}** voci canoniche"
    )

    # ── 5. Pulizia colonne H-M (area di lavoro) ───────────────────────────────
    for r in range(5, ws.max_row + 200):
        for col in (COL_H, COL_I, COL_J, COL_K, COL_L, COL_M):
            ws.cell(row=r, column=col).value = None

    # ── 6. Riscrittura allineata a col A (gruppi sommati) ─────────────────────
    matched_hi   = 0
    unmatched_hi = 0
    matched_jm   = 0
    unmatched_jm = 0
    unmatched_h_names: List[str] = []

    used_f1_indices: set = set()
    used_f2_indices: set = set()
    audit_f1: List[tuple] = []  # (raw_entries, k_sum, l_sum, m_sum, target)
    audit_f2: List[tuple] = []  # (raw_entries, p_sum, target)
    last_row = 5

    for r in range(5, ws.max_row + 100):
        val_a = ws.cell(r, COL_A).value
        if not val_a or not str(val_a).strip():
            continue
        last_row = r
        can_a = canonical_name(val_a)

        # Gruppo H:I (Sindrinn) — somma tutte le voci con stesso nome canonico
        h_grp = h_groups.get(can_a)
        if h_grp:
            h_name = h_grp['raw'][0][0]
            i_v = int(round(h_grp['p'])) if h_grp['has_num'] else None
            ws.cell(r, COL_H).value = h_name
            ws.cell(r, COL_I).value = i_v
            ws.cell(r, COL_B).value = i_v
            ws.cell(r, COL_H).fill = GREEN_FILL
            ws.cell(r, COL_I).fill = GREEN_FILL
            used_f2_indices.update(h_grp['indices'])
            matched_hi += 1
            if len(h_grp['raw']) > 1 or any(is_aliased(e[0]) for e in h_grp['raw']):
                audit_f2.append((h_grp['raw'], i_v, str(val_a)))
        else:
            ws.cell(r, COL_H).fill = RED_FILL
            ws.cell(r, COL_I).fill = RED_FILL
            unmatched_hi += 1

        # Gruppo J:M (Effettivo) — somma tutte le voci con stesso nome canonico
        j_grp = j_groups.get(can_a)
        if j_grp:
            j_name = j_grp['raw'][0][0]
            if j_grp['has_num']:
                k_v = int(round(j_grp['k']))
                l_v = int(round(j_grp['l']))
                m_v = int(round(j_grp['m']))
            else:
                k_v = l_v = m_v = None
            ws.cell(r, COL_J).value = j_name
            ws.cell(r, COL_K).value = k_v
            ws.cell(r, COL_L).value = l_v
            ws.cell(r, COL_M).value = m_v
            ws.cell(r, COL_C).value = k_v
            ws.cell(r, COL_D).value = l_v
            ws.cell(r, COL_E).value = m_v
            ws.cell(r, COL_J).fill = GREEN_FILL
            ws.cell(r, COL_K).fill = GREEN_FILL
            ws.cell(r, COL_L).fill = GREEN_FILL
            ws.cell(r, COL_M).fill = GREEN_FILL
            used_f1_indices.update(j_grp['indices'])
            matched_jm += 1
            if len(j_grp['raw']) > 1 or any(is_aliased(e[0]) for e in j_grp['raw']):
                audit_f1.append((j_grp['raw'], k_v, l_v, m_v, str(val_a)))
        else:
            ws.cell(r, COL_J).fill = RED_FILL
            ws.cell(r, COL_K).fill = RED_FILL
            ws.cell(r, COL_L).fill = RED_FILL
            ws.cell(r, COL_M).fill = RED_FILL
            unmatched_jm += 1

    # ── 6b. Append in fondo — tutte le righe non abbinate (tracking per indice)
    append_row = last_row + 2

    for idx, (val_b, val_p) in enumerate(data_f2):
        if idx not in used_f2_indices:
            ws.cell(append_row, COL_H).value = val_b
            ws.cell(append_row, COL_I).value = val_p
            ws.cell(append_row, COL_H).fill = RED_FILL
            ws.cell(append_row, COL_I).fill = RED_FILL
            unmatched_h_names.append(str(val_b))
            append_row += 1

    for idx, (val_a, val_o, val_ao, val_ab) in enumerate(data_f1):
        if idx not in used_f1_indices:
            ws.cell(append_row, COL_J).value = val_a
            ws.cell(append_row, COL_K).value = val_o
            ws.cell(append_row, COL_L).value = val_ao
            ws.cell(append_row, COL_M).value = val_ab
            ws.cell(append_row, COL_J).fill = RED_FILL
            ws.cell(append_row, COL_K).fill = RED_FILL
            ws.cell(append_row, COL_L).fill = RED_FILL
            ws.cell(append_row, COL_M).fill = RED_FILL
            append_row += 1

    # ── 6c-pre. Audit trail — voci raggruppate/convertite scritte in fondo ────
    if audit_f2 or audit_f1:
        append_row += 1  # riga vuota di separazione

    for raw_entries, p_sum, target in audit_f2:
        for orig_name, val_p in raw_entries:
            ws.cell(append_row, COL_H).value = orig_name
            ws.cell(append_row, COL_I).value = val_p
            ws.cell(append_row, COL_J).value = f"Convertita in '{target}'"
            append_row += 1

    for raw_entries, k_v, l_v, m_v, target in audit_f1:
        for orig_name, val_o, val_ao, val_ab in raw_entries:
            ws.cell(append_row, COL_J).value = orig_name
            ws.cell(append_row, COL_K).value = val_o
            ws.cell(append_row, COL_L).value = val_ao
            ws.cell(append_row, COL_M).value = val_ab
            ws.cell(append_row, COL_N).value = f"Convertita in '{target}'"
            append_row += 1

    # ── 6c. Somme speciali: VENETO OVEST e PIEMONTE NORD ─────────────────────
    # Costruisce mappa canonical(col A) → numero riga per poter cercare le sedi
    a_row_map: Dict[str, int] = {}
    for r in range(5, last_row + 1):
        va = ws.cell(r, COL_A).value
        if va and str(va).strip():
            a_row_map[canonical_name(va)] = r

    def _apply_sum(target: str, sources: List[str]) -> None:
        """Somma C, D, E delle righe `sources` nella riga di `target`."""
        r_tgt = a_row_map.get(canonical_name(target))
        if r_tgt is None:
            return
        for col in (COL_B, COL_C, COL_D, COL_E):
            total = 0.0
            found = False
            for src in sources:
                r_src = a_row_map.get(canonical_name(src))
                if r_src is not None:
                    v = ws.cell(r_src, col).value
                    if v is not None:
                        try:
                            total += float(v)
                            found = True
                        except (TypeError, ValueError):
                            pass
            if found:
                ws.cell(r_tgt, col).value = int(round(total))

    # VENETO OVEST (alias VENETO NORD) = VERONA + VICENZA
    _apply_sum("VENETO OVEST", ["VERONA", "VICENZA"])
    # PIEMONTE NORD = NOVARA + VERBANO C.O. + VERCELLI
    _apply_sum("PIEMONTE NORD", ["NOVARA", "VERBANO C.O.", "VERCELLI"])

    # ── 6d. Formula colonna F: (C + D) - E ───────────────────────────────────
    for r in range(5, last_row + 1):
        va = ws.cell(r, COL_A).value
        if va and str(va).strip():
            ws.cell(r, COL_F).value = f"=C{r}+D{r}-E{r}"

    # ── 6e. Formattazione colonne B-E: font nero + allineamento centrale ──────
    _black_font = Font(color="000000")
    _center     = Alignment(horizontal="center")
    for r in range(5, last_row + 1):
        va = ws.cell(r, COL_A).value
        if va and str(va).strip():
            for col in (COL_B, COL_C, COL_D, COL_E, COL_F):
                cell = ws.cell(r, col)
                cell.font      = _black_font
                cell.alignment = _center

    # ── 6f. Grassetto B-F per le righe dove col G = 10 ───────────────────────
    _bold_black = Font(bold=True, color="000000")
    for r in range(5, last_row + 1):
        g_val = ws.cell(r, COL_G).value
        if g_val is not None and str(g_val).strip() == "10":
            for col in (COL_B, COL_C, COL_D, COL_E, COL_F):
                ws.cell(r, col).font = _bold_black

    # ── 7. Salvataggio ────────────────────────────────────────────────────────
    out_path = out_dir / "RISULTATO DI GESTIONE.xlsx"
    wb.save(out_path)
    wb.close()

    # ── 8. Feedback all'utente ────────────────────────────────────────────────
    ctx.success(f"✅ File generato: **{out_path.name}**")
    ctx.markdown("---")
    ctx.markdown("### 📊 Riepilogo confronto")

    c1, c2 = ctx.columns(2)
    with c1:
        ctx.markdown("**Gruppo H:I — Sindrinn**")
        ctx.write(f"• ✅ Abbinati (I → B, verde): **{matched_hi}**")
        ctx.write(f"• ❌ Col A senza match Sindrinn (rosso): **{unmatched_hi}**")
        n_h_fondo = len(data_f2) - len(used_f2_indices)
        ctx.write(f"• ⬇️ Scritte in fondo (non abbinate): **{n_h_fondo}**")
    with c2:
        ctx.markdown("**Gruppo J:M — Effettivo**")
        ctx.write(f"• ✅ Abbinati (K→C, L→D, M→E, verde): **{matched_jm}**")
        ctx.write(f"• ❌ Col A senza match Effettivo (rosso): **{unmatched_jm}**")
        n_j_fondo = len(data_f1) - len(used_f1_indices)
        ctx.write(f"• ⬇️ Scritte in fondo (non abbinate): **{n_j_fondo}**")

    if unmatched_h_names:
        with ctx.expander(f"⬇️ Voci Sindrinn non abbinate — scritte in fondo ({len(unmatched_h_names)})"):
            for name in unmatched_h_names:
                ctx.write(f"• {name}")

    return [out_path]


# ══════════════════════════════════════════════
# GUIDA DINAMICA
# ══════════════════════════════════════════════
def get_guide(values: Dict[str, Any]) -> str:
    return """
### 📘 Guida e Logica del Tool

Questo strumento automatizza la compilazione del prospetto **Risultato di Gestione**
combinando due sorgenti dati distinte.

#### 📌 1. FINALITÀ DEL TOOL
Centralizza e automatizza il processo di popolare il template con i dati del
consuntivo di dicembre e del tabulato economico, eseguendo il confronto e
l'abbinamento automatico tra le sedi/comuni.

#### 🚀 2. COME UTILIZZARLO
1. **Input 1 – Effettivo Dopo Totale:** Caricare **esclusivamente** il file di **dicembre dell'anno passato**.
2. **Input 2 – Tabulato Sindrinn:** Caricare il tabulato già elaborato con i dati dell'anno in corso.
3. **Esegui:** Il tool elabora, confronta e genera il file finale.
4. **Scarica** il file `.xlsx` dalla sezione output.

#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)
- **Pulizia preventiva:** Svuota le colonne B, C, D, E, F del template prima di ogni elaborazione.
- **Popolazione colonne J-M (effettivo):**
  - Colonna J ← Col A del file 1 (da riga 4)
  - Colonna K "Concomitanti" ← Col O del file 1
  - Colonna L "Deleghe" ← Col AO del file 1
  - Colonna M "Revoche" ← Col AB del file 1
  - I dati partono dalla riga 5 del template; le intestazioni sono scritte in riga 4.
- **Popolazione colonne H-I (Sindrinn):**
  - Colonna H ← Col B del file 2 (da riga 7)
  - Colonna I ← Col P del file 2
  - I dati partono dalla riga 5 del template.
- **Confronto automatico:**
  - Per ogni riga con dati in H, il nome viene normalizzato (maiuscolo, rimozione di spazi, trattini, slash, ecc.).
  - Se esiste una corrispondenza in colonna A del template → il valore di I viene scritto nella colonna B di quella riga; celle H:I colorate di **verde**.
  - Se nessuna corrispondenza → celle H:I colorate di **rosso**.

#### 📂 4. RISULTATO FINALE
File `.xlsx` con il prospetto compilato. Le colonne J-M contengono i dati effettivi
di dicembre. Le colonne H-I sono evidenziate in base all'esito del confronto
(**verde** = abbinato e importo inserito in B, **rosso** = non abbinato).
"""
