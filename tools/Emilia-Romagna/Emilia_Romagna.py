# sindrinn_step1_2_completo.py
from __future__ import annotations

import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from collections import Counter
from typing import Dict, List, Optional, Tuple, Any

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def convert_slk_to_xlsx(src: Path, dest: Path) -> Path:
    """
    Converte un file .slk in .xlsx preservando il contenuto come stringhe.
    Non altera zeri iniziali né formattazioni numeriche.
    """
    try:
        lines = src.read_text(errors="ignore").splitlines()
        wb = Workbook()
        ws = wb.active

        for line in lines:
            if not line.startswith("C;"):
                continue
            parts = line.strip().split(";")
            x = y = None
            val: Optional[str] = None
            for part in parts[1:]:
                if part.startswith("X"):
                    try:
                        x = int(part[1:])
                    except Exception:
                        x = None
                elif part.startswith("Y"):
                    try:
                        y = int(part[1:])
                    except Exception:
                        y = None
                elif part.startswith("K"):
                    raw = part[1:]
                    if raw.startswith('"') and raw.endswith('"'):
                        raw = raw[1:-1].replace('""', '"')
                    val = raw
            if x is not None and y is not None:
                ws.cell(row=y, column=x, value=val)

        wb.save(dest)
        return dest
    except Exception:
        # In caso di problemi, restituiamo il sorgente senza modifiche
        return src



# =========================
# TOOLBOX CONFIGURATION
# =========================
TOOL = {'id': 'sindrinn_normalizer',
 'name': 'Prospetti -Genere -Categoria ecc...',
'description': (
    "#### 📌 1. FINALITÀ DEL TOOL\n"
    "Normalizza ed elabora i file di input (TXT Sindrinn o Banca Dati XLSX/SLK) per popolare i prospetti istituzionali "
    "Emilia-Romagna, calcolando automaticamente le fasce d'età, il genere e le trattenute economiche.\n\n"
    "#### 🚀 2. COME UTILIZZARLO\n"
    "1. **Input:** Carica il file Sindrinn (TXT) o la Banca Dati (XLSX). Trascina il file Template Excel che deve essere popolato.\n"
    "2. **Parametri:** Imposta l'Anno di Riferimento e i parametri economici (Minimo INPS, Aliquote).\n"
    "3. **Esecuzione:** Il tool scansiona i dati, applica le formule e compila i fogli 'genere_età', 'per_categoria' e 'per_genere_import'.\n\n"
    "#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n"
    "* **Parsing Fixed-Width:** Estrae i dati dal TXT Sindrinn basandosi su un tracciato a posizioni fisse (Categoria, Trattenuta, Anno, Genere, Sede).\n"
    "* **Algoritmo Età:** Calcola l'età basandosi sull'Anno Report e gestisce la fascia esclusiva per i nati dal 2000 in poi (Colonne AA..AF).\n"
    "* **Reverse Calculation:** Determina l'Importo Lordo partendo dalla trattenuta netta utilizzando gli scaglioni di aliquota variabili configurati.\n"
    "* **Filtro Territoriale:** Filtra automaticamente le sedi regionali normalizzandole al codice provincia (XX00).\n\n"
    "#### 📂 4. RISULTATO FINALE\n"
    "File Excel Template popolato con tabelle pivot, riepiloghi per età/genere e log di validazione vs VBA."
),
 'inputs': [{'key': 'file_txt', 'label': 'File TXT Sindrinn', 'type': 'txt_single', 'required': False},
            {'key': 'file_xlsx',
             'label': 'File Excel Template/Destinazione (o Output VBA per confronto)',
             'type': 'xlsx_single',
             'required': True},
            {'key': 'file_banca_dati',
             'label': 'Input da File Banca Dati (XLSX/SLK)',
             'type': 'xlsx_single',
             'required': False,
             'note': '⚠️ IMPORTANTE: Caricare preferibilmente file .XLSX. Ordine colonne richiesto (dalla A):\n'
                     'A: Categoria | B: Data Nascita | C: Sesso | D: Sede | E: Trattenuta'}],
 'params': [{'key': 'report_year',
              'label': 'Anno Riferimento (Report)',
              'type': 'number',
              'default': 2025,
              'section': 'Variabili Economiche & Annuali'},
            {'key': 'pension_min',
             'label': 'Trattamento Minimo INPS (€)',
             'type': 'number',
             'default': 603.4,
             'min': 0.0,
             'max': 5000.0,
             'step': 0.01,
             'section': 'Variabili Economiche & Annuali'},
             {'key': 'aliquota_1',
              'label': 'Scaglioni Trattenute: 1° Livello (%)',
              'type': 'number',
              'default': 0.5,
              'min': 0.0,
             'max': 100.0,
             'step': 0.01,
             'section': 'Variabili Economiche & Annuali'},
            {'key': 'aliquota_2',
             'label': 'Scaglioni Trattenute: 2° Livello (%)',
             'type': 'number',
             'default': 0.4,
             'min': 0.0,
             'max': 100.0,
             'step': 0.01,
             'section': 'Variabili Economiche & Annuali'},
            {'key': 'aliquota_3',
             'label': 'Scaglioni Trattenute: 3° Livello (%)',
             'type': 'number',
             'default': 0.35,
             'min': 0.0,
             'max': 100.0,
             'step': 0.01,
             'section': 'Variabili Economiche & Annuali'},
            {'key': 'enable_maggiorazione',
             'label': 'Abilita Logica Maggiorazione Sottominimo (Legacy)',
             'type': 'checkbox',
             'default': False,
             'section': 'Variabili Economiche & Annuali',
             'help': 'Se ATTIVO: Usa il coefficiente 1.022 nei calcoli inversi (logica storica per minime maggiorate).\n'
                     'Se SPENTO (Default): Usa il calcolo standard puro (Trattenuta / Aliquota) senza correttivi.'},
            {'key': 'template_start_row',
             'label': 'Riga Inizio Dati (Template)',
             'type': 'number',
             'default': 5,
             'section': 'Configurazioni Avanzate'},
            {'key': 'compare_vba',
             'label': 'Confronta con dati precedenti (es. Output VBA)',
             'type': 'checkbox',
             'default': False,
             'section': 'Configurazioni Avanzate',
             'help': 'Se attivo, confronta i risultati con il file Excel in input.'},
            {'key': 'normalize_emilia',
             'label': 'Normalizzazione Sedi Emilia-Romagna (Filtro Default)',
             'type': 'checkbox',
             'default': True,
             'section': 'Configurazioni Avanzate',
             'help': 'Normalizza i codici (XX00) e filtra solo sedi Emilia-Romagna.'},
            {'key': 'mapping_special',
             'label': 'Mappatura Sedi Speciali',
             'type': 'textarea',
             'default': '3802=3802  # AVEZZANO\n2202=2202  # Vibo Valentia\n2203=2203  # CROTONE\n1301=1301  # IMOLA\n3201=3201  # RIMINI\n7006=7006  # CIVITAVECCHIA\n4927=4927  # LODI\n0690=0690  # FERMO\n0901=0901  # BAT\n1701=1701  # CARBONIA IGLESIS\n7390=7390  # OLBIA\n3001=3001  # PRATO\n5290=5290  # VERBANIA C. OSSOLA\n8901=8901  # BIELLA\n3290=3200  # FORLI CESENA\n4901=2400  # MONZA -> LARIO BRIANZA\n4995=2400  # DESIO -> LARIO BRIANZA\n4909=2400  # SEREGNO -> LARIO BRIANZA\n9999=9999  # NAZIONALE',
             'section': 'Configurazioni Avanzate',
             'help': 'Override mapping sedi.'}]}


# =========================
# TRACCIATO SINDRINN (1-based)
# =========================
SCHEMA = {
    "categoria":      (1,   3),
    "trattenuta":     (33,  7),
    "anno_nascita":   (157, 4),
    "genere":         (295, 1),
    "sede_gestione":  (318, 4),
}

SHEET_NAME_DEFAULT = "Da Eliminare"
SHEET_ETA = "genere_età"
SHEET_CATEG = "per_categoria"
SHEET_IMPORTI = "per_genere_import"
SHEET_GENERE = "associati_per_gen"
SHEET_LOG = "Log"

# =========================
# STEP 2 - NORMALIZZAZIONE SEDI
# =========================
def parse_mapping_config(text: str) -> Dict[str, str]:
    """Converte il testo di configurazione in un dizionario."""
    mapping = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
        
        if "=" in line:
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip()
            if k and v:
                mapping[k] = v
    return mapping


# =========================
# HELPER DINAMICI (UI)
# =========================
def get_preview_calcolo(values: Dict[str, Any]) -> str:
    """Funzione helper chiamata da app.py per mostrare anteprima calcolo."""
    try:
        p_min = values.get("pension_min", 603.4)
        coeff = values.get("coeff_maggiorazione", 1.022)
        
        # Gestione robusta tipi (se arrivano stringhe da UI parziale)
        p_min = float(p_min) if p_min is not None else 0.0
        coeff = float(coeff) if coeff is not None else 1.0
        
        risultato = p_min * coeff
        
        # Formattazione
        return f"**Anteprima Calcolo:** {p_min:.2f} € * {coeff:.4f} = **{risultato:.2f} €** (Soglia Maggiorata)"
    except Exception:
        return ""

def find_last_code_row(ws, start_row: int) -> int:
    """
    Trova l'ultima riga contenente un codice sede (4 cifre) in colonna A.
    Scan dalla start_row fino a ws.max_row.
    """
    last_row = start_row - 1
    max_r = ws.max_row
    if not max_r: return last_row
    
    for r in range(start_row, max_r + 1):
        val = ws.cell(row=r, column=1).value
        if val is None:
            continue
        val_str = str(val).strip()
        if not val_str:
            continue
        val_lower = val_str.lower()
        if val_lower.startswith("totale") or val_lower.startswith("total"):
            continue
        s = _digits_only(val_str)
        if len(s) == 4:
            last_row = r
        elif str(val).lower().startswith("totale"):
            # Se troviamo la riga totale esistente, potremmo fermarci o usarla come riferimento?
            # L'utente dice: footer è last_code_row + 1. Quindi ignoriamo la presenza di "TOTALE"
            # come dato, ma continuiamo a cercare codici sotto (improbabile).
            pass
            
    return last_row

def _write_sum_formula(ws, target_row, target_col, range_start_row, range_end_row, col_idx=None):
    """Writing formula helper: =SUM(X5:X10)"""
    from openpyxl.utils import get_column_letter
    c_letter = get_column_letter(col_idx if col_idx else target_col)
    f = f"=SUM({c_letter}{range_start_row}:{c_letter}{range_end_row})"
    ws.cell(row=target_row, column=target_col).value = f





def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")

def _norm_code_from_cell(v) -> str:
    """
    Normalizza un valore letto da Excel in un codice a 4 cifre.
    Estrae il primo gruppo di cifre trovato.
    """
    if v is None:
        return ""
    if isinstance(v, int):
        s = str(v)
    elif isinstance(v, float):
        s = str(int(v)) if v.is_integer() else str(v)
    else:
        s = str(v)
    matches = re.findall(r"\d+", s)
    if not matches:
        return ""
    return matches[0].zfill(4)[-4:]

def to_cp(code4: str) -> str:
    """Regola CP: imposta 3° e 4° carattere a '00' -> XX00."""
    if len(code4) != 4 or not code4.isdigit():
        return ""
    return code4[:2] + "00"

def normalize_sede_to_cp(code_raw: str, special_map: Dict[str, str]) -> str:
    """Normalizza la sede e restituisce il CODICE PROVINCIA (CP).
    1) pulizia
    2) override da special_map
    3) fallback a CP (XX00)
    """
    d = _digits_only(code_raw)
    if not d:
        return ""
    code = d.zfill(4)[-4:]
    
    code = special_map.get(code, code)
    return to_cp(code)


# =========================
# STEP 1 - PARSING FIXED-WIDTH
# =========================
def _slice_1based(line: str, start_1based: int, length: int) -> str:
    start0 = start_1based - 1
    end0 = start0 + length
    if start0 >= len(line):
        return ""
    return line[start0:end0]

def norm_categoria(raw: str) -> str:
    s = (raw or "").strip()
    d = _digits_only(s)
    if d:
        return d.zfill(3)[:3]
    return s[:3]

def norm_anno_nascita(raw: str) -> str:
    d = _digits_only((raw or "").strip())
    return d[:4] if len(d) >= 4 else ""

def norm_genere(raw: str) -> str:
    s = (raw or "").strip().upper()
    return s if s in ("M", "F") else ""

def norm_trattenuta_text(raw: str) -> str:
    d = _digits_only((raw or "").strip())
    if not d:
        return ""
    if len(d) > 7:
        d = d[-7:]
    d = d.zfill(7)
    int_part = d[:-2].lstrip("0") or "0"
    dec_part = d[-2:]
    return f"{int_part},{dec_part}"

# =========================
# Conversione Numerica Robusta (IT <-> Float)
# =========================
def it_text_to_float(s: str) -> Optional[float]:
    """Converte testo italiano (es. '1.200,50' o '42,10') in float (1200.5, 42.1).
    Gestisce punti migliaia, virgola decimale e spazi.
    Ritorna None se stringa vuota o non valida.
    """
    if not s:
        return None
    # Rimuove spazi e punti (separatori migliaia)
    clean = s.strip().replace(" ", "").replace(".", "")
    # Sostituisce la virgola decimale con punto
    clean = clean.replace(",", ".")
    try:
        val = float(clean)
        return round(val, 2)
    except ValueError:
        return None

def float_to_it_text(x: Optional[float], decimals: int = 2) -> str:
    """Trasforma float -> testo italiano con virgola e N decimali fissi.
    Es: 1234.567 -> '1234,57'
    Se None, ritorna stringa vuota.
    """
    if x is None:
        return ""
    # Formattazione fixed-point standard (il punto si userà per lo split)
    s = f"{x:.{decimals}f}"
    return s.replace(".", ",")

def trattenuta_to_float(tr_text: str) -> Optional[float]:
    return it_text_to_float(tr_text)


# =========================
# Importo Lordo
# =========================
def calcola_importo_lordo_da_trattenuta(trattenuta: Optional[float], pension_min: float = 603.4, uplift_pct: float = 2.2) -> str:
    if trattenuta is None:
        return ""

    trattMinimo = pension_min
    fattore_magg = 1 + (uplift_pct / 100.0)
    
    doppioTrattMinimo = 2 * trattMinimo
    coeffMaggiorato = 0.005 * fattore_magg

    soglia1 = 0.005 * trattMinimo * fattore_magg
    soglia2 = 0.005 * trattMinimo + 0.004 * (doppioTrattMinimo - trattMinimo)

    if trattenuta <= soglia1:
        importoLordo = trattenuta / coeffMaggiorato
    elif trattenuta <= soglia2:
        importoLordo = trattMinimo + 250 * (trattenuta - 0.005 * trattMinimo)
    else:
        importoLordo = doppioTrattMinimo + (trattenuta - soglia2) / 0.0035

    if importoLordo <= trattMinimo:
        importoLordo *= fattore_magg

    return float_to_it_text(importoLordo, 2)


# =========================
# Record pulito
# =========================
def calcola_importo_lordo_da_trattenuta(trattenuta: Optional[float], 
                                          pension_min: float = 603.4, 
                                          coeff_maggiorazione: float = 1.022,
                                          aliq_1_pct: float = 0.50,
                                          aliq_2_pct: float = 0.40,
                                          aliq_3_pct: float = 0.35) -> str:
    """
    Versione aggiornata con parametri aliquote dinamici (ridefinizione).
    """
    if trattenuta is None:
        return ""

    trattMinimo = pension_min
    # Coefficiente Input diretto
    fattore_magg = coeff_maggiorazione
    
    # Coefficienti (da % a decimale)
    coeff_1 = aliq_1_pct / 100.0
    coeff_2 = aliq_2_pct / 100.0
    coeff_3 = aliq_3_pct / 100.0

    doppioTrattMinimo = 2 * trattMinimo
    coeffMaggiorato = coeff_1 * fattore_magg

    # Calcolo soglie basato sui nuovi coefficienti
    soglia1 = coeff_1 * trattMinimo * fattore_magg
    soglia2 = coeff_1 * trattMinimo + coeff_2 * (doppioTrattMinimo - trattMinimo)

    if trattenuta <= soglia1:
        # Avoid division by zero
        if coeffMaggiorato == 0: return float_to_it_text(0.0, 2)
        importoLordo = trattenuta / coeffMaggiorato
    elif trattenuta <= soglia2:
        if coeff_2 == 0: return float_to_it_text(trattMinimo, 2)
        numeratore = trattenuta - (coeff_1 * trattMinimo)
        importoLordo = trattMinimo + (numeratore / coeff_2)
    else:
        if coeff_3 == 0: return float_to_it_text(doppioTrattMinimo, 2)
        importoLordo = doppioTrattMinimo + (trattenuta - soglia2) / coeff_3

    # REMOVED UPLIFT APPLICATION AS PER USER REQUEST (28/01/2026)
    # The logic required is to use the "normal" gross amount for brackets.
    # if importoLordo <= trattMinimo:
    #    importoLordo *= fattore_magg

    return float_to_it_text(importoLordo, 2)


@dataclass
class CleanRow:
    categoria: str
    anno_nascita: str
    genere: str
    sede_cp: str
    trattenuta_text: str
    importo_lordo_text: str

def leggi_sindrinn(txt_path: Path, encoding: str, special_map: Dict[str, str], 
                   pension_min: float = 603.4, coeff_maggiorazione: float = 1.022,
                   a1: float = 0.50, a2: float = 0.40, a3: float = 0.35) -> List[CleanRow]:
    out: List[CleanRow] = []
    with txt_path.open("r", encoding=encoding, errors="replace") as f:
        for line in f:
            line = line.rstrip("\r\n")

            cat_raw = _slice_1based(line, *SCHEMA["categoria"])
            tra_raw = _slice_1based(line, *SCHEMA["trattenuta"])
            ann_raw = _slice_1based(line, *SCHEMA["anno_nascita"])
            gen_raw = _slice_1based(line, *SCHEMA["genere"])
            sed_raw = _slice_1based(line, *SCHEMA["sede_gestione"])

            categoria = norm_categoria(cat_raw)
            anno = norm_anno_nascita(ann_raw)
            genere = norm_genere(gen_raw)

            # Lasciamo la sede GREZZA (solo pulizia, nessuna normalizzazione)
            # La normalizzazione verrà applicata successivamente in run()
            sede_base = _digits_only((sed_raw or "").strip())
            sede_cp = sede_base.zfill(4)[-4:] if sede_base else ""

            trat_text = norm_trattenuta_text(tra_raw)
            trat_float = trattenuta_to_float(trat_text)
            lordo_text = calcola_importo_lordo_da_trattenuta(
                trat_float, pension_min, coeff_maggiorazione, a1, a2, a3
            )

            out.append(CleanRow(
                categoria=categoria,
                anno_nascita=anno,
                genere=genere,
                sede_cp=sede_cp,
                trattenuta_text=trat_text,
                importo_lordo_text=lordo_text
            ))
    return out


def prova_encoding(txt_path: Path, special_map: Dict[str, str], 
                   pension_min: float, coeff_maggiorazione: float, 
                   a1: float, a2: float, a3: float) -> Tuple[List[CleanRow], str]:
    for enc in ("cp1252", "latin1", "utf-8"):
        try:
            rows = leggi_sindrinn(txt_path, enc, special_map, pension_min, coeff_maggiorazione, a1, a2, a3)
            return rows, enc
        except Exception:
            continue
    raise RuntimeError("Impossibile leggere il TXT con cp1252/latin1/utf-8.")


def leggi_banca_dati_xlsx(path: Path,
                          pension_min: float,
                          coeff_maggiorazione: float,
                          a1: float,
                          a2: float,
                          a3: float
                          ) -> List[CleanRow]:
    """
    Lettura file banca dati (XLSX/SLK già convertito in XLSX).
    Colonne attese:
    A: Categoria
    B: Data nascita (usa solo l'anno)
    C: Sesso/Genere
    D: Sede
    E: Importo trattenuta
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out: List[CleanRow] = []

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cat_raw, dob_raw, sex_raw, sede_raw, trat_raw = (row_vals + (None,) * 5)[:5]

        categoria = norm_categoria(str(cat_raw or ""))

        # Estrai anno da data/numero/testo
        anno = ""
        if hasattr(dob_raw, "year"):
            anno = str(dob_raw.year)
        else:
            anno_digits = _digits_only(str(dob_raw or ""))
            if len(anno_digits) >= 4:
                anno = anno_digits[-4:]

        genere = norm_genere(str(sex_raw or ""))
        sede_base = _digits_only(str(sede_raw or "")).zfill(4)[-4:]
        if isinstance(trat_raw, (int, float)):
            trat_float = float(trat_raw)
            trat_text = float_to_it_text(trat_float)
        else:
            trat_text = norm_trattenuta_text(str(trat_raw) if trat_raw is not None else "")
            trat_float = trattenuta_to_float(trat_text)

        lordo_text = calcola_importo_lordo_da_trattenuta(
            trat_float, pension_min, coeff_maggiorazione, a1, a2, a3
        )

        out.append(CleanRow(
            categoria=categoria,
            anno_nascita=anno,
            genere=genere,
            sede_cp=sede_base,
            trattenuta_text=trat_text,
            importo_lordo_text=lordo_text,
        ))

    return out


# =========================
# Excel Helpers
# =========================
def _parse_brackets(s: str) -> List[float]:
    if not s:
        return []
    try:
        s = s.replace("\n", ",").replace(";", ",")
        parts = [p.strip() for p in s.split(",") if p.strip()]
        vals = [float(p) for p in parts]
        return sorted(vals)
    except Exception:
        return []

def _write_footer_totals(ws, start_row: int, end_row: int, cols: List[int], label_col: int = 1):
    """
    Scrive una riga 'TOTALE' in fondo e somma le colonne specificate.
    """
    if end_row < start_row:
        return
    r = end_row + 1
    ws.cell(row=r, column=label_col).value = "TOTALE"
    for c in cols:
        col_letter = get_column_letter(c)
        ws.cell(row=r, column=c).value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"



def get_or_create_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

def write_headers(ws):
    ws["A1"].value = "Categoria"
    ws["C1"].value = "Anno Nascita"
    ws["D1"].value = "Maschi/Femmine"
    ws["E1"].value = "Sede"
    ws["F1"].value = "Trattenuta"
    ws["G1"].value = "Importo Lordo"
    ws["H1"].value = "Minimo Utilizzato"
    ws["I1"].value = "Minimo Maggiorato"

def clear_values(ws):
    max_row = ws.max_row or 1
    if max_row < 2:
        return
    for r in range(2, max_row + 1):
        for c in range(1, 10):  # A..I
            ws.cell(row=r, column=c).value = None

def dump_rows(ws, rows: List[CleanRow], pension_min: float, coeff_maggiorazione: float):
    # Pre-calcolo valori audit (stringhe o float? Excel preferisce numeri, ma qui usiamo text per coerenza?)
    # Usiamo numeri per Excel, formattati
    min_val = pension_min
    magg_val = pension_min * coeff_maggiorazione
    
    r = 2
    for row in rows:
        ws.cell(row=r, column=1).value = row.categoria
        ws.cell(row=r, column=3).value = row.anno_nascita
        ws.cell(row=r, column=4).value = row.genere
        ws.cell(row=r, column=5).value = row.sede_cp
        ws.cell(row=r, column=6).value = row.trattenuta_text
        ws.cell(row=r, column=7).value = row.importo_lordo_text
        
        # Colonne Audit
        ws.cell(row=r, column=8).value = min_val
        ws.cell(row=r, column=9).value = magg_val
        
        r += 1

def dump_banca_dati_sheet(
    wb,
    rows: List[CleanRow],
    pension_min: float,
    coeff_maggiorazione: float,
    sheet_name: str = "FileRipu_E.R_Da_BD",
):
    """
    Scrive un riepilogo dei record provenienti dal file Banca Dati.
    Colonne: Categoria, Anno, Genere, Sede, Trattenuta, Importo Lordo,
    Pensione Minima, Pensione Maggiorata.
    """
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)

    headers = [
        "Categoria",
        "Anno Nascita",
        "Genere",
        "Sede",
        "Trattenuta",
        "Importo Lordo",
        "Pensione Minima",
        "Pensione Maggiorata",
    ]
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=idx).value = h

    min_val_str = float_to_it_text(pension_min, 2)
    # Maggiorazione Disattivata (28/01/2026) -> Non mostriamo valori
    magg_val_str = "" 

    r = 2
    for row in rows:
        ws.cell(row=r, column=1).value = row.categoria
        ws.cell(row=r, column=2).value = row.anno_nascita
        ws.cell(row=r, column=3).value = row.genere
        ws.cell(row=r, column=4).value = row.sede_cp
        ws.cell(row=r, column=5).value = row.trattenuta_text
        ws.cell(row=r, column=6).value = row.importo_lordo_text
        ws.cell(row=r, column=7).value = min_val_str
        ws.cell(row=r, column=8).value = magg_val_str
        r += 1


# =========================
# BLOCCO 2 - Distribuzione_per_genere_età (foglio esistente)
# =========================

def _eta_bucket_idx(eta: int) -> int:
    # 0..7
    if eta <= 60:
        return 0
    if 61 <= eta <= 65:
        return 1
    if 66 <= eta <= 70:
        return 2
    if 71 <= eta <= 75:
        return 3
    if 76 <= eta <= 80:
        return 4
    if 81 <= eta <= 85:
        return 5
    if 86 <= eta <= 90:
        return 6
    return 7  # >=91


def _calc_eta(anno_nascita_text: str, anno_rif: int) -> Optional[int]:
    d = _digits_only(anno_nascita_text)
    if len(d) != 4:
        return None
    try:
        return anno_rif - int(d)
    except ValueError:
        return None


def _build_sede_row_map(ws, start_row: int = 5, code_col: int = 1) -> Dict[str, int]:
    """
    Legge colonna A dal start_row in poi e crea mappa codice_sede(4 cifre) -> riga.
    Usa la PRIMA occorrenza se ci sono duplicati.
    """
    mp: Dict[str, int] = {}
    max_row = ws.max_row or start_row
    for r in range(start_row, max_row + 1):
        v = ws.cell(row=r, column=code_col).value
        # Gestione robusta: converte in stringa, pulisce, zfill
        # Se il valore in Excel è numero 1300 -> str -> "1300"
        if v is None:
            continue
        v_str = str(v).strip()
        if not v_str:
            continue
        v_lower = v_str.lower()
        if v_lower.startswith("totale") or v_lower.startswith("total"):
            continue
        code = _digits_only(v_str)
        if not code:
            continue
        code = code.zfill(4)[-4:]
        
        # Usa la PRIMA occorrenza (evita shift se ci sono duplicati)
        if code not in mp:
            mp[code] = r
        # else:
        #     print(f"DEBUG: Codice {code} duplicato trovato a riga {r}, ignorato (usando riga {mp[code]})")
    
    return mp


def _clear_eta_output(ws, start_row: int = 5, start_col: int = 3, end_col: int = 26):
    """
    Pulisce SOLO i valori nell'area dati C..Z (default) dalla riga start_row all'ultima riga usata.
    Non tocca stili/formattazione.
    """
    max_row = ws.max_row or start_row
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def compila_distribuzione_per_genere_eta(ws, rows: List["CleanRow"], anno_rif: int = 2025, start_row: int = 5):
    """
    Compila il foglio esistente 'Distribuzione_per_genere_età'.
    Logica aggiornata (28/01/2026):
    - Fasce 0-7 (Standard) basate su Età -> Colonne C..Z
    - Fascia 8 (Nati >= 2000) Esclusiva -> Colonne AA..AC
    - Totali Generali traslati in AD..AF
    """
    # 1) mappa sede -> riga
    sede_to_row = _build_sede_row_map(ws, start_row=start_row, code_col=1)

    # 2) pulizia valori area output (C..AF) -> esteso fino a colonna 32
    # C=3 ... Z=26, AA=27, AB=28, AC=29, AD=30, AE=31, AF=32
    _clear_eta_output(ws, start_row=start_row, start_col=3, end_col=32)

    # 3) aggregazione in memoria: agg[sede][bucket] = [U, D]
    # Bucket 0-7: Fasce Età Standard
    # Bucket 8: Nati >= 2000
    agg: Dict[str, List[List[int]]] = {}
    
    for r in rows:
        sede = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if sede not in sede_to_row:
            continue

        # Logica Determinazione Bucket
        anno_nascita_val = None
        if r.anno_nascita and r.anno_nascita.isdigit():
            anno_nascita_val = int(r.anno_nascita)
        
        target_bucket = -1
        
        # 1. Check Prioritario: Nati dal 2000 in poi
        if anno_nascita_val is not None and anno_nascita_val >= 2000:
            target_bucket = 8
        else:
            # 2. Check Standard Età
            eta = _calc_eta(r.anno_nascita, anno_rif=anno_rif)
            if eta is not None:
                target_bucket = _eta_bucket_idx(eta)
        
        if target_bucket == -1:
            continue

        if sede not in agg:
            agg[sede] = [[0, 0] for _ in range(9)]  # 9 buckets totali

        if r.genere == "M":
            agg[sede][target_bucket][0] += 1
        elif r.genere == "F":
            agg[sede][target_bucket][1] += 1

    # 4) scrittura su foglio
    from openpyxl.utils import get_column_letter

    final_tot_col_u = 30 # AD
    final_tot_col_d = 31 # AE
    final_tot_col_t = 32 # AF

    for sede, row_out in sede_to_row.items():
        buckets = agg.get(sede)
        
        # --- A) Fasce Standard 0-7 (Cols C..Z) ---
        col = 3  # Start C
        for b in range(8):
            if buckets:
                u = buckets[b][0]
                d = buckets[b][1]
                ws.cell(row=row_out, column=col).value = u
                ws.cell(row=row_out, column=col + 1).value = d
            
            cu = get_column_letter(col)
            cd = get_column_letter(col + 1)
            # Totale Fascia (M+F)
            ws.cell(row=row_out, column=col + 2).value = f"={cu}{row_out}+{cd}{row_out}"
            col += 3
        
        # --- B) Fascia Speciale >= 2000 (Cols AA..AC) ---
        # Colonna attuale 'col' dovrebbe essere 27 (AA) dopo il loop (3 + 8*3 = 27)
        col_special = 27 
        if buckets:
            u_spec = buckets[8][0]
            d_spec = buckets[8][1]
            ws.cell(row=row_out, column=col_special).value = u_spec
            ws.cell(row=row_out, column=col_special + 1).value = u_spec # Attenzione: bug fix required in typing? No, u_spec, d_spec
            ws.cell(row=row_out, column=col_special + 1).value = d_spec

        c_spec_u = get_column_letter(col_special)
        c_spec_d = get_column_letter(col_special + 1)
        ws.cell(row=row_out, column=col_special + 2).value = f"={c_spec_u}{row_out}+{c_spec_d}{row_out}"

        # --- C) Totali Generali (AD..AF) ---
        # Somma di tutti i bucket 0..8
        # Uomini: cols 3, 6, 9... + 27
        refs_u = [f"{get_column_letter(3 + b*3)}{row_out}" for b in range(9)]
        refs_d = [f"{get_column_letter(4 + b*3)}{row_out}" for b in range(9)]
        
        ws.cell(row=row_out, column=final_tot_col_u).value = "=" + "+".join(refs_u)
        ws.cell(row=row_out, column=final_tot_col_d).value = "=" + "+".join(refs_d)
        
        c_tot_u = get_column_letter(final_tot_col_u)
        c_tot_d = get_column_letter(final_tot_col_d)
        ws.cell(row=row_out, column=final_tot_col_t).value = f"={c_tot_u}{row_out}+{c_tot_d}{row_out}"

    # 5) Footer Totals (Formulas)
    last_code_r = find_last_code_row(ws, start_row)
    footer_row = last_code_r + 1
    
    # Formula Footer: Sum C..AF (3..32)
    for c in range(3, 33):
        _write_sum_formula(ws, footer_row, c, start_row, last_code_r, c)



# =========================
# BLOCCO 3 - Distribuzione_per_categoria (foglio esistente)
# =========================

def _get_group_categoria(cat_code: str) -> int:
    """
    Replica GetGroup (VBA):
    ritorna gruppo 1..8 in base alla categoria (stringa 3 cifre).
    """
    code = (cat_code or "").strip()
    d = _digits_only(code)
    if d:
        code = d.zfill(3)[-3:]

    if code in {"018", "088"}:
        return 1
    if code in {"019", "089"}:
        return 2
    if code in {"020", "090"}:
        return 3
    if code in {"015", "085", "016", "086", "017", "087"}:
        return 4
    if code in {"021", "091", "022", "092", "023", "093"}:
        return 5
    if code in {"001", "004", "007", "002", "005", "008", "003", "006", "009"}:
        return 6

    # Else: se numerico > 199 => gruppo 7, altrimenti 8
    if _digits_only(code):
        try:
            if int(code) > 199:
                return 7
        except ValueError:
            pass
    return 8


def _clear_categorie_output(ws, start_row: int = 5, start_col: int = 3, end_col: int = 32):
    """
    Pulisce SOLO i valori in C..AF dalla riga start_row in poi.
    (C=3, AF=32)
    """
    max_row = ws.max_row or start_row
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def compila_distribuzione_per_categoria(ws, rows: List["CleanRow"], start_row: int = 5):
    """
    Compila il foglio esistente 'Distribuzione_per_categoria':
    - chiave: codice sede in colonna A (da start_row)
    - scrive conteggi (come testo) in blocchi da 3 colonne per gruppi 1..8
    - imposta formule Totale (M+F) per ogni gruppo e totali L-M-N per gruppi 1..3
    """
    sede_to_row = _build_sede_row_map(ws, start_row=start_row, code_col=1)

    # 1) pulizia area output
    _clear_categorie_output(ws, start_row=start_row, start_col=3, end_col=32)  # C..AF

    # 2) aggregazione: agg[(sede, gruppo)] = [M, F]
    agg: Dict[tuple, List[int]] = {}

    for r in rows:
        sede = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if not sede or sede not in sede_to_row:
            continue

        gruppo = _get_group_categoria(r.categoria)
        key = (sede, gruppo)
        if key not in agg:
            agg[key] = [0, 0]

        if r.genere == "M":
            agg[key][0] += 1
        elif r.genere == "F":
            agg[key][1] += 1

    # 3) scrittura conteggi + formule Totale per gruppo
    from openpyxl.utils import get_column_letter
    group_m_cols = [3, 6, 9, 15, 18, 21, 24, 27]
    group_f_cols = [c + 1 for c in group_m_cols]
    group_t_cols = [c + 2 for c in group_m_cols]

    for (sede, gruppo), (m_cnt, f_cnt) in agg.items():
        row_out = sede_to_row[sede]

        # colonne base gruppi:
        # gruppi 1-3: C(3), F(6), I(9)
        # gruppi 4-8: O(15), R(18), U(21), X(24), AA(27)
        if gruppo <= 3:
            col_base = 3 + (gruppo - 1) * 3
        else:
            col_base = 15 + (gruppo - 4) * 3

        # M e F come numeri
        ws.cell(row=row_out, column=col_base).value = m_cnt
        ws.cell(row=row_out, column=col_base + 1).value = f_cnt

        # Totali per gruppo impostati con formule nella passata successiva

    # 4) Totali riga con formule
    for sede, row_out in sede_to_row.items():
        for m_col, f_col, t_col in zip(group_m_cols, group_f_cols, group_t_cols):
            ws.cell(row=row_out, column=t_col).value = f"={get_column_letter(m_col)}{row_out}+{get_column_letter(f_col)}{row_out}"

        refs_m_1_3 = [f"{get_column_letter(c)}{row_out}" for c in group_m_cols[:3]]
        refs_f_1_3 = [f"{get_column_letter(c)}{row_out}" for c in group_f_cols[:3]]
        ws.cell(row=row_out, column=12).value = "=" + "+".join(refs_m_1_3)
        ws.cell(row=row_out, column=13).value = "=" + "+".join(refs_f_1_3)
        ws.cell(row=row_out, column=14).value = f"={get_column_letter(12)}{row_out}+{get_column_letter(13)}{row_out}"

        refs_m_all = [f"{get_column_letter(c)}{row_out}" for c in group_m_cols]
        refs_f_all = [f"{get_column_letter(c)}{row_out}" for c in group_f_cols]
        ws.cell(row=row_out, column=30).value = "=" + "+".join(refs_m_all)
        ws.cell(row=row_out, column=31).value = "=" + "+".join(refs_f_all)
        ws.cell(row=row_out, column=32).value = f"={get_column_letter(30)}{row_out}+{get_column_letter(31)}{row_out}"

    last_code_r = find_last_code_row(ws, start_row)
    footer_row = last_code_r + 1
    for c in range(3, 33):
        _write_sum_formula(ws, footer_row, c, start_row, last_code_r, c)


# =========================
# BLOCCO 4 - Distribuzione per importo (foglio esistente)
# =========================

FASCE_MIN = [603.4, 1206.8, 1810.2, 2413.6, 3017.0, 3620.4, 3946.18]  # 8 fasce totali (0..7)

def _fascia_importo(importo: float, fasce: List[float] = None) -> int:
    """
    Ritorna fascia 0..7
    """
    val = round(importo, 2)
    soglie = fasce if fasce is not None else FASCE_MIN
    
    # Se fasce custom, possono essere diverse da 7 soglie?
    # L'utente ha chiesto di validare che siano 7 nel run().
    # Quindi assumiamo che soglie sia corretto.
    
    for i, soglia in enumerate(soglie):
        if val <= soglia:
            return i
    return len(soglie) # Ultima fascia (es 7 se soglie sono 7, indici 0..6)


def _clear_importi_output(ws, start_row: int = 5, start_col: int = 3, end_col: int = 29):
    """
    Pulisce SOLO i valori nell'area C..AC dalla riga start_row in poi.
    (8 fasce * 3 colonne + totali riga = C..AC)
    """
    max_row = ws.max_row or start_row
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None


def _set_totali_formule_importi(ws, start_row: int = 5):
    """
    Imposta le formule del totale (T) per ogni fascia:
    T = M + F (col+2 = col + col+1)
    Non tocca M e F.
    """
    max_row = ws.max_row or start_row
    for r in range(start_row, max_row + 1):
        # Filtro base: se colonna 1 non è un codice sede valido, salto (intestazioni/vuote)
        if not _digits_only(str(ws.cell(row=r, column=1).value) if ws.cell(row=r, column=1).value else ""):
            continue
        col = 3  # C
        for _ in range(8):
            c_m = ws.cell(row=r, column=col).coordinate
            c_f = ws.cell(row=r, column=col + 1).coordinate
            ws.cell(row=r, column=col + 2).value = f"=SUM({c_m},{c_f})"
            col += 3


def compila_distribuzione_per_importo(ws, rows: List["CleanRow"], fasce: List[float] = None, start_row: int = 5):
    """
    Compila il foglio distribuzione per importo.
    """
    from openpyxl.cell.cell import MergedCell
    
    num_fasce = len(fasce) + 1 if fasce else 8

    # 1) Aggregazione: agg[sede][fascia] = [M, F]
    agg: Dict[str, List[List[int]]] = {}

    for r in rows:
        # Normalizza codice sede
        sede = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if not sede:
            continue

        # Converti importo lordo
        imp = it_text_to_float(r.importo_lordo_text)
        if imp is None:
            continue

        # Determina fascia importo
        fascia = _fascia_importo(imp, fasce)
        if fascia >= num_fasce: fascia = num_fasce - 1 # Safety cap
        
        # Inizializza struttura se necessario
        if sede not in agg:
            agg[sede] = [[0, 0] for _ in range(num_fasce)]

        # Conta per genere
        if r.genere == "M":
            agg[sede][fascia][0] += 1
        elif r.genere == "F":
            agg[sede][fascia][1] += 1

    # 2) Pulizia area output (C..Z)
    _clear_importi_output(ws, start_row=start_row, start_col=3, end_col=3 + (num_fasce*3) + 2)

    # 3) Scrittura: scorri TUTTE le righe del foglio
    max_row = ws.max_row or 5
    for r in range(5, max_row + 1):  # Da riga 5 in poi
        # Leggi e normalizza il codice in colonna A
        cell_value = ws.cell(row=r, column=1).value
        code = _norm_code_from_cell(cell_value)
        
        if not code:
            continue  # Riga senza codice, salta
        
        # Scrivi conteggi come numeri
        cell_c = ws.cell(row=r, column=3)
        cell_d = ws.cell(row=r, column=4)
        cell_e = ws.cell(row=r, column=5)

        f_row = f"=C{r}+D{r}"

        # Se questo codice e nei dati aggregati, scrivi i conteggi
        if code in agg:
            m_cnt, f_cnt = agg[code]
            if not isinstance(cell_c, MergedCell):
                cell_c.value = m_cnt
            if not isinstance(cell_d, MergedCell):
                cell_d.value = f_cnt
        if not isinstance(cell_e, MergedCell):
            cell_e.value = f_row # FORMULA

    # 4) Footer Totals (Formulas)
    last_code_r = find_last_code_row(ws, start_row)
    footer_row = last_code_r + 1
    
    # Sum Cols C, D, E (3, 4, 5)
    for c in [3, 4, 5]:
        _write_sum_formula(ws, footer_row, c, start_row, last_code_r, c)
    
    # _write_footer_totals(ws, start_row, max_row, [3, 4, 5]) -> REMOVED


# =========================
# BLOCCO LOG - Foglio esistente "Log"
# =========================

# =========================
# BLOCCO 4bis - Distribuzione associati per genere (foglio esistente)
# =========================

def _clear_genere_output(ws, start_row: int = 5, start_col: int = 3, end_col: int = 5):
    """
    Pulisce SOLO le colonne C e D (Maschi, Femmine) da start_row in poi.
    NON tocca la colonna E (Totale con formule).
    NON tocca le intestazioni (righe < 5).
    """
    from openpyxl.cell.cell import MergedCell
    if start_row < 5:
        start_row = 5
    max_row = ws.max_row or start_row
    if max_row < start_row:
        return
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col):
            cell = ws.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def compila_distribuzione_associati_genere(ws, rows: List["CleanRow"], start_row: int = 5):
    """
    Compila il foglio distribuzione per genere.
    """
    from openpyxl.cell.cell import MergedCell

    agg: Dict[str, List[int]] = {}
    for r in rows:
        sede = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if not sede:
            continue
        if sede not in agg:
            agg[sede] = [0, 0]
        if r.genere == "M":
            agg[sede][0] += 1
        elif r.genere == "F":
            agg[sede][1] += 1

    _clear_genere_output(ws, start_row=start_row)

    max_row = ws.max_row or 5
    for r in range(start_row, max_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        code = _norm_code_from_cell(cell_value)
        if not code:
            continue

        cell_c = ws.cell(row=r, column=3)
        cell_d = ws.cell(row=r, column=4)
        cell_e = ws.cell(row=r, column=5)
        f_row = f"=C{r}+D{r}"

        if code in agg:
            m_cnt, f_cnt = agg[code]
            if not isinstance(cell_c, MergedCell):
                cell_c.value = m_cnt
            if not isinstance(cell_d, MergedCell):
                cell_d.value = f_cnt
        if not isinstance(cell_e, MergedCell):
            cell_e.value = f_row

    last_code_r = find_last_code_row(ws, start_row)
    footer_row = last_code_r + 1
    for c in [3, 4, 5]:
        _write_sum_formula(ws, footer_row, c, start_row, last_code_r, c)


# =========================
# BLOCCO LOG - Foglio esistente "Log"
# =========================

def _clear_log_sheet(ws, max_rows: int = 200, max_cols: int = 12):
    """Pulisce SOLO i valori in un'area del foglio Log (non stili)."""
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            ws.cell(row=r, column=c).value = None


def _count_invalid_genere(rows: List["CleanRow"]) -> int:
    return sum(1 for r in rows if (r.genere or "") not in ("M", "F"))


def _count_invalid_anno(rows: List["CleanRow"]) -> int:
    return sum(1 for r in rows if len(_digits_only(r.anno_nascita)) != 4)


def _count_invalid_categoria(rows: List["CleanRow"]) -> int:
    return sum(1 for r in rows if not _digits_only(r.categoria))


def _count_missing_sede(rows: List["CleanRow"]) -> int:
    return sum(1 for r in rows if not _digits_only(r.sede_cp))


def _count_invalid_importo_lordo(rows: List["CleanRow"]) -> int:
    # usa la funzione del blocco importi
    bad = 0
    for r in rows:
        s = (r.importo_lordo_text or "").strip()
        if not s:
            bad += 1
            continue
        if it_text_to_float(s) is None:
            bad += 1
    return bad


def _audit_sheet_coverage(ws, rows: List["CleanRow"], start_row: int) -> Dict[str, Any]:
    """
    Analisi approfondita della copertura per un foglio.
    Ritorna:
    - 'present_rows': volume record con sede mappata
    - 'absent_rows': volume record persi (sede mancante)
    - 'missing_codes': lista [(codice, count), ...] dei top missing
    """
    # 1. Quali sedi sono disponibili nel Template?
    sede_to_row = _build_sede_row_map(ws, start_row=start_row, code_col=1)
    available_sedi = set(sede_to_row.keys())
    
    # 2. Contiamo le sedi nei dati di input (già filtrati Emilia)
    input_sedi_counts = Counter()
    for r in rows:
        code = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if code:
            input_sedi_counts[code] += 1
            
    # 3. Incrocio
    present_vol = 0
    absent_vol = 0
    missing_counts = Counter()
    
    for code, count in input_sedi_counts.items():
        if code in available_sedi:
            present_vol += count
        else:
            absent_vol += count
            missing_counts[code] = count
            
    return {
        "present_rows": present_vol,
        "absent_rows": absent_vol,
        "missing_codes": missing_counts.most_common(5)
    }


def scrivi_log(
    wb,
    *,
    file_info: Dict[str, str],
    stats: Dict[str, int],
    quality_checks: Dict[str, int],
    rows_filtered: List["CleanRow"],
    emilia_codes: set,
    special_map: Dict[str, str],
    start_row_templates: int = 5,
    anno_rif: int = 2025,
    pension_min: float = 603.4,
    fasce_importo: Optional[List[float]] = None,
):
    """
    Scrive un riepilogo nel foglio esistente 'Log' con formattazione potenziata.
    """
    if SHEET_LOG not in wb.sheetnames:
        return

    ws = wb[SHEET_LOG]
    _clear_log_sheet(ws)
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD") # Blue
    
    subhead_font = Font(bold=True, italic=True)
    subhead_fill = PatternFill("solid", fgColor="DCE6F1") # Light Blue
    
    warn_fill = PatternFill("solid", fgColor="FFF2CC") # Yellow
    good_fill = PatternFill("solid", fgColor="E2EFDA") # Green
    bad_fill = PatternFill("solid", fgColor="FCE4D6") # Red
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def write_pair(r, label, value, fill=None, is_header=False, wrap_value=True):
        c1 = ws.cell(row=r, column=1)
        c2 = ws.cell(row=r, column=2)
        c1.value = label
        c2.value = value
        
        c1.border = thin_border
        c2.border = thin_border
        
        if is_header:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        elif fill:
            c1.fill = fill
            c2.fill = fill
        
        # Alignment
        c1.alignment = Alignment(vertical='center', wrap_text=True) # Wrap label too
        c2.alignment = Alignment(vertical='center', wrap_text=wrap_value)
            
        return r + 1

    r = 1
    # Placeholder per Sommario Esecutivo (scritto alla fine)
    summary_row = r
    r += 2 # Lasciamo spazio per l'esito
    
    r = write_pair(r, "📄 LOG ESECUZIONE SINDRINN", "", is_header=True)
    r += 1
    
    # Global Check Flag
    global_status_ok = True

    # File Info & Parametri
    r = write_pair(r, "📂 Input & Configurazioni", "", fill=subhead_fill)
    for k, v in file_info.items():
        r = write_pair(r, k, v)
    
    # Liste Codici (Espandibili)
    r = write_pair(r, "📋 Liste Codici & Mapping", "", fill=subhead_fill)
    
    # Emilia Codes
    emilia_sorted = sorted(list(emilia_codes))
    emilia_str = ", ".join(emilia_sorted)
    r = write_pair(r, f"Codici Emilia Ammessi ({len(emilia_sorted)})", emilia_str)
    
    # Special Map
    if special_map:
        map_items = [f"{k}->{v}" for k, v in special_map.items()]
        map_str = "; ".join(map_items)
        r = write_pair(r, f"Mapping Speciali ({len(map_items)})", map_str)
    else:
        write_pair(r, "Mapping Speciali", "Nessuno")
    r += 1

    # Parametri Calcolo fasce
    # Parametri Calcolo fasce & Distribuzione
    if fasce_importo:
        r = write_pair(r, "EXTRA: Verifica Distribuzione per Importo", f"Basata su Minimo {pension_min} €", is_header=True)
        r += 1
        
        # Calcolo distribuzione in memoria (con split Genere)
        counts_m = Counter()
        counts_f = Counter()
        
        for row in rows_filtered:
            val = it_text_to_float(row.importo_lordo_text)
            target_idx = -1
            
            if val is not None:
                # Logica _fascia_importo replicata localmente per sicurezza
                found = False
                for i, soglia in enumerate(fasce_importo):
                    if val <= soglia:
                        target_idx = i
                        found = True
                        break
                if not found:
                    target_idx = len(fasce_importo) # Ultima fascia (over max)
            
            if target_idx != -1:
                if row.genere == 'M':
                    counts_m[target_idx] += 1
                else:
                    counts_f[target_idx] += 1
            # else: row skipped/invalid import logic, treated elsewhere in quality checks

        # Scrittura Tabella
        # Intestazione custom
        c1 = ws.cell(row=r, column=1)
        c2 = ws.cell(row=r, column=2)
        c1.value = "Fascia / Soglia"
        c2.value = "Limite (€) | M | F | TOT"
        c1.fill = subhead_fill
        c2.fill = subhead_fill
        c1.font = subhead_font
        c2.font = subhead_font
        c1.alignment = Alignment(horizontal='center')
        c2.alignment = Alignment(horizontal='center')
        r += 1
        
        grand_tot_calc = 0

        for i, soglia in enumerate(fasce_importo):
            fascia_name = f"Fascia {i}"
            mult_desc = f"{i+1}x Minimo" if i < 6 else "Massimale"
            label = f"{fascia_name} (≤ {mult_desc})"
            
            m = counts_m[i]
            f = counts_f[i]
            tot = m + f
            grand_tot_calc += tot
            
            val_fmt = f"≤ {soglia:,.2f} €"
            content = f"{val_fmt}  |  M: {m}  F: {f}  ->  Tot: {tot}"
            
            r = write_pair(r, label, content)
        
        # Ultima fascia (oltre massimale)
        last_idx = len(fasce_importo)
        m_last = counts_m[last_idx]
        f_last = counts_f[last_idx]
        tot_last = m_last + f_last
        grand_tot_calc += tot_last
        
        label_last = f"Fascia {last_idx} (Oltre Massimale)"
        content_last = f"> {fasce_importo[-1]:,.2f} €  |  M: {m_last}  F: {f_last}  ->  Tot: {tot_last}"
        r = write_pair(r, label_last, content_last)
        
        # VERIFICA TOTALE
        # Confrontiamo la somma della tabella con i record filtrati totali
        # Dovrebbero combaciare se tutti hanno un importo valido.
        
        tot_expected = len(rows_filtered)
        diff = tot_expected - grand_tot_calc
        
        if diff == 0:
            check_msg = f"OK (Tabella: {grand_tot_calc} == Totale: {tot_expected})"
            check_fill = good_fill
        else:
            check_msg = f"ERRORE (Tabella: {grand_tot_calc} != Totale: {tot_expected}, Diff: {diff})"
            check_fill = bad_fill
            global_status_ok = False
            
        r += 1
        r = write_pair(r, "VERIFICA QUADRATURA IMPORTI", check_msg, fill=check_fill)
        r += 1

    # =========================================
    # EXTRA 5: Verifica Distribuzione Età/Genere
    # =========================================
    r = write_pair(r, "EXTRA: Verifica Distribuzione Età/Genere", f"Riferimento Anno {anno_rif}", is_header=True)
    r += 1

    # 1. Calcolo in memoria
    eta_m = Counter()
    eta_f = Counter()

    for row in rows_filtered:
        # Replicate logic from compila_distribuzione_per_genere_eta
        anno_val = int(row.anno_nascita) if (row.anno_nascita and row.anno_nascita.isdigit()) else None
        target = -1
        
        if anno_val is not None and anno_val >= 2000:
            target = 8
        else:
            eta = _calc_eta(row.anno_nascita, anno_rif=anno_rif)
            if eta is not None:
                # Helper _eta_bucket_idx is global
                if eta <= 60: target = 0
                elif 61 <= eta <= 65: target = 1
                elif 66 <= eta <= 70: target = 2
                elif 71 <= eta <= 75: target = 3
                elif 76 <= eta <= 80: target = 4
                elif 81 <= eta <= 85: target = 5
                elif 86 <= eta <= 90: target = 6
                else: target = 7 # >90

        if target != -1:
            if row.genere == 'M':
                eta_m[target] += 1
            else:
                eta_f[target] += 1
        else:
            eta_f[-1] += 1 # Invalid/Skipped

    # 2. Scrittura Tabella
    c1 = ws.cell(row=r, column=1)
    c2 = ws.cell(row=r, column=2)
    c1.value = "Fascia Età"
    c2.value = "Dettaglio | M | F | TOT"
    c1.fill = subhead_fill
    c2.fill = subhead_fill
    c1.font = subhead_font
    c2.font = subhead_font
    c1.alignment = Alignment(horizontal='center')
    c2.alignment = Alignment(horizontal='center')
    r += 1

    bucket_labels = [
        "Fino a 60 anni",
        "61 - 65 anni",
        "66 - 70 anni",
        "71 - 75 anni",
        "76 - 80 anni",
        "81 - 85 anni",
        "86 - 90 anni",
        "Oltre 90 anni",
        "Nati >= 2000 (Special)"
    ]

    grand_total_eta = 0
    # Buckets 0-8
    for b in range(9):
        label = bucket_labels[b] if b < len(bucket_labels) else f"Bucket {b}"
        m = eta_m[b]
        f = eta_f[b]
        tot = m + f
        grand_total_eta += tot
        
        content = f"M: {m}   F: {f}   ->   Tot: {tot}"
        r = write_pair(r, label, content)

    # Verifica Totale
    diff_eta = len(rows_filtered) - grand_total_eta
    if diff_eta == 0:
         msg_eta = f"OK (Tabella: {grand_total_eta} == Totale)"
         fill_eta = good_fill
    else:
         msg_eta = f"ERRORE (Tabella: {grand_total_eta} != Totale, Diff: {diff_eta})"
         fill_eta = bad_fill
         global_status_ok = False
    
    r += 1
    r = write_pair(r, "VERIFICA QUADRATURA ETÀ", msg_eta, fill=fill_eta)
    r += 1

    # Stats Elaborazione (Volume & Totali)
    
    # Calcolo totali M/F e Importi sui filtrati
    tot_m = sum(1 for row in rows_filtered if row.genere == 'M')
    tot_f = sum(1 for row in rows_filtered if row.genere == 'F')
    # Removed Totale Importo Lordo as requested
    
    r = write_pair(r, "⚙️ Statistiche Elaborazione (Globali)", "", fill=subhead_fill)
    r = write_pair(r, "📥 Record Totali Letti", stats['total'])
    
    filt_msg = f"{stats['filtered']} (su {stats['total']})"
    fill_filt = bad_fill if stats['filtered'] == 0 else good_fill
    r = write_pair(r, "✅ Record Totali Emilia (Filtrati)", filt_msg, fill=fill_filt)

    r = write_pair(r, "👥 Distribuzione Genere", f"M: {tot_m} | F: {tot_f}")

    norm_msg = f"{stats['normalized']} (Normalizzati XX00)" if stats['normalized'] > 0 else "0"
    r = write_pair(r, "🔄 Normalizzazioni CP", norm_msg, fill=warn_fill if stats['normalized']>0 else None)
    
    spec_msg = f"{stats['special']} (da Mappa Speciale)"
    r = write_pair(r, "🌟 Casi Speciali", spec_msg, fill=warn_fill if stats['special']>0 else None)
    
    r = write_pair(r, "⏹️ Invariati", stats['unchanged'])
    r += 1
    
    # Qualità Dati
    r = write_pair(r, "⚠️ Controllo Qualità (sui filtrati)", "", fill=subhead_fill)
    r = write_pair(r, "Genere Invalid (NO M/F)", quality_checks['inv_gen'], fill=bad_fill if quality_checks['inv_gen']>0 else good_fill)
    r = write_pair(r, "Anno Invalid (NO 4 cifre)", quality_checks['inv_ann'], fill=bad_fill if quality_checks['inv_ann']>0 else good_fill)
    r = write_pair(r, "Categoria Invalid (Non num)", quality_checks['inv_cat'], fill=bad_fill if quality_checks['inv_cat']>0 else good_fill)
    r = write_pair(r, "Sede Invalid (Non num)", quality_checks['miss_sede'], fill=bad_fill if quality_checks['miss_sede']>0 else good_fill)
    r = write_pair(r, "Importo Invalid (Err conv)", quality_checks['inv_lordo'], fill=bad_fill if quality_checks['inv_lordo']>0 else good_fill)
    r += 1

    # Copertura Fogli (Audit Dettagliato)
    r = write_pair(r, "📊 Audit Copertura & Posizionamento", "", fill=subhead_fill)
    
    readable_names = {
        SHEET_NAME_DEFAULT: "Dati Base (Da Eliminare)",
        SHEET_ETA: "Genere/Età",
        SHEET_CATEG: "Per Categoria",
        SHEET_IMPORTI: "Importi",
        SHEET_GENERE: "Associati per Genere"
    }
    
    sheets_to_check = [SHEET_NAME_DEFAULT, SHEET_ETA, SHEET_CATEG, SHEET_IMPORTI, SHEET_GENERE]

    for sname in sheets_to_check:
        display_name = readable_names.get(sname, sname)
        if sname in wb.sheetnames:
            ws_sheet = wb[sname]
            audit = _audit_sheet_coverage(ws_sheet, rows_filtered, start_row=start_row_templates)
            ok = audit['present_rows']
            ko = audit['absent_rows']
            missing = audit['missing_codes']
            
            # Status line
            if ok + ko == 0:
                 val = "⚠️ Template/Dati Vuoti"
                 fill_val = warn_fill
            else:
                icon = "🟢" if ko == 0 else "🟠"
                if ko == 0:
                     val = f"{icon} Posizionati: {ok} | Persi: 0"
                     fill_val = good_fill
                else:
                     pct = (ko / (ok+ko)) * 100
                     val = f"{icon} Posizionati: {ok} | PERSI: {ko} ({pct:.1f}%)"
                     fill_val = warn_fill
            
            r = write_pair(r, f"Foglio: {display_name}", val, fill=fill_val)
            
            # --- DETTAGLI POSIZIONAMENTO (Dati verificati) ---
            if ok > 0:
                # Recuperiamo le sedi disponibili per filtrare
                sede_to_row = _build_sede_row_map(ws_sheet, start_row=start_row_templates, code_col=1)
                
                # Sottogruppo di record che SONO davvero finiti in questo foglio
                placed_rows = [row for row in rows_filtered 
                               if _digits_only(str(row.sede_cp or "")).zfill(4)[-4:] in sede_to_row]
                
                if sname == SHEET_ETA:
                    # Breakdown Età
                    buk_counts = Counter()
                    for row in placed_rows:
                        eta = _calc_eta(row.anno_nascita, anno_rif)
                        if eta is not None:
                            idx = _eta_bucket_idx(eta)
                            buk_name = f"Fascia {idx+1}" # (es. <60, 61-65...)
                            # Mapping nomi fasce approssimativo per leggibilità
                            labels = ["≤60", "61-65", "66-70", "71-75", "76-80", "81-85", "86-90", "≥91"]
                            if 0 <= idx < len(labels): buk_name = labels[idx]
                            buk_counts[buk_name] += 1
                        else:
                            buk_counts["N/D"] += 1
                    
                    # Top 4 fasce + M/F
                    top_b = buk_counts.most_common(5)
                    desc = " | ".join([f"{k}: {v}" for k, v in top_b])
                    pf = sum(1 for x in placed_rows if x.genere=='F')
                    pm = len(placed_rows) - pf
                    r = write_pair(r, "   ↳ Distribuzione (Piazzati)", f"F: {pf}, M: {pm} || Età: {desc}")
                
                elif sname == SHEET_CATEG:
                    # Breakdown Categoria
                    cat_counts = Counter(r.categoria for r in placed_rows)
                    top_c = cat_counts.most_common(5)
                    desc = ", ".join([f"Cat {k}: {v}" for k, v in top_c])
                    r = write_pair(r, "   ↳ Top Categorie (Piazzati)", desc)
                
                elif sname == SHEET_IMPORTI and fasce_importo:
                    # Breakdown Importi
                    # Logica fasce semplice (copiata da _imp_bucket_idx se esistesse, ma qui replico in breve)
                    # Fasce: < F0, F0-F1, ...
                    imp_counts = Counter()
                    thresholds = sorted(fasce_importo)
                    for row in placed_rows:
                        val = it_text_to_float(row.importo_lordo_text) or 0.0
                        # Trova fascia
                        found = False
                        for i, th in enumerate(thresholds):
                            if val <= th:
                                imp_counts[f"≤{int(th)}"] += 1
                                found = True
                                break
                        if not found:
                             imp_counts[f">{int(thresholds[-1])}"] += 1
                    
                    # Ordine visualizzazione (custom sort?)
                    # Most common è ok
                    desc = " | ".join([f"{k}: {v}" for k, v in imp_counts.most_common(5)])
                    r = write_pair(r, "   ↳ Fasce Importo (Piazzati)", desc)

            # Detail line for missing
            if missing:
                missing_str = ", ".join([f"{c}({n})" for c, n in missing])
                r = write_pair(r, "   ↳ Sedi Mancanti (Top 5)", f"Codici non in template: {missing_str}", fill=bad_fill)

        else:
            write_pair(r, f"Foglio: {display_name}", "❌ NON TROVATO", fill=bad_fill)
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 80

    # =========================================
    # RICONCILIAZIONE SEDE x SEDE (Audit Completo 4 Fogli)
    # =========================================
    r += 1
    r = write_pair(r, "🔍 RICONCILIAZIONE COMPLETA (Memoria vs Fogli)", "", fill=subhead_fill)
    r = write_pair(r, "Legenda", "✅=Corrisponde | ⚠️=Differenza (Es. dati mancanti/invalidi) | ❌=Sede mancante nel foglio")

    # 1. Definiamo i fogli da controllare e le loro logiche
    # (NomeFoglio, LabelBreve) - tutti scrivono count da col 3 in poi
    sheets_config = [
        (SHEET_GENERE, "Gen"),
        (SHEET_ETA, "Eta"),
        (SHEET_CATEG, "Cat"),
        (SHEET_IMPORTI, "Imp")
    ]
    
    # Pre-calcola mappe (Sede -> Row) per tutti i fogli
    maps = {}
    worksheets = {}
    for sname, label in sheets_config:
        if sname in wb.sheetnames:
            ws_s = wb[sname]
            worksheets[sname] = ws_s
            maps[sname] = _build_sede_row_map(ws_s, start_row=start_row_templates, code_col=1)
        else:
            maps[sname] = None # Foglio mancante

    # 2. Conteggi attesi (Memoria: Totale Record per Sede)
    counts_mem = Counter()
    for row in rows_filtered:
        cd = _digits_only(str(row.sede_cp or "")).zfill(4)[-4:]
        if cd: counts_mem[cd] += 1
        
    sorted_sedi = sorted(counts_mem.keys())
    
    # 3. Iterazione per Sede
    for sede in sorted_sedi:
        mem_val = counts_mem[sede]
        
        results = []
        all_ok = True
        
        for sname, label in sheets_config:
            if maps[sname] is None:
                results.append(f"{label}:🚫(NoSheet)")
                continue
                
            row_idx = maps[sname].get(sede)
            if not row_idx:
                results.append(f"{label}:❌") # Sede non nel template
                all_ok = False
            else:
                # Somma riga (Col 3 -> End)
                ws_s = worksheets[sname]
                row_sum = 0
                # Ottimizzazione: leggi celle da col 3 fino a max_column (non troppo oltre)
                # Max colonne ragionevole: 50?
                # Per essere sicuri, iteriamo sulle celle con valori
                # iter_cols è lento riga per riga, meglio cell direct
                # Assumiamo max 60 colonne (Importi/Età possono essere larghe)
                for c_idx in range(3, 80): 
                    val = ws_s.cell(row=row_idx, column=c_idx).value
                    if isinstance(val, (int, float)):
                        row_sum += int(val)
                
                diff = mem_val - row_sum
                if diff == 0:
                    results.append(f"{label}:✅")
                else:
                    # Se diff > 0, mancano record (es. invalidi)
                    results.append(f"{label}:⚠️{-diff}")
                    all_ok = False
        
        status_str = " | ".join(results)
        final_msg = f"Tot: {mem_val} -> {status_str}"
        fill_cell = None if all_ok else bad_fill
        
        write_pair(r, f"Sede {sede}", final_msg, fill=fill_cell)
        r += 1

    # =========================================
    # Modifiche del 28/01/2026 - Audit Gruppi Categorie
    # =========================================
    r += 1
    r = write_pair(r, "EXTRA: Modifiche del 28/01/2026", "Riferito al foglio per_categoria", is_header=True)
    r += 1
    
    # 1. Calcolo atteso in memoria
    mem_group_counts = Counter()
    for row in rows_filtered:
        grp = _get_group_categoria(row.categoria)
        mem_group_counts[grp] += 1
        
    # 2. Lettura dal foglio 'Distribuzione_per_categoria'
    sheet_cat_name = SHEET_CATEG
    excel_group_counts = Counter()
    
    if sheet_cat_name in wb.sheetnames:
        ws_cat = wb[sheet_cat_name]
        
        # Mappa colonne per gruppo
        group_cols = {}
        for g in range(1, 9):
            if g <= 3:
                c = 3 + (g - 1) * 3
            else:
                c = 15 + (g - 4) * 3
            group_cols[g] = (c, c+1) # (Maschi, Femmine)
            
        cat_map = _build_sede_row_map(ws_cat, start_row=start_row_templates, code_col=1)
        
        for sede_code, row_idx in cat_map.items():
            for g in range(1, 9):
                c_m, c_f = group_cols[g]
                val_m = ws_cat.cell(row=row_idx, column=c_m).value
                val_f = ws_cat.cell(row=row_idx, column=c_f).value
                
                try: m = int(val_m) if val_m is not None else 0
                except: m = 0
                try: f = int(val_f) if val_f is not None else 0
                except: f = 0
                
                excel_group_counts[g] += (m + f)
    else:
        r = write_pair(r, "Errore", f"Foglio {sheet_cat_name} non trovato", fill=bad_fill)

    # 3. Scrittura Confronto con Subtotali
    r = write_pair(r, "Gruppo", "Memoria vs Excel (Diff)", fill=subhead_fill)
    
    # Helper per riga confronto
    def _write_check_line(row_idx, label, val_mem, val_xl, indent=False):
        diff = val_mem - val_xl
        status_icon = "✅" if diff == 0 else "❌"
        diff_str = f" ({diff:+d})" if diff != 0 else ""
        msg = f"Mem: {val_mem} | XL: {val_xl} -> {status_icon}{diff_str}"
        fill_c = None if diff == 0 else bad_fill
        
        lbl = f"   {label}" if indent else label
        return write_pair(row_idx, lbl, msg, fill=fill_c)

    # Accumulatori
    sub_1_3_mem = 0
    sub_1_3_xl = 0
    sub_4_8_mem = 0
    sub_4_8_xl = 0
    
    # Ciclo Gruppi 1-3
    for g in range(1, 4):
        m = mem_group_counts[g]
        x = excel_group_counts[g]
        sub_1_3_mem += m
        sub_1_3_xl += x
        r = _write_check_line(r, f"Gruppo {g}", m, x, indent=True)
    
    # Subtotale 1-3
    r = _write_check_line(r, "TOTALE Gruppi 1-3", sub_1_3_mem, sub_1_3_xl)
    r += 1 # Spaziatura
    
    # Ciclo Gruppi 4-8
    for g in range(4, 9):
        m = mem_group_counts[g]
        x = excel_group_counts[g]
        sub_4_8_mem += m
        sub_4_8_xl += x
        r = _write_check_line(r, f"Gruppo {g}", m, x, indent=True)
        
    # Subtotale 4-8
    r = _write_check_line(r, "TOTALE Gruppi 4-8", sub_4_8_mem, sub_4_8_xl)
    r += 1 # Spaziatura
        
    # Totale Generale
    tot_mem = sub_1_3_mem + sub_4_8_mem
    tot_xl = sub_1_3_xl + sub_4_8_xl
    r = _write_check_line(r, "TOTALE GENERALE", tot_mem, tot_xl)
    
    diff_tot = tot_mem - tot_xl
    if diff_tot != 0: global_status_ok = False
    
    # =========================================
    # SCRITTURA SOMMARIO ESECUTIVO (IN CIMA)
    # =========================================
    if global_status_ok:
        msg = "ESITO: OK (Tutti i controlli passati)"
        fill = good_fill
    else:
        msg = "ESITO: ATTENZIONE (Trovate differenze/errori)"
        fill = bad_fill
    
    # Scrittura in riga summary_row (fissata all'inizio)
    # Usiamo write_pair ma forzando la riga
    c1 = ws.cell(row=summary_row, column=1)
    c2 = ws.cell(row=summary_row, column=2)
    c1.value = "STATO ANALISI"
    c2.value = msg
    
    # Style Big
    big_font = Font(bold=True, size=14)
    c1.font = big_font
    c2.font = big_font
    c1.fill = fill
    c2.fill = fill
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c2.alignment = Alignment(horizontal='center', vertical='center')

    # =========================================
    # EXTRA 2: VALIDAZIONE SOMME IMPORTI (Trattenuta/Lordo)
    # =========================================
    r += 2
    r = write_pair(r, "EXTRA: Validazione Importi", "Somme Totali (Memoria vs Excel)", is_header=True)
    r += 1
    
    # Calcolo Memoria
    tot_trat_mem = sum(trattenuta_to_float(r.trattenuta_text) or 0.0 for r in rows_filtered)
    tot_lordo_mem = sum(it_text_to_float(r.importo_lordo_text) or 0.0 for r in rows_filtered)
    
    # Lettura Excel (somma da colonna F e G del foglio default o simile? No, meglio usare FileRipu_E.R_Da_BD se c'è, o Default)
    # Usiamo 'FileRipu_E.R_Da_BD' (col 5 e 6) se esiste, altrimenti SHEET_NAME_DEFAULT (col 6 e 7)
    
    sheet_src = "FileRipu_E.R_Da_BD" if "FileRipu_E.R_Da_BD" in wb.sheetnames else SHEET_NAME_DEFAULT
    col_tr = 5 if sheet_src == "FileRipu_E.R_Da_BD" else 6
    col_lor = 6 if sheet_src == "FileRipu_E.R_Da_BD" else 7
    
    tot_trat_xl = 0.0
    tot_lordo_xl = 0.0
    
    if sheet_src in wb.sheetnames:
        ws_src = wb[sheet_src]
        max_r_src = ws_src.max_row or 2
        # Start row 2 (header in 1) or 5 checks
        start_check = 2 if sheet_src == "FileRipu_E.R_Da_BD" else 5
        
        for i in range(start_check, max_r_src + 1):
             # Skip formula footer if present
             val_1 = ws_src.cell(row=i, column=1).value
             if str(val_1).lower().startswith("totale"): break
             
             t_val = ws_src.cell(row=i, column=col_tr).value
             l_val = ws_src.cell(row=i, column=col_lor).value
             
             # Clean and sum
             try: 
                if isinstance(t_val, (int, float)): tot_trat_xl += float(t_val)
                else: tot_trat_xl += (it_text_to_float(str(t_val)) or 0.0)
             except: pass
             
             try:
                if isinstance(l_val, (int, float)): tot_lordo_xl += float(l_val)
                else: tot_lordo_xl += (it_text_to_float(str(l_val)) or 0.0)
             except: pass
    
    # Check Trattenuta
    diff_tr = abs(tot_trat_mem - tot_trat_xl)
    status_tr = "✅" if diff_tr < 1.0 else "⚠️" # Tolleranza 1 euro arrotondamenti
    msg_tr = f"Mem: {tot_trat_mem:,.2f} | XL: {tot_trat_xl:,.2f} -> {status_tr} (Diff: {diff_tr:.2f})"
    r = write_pair(r, "Totale Trattenute", msg_tr, fill=None if diff_tr < 1.0 else warn_fill)
    if diff_tr >= 1.0: global_status_ok = False
    
    # Check Lordo
    diff_lo = abs(tot_lordo_mem - tot_lordo_xl)
    status_lo = "✅" if diff_lo < 1.0 else "⚠️"
    msg_lo = f"Mem: {tot_lordo_mem:,.2f} | XL: {tot_lordo_xl:,.2f} -> {status_lo} (Diff: {diff_lo:.2f})"
    r = write_pair(r, "Totale Lordo", msg_lo, fill=None if diff_lo < 1.0 else warn_fill)
    if diff_lo >= 1.0: global_status_ok = False

    # =========================================
    # EXTRA 3: INTEGRITÀ FORMULE (Check esistenza =SUM)
    # =========================================
    r += 1
    r = write_pair(r, "EXTRA: Integrità Formule", "Verifica presenza =SUM", is_header=True)
    r += 1
    
    sheets_to_audit = [SHEET_CATEG, SHEET_ETA, SHEET_IMPORTI, SHEET_GENERE]
    broken_formulas = 0
    
    for sname in sheets_to_audit:
        if sname not in wb.sheetnames: continue
        ws_s = wb[sname]
        # Check Totale Riga for first few filled rows
        row_map = _build_sede_row_map(ws_s, start_row=start_row_templates, code_col=1)
        checked = 0
        errors_in_sheet = 0
        
        # Determine Tot column based on sheet logic
        tot_col = -1
        if sname == SHEET_GENERE: tot_col = 5
        elif sname == SHEET_ETA: tot_col = 29 # AC
        elif sname == SHEET_CATEG: tot_col = 32 # AF
        elif sname == SHEET_IMPORTI: 
            # Last col depends on buckets... assuming default 8 buckets -> 3+24+2 = 29 (AC)? 
            # No, logic in compila importi is dynamic. Let's skip Importi precise col check or assume standard 32?
            # Let's check a standard cell that MUST be formula in all sheets: The first group total?
            # Better: Check Footer Row! Footer must be sums.
            pass
            
        # Let's verify Footer specifically, it's critical
        last_code_r = find_last_code_row(ws_s, start_row_templates)
        footer_r = last_code_r + 1
        # Check col 3 (C) in footer
        f_val = ws_s.cell(row=footer_r, column=3).value
        is_formula = str(f_val).startswith("=") if f_val else False
        
        status_form = "✅ OK" if is_formula else "❌ MANCANTE"
        fill_form = None if is_formula else bad_fill
        if not is_formula: 
            broken_formulas += 1
            global_status_ok = False
            
        write_pair(r, f"Footer Formule ({sname})", status_form, fill=fill_form)
        r += 1

    # =========================================
    # EXTRA 4: EDGE CASES (Fasce Residuali)
    # =========================================
    r += 1
    r = write_pair(r, "EXTRA: Warning Dati", "Controllo Fasce Estreme", is_header=True)
    r += 1
    
    # Oltre 90 anni
    over_90 = sum(1 for row in rows_filtered if _calc_eta(row.anno_nascita, anno_rif) is not None and _calc_eta(row.anno_nascita, anno_rif) >= 91)
    # Importo massimale (fascia 7)
    max_imp_cnt = 0
    if fasce_importo:
        last_thresh = fasce_importo[-1]
        max_imp_cnt = sum(1 for row in rows_filtered if (it_text_to_float(row.importo_lordo_text) or 0) > last_thresh)
        
    warn_90 = "⚠️" if over_90 > 0 else "✅"
    write_pair(r, f"Età ≥ 91 anni", f"{over_90} casi {warn_90}", fill=warn_fill if over_90 > 0 else None)
    r += 1
    
    warn_imp = "⚠️" if max_imp_cnt > 0 else "✅"
    if fasce_importo:
        write_pair(r, f"Importo > {int(fasce_importo[-1])}€", f"{max_imp_cnt} casi {warn_imp}", fill=warn_fill if max_imp_cnt > 0 else None)
    
    # RE-UPDATE SUMMARY AT TOP WITH NEW GLOBAL STATUS
    if not global_status_ok:
        c2 = ws.cell(row=summary_row, column=2)
        c2.value = "ESITO: ATTENZIONE (Trovate differenze/errori)"
        c2.fill = bad_fill



# =========================
# HELPER PER VALIDAZIONE (SNAPSHOT)
# =========================
def _get_sheet_snapshot(wb, sheet_name: str, start_row: int = 1, end_col: int = 50) -> Dict[Tuple[int, int], str]:
    """
    Legge un foglio e ritorna una mappa {(r, c): valore_str_pulito}.
    Ignora righe vuote o fuori dal range utile.
    
    CRITERIO DI SCANSIONE (Richiesta Utente):
    - Inizia da start_row.
    - Cerca in Colonna A l'ultimo codice sede valido (4 cifre).
    - Considera valide SOLO le righe fino a quel punto (escluso Footer/Totali).
    - Ignora formule ("=").
    """
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    snap = {}
    
    # 1. Identifica l'ultima riga dati valida
    effective_max_row = start_row - 1
    raw_max = ws.max_row or start_row
    
    # Scansioniamo la colonna A per trovare l'ultima riga con codice 4 cifre
    for r in range(start_row, raw_max + 1):
        val = ws.cell(row=r, column=1).value
        # Normalizza e verifica se è codice 4 cifre (richiesto da utente)
        code = _digits_only(str(val) if val else "")
        if len(code) == 4: 
            # Diciamo che se sembra un codice numerico valido (4 cifre), è una riga dati.
            # Se è "Totale" non è digits only (di solito).
            effective_max_row = r
        elif str(val).lower().strip().startswith("total"):
             # Se troviamo "Totale" esplicitamente, ci fermiamo PRIMA di questa riga?
             # No, loop continua, ma effective_max_row non si aggiorna.
             pass

    # 2. Leggi solo il range dati effettivo
    for r in range(start_row, effective_max_row + 1):
        for c in range(1, end_col + 1):
            val = ws.cell(row=r, column=c).value
            
            # Normalizzazioni base
            if val is None:
                s_val = ""
            else:
                s_val = str(val).strip()
            
            # Ignora formule
            if s_val.startswith("="):
                continue
                
            snap[(r, c)] = s_val # manteniamo case original? Utente vuole normalizzazione strong dopo.
            
    return snap


def _compare_snapshots(old_snap: Dict, new_snap: Dict) -> List[str]:
    diffs = []
    # Chiavi totali
    all_keys = set(old_snap.keys()) | set(new_snap.keys())
    
    def normalize_val(v):
        """
        Normalizza valore per confronto:
        - None, "", "0", 0 -> "0"
        - Numeri stringa "1.0", "1,0" -> float -> int se intero
        """
        if v is None:
            return 0.0
        s = str(v).strip().replace(",", ".")
        if not s:
            return 0.0
        try:
            return float(s)
        except:
            return s.lower() # Fallback testo lowercase

    for k in sorted(list(all_keys)):
        val_old = old_snap.get(k)
        val_new = new_snap.get(k)
        
        # Ignora header se per caso finiti qui (ma _get filtra per start_row)
        
        n_old = normalize_val(val_old)
        n_new = normalize_val(val_new)
        
        # Confronto
        match = False
        if isinstance(n_old, float) and isinstance(n_new, float):
            if abs(n_old - n_new) < 0.001:
                match = True
        else:
            if n_old == n_new:
                match = True
        
        if not match:
             from openpyxl.utils import get_column_letter
             addr = f"{get_column_letter(k[1])}{k[0]}"
             # Formatta messaggio user-friendly
             v_old_repr = val_old if val_old else "(vuoto)"
             v_new_repr = val_new if val_new else "(vuoto)"
             diffs.append(f"Cella {addr}: Nuovo '{v_new_repr}' != Orig '{v_old_repr}'")
                
    return diffs


def _mostra_risultati_validazione(results_data):
    """Mostra i risultati della validazione salvati in session_state"""
    st.markdown("---")
    st.subheader("📊 Risultati Validazione vs File Originale")
    
    validation_results = results_data.get("results", [])
    total_diffs = results_data.get("total_diffs", 0)
    total_sheets = len(validation_results)
    
    # Contiamo errori reali e warning
    errors = sum(1 for r in validation_results if r["status"] == "error")
    warnings = sum(1 for r in validation_results if r["status"] == "warning")
    
    if total_sheets == 0:
        st.warning("⚠️ Nessun foglio trovato per il confronto (nomi fogli diversi?).")
    else:
        # Sommario
        if total_diffs == 0:
            st.success(f"🏆 **VALIDAZIONE PERFETTA** su {total_sheets} fogli!")
        elif errors == 0 and warnings > 0:
             st.warning(f"⚠️ **Trovate {total_diffs} differenze** (Probabili correzioni VBA fixato)")
        else:
            st.error(f"❌ Trovate **{total_diffs}** differenze totali ({errors} fogli con errori critici)")
        
        # Dettagli
        for result in validation_results:
            if result["status"] == "success":
                st.success(f"✅ **{result['sheet']}**: MATCH 100% (Dati identici)")
            elif result["status"] == "warning":
                st.warning(f"⚠️ **{result['sheet']}**: {result['count']} differenze (Miglioria rispetto al VBA)")
                with st.expander(f"🔍 Dettagli correzioni - {result['sheet']}", expanded=False):
                    msg_info = "Miglioria nota rispetto al VBA."
                    if result['sheet'] == SHEET_GENERE:
                        msg_info = "Correzione Bug VBA: le righe erano scambiate, ora sono corrette."
                    elif result['sheet'] == SHEET_IMPORTI:
                        msg_info = "Correzione Parametri: Python usa soglie aggiornate (2025)."
                    else:
                        msg_info = "Miglioria Tecnica: Sostituite formule Excel con valori statici (M+F) per visibilità immediata."
                    st.info(msg_info)
                    for i, diff in enumerate(result['details'][:100], 1):
                        st.text(f"{i}. {diff}")
                    if len(result['details']) > 100:
                        st.caption(f"... e altre {len(result['details']) - 100} differenze nascoste")
            else:
                st.error(f"❌ **{result['sheet']}**: {result['count']} differenze trovate")
                with st.expander(f"🔍 Dettagli differenze - {result['sheet']}", expanded=False):
                    for i, diff in enumerate(result['details'][:100], 1):
                        st.text(f"{i}. {diff}")
                    if len(result['details']) > 100:
                        st.caption(f"... e altre {len(result['details']) - 100} differenze nascoste")
    
    st.markdown("---")


def _genera_html_risultati(validation_results, total_diffs):
    """Genera HTML compatto per la visualizzazione dei risultati"""
    
    total_sheets = len(validation_results)
    errors = sum(1 for r in validation_results if r["status"] == "error")
    warnings = sum(1 for r in validation_results if r["status"] == "warning")
    
    # Sommario
    if total_diffs == 0:
        html = '<div style="padding: 15px; background: #d4edda; border-left: 4px solid #28a745; margin: 10px 0;">'
        html += '<h4 style="color: #155724; margin: 0;">🏆 VALIDAZIONE PERFETTA</h4>'
        html += f'<p style="color: #155724; margin: 5px 0;">Tutti i {total_sheets} fogli sono identici!</p>'
        html += '</div>'
    elif errors == 0:
        html = '<div style="padding: 15px; background: #fff3cd; border-left: 4px solid #ffc107; margin: 10px 0;">'
        html += '<h4 style="color: #856404; margin: 0;">⚠️ DIFFERENZE ATTESE (Fix VBA)</h4>'
        html += f'<p style="color: #856404; margin: 5px 0;">{total_diffs} differenze su {warnings} fogli (Correzioni)</p>'
        html += '</div>'
    else:
        html = '<div style="padding: 15px; background: #f8d7da; border-left: 4px solid #dc3545; margin: 10px 0;">'
        html += '<h4 style="color: #721c24; margin: 0;">❌ DIFFERENZE CRITICHE</h4>'
        html += f'<p style="color: #721c24; margin: 5px 0;">{total_diffs} differenze trovate</p>'
        html += '</div>'
    
    # Dettagli per foglio
    for result in validation_results:
        sheet_name = result['sheet']
        status = result['status']
        count = result['count']
        
        if status == "success":
            html += f'<div style="padding: 10px; background: #d1ecf1; border-left: 3px solid #17a2b8; margin: 5px 0;">'
            html += f'<strong>✅ {sheet_name}</strong>: MATCH 100%'
            html += '</div>'
        elif status == "warning":
            html += f'<div style="padding: 10px; background: #fff3cd; border-left: 3px solid #ffc107; margin: 5px 0;">'
            html += f'<strong>⚠️ {sheet_name}</strong>: {count} differenze (Miglioria)'
            html += '</div>'
        else:
            html += f'<div style="padding: 10px; background: #f8d7da; border-left: 3px solid #dc3545; margin: 5px 0;">'
            html += f'<strong>❌ {sheet_name}</strong>: {count} differenze'
            html += '</div>'
    
    return html


# =========================
# RUNNER
# =========================
def compila_distribuzione_per_importo_v3(ws, rows: List["CleanRow"], fasce: List[float] = None, start_row: int = 5):
    """
    Compila il foglio distribuzione per importo (V3).
    """
    from openpyxl.cell.cell import MergedCell
    
    num_fasce = len(fasce) + 1 if fasce else 8

    agg: Dict[str, List[List[int]]] = {}

    for r in rows:
        sede = _digits_only(str(r.sede_cp or "")).zfill(4)[-4:]
        if not sede: continue
        imp = it_text_to_float(r.importo_lordo_text)
        if imp is None: continue
        fascia = _fascia_importo(imp, fasce)
        if fascia >= num_fasce: fascia = num_fasce - 1
        
        if sede not in agg:
            agg[sede] = [[0, 0] for _ in range(num_fasce)]
        if r.genere == "M": agg[sede][fascia][0] += 1
        elif r.genere == "F": agg[sede][fascia][1] += 1

    # 2) Pulizia
    _clear_importi_output(ws, start_row=start_row, start_col=3, end_col=3 + (num_fasce*3) + 2)

    # 3) Scrittura
    max_row = ws.max_row or start_row
    from openpyxl.utils import get_column_letter
    cols_m = [3 + i * 3 for i in range(num_fasce)]
    cols_f = [4 + i * 3 for i in range(num_fasce)]
    sc = 3 + num_fasce * 3

    for r in range(start_row, max_row + 1):
        cell_value = ws.cell(row=r, column=1).value
        code = _norm_code_from_cell(cell_value)
        if not code:
            continue

        buckets = agg.get(code)
        col = 3
        for idx in range(num_fasce):
            if buckets:
                m = buckets[idx][0]
                f = buckets[idx][1]
                if not isinstance(ws.cell(row=r, column=col), MergedCell):
                    ws.cell(row=r, column=col).value = m
                if not isinstance(ws.cell(row=r, column=col + 1), MergedCell):
                    ws.cell(row=r, column=col + 1).value = f

            if not isinstance(ws.cell(row=r, column=col + 2), MergedCell):
                fm = get_column_letter(col)
                ff = get_column_letter(col + 1)
                ws.cell(row=r, column=col + 2).value = f"={fm}{r}+{ff}{r}"
            col += 3

        refs_m = [f"{get_column_letter(c)}{r}" for c in cols_m]
        refs_f = [f"{get_column_letter(c)}{r}" for c in cols_f]
        ws.cell(row=r, column=sc).value = "=" + "+".join(refs_m)
        ws.cell(row=r, column=sc + 1).value = "=" + "+".join(refs_f)
        ws.cell(row=r, column=sc + 2).value = f"={get_column_letter(sc)}{r}+{get_column_letter(sc + 1)}{r}"

    # 4) Footer (Formulas)
    last_code_r = find_last_code_row(ws, start_row)
    footer_row = last_code_r + 1
    
    # Sum all columns from 3 up to sc+2
    last_col_idx = (3 + num_fasce * 3) + 2
    for c in range(3, last_col_idx + 1):
        _write_sum_formula(ws, footer_row, c, start_row, last_code_r, c)

# =========================
# RUNNER
# =========================
def run(file_txt: Optional[Path], file_xlsx: Path, mapping_special: str, out_dir: Path, 
        file_banca_dati: Optional[Path] = None,
        compare_vba: bool = False, normalize_emilia: bool = True,
        report_year: int = 2025, pension_min: float = 603.4, 
        # coeff_maggiorazione REMOVED from UI, kept here for compatibility but overwritten logic
        coeff_maggiorazione: float = 1.022, 
        aliquota_1: float = 0.50, aliquota_2: float = 0.40, aliquota_3: float = 0.35,
        import_brackets: str = "", template_start_row: int = 5,
        enable_maggiorazione: bool = False) -> List[Path]:
    """
    Runner con parametri.
    """
    # Determinazione coefficiente effettivo
    # Se 'enable_maggiorazione' è True (Legacy), usiamo 1.022.
    # Se False (Default), usiamo 1.0 (calcolo puro).
    effective_coeff = 1.022 if enable_maggiorazione else 1.0
    
    st.write(f"DEBUG: Parametro compare_vba ricevuto: {compare_vba}")
    # Gestione eventuale file banca dati (xlsx o slk)
    banca_out: Optional[Path] = None
    if file_banca_dati is not None:
        banca_src = Path(file_banca_dati)
        out_dir.mkdir(parents=True, exist_ok=True)
        if banca_src.suffix.lower() == ".slk":
            banca_out = out_dir / f"{banca_src.stem}.xlsx"
            banca_out = convert_slk_to_xlsx(banca_src, banca_out)
        else:
            banca_out = out_dir / banca_src.name
            shutil.copy(banca_src, banca_out)

    file_xlsx_path = Path(file_xlsx)
    file_txt_path = Path(file_txt) if file_txt else None
    # Se l'utente carica il file banca dati, ha la precedenza sul TXT
    use_banca = banca_out is not None

    if compare_vba:
        st.write("🔍 Debug: Validazione attivata. Inizio analisi...")

    if not file_xlsx_path.exists():
        raise FileNotFoundError(f"File XLSX non trovato: {file_xlsx_path}")
    if not use_banca and (file_txt_path is None or not file_txt_path.exists()):
        raise FileNotFoundError(f"File TXT non trovato: {file_txt}")

    # Parse brackets
    brackets = _parse_brackets(import_brackets)
    if brackets and len(brackets) != 7:
        st.error(f"Errore: import_brackets deve contenere esattamente 7 valori. Trovati {len(brackets)}. Uso default.")
        brackets = None
    
    if not brackets:
        # Calcolo dinamico basato sul minimo (richiesta utente)
        # Fasce: 1x, 2x, 3x, 4x, 5x, 6x Minimo.
        # L'ultima soglia storica era 3946.18 (circa 6.5 volte il minimo 2021). 
        # Manteniamo 3946.18 fisso o lo scaliamo? 
        # Nel dubbio, ricalcoliamo le prime 6 e teniamo la 7ima fissa se è superiore, altrimenti la adattiamo?
        # Per coerenza con "doppio, triplo", generiamo i multipli. 
        # La 7ima soglia è un massimale specifico. Se non abbiamo info, lasciamo il valore storico hardcoded o lo leghiamo al minimo?
        # Manteniamo la soglia fissa originale per l'ultimo scaglione per compatibilità, a meno che i multipli non la superino.
        
        # Generiamo le prime 6 fasce dinamicamente
        multiples = [1, 2, 3, 4, 5, 6]
        dyn_brackets = [round(pension_min * m, 2) for m in multiples]
        
        # Aggiungiamo la 7ima. Storicamente 3946.18.
        # Se il minimo è cambiato drasticamente, questo valore potrebbe non avere senso.
        # Tuttavia, per non rompere la logica esistente se non richiesto esplicitamente, appendiamo il valore fisso (o lo ricalcoliamo se necessario).
        # Useremo il vecchio valore hardcoded come default per l'ultimo bucket.
        dyn_brackets.append(3946.18) 
        
        real_brackets = dyn_brackets
    else:
        real_brackets = brackets

    # ================= VALIDAZIONE PRE-EXEC =================
    snapshots_old = {}
    if compare_vba:
        try:
            st.info("🟡 Lettura dati file Input (baseline)...")
            wb_old = load_workbook(file_xlsx_path, data_only=False)
            # Case-insensitive exclusion set
            exclude_lower = {n.lower() for n in {SHEET_NAME_DEFAULT, SHEET_LOG}}
            
            found_sheets = 0
            for sname in wb_old.sheetnames:
                if sname.lower() in exclude_lower: continue
                # Usa template_start_row per ignorare header
                snapshots_old[sname] = _get_sheet_snapshot(wb_old, sname, start_row=template_start_row)
                found_sheets += 1
            wb_old.close()
            st.write(f"🔍 Debug: Letti {found_sheets} fogli dal file di input per confronto.")
        except Exception as e:
            st.error(f"❌ Errore critico lettura file input: {e}")
            compare_vba = False 

    # ================= MAPPING & LOGIC (V2) =================
    
    VALID_EMILIA_CODES = {
        "1300", "2900", "3200", "5000", "5600", "6100", "6600", "6800",
        "1301", "3201"
    }

    special_map = parse_mapping_config(mapping_special)

    out_xlsx = out_dir / file_xlsx_path.name
    shutil.copy2(file_xlsx_path, out_xlsx)

    # 1. Lettura dati (passando params calcolo lordo)
    if use_banca:
        fonte = banca_out or Path(file_banca_dati)
        rows = leggi_banca_dati_xlsx(
            fonte,
            pension_min,
            effective_coeff,
            aliquota_1,
            aliquota_2,
            aliquota_3,
        )
        enc = "banca_dati"
    else:
        rows, enc = prova_encoding(file_txt_path, special_map, pension_min, effective_coeff, aliquota_1, aliquota_2, aliquota_3)
        fonte = file_txt_path

    total_rows = len(rows)

    # 2. Elaborazione: Mappatura + Filtro
    filtered_rows = []
    
    # Stats Counters
    stats_norm = 0
    stats_special = 0
    stats_unchanged = 0
    
    for r in rows:
        orig = r.sede_cp
        mapped = orig
        
        is_modified = False
        is_special = False

        if orig in special_map:
            mapped = special_map[orig]
            is_special = True
            is_modified = True
        elif normalize_emilia:
             mapped = to_cp(orig)
             if mapped != orig:
                 is_modified = True
        
        if is_special:
            stats_special += 1
        elif is_modified:
            stats_norm += 1
        else:
            stats_unchanged += 1

        r.sede_cp = mapped
        if mapped in VALID_EMILIA_CODES:
            filtered_rows.append(r)

    filtered_rows_count = len(filtered_rows)
    print(f"Letti {total_rows} record. Filtrati {filtered_rows_count} record (Emilia).")
    
    rows = filtered_rows
    
    wb = load_workbook(out_xlsx)

    if use_banca:
        dump_banca_dati_sheet(wb, rows, pension_min, coeff_maggiorazione)

    # 3. Dati Base
    ws = get_or_create_sheet(wb, SHEET_NAME_DEFAULT)
    write_headers(ws)
    clear_values(ws)
    dump_rows(ws, rows, pension_min, coeff_maggiorazione)

    # 4. Età
    if SHEET_ETA in wb.sheetnames:
        ws_eta = wb[SHEET_ETA]
        compila_distribuzione_per_genere_eta(ws_eta, rows, anno_rif=report_year, start_row=template_start_row)
    else:
        print(f"ATTENZIONE: Foglio '{SHEET_ETA}' non trovato. Salto.")

    # 5. Categ
    if SHEET_CATEG in wb.sheetnames:
        ws_categ = wb[SHEET_CATEG]
        compila_distribuzione_per_categoria(ws_categ, rows, start_row=template_start_row)
    else:
        print(f"ATTENZIONE: Foglio '{SHEET_CATEG}' non trovato. Salto.")

    # 6. Importi
    if SHEET_IMPORTI in wb.sheetnames:
        ws_imp = wb[SHEET_IMPORTI]
        compila_distribuzione_per_importo_v3(ws_imp, rows, fasce=real_brackets, start_row=template_start_row)
    else:
        print(f"ATTENZIONE: Foglio '{SHEET_IMPORTI}' non trovato. Salto.")

    # 7. Genere
    if SHEET_GENERE in wb.sheetnames:
        ws_gen = wb[SHEET_GENERE]
        compila_distribuzione_associati_genere(ws_gen, rows, start_row=template_start_row)
    else:
        print(f"ATTENZIONE: Foglio '{SHEET_GENERE}' non trovato. Salto.")

    # 8. Log enhanced
    val_pension_magg = pension_min * coeff_maggiorazione
    info_file = {
        "File Input": use_banca and (banca_out.name if banca_out else file_banca_dati) or fonte.name,
        "Tipo Input": "Banca Dati (Excel/SLK)" if use_banca else "TXT Sindrinn",
        "Encoding": enc,
        "File Output": out_xlsx.name,
        "Anno Riferimento": str(report_year),
        "Pensione Min": f"{pension_min:.2f} €",
        "Coeff. Maggiorazione": f"{coeff_maggiorazione}",
        "Pensione Maggiorata": f"{val_pension_magg:.2f} €",
    }
    
    stats_elab = {
        "total": total_rows,
        "filtered": filtered_rows_count,
        "normalized": stats_norm,
        "special": stats_special,
        "unchanged": stats_unchanged
    }
    
    q_checks = {
        "inv_gen": _count_invalid_genere(rows),
        "inv_ann": _count_invalid_anno(rows),
        "inv_cat": _count_invalid_categoria(rows),
        "miss_sede": _count_missing_sede(rows),
        "inv_lordo": _count_invalid_importo_lordo(rows),
    }

    scrivi_log(
        wb,
        file_info=info_file,
        stats=stats_elab,
        quality_checks=q_checks,
        rows_filtered=rows,
        emilia_codes=VALID_EMILIA_CODES,
        special_map=special_map,
        start_row_templates=template_start_row,
        anno_rif=report_year,
        pension_min=pension_min,
        fasce_importo=real_brackets
    )

    wb.save(out_xlsx)
    
    # ================= VALIDAZIONE POST-EXEC =================
    if compare_vba:
        validation_key = "emilia_validation_results"
        
        try:
            wb_new = load_workbook(out_xlsx, data_only=False)
            total_diffs = 0
            validation_results = []
            
            for sname, old_snap in snapshots_old.items():
                # Usa template_start_row
                new_snap = _get_sheet_snapshot(wb_new, sname, start_row=template_start_row)
                diffs = _compare_snapshots(old_snap, new_snap)
                
                if diffs:
                    # CLASSIFICAZIONE INTELLIGENTE STATO (Fix richiesto da utente)
                    # Se ci sono differenze reali (dopo normalizzazione), sono ERRORI potenziali.
                    # Ma abbiamo detto che alcune sono "Migliorie".
                    # Tuttavia, con la logica "Static Values", dovremmo avere 0 differenze se tutto coincide.
                    # Se ne troviamo, sono differenze VERE sui dati.
                    status = "error"
                    
                    total_diffs += len(diffs)
                    validation_results.append({
                        "sheet": sname,
                        "status": status,
                        "count": len(diffs),
                        "details": diffs
                    })
                else:
                    validation_results.append({
                        "sheet": sname,
                        "status": "success",
                        "count": 0,
                        "details": []
                    })
            
            wb_new.close()
            
            # Salvataggio stato
            final_data = {
                "results": validation_results,
                "total_diffs": total_diffs
            }
            st.session_state[validation_key] = final_data
            
            # Visualizzazione
            _mostra_risultati_validazione(final_data)
            
        except Exception as e:
            st.error(f"❌ **Errore durante confronto finale:** {e}")
            st.session_state[validation_key] = {
                "results": [],
                "total_diffs": 0,
                "error": str(e)
            }

    return [out_xlsx]

