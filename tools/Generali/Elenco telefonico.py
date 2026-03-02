# pulizia_elenco.py
from __future__ import annotations

import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ----------------------------
# Config (modificabile)
# ----------------------------
FINAL_HEADERS = ["Cognome", "Nome", "Telefono", "Telefono2", "Cellulare", "E-mail", "Note"]

# Suffissi dominio da svuotare (match per suffisso: es. roma.cna.it -> cna.it)
BLOCKED_DOMAIN_SUFFIXES_DEFAULT = [
    "epasa.it",
    "cna.it",
    "itaco.it",
    "inps.it",
    "enpals.it",
    "inpdap.it",
    "caf.it",
]

START_ROW = 2  # riga 1 = intestazione

def _slug_key(s: str) -> str:
    """Genera una chiave sicura per i parametri (solo lettere, numeri, underscore)."""
    return re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")

TOOL = {'id': 'pulizia_elenco_v3',
 'name': 'Pulizia Elenco Telefonico',
 'description': (
    "#### 📌 1. FINALITÀ DEL TOOL\n"
    "Normalizza ed organizza liste di contatti provenienti da export Excel o SLK, gestendo la pulizia di anagrafiche, "
    "telefoni ed indirizzi email per campagne marketing o consultazione professionale.\n\n"
    "#### 🚀 2. COME UTILIZZARLO\n"
    "1. **Caricamento:** Inserisci il file sorgente (es. Export EPASA/CNA).\n"
    "2. **Mappatura:** Indica se vuoi rinominare le colonne nel file di output pulito.\n"
    "3. **Filtri:** Definisci eventuali domini email da escludere (es. domini aziendali).\n\n"
    "#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n"
    "* **Smart Title Case:** Converte nomi e cognomi in formato titolo corretto, gestendo particelle (De, Di) e apostrofi (D'Angelo).\n"
    "* **Phone Sanitizer:** Rimuove prefissi internazionali (+39/0039), spazi e caratteri non numerici, validando la lunghezza minima per fissi e cellulari.\n"
    "* **Email Recover:** Corregge errori comuni di digitazione (es. `,it` -> `.it`) e recupera email prive di `@` se il dominio è noto.\n"
    "* **Audit degli Scarti:** Genera un file parallelo 'TOTALI' che elenca ogni riga eliminata specificando il motivo (es. data incompleta, dominio bloccato).\n\n"
    "#### 📂 4. RISULTATO FINALE\n"
    "File Excel 'Pulito' con layout centrato e intestazioni evidenziate + File 'Scarti' per controllo qualità."
),
 'inputs': [{'key': 'input_file', 'label': 'File Elenco (.xlsx, .slk)', 'type': 'xlsx_single', 'required': True}],
 'params': [{'key': 'blocked_suffixes',
             'label': 'Domini Email Bloccati',
             'type': 'textarea',
             'default': 'epasa.it, cna.it, itaco.it, inps.it, enpals.it, inpdap.it, caf.it'},
            {'key': 'clean_sheet_label', 'label': 'Nome Foglio (File Pulito)', 'type': 'text', 'default': 'Pulito'},
            {'key': 'scarti_sheet_label',
             'label': 'Nome Foglio (File Scarti - Totali)',
             'type': 'text',
             'default': 'TOTALI'},
            {'key': 'rename_cognome', 'label': '✏️ Rinomina "Cognome"', 'type': 'text', 'default': 'Cognome'},
            {'key': 'rename_nome', 'label': '✏️ Rinomina "Nome"', 'type': 'text', 'default': 'Nome'},
            {'key': 'rename_telefono', 'label': '✏️ Rinomina "Telefono"', 'type': 'text', 'default': 'Telefono'},
            {'key': 'rename_telefono2', 'label': '✏️ Rinomina "Telefono2"', 'type': 'text', 'default': 'Telefono2'},
            {'key': 'rename_cellulare', 'label': '✏️ Rinomina "Cellulare"', 'type': 'text', 'default': 'Cellulare'},
            {'key': 'rename_e_mail', 'label': '✏️ Rinomina "E-mail"', 'type': 'text', 'default': 'E-mail'},
            {'key': 'rename_note', 'label': '✏️ Rinomina "Note"', 'type': 'text', 'default': 'Note'}]}

# Aggiunge parametri statici di rinomina colonne (Soluzione Definitiva)
existing_keys = {p['key'] for p in TOOL['params']}
for h in FINAL_HEADERS:
    k = f'rename_{_slug_key(h)}'
    if k not in existing_keys:
        TOOL['params'].append({
            'key': k,
            'label': f'✏️ Rinomina "{h}"',
            'type': 'text',
            'default': h
        })

# ----------------------------
# Utility: conversione SLK -> XLSX
# ----------------------------
def parse_sylk_to_matrix(file_input: Any, max_rows: Optional[int] = None) -> List[List[Any]]:
    """Legge un file SYLK (.slk) e restituisce una matrice di dati."""
    content = ""
    
    # Caso 1: Path o stringa (file su disco)
    if isinstance(file_input, (str, Path)):
        with open(file_input, 'r', encoding='latin1', errors='replace') as f:
            content = f.read()
    # Caso 2: Oggetto file-like (BytesIO, UploadedFile in memoria)
    else:
        if hasattr(file_input, "seek"):
            file_input.seek(0)
        raw = file_input.read()
        if isinstance(raw, bytes):
            content = raw.decode('latin1', errors='replace')
        else:
            content = raw
        if hasattr(file_input, "seek"):
            file_input.seek(0)
    
    lines = content.splitlines()
    if not lines or not lines[0].startswith("ID;"):
        raise ValueError("Not a SYLK file")

    data = {}
    max_r = 0
    max_c = 0
    
    # State (SYLK è stateful: ricorda riga/colonna precedente)
    curr_r = 0
    curr_c = 0

    for line in lines:
        line = line.strip()
        if not line: continue
        parts = line.split(';')
        record_type = parts[0]
        
        if record_type == 'C':
            val = None
            for p in parts[1:]:
                if not p: continue
                code = p[0]
                rest = p[1:]
                if code == 'Y': 
                    curr_r = int(rest)
                elif code == 'X': 
                    curr_c = int(rest)
                    # Ottimizzazione: se stiamo leggendo solo l'header (max_rows=1) e andiamo oltre
                    if max_rows is not None and curr_r > max_rows:
                        break
                elif code == 'K':
                    val = rest
                    if val.startswith('"') and val.endswith('"'): val = val[1:-1]
            
            max_r = max(max_r, curr_r)
            max_c = max(max_c, curr_c)

            if val is not None:
                data[(curr_r, curr_c)] = val
        
        if max_rows is not None and max_r > max_rows:
            break

    matrix = []
    for r in range(1, max_r + 1):
        row_vals = []
        for c in range(1, max_c + 1):
            row_vals.append(data.get((r, c), None))
        matrix.append(row_vals)
    return matrix

def convert_slk_to_xlsx(input_path: Path, output_path: Optional[Path] = None) -> Path:
    """
    Converte .slk in .xlsx.
    Strategia:
      1) Windows + Excel installato -> win32com (più affidabile)
      2) LibreOffice headless (soffice) se disponibile
    """
    if input_path.suffix.lower() != ".slk":
        return input_path

    if output_path is None:
        output_path = input_path.with_suffix(".xlsx")

    rows = []
    # 1. Tentativo parser SYLK nativo
    try:
        rows = parse_sylk_to_matrix(input_path)
    except Exception:
        # Fix robustezza: se sembrava un SYLK (inizia con ID;) ma è fallito, 
        # NON provare a leggerlo come CSV, è probabilmente corrotto o un SYLK complesso non supportato.
        # Evitiamo di generare Excel spazzatura.
        try:
            with open(input_path, 'r', encoding='latin1', errors='replace') as f:
                first = f.readline().strip()
                if first.startswith("ID;"):
                    raise RuntimeError("File SYLK non valido o formato non supportato.")
        except Exception:
            # Se l'eccezione è RuntimeError (lanciata sopra), la lasciamo salire.
            # Altrimenti (errore IO generico), proseguiamo al fallback.
            if sys.exc_info()[0] == RuntimeError:
                raise

        # 2. Fallback: CSV con punto e virgola (spesso i file .slk sono solo csv rinominati)
        try:
            with open(input_path, 'r', encoding='latin1', errors='replace') as f:
                reader = csv.reader(f, delimiter=';')
                rows = list(reader)
        except Exception as e:
            raise RuntimeError(f"Impossibile convertire .slk (né come SYLK né come CSV): {e}")

    # Creazione Excel con openpyxl
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    
    wb.save(output_path)
    return output_path


# ----------------------------
# Normalizzazione: nome/cognome
# ----------------------------
PARTICLES = {
    "de", "di", "del", "dello", "della", "dei", "degli", "delle", "da",
    "la", "lo", "le", "van", "von",
}

def normalize_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip())

def smart_title(s: str) -> str:
    """
    Title case "intelligente": gestisce apostrofi e trattini.
    Esempio: d'angelo -> D'Angelo, de-santis -> De-Santis
    """
    s = normalize_spaces(s)
    if not s:
        return ""

    def cap_token(tok: str) -> str:
        if not tok:
            return tok
        # Mantieni tutto minuscolo e poi capitalizza segmenti
        lower = tok.lower()

        # Split su apostrofi " ' " e " ’ "
        for apost in ("'", "’"):
            if apost in lower:
                parts = lower.split(apost)
                parts = [p[:1].upper() + p[1:] if p else "" for p in parts]
                return apost.join(parts)

        # Split su trattino
        if "-" in lower:
            parts = lower.split("-")
            parts = [p[:1].upper() + p[1:] if p else "" for p in parts]
            return "-".join(parts)

        return lower[:1].upper() + lower[1:]

    tokens = s.split(" ")
    # Non forziamo particelle minuscole per cognomi: in Italia spesso sono capitalizzate
    return " ".join(cap_token(t) for t in tokens)

def split_surname_name_it(fullname: str) -> Tuple[str, str, Optional[str]]:
    """
    Euristica italiana light.
    Ritorna: (cognome, nome, warning_reason|None)
    """
    fullname = normalize_spaces(fullname)
    if not fullname:
        return "", "", "NOMINATIVO_VUOTO"

    tokens = fullname.split(" ")
    if len(tokens) == 1:
        return tokens[0], "", "NOMINATIVO_SENZA_SPAZIO"

    # Caso classico: "De Santis Marco" -> cognome 2 token se primo è particella
    t0 = tokens[0].lower()
    t1 = tokens[1].lower()

    if t0 in PARTICLES and len(tokens) >= 3:
        surname = " ".join(tokens[:2])
        name = " ".join(tokens[2:])
        return surname, name, None

    # Caso: "Rossi De Santis Marco" (secondo token particella) -> cognome 3 token
    # Attenzione anche a d' / d’
    if (t1 in PARTICLES or t1.startswith("d'") or t1.startswith("d’")) and len(tokens) >= 4:
        surname = " ".join(tokens[:3])
        name = " ".join(tokens[3:])
        return surname, name, None

    # Default: split al primo spazio (cognome=primo token)
    surname = tokens[0]
    name = " ".join(tokens[1:])
    return surname, name, None


# ----------------------------
# Normalizzazione: telefoni
# ----------------------------
PHONE_DIGITS_RE = re.compile(r"\D+")

@dataclass
class PhoneResult:
    normalized: str
    issue: Optional[str] = None
    original_digits: str = ""

def strip_country_prefix(digits: str) -> str:
    # 0039 -> rimuovi
    if digits.startswith("0039"):
        return digits[4:]
    # 39 + 10 cifre -> rimuovi
    if digits.startswith("39") and len(digits) == 12:
        rest = digits[2:]
        if len(rest) == 10:
            return rest
    return digits

def normalize_phone_generic(raw: Any) -> PhoneResult:
    """
    Per Telefono / Telefono2 (C,D):
    - accetta fissi: iniziano con 0 e lunghezza 8-11
    - accetta anche 10 cifre che iniziano con 3 (se l'utente ha messo un mobile qui)
    - altrimenti svuota e segnala
    """
    if raw is None:
        return PhoneResult("", None, "")

    s = str(raw).strip()
    if not s:
        return PhoneResult("", None, "")

    digits = PHONE_DIGITS_RE.sub("", s)
    digits = strip_country_prefix(digits)
    if not digits:
        return PhoneResult("", "TEL_VUOTO_DOPO_PULIZIA", "")

    # fissi plausibili
    if digits.startswith("0") and 8 <= len(digits) <= 11:
        return PhoneResult(digits, None, digits)

    # mobile plausibile infilato in telefono
    if len(digits) == 10 and digits.startswith("3"):
        return PhoneResult(digits, None, digits)

    return PhoneResult("", f"TEL_FORMATO_NON_PLAUSIBILE_{len(digits)}", digits)

def normalize_mobile(raw: Any) -> PhoneResult:
    """
    Per Cellulare (E):
    - normalizza cifre e prefisso paese
    - se 10 cifre: accetta; se non inizia con 3 segnala warning ma tiene
    - se 9 cifre: scarta (dato incompleto) e segnala
    - altro: scarta e segnala
    """
    if raw is None:
        return PhoneResult("", None, "")

    s = str(raw).strip()
    if not s:
        return PhoneResult("", None, "")

    digits = PHONE_DIGITS_RE.sub("", s)
    digits = strip_country_prefix(digits)
    if not digits:
        return PhoneResult("", "CELL_VUOTO_DOPO_PULIZIA", "")

    if len(digits) == 10:
        if not digits.startswith("3"):
            return PhoneResult(digits, "CELL_NON_INIZIA_PER_3", digits)
        return PhoneResult(digits, None, digits)

    if len(digits) == 9:
        return PhoneResult("", "CELL_9_CIFRE_INCOMPLETO", digits)

    return PhoneResult("", f"CELL_LUNGHEZZA_{len(digits)}_NON_VALIDA", digits)


# ----------------------------
# Normalizzazione: email
# ----------------------------
EMAIL_RE = re.compile(r"^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$", re.IGNORECASE)
COMMON_DOMAINS = [
    "gmail.com", "libero.it", "tiscali.it", "yahoo.it", "hotmail.com", "hotmail.it",
    "virgilio.it", "alice.it", "outlook.com", "outlook.it", "icloud.com", "tin.it",
    "fastwebnet.it", "cna.it", "live.it", "live.com", "me.com", "msn.com"
]

def fix_common_email_typos(s: str) -> str:
    s = s.strip()
    # 1. Fix punteggiatura finale (,it -> .it)
    s = re.sub(r",it$", ".it", s, flags=re.IGNORECASE)
    s = re.sub(r",com$", ".com", s, flags=re.IGNORECASE)

    # 2. Fix domini ripetuti (es. @gmail@gmail.com -> @gmail.com)
    # Cerca pattern (@testo) ripetuto due volte
    s = re.sub(r"(@[a-zA-Z0-9.-]+)\1", r"\1", s, flags=re.IGNORECASE)

    # 3. Fix @ mancante per domini noti
    if "@" not in s:
        s_lower = s.lower()
        for d in COMMON_DOMAINS:
            if s_lower.endswith(d) and len(s) > len(d):
                # Inserisce @ prima del dominio
                s = s[:-len(d)] + "@" + s[-len(d):]
                break
    return s

@dataclass
class EmailResult:
    normalized: str
    issue: Optional[str] = None
    domain: str = ""
    was_fixed: bool = False

def extract_first_email_candidate(s: str) -> str:
    s = s.strip()
    if not s:
        return ""
    # Gestione formato: Nome Cognome <mail@dominio>
    if "<" in s and ">" in s:
        inside = re.findall(r"<([^>]+)>", s)
        if inside:
            s = inside[0].strip()

    # Split su separatori comuni e prendi il primo token che contiene '@'
    parts = re.split(r"[;, \t\r\n]+", s)
    for p in parts:
        if "@" in p:
            return p.strip()
    # se non trova, prova tutta la stringa
    return s.strip()

def normalize_email(raw: Any) -> EmailResult:
    if raw is None:
        return EmailResult("", None, "")

    s = str(raw).strip()
    if not s:
        return EmailResult("", None, "", False)

    # Applica fix automatici prima della validazione
    s_fixed = fix_common_email_typos(s)
    # Se la stringa è cambiata, segniamo che è stata "riparata"
    was_fixed = (s_fixed != s)

    cand = extract_first_email_candidate(s_fixed)
    cand = cand.replace(" ", "").lower()
    if not cand:
        return EmailResult("", "EMAIL_VUOTA_DOPO_PULIZIA", "", was_fixed)

    if not EMAIL_RE.match(cand):
        return EmailResult("", "EMAIL_NON_VALIDA", "", was_fixed)

    domain = cand.split("@", 1)[1].lower()
    return EmailResult(cand, None, domain, was_fixed)

def apply_domain_filter(email_res: EmailResult, blocked_suffixes: List[str]) -> EmailResult:
    if not email_res.normalized:
        return email_res
    dom = email_res.domain
    for suf in blocked_suffixes:
        suf = suf.lower()
        if dom.endswith(suf):
            # Se bloccata, resettiamo anche il flag fixed perché tanto la buttiamo
            return EmailResult("", "EMAIL_DOMINIO_BLOCCATO", dom, False)
    return email_res


# ----------------------------
# Excel helpers
# ----------------------------
def get_last_row(ws, cols: List[int], start_row: int) -> int:
    """
    Trova l'ultima riga "vera" guardando le colonne indicate.
    """
    max_row = ws.max_row
    last = start_row - 1
    for r in range(max_row, start_row - 1, -1):
        for c in cols:
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                return r
    return last

def format_header_row(ws, num_cols: int, font_size: int = 11, row_height: float = None):
    fill = PatternFill("solid", fgColor="FFFF00")
    font = Font(bold=True, size=font_size)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if row_height:
        ws.row_dimensions[1].height = row_height

    for c in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is None or str(cell.value).strip() == "":
            continue
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def set_column_widths(ws, widths: Dict[int, int]):
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def autofit_columns(ws):
    """Adatta larghezza colonne al contenuto (con limiti min/max)."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    val = str(cell.value)
                    # Gestione multiline: prendiamo la riga più lunga
                    lines = val.split('\n')
                    curr_len = max(len(l) for l in lines) if lines else 0
                    if curr_len > max_length:
                        max_length = curr_len
            except:
                pass
        
        # Calcolo larghezza euristica
        adjusted_width = (max_length + 2) * 1.1
        if adjusted_width > 60: adjusted_width = 60
        if adjusted_width < 12: adjusted_width = 12
        
        ws.column_dimensions[col_letter].width = adjusted_width

# ----------------------------
# Scarti: struttura
# ----------------------------
def scarti_headers() -> List[str]:
    return [
        "RigaOriginale",
        "Motivo",
        "EsitoRiga",
        "A_Orig_Nominativo",
        "G_Orig_Telefono",
        "H_Orig_Telefono2",
        "I_Orig_Cellulare",
        "J_Orig_Email",
        "A_Final_Cognome",
        "B_Final_Nome",
        "C_Final_Telefono",
        "D_Final_Telefono2",
        "E_Final_Cellulare",
        "F_Final_Email",
    ]

def add_scarto(bucket: Dict[str, List[List[Any]]], sheet_name: str, row_data: List[Any]):
    bucket.setdefault(sheet_name, []).append(row_data)


# ----------------------------
# Core pipeline
# ----------------------------
def process_file(
    input_path: Path,
    output_clean_path: Path,
    output_scarti_path: Path,
    sheet_name: Optional[str],
    blocked_suffixes: List[str],
    renames: Dict[str, str] = None,
    extra_columns: List[str] = None,
    clean_sheet_label: str = "Pulito",
    scarti_sheet_label: str = "TOTALI",
    progress_callback=None,
) -> None:
    if renames is None:
        renames = {}
    # Conversione se serve
    if progress_callback: progress_callback(0, 100, "Lettura e conversione file...")
    src_path = convert_slk_to_xlsx(input_path)

    wb = load_workbook(src_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Colonne input: A=1, G=7, H=8, I=9, J=10
    last_row = get_last_row(ws, cols=[1, 7, 8, 9, 10], start_row=START_ROW)
    if last_row < START_ROW:
        raise RuntimeError("Nessun dato trovato (oltre l'intestazione).")

    # Mappatura colonne extra (Nome -> Indice 1-based)
    extra_col_indices = []
    if extra_columns:
        # Leggiamo l'intestazione (riga 1) per trovare gli indici, gestendo duplicati
        header_map = {}
        seen_headers = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=c).value
            if val:
                h_raw = str(val).strip()
                h_unique = h_raw
                count = 1
                while h_unique in seen_headers:
                    count += 1
                    h_unique = f"{h_raw} ({count})"
                seen_headers[h_unique] = True
                header_map[h_unique] = c
        
        for col_name in extra_columns:
            if col_name in header_map:
                extra_col_indices.append((col_name, header_map[col_name]))

    # Crea workbook pulito separato (più sicuro)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = (clean_sheet_label.strip() or "Pulito")[:31]

    # Header
    # Intestazioni standard
    current_headers = list(FINAL_HEADERS)
    # Aggiungi intestazioni extra
    if extra_columns:
        current_headers.extend(extra_columns)

    for i, h in enumerate(current_headers, start=1):
        val = renames.get(h, h)
        out_ws.cell(row=1, column=i, value=val)
    format_header_row(out_ws, num_cols=len(current_headers), font_size=13, row_height=35)

    # Scarti bucket
    scarti: Dict[str, List[List[Any]]] = {}

    # Accumulo righe buone (poi ordino)
    good_rows: List[Tuple[str, str, List[Any]]] = []  # (keyA, keyB, row_values A..F)

    total_rows = last_row - START_ROW + 1
    processed_count = 0

    for r in range(START_ROW, last_row + 1):
        processed_count += 1
        if progress_callback and processed_count % 20 == 0: # Aggiorna ogni 20 righe per velocità
            progress_callback(processed_count, total_rows, f"Elaborazione riga {processed_count}/{total_rows}")

        a_raw = ws.cell(r, 1).value
        g_raw = ws.cell(r, 7).value
        h_raw = ws.cell(r, 8).value
        i_raw = ws.cell(r, 9).value
        j_raw = ws.cell(r, 10).value

        a_str = "" if a_raw is None else str(a_raw)
        g_str = "" if g_raw is None else str(g_raw)
        h_str = "" if h_raw is None else str(h_raw)
        i_str = "" if i_raw is None else str(i_raw)
        j_str = "" if j_raw is None else str(j_raw)

        # 1) Split nome
        cognome_raw, nome_raw, name_issue = split_surname_name_it(a_str)
        cognome = smart_title(cognome_raw)
        nome = smart_title(nome_raw)

        # 2) Telefoni
        tel1 = normalize_phone_generic(g_raw)
        tel2 = normalize_phone_generic(h_raw)
        cell = normalize_mobile(i_raw)

        # 3) Email + filtro dominio
        email = normalize_email(j_raw)
        email = apply_domain_filter(email, blocked_suffixes)

        # 5) Pre-calcolo esiti per logging e decisione finale
        tel1_final = tel1.normalized
        tel2_final = tel2.normalized
        cell_final = cell.normalized
        email_final = email.normalized

        # La riga viene mantenuta se almeno un campo di contatto è valido (pieno)
        is_kept = bool(tel1_final or tel2_final or cell_final or email_final)
        esito_str = "MANTENUTA (Campo pulito)" if is_kept else "ELIMINATA"

        # Preparazione riga base per scarti/audit
        base_row = [
            r,
            "",  # motivo placeholder
            esito_str,
            a_str,
            g_str,
            h_str,
            i_str,
            j_str,
            cognome,
            nome,
            tel1_final,
            tel2_final,
            cell_final,
            email_final,
        ]

        # SE LA RIGA È MANTENUTA (ha almeno un dato valido), VA NEL FILE BUONO.
        # Ma registriamo comunque le anomalie in fogli dedicati "_MANTENUTE" per audit.
        if is_kept:
            # Audit anomalie su righe mantenute
            if name_issue:
                row = base_row.copy()
                row[1] = name_issue
                add_scarto(scarti, "NOMINATIVO_PROBLEMI_MANTENUTE", row)
            
            if tel1.issue:
                row = base_row.copy()
                row[1] = tel1.issue
                add_scarto(scarti, "TELEFONO_ANOMALO_MANTENUTE", row)

            if tel2.issue:
                row = base_row.copy()
                row[1] = tel2.issue
                add_scarto(scarti, "TELEFONO2_ANOMALO_MANTENUTE", row)

            if cell.issue:
                row = base_row.copy()
                row[1] = cell.issue
                add_scarto(scarti, "CELLULARE_ANOMALO_MANTENUTE", row)

            if email.issue and email.issue != "EMAIL_VUOTA_DOPO_PULIZIA":
                row = base_row.copy()
                row[1] = f"{email.issue} ({email.domain})" if email.domain else email.issue
                add_scarto(scarti, "EMAIL_ANOMALA_MANTENUTE", row)

            notes = []
            if email.was_fixed:
                notes.append("Email recuperata (typo/formato)")
            
            # Costruzione riga finale
            note_str = "; ".join(notes)
            row_values = [cognome, nome, tel1_final, tel2_final, cell_final, email_final, note_str]
            
            # Aggiungi valori colonne extra (copia 1:1)
            for _, col_idx in extra_col_indices:
                val = ws.cell(row=r, column=col_idx).value
                row_values.append(val)

            good_rows.append((cognome.casefold(), nome.casefold(), row_values))
            continue

        # --- SE SIAMO QUI, LA RIGA È STATA COMPLETAMENTE ELIMINATA ---

        # Raccoglie tutti i problemi riscontrati per questa riga
        issues_found = []

        if name_issue:
            row = base_row.copy()
            row[1] = name_issue
            add_scarto(scarti, "NOMINATIVO_PROBLEMI", row)
            issues_found.append(f"Nome: {name_issue}")

        if tel1.issue:
            row = base_row.copy()
            row[1] = tel1.issue
            add_scarto(scarti, "TELEFONO_ANOMALO", row)
            issues_found.append(f"Tel1: {tel1.issue}")

        if tel2.issue:
            row = base_row.copy()
            row[1] = tel2.issue
            add_scarto(scarti, "TELEFONO2_ANOMALO", row)
            issues_found.append(f"Tel2: {tel2.issue}")

        if cell.issue:
            row = base_row.copy()
            row[1] = cell.issue
            add_scarto(scarti, "CELLULARE_ANOMALO", row)
            issues_found.append(f"Cell: {cell.issue}")

        if email.issue == "EMAIL_NON_VALIDA":
            row = base_row.copy()
            row[1] = email.issue
            add_scarto(scarti, "EMAIL_NON_VALIDA", row)
            issues_found.append(f"Email: {email.issue}")
        elif email.issue == "EMAIL_DOMINIO_BLOCCATO":
            row = base_row.copy()
            row[1] = f"{email.issue} ({email.domain})"
            add_scarto(scarti, "EMAIL_DOMINIO_BLOCCATO", row)
            issues_found.append(f"Email: {email.issue} ({email.domain})")
        elif email.issue:
            row = base_row.copy()
            row[1] = email.issue
            add_scarto(scarti, "EMAIL_ALTRI_PROBLEMI", row)
            issues_found.append(f"Email: {email.issue}")

        # 6) Aggiungi a RIGHE_ELIMINATE
        row = base_row.copy()
        row[1] = "; ".join(issues_found) if issues_found else "TUTTI_CAMPI_VUOTI"
        add_scarto(scarti, "RIGHE_ELIMINATE", row)

    # Ordina righe buone A->Z per Cognome poi Nome
    good_rows.sort(key=lambda x: (x[0], x[1]))

    # Scrivi righe buone
    if progress_callback: progress_callback(90, 100, "Scrittura file Excel...")
    align_center = Alignment(horizontal="center", vertical="center")
    out_r = 2

    for _, __, vals in good_rows:
        for c, v in enumerate(vals, start=1):
            cell = out_ws.cell(row=out_r, column=c, value=v)
            # Allinea TUTTO al centro come richiesto
            cell.alignment = align_center
        out_r += 1

    # Migliorie foglio pulito
    # Auto-adattamento larghezza colonne
    autofit_columns(out_ws)
    
    out_ws.freeze_panes = "A2"

    out_wb.save(output_clean_path)

    # ----------------------------
    # Crea workbook scarti
    # ----------------------------
    sc_wb = Workbook()
    # Rimuovi il foglio di default
    default_ws = sc_wb.active
    sc_wb.remove(default_ws)

    # Foglio TOTALI
    totals_ws = sc_wb.create_sheet((scarti_sheet_label.strip() or "TOTALI")[:31])
    totals_ws.append(["Categoria", "TotaleRighe"])
    format_header_row(totals_ws, 2)
    totals_ws.column_dimensions["A"].width = 28
    totals_ws.column_dimensions["B"].width = 14

    # Crea fogli per categoria
    for sheet, rows in sorted(scarti.items(), key=lambda kv: kv[0]):
        ws_cat = sc_wb.create_sheet(sheet[:31])  # limite Excel 31 chars
        ws_cat.append(scarti_headers())
        for row in rows:
            ws_cat.append(row)
        format_header_row(ws_cat, num_cols=len(scarti_headers()))
        ws_cat.freeze_panes = "A2"
        ws_cat.auto_filter.ref = f"A1:{get_column_letter(len(scarti_headers()))}{len(rows)+1}"

        # larghezze sensate
        set_column_widths(ws_cat, {
            1: 12, 2: 34, 3: 20, 4: 26, 5: 18, 6: 18, 7: 18, 8: 30,
            9: 22, 10: 22, 11: 16, 12: 16, 13: 16, 14: 30,
        })

        totals_ws.append([sheet, len(rows)])

    # Totali globali (righe uniche “problema” e righe eliminate)
    totals_ws.append([])
    totals_ws.append(["NOTE", ""])
    totals_ws.append(["Le righe possono comparire in più fogli se hanno più problemi.", ""])

    sc_wb.save(output_scarti_path)


def get_dynamic_params(inputs: dict, params: dict = None) -> List[dict]:
    dyn = []
    
    # 1. Opzioni per aggiungere colonne extra dal file sorgente (IN ALTO)
    file_obj = inputs.get("input_file")
    if file_obj:
        try:
            headers = []
            if hasattr(file_obj, "seek"):
                file_obj.seek(0)
            filename = getattr(file_obj, "name", str(file_obj))
            
            if filename.lower().endswith('.slk'):
                matrix = parse_sylk_to_matrix(file_obj, max_rows=1)
                if matrix:
                    headers = [str(c) for c in matrix[0] if c]
            else:
                wb = load_workbook(file_obj, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                    headers = [str(c) for c in row if c]
                    break
                wb.close()
            
            if headers:
                # Gestione duplicati per il menu (es. "Telefono", "Telefono") -> "Telefono", "Telefono (2)"
                unique_headers = []
                seen = {}
                for h in headers:
                    h_clean = h
                    c = 1
                    while h_clean in seen:
                        c += 1
                        h_clean = f"{h} ({c})"
                    seen[h_clean] = True
                    unique_headers.append(h_clean)

                dyn.append({
                    'key': 'extra_columns',
                    'label': '✨ **SELEZIONA COLONNE EXTRA** (dal file originale)',
                    'type': 'multiselect',
                    'options': unique_headers,
                    'default': [],
                    'help': 'Seleziona colonne aggiuntive da riportare nel file pulito (copiate 1:1).'
                })
        except Exception:
            pass
            
    # I campi di rinomina sono ora statici in TOOL['params'], qui gestiamo solo le extra
    return dyn

def run(input_file: Path, blocked_suffixes: str, out_dir: Path, 
        sheet_name: str = "",
        clean_sheet_label: str = "Pulito", scarti_sheet_label: str = "TOTALI",
        clean_filename: str = "", scarti_filename: str = "",
        **kwargs) -> List[Path]:
    if not input_file:
        raise ValueError("File di input mancante.")

    out_dir.mkdir(parents=True, exist_ok=True)
    blocked = [x.strip().lower() for x in re.split(r"[,\n;]+", blocked_suffixes) if x.strip()]
    if not blocked:
        blocked = BLOCKED_DOMAIN_SUFFIXES_DEFAULT

    c_name = clean_filename.strip() if clean_filename.strip() else f"Clean_{input_file.stem}"
    if not c_name.lower().endswith(".xlsx"): c_name += ".xlsx"
    s_name = scarti_filename.strip() if scarti_filename.strip() else f"Scarti_{input_file.stem}"
    if not s_name.lower().endswith(".xlsx"): s_name += ".xlsx"

    clean_path = out_dir / c_name
    scarti_path = out_dir / s_name

    renames = {}
    for h in FINAL_HEADERS:
        val = kwargs.get(f'rename_{_slug_key(h)}')
        if val:
            renames[h] = str(val).strip()

    extra_columns = kwargs.get('extra_columns', [])
    progress_callback = kwargs.get('progress_callback')

    try:
        process_file(
            input_path=input_file,
            output_clean_path=clean_path,
            output_scarti_path=scarti_path,
            sheet_name=sheet_name if sheet_name.strip() else None,
            blocked_suffixes=blocked,
            renames=renames,
            extra_columns=extra_columns,
            clean_sheet_label=clean_sheet_label,
            scarti_sheet_label=scarti_sheet_label,
            progress_callback=progress_callback,
        )
    except PermissionError as e:
        raise RuntimeError(f"Errore di permesso: {e}\n\n⚠️ Chiudi i file Excel se sono aperti!")

    return [clean_path, scarti_path]
