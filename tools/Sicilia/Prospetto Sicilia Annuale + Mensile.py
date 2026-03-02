#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script per:
- Selezionare 12 TXT (record a larghezza fissa)
- Estrarre campi a posizioni fisse
- Filtrare record:
    - tiene sempre se PR in {AG,CL,CT,EN,ME,PA,RG,SR,TP}
    - se PR NON è tra quelle, tiene comunque se SedeGestione (pos 345, len 4) è nella whitelist codici
    - Per i record con Comune mancante: lookup da tabella codici catastali (Excel)
    - Comune è in colonna A
    - Codice catastale (4 char) viene cercato in una colonna del file (auto-detect); spesso è in colonna D
    - se non trovato => "Non Trovato"
- Output Excel "completo": un foglio per provincia (tutte le PR presenti dopo il filtro),
    Comune in colonna A + colonne dinamiche per Codice Funzione (pos 7),
    + riga bianca + riga totali provincia
- Output extra: 12 Excel mensili SOLO per PA e TP (uno per ogni TXT selezionato),
    stessi criteri e stessi totali.

Dipendenze:
    pip install pandas openpyxl
"""

import os
import re
import zipfile
from pathlib import Path
from typing import List, Optional

# heavy imports (pandas, openpyxl) are intentionally performed inside `run()`
# to avoid import-time failures when the toolbox loads the file.

TOOL = {'id': 'prospetto_sicilia_annuale_mensile',
 'name': 'Prospetto per Comuni  (Filippello)',
 'description': (
    "#### 📌 1. FINALITÀ DEL TOOL\n"
    "Analizza i flussi mensili INPS (TXT larga fissa) della Sicilia per produrre report dettagliati per Comune e Provincia, "
    "integrando la decodifica automatica dei codici catastali.\n\n"
    "#### 🚀 2. COME UTILIZZARLO\n"
    "1. **File:** Carica uno o più TXT mensili (es. 12 file per l'anno intero).\n"
    "2. **Opzioni:** Scegli se vuoi il report annuale consolidato o i singoli file mensili.\n"
    "3. **Filtri:** Seleziona le province specifiche da includere nei report mensili.\n\n"
    "#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n"
    "* **Tracciato Fisso:** Estrae i Codici Funzione, Comune e Sede da posizioni specifiche nel record TXT.\n"
    "* **Comune Lookup:** Se il nome comune è mancante, utilizza una tabella di corrispondenza integrata (Codici Catastali) per assegnare il nome corretto al record.\n"
    "* **Filtro Sede/Provincia:** Applica una whitelist di 'Sedi Gestione' valida per la Sicilia (es. 0100, 1800, ecc.) per scartare record extra-regionali.\n"
    "* **Pivot Automatico:** Raggruppa i dati per Comune/Codice Funzione e calcola i totali provinciali con layout pronto per la stampa.\n\n"
    "#### 📂 4. RISULTATO FINALE\n"
    "Report Excel Annuale (multi-foglio per provincia) e/o 12 Excel mensili con totali e formattazione professionale."
),
 'inputs': [{'key': 'file_txt', 'label': 'File TXT', 'type': 'txt_multi', 'required': True}],
 'params': [{'key': 'export_annual',
             'label': 'Genera Prospetto Annuale',
             'type': 'radio',
             'options': ['SI', 'NO'],
             'default': 'SI'},
            {'key': 'export_monthly',
             'label': 'Genera File Mensili',
             'type': 'radio',
             'options': ['SI', 'NO'],
             'default': 'SI'},
            {'key': 'provinces_selected',
             'label': 'Province da includere (mensili)',
             'type': 'multiselect',
             'options': ['AG', 'CL', 'CT', 'EN', 'ME', 'PA', 'RG', 'SR', 'TP'],
             'default': ['AG', 'CL', 'CT', 'EN', 'ME', 'PA', 'RG', 'SR', 'TP']}]}


# -----------------------------
# Config posizioni (1-based)
# -----------------------------
POS_COD_FUN = (7, 1)
POS_COMUNE = (235, 36)
POS_PROVINCIA = (271, 2)
POS_CAP = (274, 5)
POS_COD_CATASTALE = (339, 4)
POS_COD_ESTERO = (343, 2)
POS_SEDE_GESTIONE = (345, 4)

MIN_LINE_LEN = POS_SEDE_GESTIONE[0] - 1 + POS_SEDE_GESTIONE[1]  # fino a pos 348 inclusa


ALLOWED_PROVINCES = {"AG", "CL", "CT", "EN", "ME", "PA", "RG", "SR", "TP"}

SEDE_WHITELIST = {
    "0100", "0101",
    "0190", "0191", "0192",
    "1800", "1890",
    "2100",
    "2190", "2191", "2192", "2193", "2194", "2196",
    "2800", "2890", "2891",
    "4800", "4890", "4891", "4892", "4893", "4894",
    "5500", "5501", "5502",
    "5590", "5591", "5592", "5593", "5594", "5596",
    "6500", "6590", "6591",
    "7600", "7601", "7690", "7691",
    "8200", "8290", "8291", "8292", "8293",
}


# -----------------------------
# Utility
# -----------------------------
def field(line: str, pos1: int, length: int) -> str:
    """Estrae substring usando posizioni 1-based."""
    start = max(pos1 - 1, 0)
    end = start + length
    return line[start:end]


def normalize_comune(s: str) -> str:
    s = (s or "").strip().upper()
    # collassa spazi multipli
    s = " ".join(s.split())
    return s


def normalize_provincia(s: str) -> str:
    return (s or "").strip().upper()


def normalize_code(s: str) -> str:
    return (s or "").strip().upper()


def normalize_sede(s: str) -> str:
    s = (s or "").strip()
    if s.isdigit():
        s = s.zfill(4)
    return s


def safe_sheet_name(name: str) -> str:
    # Excel: max 31 char, niente []:*?/\
    bad = r'[\[\]\:\*\?\/\\]'
    name = re.sub(bad, "_", name)
    name = name.strip()
    if not name:
        name = "SHEET"
    return name[:31]


def infer_month_key_from_filename(path: str, index_1based: int) -> str:
    """
    Tenta di ricavare YYYY-MM dal nome file (es. 202601, 2026-01, 2026_01).
    Se non trova nulla => M01..M12 in base all'ordine di selezione.
    """
    name = os.path.basename(path)
    m = re.search(r"(20\d{2})[^\d]?([01]\d)", name)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return f"{year}-{month:02d}"
    return f"M{index_1based:02d}"


def sort_codefun_columns(cols):
    """
    Ordina i codici funzione:
    - numerici prima (0,1,2,...)
    - poi alfanumerici (A,B,...) in ordine lessicografico.
    """
    def key(c):
        c = str(c)
        if c.isdigit():
            return (0, int(c), "")
        return (1, 9999, c)
    return sorted(cols, key=key)


# -----------------------------
# Parsing TXT
# -----------------------------
def parse_txt_file(txt_path: str, month_key: str):
    global EMBEDDED_CATASTO
    records = []
    stats = {
        "read_lines": 0,
        "kept_allowed_pr": 0,
        "kept_by_sede": 0,
        "discarded": 0,
        "short_lines_padded": 0,
    }

    # latin-1 è spesso la scelta più robusta per file “vecchi” (evita crash su caratteri strani)
    # Apriamo SOLO in lettura; non scriviamo mai sui file di input.
    with open(txt_path, "r", encoding="latin-1", errors="replace") as f:
        for orig_line in f:
            stats["read_lines"] += 1
            orig_line = orig_line.rstrip("\r\n")
            line = orig_line
            if len(line) < MIN_LINE_LEN:
                stats["short_lines_padded"] += 1
                # padding usato solo in memoria per l'estrazione, non si scrive sul file
                line = line.ljust(MIN_LINE_LEN)

            cod_fun = normalize_code(field(line, *POS_COD_FUN))
            comune = normalize_comune(field(line, *POS_COMUNE))
            pr = normalize_provincia(field(line, *POS_PROVINCIA))
            cap = normalize_code(field(line, *POS_CAP))
            cod_cat = normalize_code(field(line, *POS_COD_CATASTALE))
            cod_est = normalize_code(field(line, *POS_COD_ESTERO))
            sede = normalize_sede(field(line, *POS_SEDE_GESTIONE))

            # regola filtro
            keep = False
            if pr in ALLOWED_PROVINCES:
                keep = True
                stats["kept_allowed_pr"] += 1
            else:
                if sede in SEDE_WHITELIST:
                    keep = True
                    stats["kept_by_sede"] += 1

            if not keep:
                stats["discarded"] += 1
                continue

            if not pr:
                pr = "NO_PR"

            if not cod_fun:
                cod_fun = "?"  # evita colonne vuote

            records.append({
                "source_file": os.path.basename(txt_path),
                "month_key": month_key,
                "provincia": pr,
                "comune": comune,  # potrebbe essere ""
                "cod_funzione": cod_fun,
                "cap": cap,
                "cod_catastale": cod_cat,
                "cod_estero": cod_est,
                "sede_gestione": sede,
            })

        EMBEDDED_CATASTO = {
            "A089": "AGRIGENTO",
    "A181": "ALESSANDRIA DELLA ROCCA",
    "A351": "ARAGONA",
    "A896": "BIVONA",
    "B275": "BURGIO",
    "B377": "CALAMONACI",
    "B427": "CALTABELLOTTA",
    "B460": "CAMASTRA",
    "B486": "CAMMARATA",
    "B520": "CAMPOBELLO DI LICATA",
    "B602": "CANICATTÌ",
    "C275": "CASTELTERMINI",
    "C341": "CASTROFILIPPO",
    "C356": "CATTOLICA ERACLEA",
    "C668": "CIANCIANA",
    "C928": "COMITINI",
    "D514": "FAVARA",
    "E209": "GROTTE",
    "E390": "JOPPOLO GIANCAXIO",
    "E431": "LAMPEDUSA E LINOSA",
    "E573": "LICATA",
    "E714": "LUCCA SICULA",
    "F126": "MENFI",
    "F414": "MONTALLEGRO",
    "F655": "MONTEVAGO",
    "F845": "NARO",
    "G282": "PALMA DI MONTECHIARO",
    "F299": "PORTO EMPEDOCLE",
    "H148": "RACALMUTO",
    "H159": "RAFFADALI",
    "H194": "RAVANUSA",
    "H205": "REALMONTE",
    "H269": "RIBERA",
    "H743": "SAMBUCA DI SICILIA",
    "H778": "SAN BIAGIO PLATANI",
    "H914": "SAN GIOVANNI GEMINI",
    "I185": "SANTA ELISABETTA",
    "I224": "SANTA MARGHERITA DI BELICE",
    "I290": "SANT'ANGELO MUXARO",
    "I356": "SANTO STEFANO QUISQUINA",
    "I533": "SCIACCA",
    "I723": "SICULIANA",
    "L944": "VILLAFRANCA SICULA",

    "A049": "ACQUAVIVA PLATANI",
    "A957": "BOMPENSIERE",
    "B302": "BUTERA",
    "B429": "CALTANISSETTA",
    "B537": "CAMPOFRANCO",
    "D267": "DELIA",
    "D960": "GELA",
    "E953": "MARIANOPOLI",
    "F065": "MAZZARINO",
    "E618": "MILENA",
    "F489": "MONTEDORO",
    "F830": "MUSSOMELI",
    "F899": "NISCEMI",
    "H245": "RESUTTANO",
    "H281": "RIESI",
    "H792": "SAN CATALDO",
    "I169": "SANTA CATERINA VILLARMOSA",
    "I644": "SERRADIFALCO",
    "I824": "SOMMATINO",
    "L016": "SUTERA",
    "L609": "VALLELUNGA PRATAMENO",
    "L959": "VILLALBA",

    "A025": "ACI BONACCORSI",
    "A026": "ACI CASTELLO",
    "A027": "ACI CATENA",
    "A028": "ACIREALE",
    "A029": "ACI SANT'ANTONIO",
    "A056": "ADRANO",
    "A766": "BELPASSO",
    "A841": "BIANCAVILLA",
    "B202": "BRONTE",
    "B384": "CALATABIANO",
    "B428": "CALTAGIRONE",
    "B561": "CAMPOROTONDO ETNEO",
    "C091": "CASTEL DI IUDICA",
    "C297": "CASTIGLIONE DI SICILIA",
    "C351": "CATANIA",
    "D623": "FIUMEFREDDO DI SICILIA",
    "E017": "GIARRE",
    "E133": "GRAMMICHELE",
    "E156": "GRAVINA DI CATANIA",
    "E578": "LICODIA EUBEA",
    "E602": "LINGUAGLOSSA",
    "E854": "MALETTO",
    "F004": "MASCALI",
    "F005": "MASCALUCIA",
    "F209": "MILITELLO IN VAL DI CATANIA",
    "F214": "MILO",
    "F217": "MINEO",
    "F231": "MIRABELLA IMBACCARI",
    "F250": "MISTERBIANCO",
    "F781": "MOTTA SANT'ANASTASIA",
    "F890": "NICOLOSI",
    "G253": "PALAGONIA",
    "G371": "PATERNÒ",
    "G402": "PEDARA",
    "G597": "PIEDIMONTE ETNEO",
    "H154": "RADDUSA",
    "H168": "RAMACCA",
    "H175": "RANDAZZO",
    "H325": "RIPOSTO",
    "H805": "SAN CONO",
    "H922": "SAN GIOVANNI LA PUNTA",
    "H940": "SAN GREGORIO DI CATANIA",
    "I035": "SAN MICHELE DI GANZARIA",
    "I098": "SAN PIETRO CLARENZA",
    "I202": "SANT'AGATA LI BATTIATI",
    "I216": "SANT'ALFIO",
    "I240": "SANTA MARIA DI LICODIA",
    "I314": "SANTA VENERINA",
    "I548": "SCORDIA",
    "L355": "TRECASTAGNI",
    "L369": "TREMESTIERI ETNEO",
    "L658": "VALVERDE",
    "L828": "VIAGRANDE",
    "M100": "VIZZINI",
    "M139": "ZAFFERANA ETNEA",
    "M271": "MAZZARRONE",
    "M283": "MANIACE",
    "M287": "RAGALNA",

    "A070": "AGIRA",
    "A098": "AIDONE",
    "A478": "ASSORO",
    "A676": "BARRAFRANCA",
    "B381": "CALASCIBETTA",
    "C353": "CATENANUOVA",
    "C471": "CENTURIPE",
    "C480": "CERAMI",
    "C342": "ENNA",
    "D849": "GAGLIANO CASTELFERRATO",
    "E536": "LEONFORTE",
    "F892": "NICOSIA",
    "F900": "NISSORIA",
    "G580": "PIAZZA ARMERINA",
    "G624": "PIETRAPERZIA",
    "H221": "REGALBUTO",
    "I891": "SPERLINGA",
    "L448": "TROINA",
    "L583": "VALGUARNERA CAROPEPE",
    "M011": "VILLAROSA",

    "A177": "ALCARA LI FUSI",
    "A194": "ALÌ",
    "A201": "ALÌ TERME",
    "A313": "ANTILLO",
    "A638": "BARCELLONA POZZO DI GOTTO",
    "A698": "BASICÒ",
    "B198": "BROLO",
    "B660": "CAPIZZI",
    "B666": "CAPO D'ORLANDO",
    "B695": "CAPRI LEONE",
    "B804": "CARONIA",
    "B918": "CASALVECCHIO SICULO",
    "C094": "CASTEL DI LUCIO",
    "C051": "CASTELL'UMBERTO",
    "C210": "CASTELMOLA",
    "C347": "CASTROREALE",
    "C568": "CESARÒ",
    "C956": "CONDRÒ",
    "D474": "FALCONE",
    "D569": "FICARRA",
    "D622": "FIUMEDINISI",
    "D635": "FLORESTA",
    "D661": "FONDACHELLI-FANTINA",
    "D733": "FORZA D'AGRÓ",
    "D765": "FRANCAVILLA DI SICILIA",
    "D793": "FRAZZANÒ",
    "D824": "FURCI SICULO",
    "D825": "FURNARI",
    "D844": "GAGGI",
    "D861": "GALATI MAMERTINO",
    "D885": "GALLODORO",
    "E014": "GIARDINI-NAXOS",
    "E043": "GIOIOSA MAREA",
    "E142": "GRANITI",
    "E233": "GUALTIERI SICAMINO'",
    "E374": "ITALA",
    "E523": "LENI",
    "E555": "LETOJANNI",
    "E571": "LIBRIZZI",
    "E594": "LIMINA",
    "E606": "LIPARI",
    "E674": "LONGI",
    "E855": "MALFA",
    "E869": "MALVAGNA",
    "E876": "MANDANICI",
    "F066": "MAZZARRÀ SANT'ANDREA",
    "F147": "MERÌ",
    "F158": "MESSINA",
    "F206": "MILAZZO",
    "F210": "MILITELLO ROSMARINO",
    "F242": "MIRTO",
    "F251": "MISTRETTA",
    "F277": "MOIO ALCANTARA",
    "F359": "MONFORTE SAN GIORGIO",
    "F368": "MONGIUFFI MELIA",
    "F395": "MONTAGNAREALE",
    "F400": "MONTALBANO ELICONA",
    "F772": "MOTTA CAMASTRA",
    "F773": "MOTTA D'AFFERMO",
    "F848": "NASO",
    "F901": "NIZZA DI SICILIA",
    "F951": "NOVARA DI SICILIA",
    "G036": "OLIVERI",
    "G209": "PACE DEL MELA",
    "G234": "PAGLIARA",
    "G377": "PATTI",
    "G522": "PETTINEO",
    "G699": "PIRAINO",
    "H151": "RACCUJA",
    "H228": "REITANO",
    "H405": "ROCCAFIORITA",
    "H418": "ROCCALUMERA",
    "H380": "ROCCAVALDINA",
    "H455": "ROCCELLA VALDEMONE",
    "H479": "RODÌ MILICI",
    "H519": "ROMETTA",
    "H842": "SAN FILIPPO DEL MELA",
    "H850": "SAN FRATELLO",
    "H982": "SAN MARCO D'ALUNZIO",
    "I084": "SAN PIER NICETO",
    "I086": "SAN PIERO PATTI",
    "I147": "SAN SALVATORE DI FITALIA",
    "I184": "SANTA DOMENICA VITTORIA",
    "I199": "SANT'AGATA DI MILITELLO",
    "I215": "SANT'ALESSIO SICULO",
    "I220": "SANTA LUCIA DEL MELA",
    "I254": "SANTA MARINA SALINA",
    "I283": "SANT'ANGELO DI BROLO",
    "I311": "SANTA TERESA DI RIVA",
    "I328": "SAN TEODORO",
    "I370": "SANTO STEFANO DI CAMASTRA",
    "I420": "SAPONARA",
    "I477": "SAVOCA",
    "I492": "SCALETTA ZANCLEA",
    "I747": "SINAGRA",
    "I881": "SPADAFORA",
    "L042": "TAORMINA",
    "L271": "TORREGROTTA",
    "L308": "TORTORICI",
    "L431": "TRIPI",
    "L478": "TUSA",
    "L482": "UCRIA",
    "L561": "VALDINA",
    "L735": "VENETICO",
    "L950": "VILLAFRANCA TIRRENA",
    "M210": "TERME VIGLIATORE",
    "M211": "ACQUEDOLCI",
    "M286": "TORRENOVA",

    "A195": "ALIA",
    "A202": "ALIMENA",
    "A203": "ALIMINUSA",
    "A229": "ALTAVILLA MILICIA",
    "A239": "ALTOFONTE",
    "A546": "BAGHERIA",
    "A592": "BALESTRATE",
    "A719": "BAUCINA",
    "A764": "BELMONTE MEZZAGNO",
    "A882": "BISACQUINO",
    "A946": "BOLOGNETTA",
    "A958": "BOMPIETRO",
    "A991": "BORGETTO",
    "B315": "CACCAMO",
    "B430": "CALTAVUTURO",
    "B533": "CAMPOFELICE DI FITALIA",
    "B532": "CAMPOFELICE DI ROCCELLA",
    "B535": "CAMPOFIORITO",
    "B556": "CAMPOREALE",
    "B645": "CAPACI",
    "B780": "CARINI",
    "C067": "CASTELBUONO",
    "C074": "CASTELDACCIA",
    "C135": "CASTELLANA SICULA",
    "C344": "CASTRONOVO DI SICILIA",
    "C420": "CEFALÀ DIANA",
    "C421": "CEFALÙ",
    "C496": "CERDA",
    "C654": "CHIUSA SCLAFANI",
    "C696": "CIMINNA",
    "C708": "CINISI",
    "C871": "COLLESANO",
    "C968": "CONTESSA ENTELLINA",
    "D009": "CORLEONE",
    "D567": "FICARAZZI",
    "D907": "GANGI",
    "D977": "GERACI SICULO",
    "E013": "GIARDINELLO",
    "E055": "GIULIANA",
    "E074": "GODRANO",
    "E149": "GRATTERI",
    "E337": "ISNELLO",
    "E350": "ISOLA DELLE FEMMINE",
    "E459": "LASCARI",
    "E541": "LERCARA FRIDDI",
    "E957": "MARINEO",
    "F184": "MEZZOJUSO",
    "F246": "MISILMERI",
    "F377": "MONREALE",
    "F544": "MONTELEPRE",
    "F553": "MONTEMAGGIORE BELSITO",
    "G263": "PALAZZO ADRIANO",
    "G273": "PALERMO",
    "G348": "PARTINICO",
    "G510": "PETRALIA SOPRANA",
    "G511": "PETRALIA SOTTANA",
    "G543": "PIANA DEGLI ALBANESI",
    "G792": "POLIZZI GENEROSA",
    "G797": "POLLINA",
    "H070": "PRIZZI",
    "H422": "ROCCAMENA",
    "H428": "ROCCAPALUMBA",
    "H797": "SAN CIPIRELLO",
    "H933": "SAN GIUSEPPE JATO",
    "I028": "SAN MAURO CASTELVERDE",
    "I174": "SANTA CRISTINA GELA",
    "I188": "SANTA FLAVIA",
    "I534": "SCIARA",
    "I541": "SCLAFANI BAGNI",
    "L112": "TERMINI IMERESE",
    "L131": "TERRASINI",
    "L282": "TORRETTA",
    "L317": "TRABIA",
    "L332": "TRAPPETO",
    "L519": "USTICA",
    "L603": "VALLEDOLMO",
    "L740": "VENTIMIGLIA DI SICILIA",
    "L837": "VICARI",
    "L916": "VILLABATE",
    "L951": "VILLAFRATI",
    "I538": "SCILLATO",
    "M268": "BLUFI",
    "A014": "ACATE",
    "C612": "CHIARAMONTE GULFI",
    "C927": "COMISO",
    "E016": "GIARRATANA",
    "E366": "ISPICA",
    "F258": "MODICA",
    "F610": "MONTEROSSO ALMO",
    "G953": "POZZALLO",
    "H163": "RAGUSA",
    "I178": "SANTA CROCE CAMERINA",
    "I535": "SCICLI",
    "M088": "VITTORIA",
    "A494": "AUGUSTA",
    "A522": "AVOLA",
    "B237": "BUCCHERI",
    "B287": "BUSCEMI",
    "B603": "CANICATTINI BAGNI",
    "B787": "CARLENTINI",
    "C006": "CASSARO",
    "D540": "FERLA",
    "D636": "FLORIDIA",
    "D768": "FRANCOFONTE",
    "E532": "LENTINI",
    "F107": "MELILLI",
    "F943": "NOTO",
    "G211": "PACHINO",
    "G267": "PALAZZOLO ACREIDE",
    "H574": "ROSOLINI",
    "I754": "SIRACUSA",
    "I785": "SOLARINO",
    "I864": "SORTINO",
    "M257": "PORTOPALO DI CAPO PASSERO",
    "M279": "PRIOLO GARGALLO",
    "A176": "ALCAMO",
    "B288": "BUSETO PALIZZOLO",
    "B385": "CALATAFIMI-SEGESTA",
    "B521": "CAMPOBELLO DI MAZARA",
    "C130": "CASTELLAMMARE DEL GOLFO",
    "C286": "CASTELVETRANO",
    "D234": "CUSTONACI",
    "D423": "ERICE",
    "D518": "FAVIGNANA",
    "E023": "GIBELLINA",
    "E974": "MARSALA",
    "F061": "MAZARA DEL VALLO",
    "G208": "PACECO",
    "G315": "PANTELLERIA",
    "G347": "PARTANNA",
    "G767": "POGGIOREALE",
    "H688": "SALAPARUTA",
    "H700": "SALEMI",
    "I291": "SANTA NINFA",
    "I407": "SAN VITO LO CAPO",
    "L331": "TRAPANI",
    "G319": "VALDERICE",
    "M081": "VITA",
    "M281": "PETROSINO",
    "M432": "MISILISCEMI",
}

    return records, stats

DEFAULT_CATAS_FILES = [
    "codici_catastali.xlsx",
    "codici_catastali.xls",
    "codici_catastali.csv",
]
def build_catasto_mapping(catasto_xlsx_path: str, target_codes: set):
    """
    Legge l'Excel "codici catastali".
    - Comune: colonna A (1a colonna)
    - Colonna del codice catastale: auto-detect (massimo overlap con i codici del TXT); spesso si trova in colonna D
    """
    import pandas as pd

    df = pd.read_excel(catasto_xlsx_path, dtype=str, engine="openpyxl")
    if df.shape[1] < 4:
        raise ValueError("La tabella codici catastali deve avere almeno 4 colonne (Comune in colonna A e codice catastale in D).")

    comune_col = df.columns[0]  # colonna A
    # normalizza colonne
    norm_series = lambda s: s.fillna("").astype(str).str.strip().str.upper()

    df_norm = df.copy()
    for col in df_norm.columns:
        df_norm[col] = norm_series(df_norm[col])

    # candidati: tutte le colonne tranne la D
    candidates = [c for c in df_norm.columns if c != comune_col]

    # scegli la colonna con maggior overlap coi codici che ci interessano
    best = None
    for col in candidates:
        vals = df_norm[col]
        match_count = vals.str.match(r"^[A-Z0-9]{4}$").sum()
        overlap = vals.isin(target_codes).sum()
        score = (overlap, match_count)
        if best is None or score > best[0]:
            best = (score, col)

    code_col = best[1]

    mapping = {}
    for _, row in df_norm.iterrows():
        code = row[code_col]
        comune = normalize_comune(row[comune_col])
        if code and re.match(r"^[A-Z0-9]{4}$", code):
            mapping[code] = comune  # può essere "" -> verrà gestito come Non Trovato

    return mapping, code_col, comune_col


def build_catasto_mapping_from_df(df, target_codes: set):
    """Stessa logica di build_catasto_mapping ma partendo da un DataFrame già caricato."""
    import pandas as pd

    if df.shape[1] < 2:
        raise ValueError("La tabella codici catastali deve avere almeno 2 colonne (Comune in colonna A).")

    comune_col = df.columns[0]
    norm_series = lambda s: s.fillna("").astype(str).str.strip().str.upper()

    df_norm = df.copy()
    for col in df_norm.columns:
        df_norm[col] = norm_series(df_norm[col])

    candidates = [c for c in df_norm.columns if c != comune_col]

    best = None
    for col in candidates:
        vals = df_norm[col]
        match_count = vals.str.match(r"^[A-Z0-9]{4}$").sum()
        overlap = vals.isin(target_codes).sum()
        score = (overlap, match_count)
        if best is None or score > best[0]:
            best = (score, col)

    code_col = best[1]

    mapping = {}
    for _, row in df_norm.iterrows():
        code = row[code_col]
        comune = normalize_comune(row[comune_col])
        if code and re.match(r"^[A-Z0-9]{4}$", code):
            mapping[code] = comune

    return mapping, code_col, comune_col


# -----------------------------
# Costruzione tabelle pivot
# -----------------------------
def build_pivot(df):
    """
    Ritorna tabella:
    index = comune
    columns = cod_funzione
    values = conteggio
    """
    import pandas as pd

    if df.empty:
        return pd.DataFrame()

    # Comune vuoto non dovrebbe più esserci; ma se c'è, lo rendiamo "Non Trovato"
    work = df.copy()
    work["comune"] = work["comune"].apply(lambda x: x if x else "Non Trovato")

    pivot = pd.crosstab(work["comune"], work["cod_funzione"])

    # ordina colonne in modo “umano”
    pivot = pivot.reindex(columns=sort_codefun_columns(pivot.columns))

    # ordina comuni alfabeticamente
    pivot = pivot.sort_index()

    return pivot


# -----------------------------
# Scrittura Excel (openpyxl)
# -----------------------------
def write_pivot_to_sheet(ws, provincia: str, pivot):
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    center = Alignment(vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    if pivot.empty:
        # Foglio vuoto: scrivi messaggio + totali
        ws["A1"] = "Nessun record per questo criterio."
        ws["A1"].font = bold

        ws["A3"] = ""  # riga bianca “di separazione”
        ws["A4"] = f"TOTALE PROVINCIA ({provincia})"
        ws["A4"].font = bold
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 35
        return

    # Header
    headers = ["Comune"] + [f"Codice {c}" for c in pivot.columns]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = center

    # Data
    r = 2
    for comune, row in pivot.iterrows():
        ws.cell(row=r, column=1, value=comune).alignment = left
        for j, c in enumerate(pivot.columns, start=2):
            ws.cell(row=r, column=j, value=int(row[c])).alignment = center
        r += 1

    # Riga bianca
    blank_row = r
    r += 1

    # Totali provincia
    totals_row = r
    ws.cell(row=totals_row, column=1, value=f"TOTALE PROVINCIA ({provincia})").font = bold
    ws.cell(row=totals_row, column=1).alignment = left

    col_sums = pivot.sum(axis=0)
    for j, c in enumerate(pivot.columns, start=2):
        cell = ws.cell(row=totals_row, column=j, value=int(col_sums[c]))
        cell.font = bold
        cell.alignment = center

    # Formattazione base
    ws.freeze_panes = "A2"

    # Auto-width semplice
    max_col = 1 + len(pivot.columns)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for rr in range(1, totals_row + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)


def write_combined_sheet_with_province_totals(ws, df_all):
    """
    Versione per il file annuale: ordina le province richieste prima,
    poi le altre; inserisce riga bianca e riga totale per provincia quando cambia sigla.
    """
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd

    bold = Font(bold=True)
    center = Alignment(vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    if df_all.empty:
        ws["A1"] = "Nessun record per questo criterio."
        ws["A1"].font = bold
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 35
        return

    # Mantieni le province così come sono; ordina quelle richieste per prime
    PROVINCE_ORDER = ["AG", "CL", "CT", "EN", "ME", "PA", "RG", "SR", "TP"]
    pivot = pd.crosstab([df_all["provincia"], df_all["comune"]], df_all["cod_funzione"])
    # Ensure all code columns 0-6 are present (even if not in data)
    required_cols = ["0", "1", "2", "3", "4", "5", "6"]
    all_cols = sorted(set(required_cols) | set(pivot.columns), key=lambda x: (x not in required_cols, required_cols.index(x) if x in required_cols else 999, x))
    pivot = pivot.reindex(columns=all_cols, fill_value=0)
    pivot = pivot.sort_index()

    # Header labels mapping (codici 0..6)
    header_label_map = {
        "0": "Concomitanti",
        "1": "Revoche",
        "2": "Deleghe",
        "3": "Eliminate",
        "4": "Trasferite A",
        "5": "Trasferite Da",
        "6": "Cambio Sede",
    }

    headers = ["Provincia", "Comune"] + [header_label_map.get(str(c), f"Codice {c}") for c in pivot.columns]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = center

    # Province order and name map
    PROV_NAME = {
        "AG": "Agrigento",
        "CL": "Caltanissetta",
        "CT": "Catania",
        "EN": "Enna",
        "ME": "Messina",
        "PA": "Palermo",
        "RG": "Ragusa",
        "SR": "Siracusa",
        "TP": "Trapani",
    }

    provinces_present = list(pivot.index.get_level_values(0).unique())
    ordered_provinces = [p for p in PROVINCE_ORDER if p in provinces_present]
    other_provs = sorted([p for p in provinces_present if p not in PROVINCE_ORDER])
    ordered_provinces += other_provs

    r = 2
    last_was_prov_total = False
    for prov in ordered_provinces:
        try:
            sub = pivot.loc[prov]
        except KeyError:
            continue

        # sub is a DataFrame indexed by comune
        for comune, row in sub.iterrows():
            ws.cell(row=r, column=1, value=prov).alignment = left
            ws.cell(row=r, column=2, value=comune).alignment = left
            for j, c in enumerate(pivot.columns, start=3):
                ws.cell(row=r, column=j, value=int(row[c])).alignment = center
            r += 1
            last_was_prov_total = False

        # If the province is one of the specified Sicily provinces, add a blank row with province totals.
        if prov in PROVINCE_ORDER:
            # Riga con totali della provincia (etichetta + valori)
            totals_row = r
            name = PROV_NAME.get(prov, prov)

            # Etichetta Totale Provincia
            ws.cell(row=totals_row, column=1, value=f"Totale {name}").font = bold
            ws.cell(row=totals_row, column=1).alignment = left

            # Totali nelle colonne dei codici funzione (da colonna 3 in poi)
            col_sums = sub.sum(axis=0)
            for j, c in enumerate(pivot.columns, start=3):
                cell = ws.cell(row=totals_row, column=j, value=int(col_sums[c]))
                cell.font = bold
                cell.alignment = center

            # Applica bordo esterno continuo sulla riga dei totali (top/bottom su tutte le celle,
            # bordi verticali solo sui primi/ultimi convergenti con le celle che contengono valori)
            last_col = 2 + len(pivot.columns)
            side = Side(style='medium', color='000000')
            border_tb = Border(top=side, bottom=side)
            border_left = Border(left=side, top=side, bottom=side)
            border_right = Border(right=side, top=side, bottom=side)
            for col_idx in range(1, last_col + 1):
                cell = ws.cell(row=totals_row, column=col_idx)
                if col_idx == 1:
                    cell.border = border_left
                elif col_idx == last_col:
                    cell.border = border_right
                else:
                    cell.border = border_tb
            r += 1
            last_was_prov_total = True
        else:
            # For other provinces do not add per-province totals; continue listing
            continue

    # Se la riga precedente è un totale provinciale, inseriamo una riga bianca di separazione
    # Inserisci Totale Altre se abbiamo elencato province non-Sicilia
    if other_provs:
        try:
            pivot_other = pivot.loc[other_provs]
        except KeyError:
            pivot_other = pivot[pivot.index.get_level_values(0).isin(other_provs)]

        totals_row = r
        ws.cell(row=totals_row, column=1, value="Totale Altre").font = bold
        ws.cell(row=totals_row, column=1).alignment = left
        col_sums = pivot_other.sum(axis=0)
        for j, c in enumerate(pivot.columns, start=3):
            cell = ws.cell(row=totals_row, column=j, value=int(col_sums[c]))
            cell.font = bold
            cell.alignment = center

        last_col = 2 + len(pivot.columns)
        side = Side(style='medium', color='000000')
        border_tb = Border(top=side, bottom=side)
        border_left = Border(left=side, top=side, bottom=side)
        border_right = Border(right=side, top=side, bottom=side)
        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row=totals_row, column=col_idx)
            if col_idx == 1:
                cell.border = border_left
            elif col_idx == last_col:
                cell.border = border_right
            else:
                cell.border = border_tb
        r += 1
        last_was_prov_total = True

    try:
        if last_was_prov_total:
            r += 1
    except NameError:
        pass

    # Totale generale finale
    ws.cell(row=r, column=1, value=f"Totale Generale").font = bold
    ws.cell(row=r, column=1).alignment = left
    col_sums = pivot.sum(axis=0)
    for j, c in enumerate(pivot.columns, start=3):
        cell = ws.cell(row=r, column=j, value=int(col_sums[c]))
        cell.font = bold
        cell.alignment = center

    # Applica bordo esterno sulla riga del Totale Generale
    last_col = 2 + len(pivot.columns)
    side = Side(style='medium', color='000000')
    border_tb = Border(top=side, bottom=side)
    border_left = Border(left=side, top=side, bottom=side)
    border_right = Border(right=side, top=side, bottom=side)
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=r, column=col_idx)
        if col_idx == 1:
            cell.border = border_left
        elif col_idx == last_col:
            cell.border = border_right
        else:
            cell.border = border_tb

    # Formattazione base
    ws.freeze_panes = "A2"

    # Auto-width semplice
    max_col = 2 + len(pivot.columns)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for rr in range(1, r + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)


def write_combined_sheet(ws, df_all):
    """
    Scrive un singolo foglio contenente tutte le province.
    Colonne: Provincia, Comune, Codice <funzione>...
    """
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import pandas as pd

    bold = Font(bold=True)
    center = Alignment(vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    if df_all.empty:
        ws["A1"] = "Nessun record per questo criterio."
        ws["A1"].font = bold
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = 35
        return

    pivot = pd.crosstab([df_all["provincia"], df_all["comune"]], df_all["cod_funzione"])
    # Ensure all code columns 0-6 are present (even if not in data)
    required_cols = ["0", "1", "2", "3", "4", "5", "6"]
    all_cols = sorted(set(required_cols) | set(pivot.columns), key=lambda x: (x not in required_cols, required_cols.index(x) if x in required_cols else 999, x))
    pivot = pivot.reindex(columns=all_cols, fill_value=0)
    pivot = pivot.sort_index()

    # Header labels mapping (codici 0..6)
    header_label_map = {
        "0": "Concomitanti",
        "1": "Revoche",
        "2": "Deleghe",
        "3": "Eliminate",
        "4": "Trasferite A",
        "5": "Trasferite Da",
        "6": "Cambio Sede",
    }

    headers = ["Provincia", "Comune"] + [header_label_map.get(str(c), f"Codice {c}") for c in pivot.columns]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.alignment = center

    # Data
    r = 2
    for (prov, comune), row in pivot.iterrows():
        ws.cell(row=r, column=1, value=prov).alignment = left
        ws.cell(row=r, column=2, value=comune).alignment = left
        for j, c in enumerate(pivot.columns, start=3):
            ws.cell(row=r, column=j, value=int(row[c])).alignment = center
        r += 1

    # Totali generali
    totals_row = r
    ws.cell(row=totals_row, column=1, value=f"Totale Generale").font = bold
    ws.cell(row=totals_row, column=1).alignment = left

    col_sums = pivot.sum(axis=0)
    for j, c in enumerate(pivot.columns, start=3):
        cell = ws.cell(row=totals_row, column=j, value=int(col_sums[c]))
        cell.font = bold
        cell.alignment = center

    # Formattazione base
    ws.freeze_panes = "A2"

    # Auto-width semplice
    max_col = 2 + len(pivot.columns)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for rr in range(1, totals_row + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)


def save_workbook_complete(df_all, out_path: str):
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet(title=safe_sheet_name("Prospetto Annuale Sicilia"))
    # per il file annuale vogliamo ordinamento e totali per provincia
    write_combined_sheet_with_province_totals(ws, df_all)

    wb.save(out_path)


def save_monthly_workbooks_pa_tp(df_all, month_order: list, out_dir: str, prefix: str = "Riepilogo_PA_TP", provinces: Optional[List[str]] = None) -> List[Path]:
    """
    Crea 1 excel per ogni mese (qui “mese” = 1 file TXT selezionato),
    contenente le province specificate in `provinces` (lista di sigle).
    Restituisce la lista di Path dei file creati.
    """
    os.makedirs(out_dir, exist_ok=True)

    provinces_list = provinces if provinces else ["PA", "TP", "ME"]

    from openpyxl import Workbook
    months_it = ["Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno","Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]

    def month_label_from_key(k: str) -> str:
        m = re.search(r"-(0[1-9]|1[0-2])$", k)
        if m:
            idx = int(m.group(1))
            return months_it[idx-1]
        m = re.match(r"M(0?[1-9]|1[0-2])$", k)
        if m:
            idx = int(m.group(1))
            return months_it[idx-1]
        m = re.search(r"(0[1-9]|1[0-2])", k)
        if m:
            idx = int(m.group(1))
            return months_it[idx-1]
        return k

    created = []
    for month_key in month_order:
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet(title=safe_sheet_name(month_key))
        subset = df_all[(df_all["month_key"] == month_key) & (df_all["provincia"].isin(provinces_list))]
        write_combined_sheet_with_province_totals(ws, subset)

        month_label = month_label_from_key(month_key)
        prov_tag = "+".join(provinces_list)
        filename = f"{month_label} {prov_tag}.xlsx"
        out_path = os.path.join(out_dir, filename)
        wb.save(out_path)
        created.append(Path(out_path))

    return created


# -----------------------------
# Main
# -----------------------------
def run(
    file_txt: List[Path],
    provinces_selected: List[str] = None,
    export_annual: str = "SI",
    export_monthly: str = "SI",
    out_dir: Path = Path("."),
) -> List[Path]:
    """
    Esegue l'elaborazione senza interazione.

    Parametri:
      - file_txt: lista di percorsi ai file TXT (obbligatoria)
      - monthly_provinces: lista di sigle province da includere nei mensili (es. ["PA","TP","ME"])
      - out_dir: cartella in cui scrivere i file

    Restituisce: lista di Path dei file creati.
    """
    if provinces_selected is None:
        provinces_selected = ["AG", "CL", "CT", "EN", "ME", "PA", "RG", "SR", "TP"]

    if not file_txt:
        raise ValueError("file_txt è obbligatorio e deve contenere almeno un file TXT.")

    print(f"--> Caricati {len(file_txt)} file TXT. Inizio elaborazione...")

    import pandas as pd

    txt_paths = [str(p) for p in file_txt]
    month_keys = [infer_month_key_from_filename(p, idx) for idx, p in enumerate(txt_paths, start=1)]

    # Parse
    all_records = []
    global_stats = {
        "read_lines": 0,
        "kept_allowed_pr": 0,
        "kept_by_sede": 0,
        "discarded": 0,
        "short_lines_padded": 0,
    }
    for p, mk in zip(txt_paths, month_keys):
        recs, st = parse_txt_file(p, mk)
        all_records.extend(recs)
        for k in global_stats:
            global_stats[k] += st.get(k, 0)

    if not all_records:
        return []

    df = pd.DataFrame(all_records)

    # Lookup comune mancante (solo se serve)
    missing_mask = df["comune"].isna() | (df["comune"].astype(str).str.strip() == "")
    code_col = None
    comune_col = None
    if missing_mask.any():
        target_codes = set(df.loc[missing_mask, "cod_catastale"].astype(str).str.strip().str.upper())
        mapping = {}
        if EMBEDDED_CATASTO:
            mapping = {k.strip().upper(): normalize_comune(v) for k, v in EMBEDDED_CATASTO.items()}
            code_col = "<embedded>"
            comune_col = "<embedded>"
        else:
            # cerca file predefiniti
            script_dir = Path(__file__).parent
            found_path = None
            for sp in [Path.cwd(), script_dir, script_dir.parent]:
                for fname in DEFAULT_CATAS_FILES:
                    pth = sp / fname
                    if pth.exists():
                        found_path = pth
                        break
                if found_path:
                    break
            if found_path:
                try:
                    if found_path.suffix.lower() in (".xlsx", ".xls"):
                        mapping, code_col, comune_col = build_catasto_mapping(str(found_path), target_codes)
                    else:
                        import pandas as pd
                        try:
                            df_cat = pd.read_csv(str(found_path), sep=';', dtype=str, encoding='utf-8', engine='python')
                        except Exception:
                            df_cat = pd.read_csv(str(found_path), dtype=str, engine='python')
                        mapping, code_col, comune_col = build_catasto_mapping_from_df(df_cat, target_codes)
                except Exception:
                    mapping = {}
                    code_col = None
                    comune_col = None

        if mapping:
            filled = (
                df.loc[missing_mask, "cod_catastale"]
                .astype(str).str.strip().str.upper()
                .map(mapping)
            )
            filled = filled.fillna("Non Trovato").apply(lambda x: x if x else "Non Trovato")
            df.loc[missing_mask, "comune"] = filled
        else:
            df.loc[missing_mask, "comune"] = "Non Trovato"

    # Salvataggi
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    created: List[Path] = []
    first_stem = Path(txt_paths[0]).stem if txt_paths else "Riepilogo"

    do_annual = True if str(export_annual).upper() == "SI" else False
    do_monthly = True if str(export_monthly).upper() == "SI" else False

    if do_annual:
        out_full = out_dir / "Prospetto Annuale Sicilia.xlsx"
        save_workbook_complete(df, str(out_full))
        created.append(out_full)

    if do_monthly:
        # build provinces list from individual checkbox params
        provinces = provinces_selected
        if not provinces:
            provinces = ["PA", "TP", "ME"]
        monthly_created = save_monthly_workbooks_pa_tp(df, month_keys, str(out_dir), prefix=f"{first_stem}_", provinces=provinces)
        created.extend(monthly_created)

    # Se sono stati creati più di 2 file, li comprime in un unico ZIP
    if len(created) > 2:
        zip_path = out_dir / "Archivio_Output_Completo.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in created:
                zf.write(p, arcname=p.name)
        return [zip_path]

    return created
