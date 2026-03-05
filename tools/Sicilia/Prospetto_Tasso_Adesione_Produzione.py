import os
import io
import sys
import json
import time
import subprocess
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import unicodedata
import difflib
import re # Risk 6: Regex validation

from core.toolkit import ctx

# Safe imports
try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "requests"])
    import requests

try:
    from dotenv import load_dotenv
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "python-dotenv"])
    from dotenv import load_dotenv

# ─── API CONFIG ──────────────────────────────────────────────────────────────
_ENV_PATH = Path(__file__).resolve().parent / "API" / ".env"
if _ENV_PATH.exists():
    load_dotenv(_ENV_PATH)
LOCATIONIQ_API_KEY = os.getenv("LOCATIONIQ_API_KEY", "")

# =============================================================================
# TOOLBOX CONFIGURATION
# =============================================================================
TOOL = {
    'id': 'prospetto_tasso_adesione_produzione',
    'name': 'Prospetto Tasso Adesione/Produz',
    'description': (
        "#### 📌 1. FINALITÀ DEL TOOL\n"
        "Elabora i tracciati **SINDRINN** e **SINDMENS** della Sicilia, identifica i comuni tramite "
        "codice catastale e popola automaticamente il **template Excel** *Sicilia_Prospetto_Variazioni* "
        "con dati aggregati per provincia, formule e formattazione professionale.\n\n"

        "#### 🚀 2. COME UTILIZZARLO\n"
        "1. **Carica File:** Seleziona uno o più file SINDRINN o SINDMENS.\n"
        "2. **Whitelist Sedi:** Verifica che i codici sede siciliani siano corretti (modificabili sotto).\n"
        "3. **Elaborazione:** Il tool riconosce il tipo di file, filtra per sede, cerca il codice "
        "catastale e associa il nome del comune.\n"
        "4. **Fallback API:** Se il comune non è nel lookup locale, interviene l'API LocationIQ.\n"
        "5. **Template:** Il tool compila il prospetto Excel con dati raggruppati per provincia "
        "siciliana, sezione *Fuori Provincia o Assente* e sezione *Estero*.\n"
        "6. **Download:** Scarica il file Excel compilato.\n\n"

        "#### 🧠 3. LOGICA DI ELABORAZIONE\n"
        "- **SINDRINN:** Codice catastale a pos. 312 (4 car.). Conta come *Rinnovo* (col. C).\n"
        "- **SINDMENS:** Codice catastale a pos. 339 (4 car.). Classificato per *cod_funzione*: "
        "0→Concomitanti (E), 2→Deleghe (G), 1→Revoche (I), 3→Eliminate (K).\n"
        "- **Lookup:** File *Elenco-comuni-italiani.xlsx* (Col U = Codice, Col G = Nome).\n"
        "- **Fallback:** Se il codice non è nel lookup → API LocationIQ → se fallisce → *Provincia non esistente*.\n"
        "- **Correzioni:** Dizionario personalizzabile per correggere nomi errati dall'API.\n\n"

        "#### � 4. TEMPLATE EXCEL\n"
        "Il prospetto viene popolato con:\n"
        "- **Raggruppamento per provincia** siciliana (AG, CL, CT, EN, ME, PA, RG, SR, TP) con subtotali.\n"
        "- **Sezione 'Fuori Provincia o Assente'** per comuni non siciliani e dati mancanti.\n"
        "- **Sezione 'Estero'** per record esteri.\n"
        "- **Formule automatiche:** Tassi (F, H, J, L, N, P) e differenze (M, O) con gestione errori.\n"
        "- **Formattazione:** Bordi, grassetto per subtotali/totali, valori centrati, 2 decimali.\n"
        "- **Totale Generale** con formule SUM che sommano i subtotali.\n\n"

        "#### 📂 5. RISULTATO FINALE\n"
        "- **Sicilia_Prospetto_Variazioni_Compilato.xlsx** — il template compilato con dati, "
        "formule e formattazione.\n"
        "- **Dashboard riepilogativa** persistente con conteggi, stato API e dettaglio elaborazione."
    ),
    'inputs': [
        {'key': 'files', 'label': 'File SINDRINN / SINDMENS', 'type': 'file_multi', 'required': True}
    ],
    'params': [
        {'key': 'preview', 'label': 'Anteprima File Caricati', 'type': 'dynamic_info', 'function': 'preview_files', 'section': 'Anteprima Riconoscimento File'},
        {'key': 'whitelist_sedi', 'label': 'Codici Sede Sicilia (Whitelist)', 'type': 'textarea', 'default': '2193, 2192, 0100, 8291, 7691, 5593, 4892, 0192, 2190, 1800, 0191, 5503, 8292, 2100, 5592, 2800, 1890, 2194, 7690, 0190, 8290, 2196, 8293, 4800, 4894, 5596, 6590, 2890, 7601, 5500, 5502, 5591, 2191, 4890, 5594, 6500, 4893, 4891, 0101, 7600, 5590, 8200, 6591', 'help': 'Inserisci i codici sede da considerare, separati da virgola.'},
        {'key': 'corrections_manager', 'label': '📝 Gestione Correzioni Nomi', 'type': 'dynamic_info', 'function': 'manage_corrections', 'section': 'Altre Opzioni'},
        {'key': 'template_manager', 'label': '📝 Gestione Template', 'type': 'dynamic_info', 'function': 'manage_template', 'section': 'Template Precaricato'},
        {'key': 'clear_cache_button', 'label': '🗑️ Svuota Cache API', 'type': 'dynamic_info', 'function': 'handle_clear_cache', 'section': 'Cache Geocoding', 'help': '⚠️ PROMEMORIA: Svuota la cache solo se hai modificato la logica API o se vuoi forzare una nuova geocodifica per tutti i comuni. La cache velocizza notevolmente l\'elaborazione riutilizzando i risultati precedenti.'},
        {'key': 'enable_validation', 'label': '✅ Attiva Validazione Dati Avanzata', 'type': 'checkbox', 'default': False, 'help': 'Se attivo, evidenzia nel file Excel i comuni duplicati e le province errate.'}
    ]
}


# ─────────────────────────────────────────────────────────────────────────────
#  GEOCODING CACHE
# ─────────────────────────────────────────────────────────────────────────────# Cache & Corrections paths
_API_DIR = Path(__file__).parent.parent / "Amministrazione" / "API" / "x-Estrazione_Deleghe_estere"
_CACHE_FILE = _API_DIR / "geocode_cache.json"
_CORRECTIONS_FILE = _API_DIR / "name_corrections.json"

# Default name corrections for common API typos/abbreviations
_DEFAULT_CORRECTIONS = {
    "RAGUS": "Ragusa",
    "MODIC": "Modica",
    "TREMESTIERI ETENEO": "Tremestieri Etneo",
    "PORTOPALO DI CAPOPASSERO": "Portopalo di Capo Passero",
    "VILLASMUNDO MELILLI": "Melilli",
    "MISTERBIAN": "Misterbianco"
}


# ─── RISK 6: CONFIGURAZIONE TRACCIATO ────────────────────────────────────────
# Configurazioni posizioni fisse (1-based come da documentazione INPS)
TRACCIATO = {
    "SINDRINN": {
        "comune": {"start": 312, "len": 4},
        "sede": {"start": 318, "len": 4},
        "addr": {"start": 161, "len": 52},
        "city": {"start": 248, "len": 30},
        "cap": {"start": 281, "len": 9},
        "fun": None
    },
    "SINDMENS": {
        "comune": {"start": 339, "len": 4},
        "sede": {"start": 345, "len": 4},
        "addr": {"start": 123, "len": 52},
        "city": {"start": 235, "len": 36},
        "cap": {"start": 274, "len": 9},
        "fun": {"start": 7, "len": 1}
    }
}


def _load_cache() -> Dict[str, str]:
    """Load geocoding cache from disk."""
    if _CACHE_FILE.exists():
        try:
            with open(_CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_cache(cache: Dict[str, str]):
    """Save geocode cache to JSON file."""
    try:
        _API_DIR.mkdir(parents=True, exist_ok=True)
        with open(_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving cache: {e}")


def _load_corrections() -> Dict[str, str]:
    """Load custom name corrections from JSON file."""
    if _CORRECTIONS_FILE.exists():
        try:
            with open(_CORRECTIONS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading corrections: {e}")
            return _DEFAULT_CORRECTIONS.copy()
    return _DEFAULT_CORRECTIONS.copy()


def _save_corrections(corrections: Dict[str, str]):
    """Save custom name corrections to JSON file."""
    try:
        _API_DIR.mkdir(parents=True, exist_ok=True)
        with open(_CORRECTIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(corrections, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving corrections: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  LOCATIONIQ GEOCODING
# ─────────────────────────────────────────────────────────────────────────────

def _geocode_single_query(query: str, api_key: str,
                          valid_names: List[str] = None,
                          max_retries: int = 3) -> Optional[str]:
    """
    Geocode a single query via LocationIQ with retry + exponential backoff.
    Returns the municipality name in UPPERCASE or None.
    If valid_names is provided, only accepts results that match a known municipality.
    """
    if not api_key or not query.strip():
        return None

    url = "https://us1.locationiq.com/v1/search.php"
    params = {
        'key': api_key,
        'q': query,
        'format': 'json',
        'accept-language': 'it',
        'addressdetails': 1,
        'limit': 1,
        'countrycodes': 'it'
    }

    for attempt in range(max_retries):
        try:
            r = requests.get(url, params=params, timeout=5)
            if r.status_code == 200:
                data = r.json()
                if data and isinstance(data, list) and len(data) > 0:
                    addr = data[0].get('address', {})
                    # Try specific municipality fields only (no county)
                    city = (addr.get('city') or addr.get('town') or
                            addr.get('village') or addr.get('municipality'))
                    if city:
                        city_upper = city.upper()
                        # Validate against known municipalities if list provided
                        if valid_names:
                            matched, found_in_excel = normalize_city_name(city_upper, valid_names)
                            if found_in_excel:
                                # Found in Excel list → valid
                                return city_upper
                            # Not in list → skip
                            return None
                        return city_upper
                return None
            elif r.status_code == 429:
                time.sleep(2 ** (attempt + 1))
            elif r.status_code == 404:
                return None
            else:
                time.sleep(1)
        except Exception:
            time.sleep(1)

    return None


def _geocode_location(address: str, city: str, zip_code: str,
                      api_key: str, cache: Dict[str, str],
                      valid_names: List[str] = None,
                      force_sicilia: bool = True) -> Optional[Tuple[str, str]]:
    """Try to geocode using Address + City + Zip combinations. Cache aware.
    Appends ', Sicilia, Italia' to restrict results to Sicily if force_sicilia=True.
    
    Returns:
        Tuple of (result, query_description) or None if all queries fail
    """
    queries = []

    p_addr = address.strip()
    p_city = city.strip()
    p_zip = zip_code.strip()
    
    # Risk 3: Make suffix optional
    suffix = ", Sicilia, Italia" if force_sicilia else ", Italia"

    # Most specific first - track which query we're using
    if p_addr and p_city:
        queries.append((f"{p_addr}, {p_city}{suffix}", "Indirizzo + Comune"))
    if p_city and p_zip:
        queries.append((f"{p_city} {p_zip}{suffix}", "Comune + CAP"))
    if p_city:
        # If not forcing Sicily, maybe try just the city name without suffix? 
        # Or keep Italia? Let's keep suffix logic simple for now.
        queries.append((f"{p_city}{suffix}", "Solo Comune"))

    for q, q_desc in queries:
        q_key = q.upper()
        if q_key in cache:
            return (cache[q_key], q_desc + " (cache)")

        res = _geocode_single_query(q, api_key, valid_names=valid_names)
        if res:
            cache[q_key] = res
            return (res, q_desc)
        time.sleep(0.5)

    return None


# ─────────────────────────────────────────────────────────────────────────────
#  PREVIEW
# ─────────────────────────────────────────────────────────────────────────────

def preview_files(params: Dict[str, Any]) -> str:
    """Shows a preview of file recognition status directly in the UI."""
    files = []

    expected_key_suffix = "_prospetto_tasso_adesione_produzione_files"

    for k in ctx.session_state.keys():
        if str(k).endswith(expected_key_suffix) and k.startswith("up_"):
            files = ctx.session_state[k]
            break

    if not files:
        return ""

    data = []
    for f in files:
        ftype = "SCONOSCIUTO"
        action = "IGNORA"
        if "SINDRINN" in f.name:
            ftype = "SINDRINN"
            action = "ELABORA"
        elif "SINDMENS" in f.name:
            ftype = "SINDMENS"
            action = "ELABORA + UNISCI"

        data.append({
            "Nome File": f.name,
            "Tipo Riconosciuto": ftype,
            "Azione Prevista": action
        })

    if data:
        ctx.dataframe(pd.DataFrame(data), use_container_width=True)

        unknown_count = sum(1 for d in data if d["Tipo Riconosciuto"] == "SCONOSCIUTO")
        if unknown_count > 0:
            ctx.warning(f"⚠️ Ci sono {unknown_count} file non riconosciuti che verranno ignorati.")

    return ""


def manage_template(params: Dict[str, Any]) -> str:
    """Displays UI for managing the Excel template (Cell C3 and E1) with auto-date."""
    import calendar
    from datetime import datetime
    
    base_path = Path(__file__).resolve().parent
    template_path = base_path / "Tabelle - Template" / "Template" / "Sicilia_Prospetto_Variazioni.xlsx"
    
    ctx.markdown("##### 📝 Configurazione Template Excel")
    
    # UI: Path + buttons
    col_path, col_btn1, col_btn2 = ctx.columns([0.85, 0.075, 0.075])
    with col_path:
        ctx.text_input("Percorso Template", value=str(template_path), disabled=True, label_visibility="collapsed")
    with col_btn1:
        if ctx.button("📂", help="Apri cartella template"):
            try:
                os.startfile(template_path.parent)
            except Exception as e:
                ctx.error(f"Errore: {e}")
    with col_btn2:
        if ctx.button("🔎", help="Apri file template"):
            try:
                os.startfile(template_path)
            except Exception as e:
                ctx.error(f"Errore: {e}")

    if not template_path.exists():
        ctx.error("❌ File template non trovato!")
        return ""
    
    # ── Auto-date calculation ────────────────────────────────────────────
    prev_year = datetime.now().year - 1
    
    # C3: always 01/01 of previous year
    auto_c3 = f"01/01/{prev_year}"
    
    # E1: "VARIAZIONI DAL {ultimo_giorno_settembre}-09-{prev_year} AL {ultimo_giorno_gennaio}-01-{prev_year}"
    # calendar.monthrange(year, month) returns (weekday_of_first_day, number_of_days)
    last_day_sep = calendar.monthrange(prev_year, 9)[1]   # 30
    last_day_jan = calendar.monthrange(prev_year, 1)[1]    # 31
    auto_e1 = (
        f"VARIAZIONI DAL {last_day_jan:02d} - 01 - {prev_year} "
        f"AL {last_day_sep:02d} - 09 - {prev_year}"
    )
    
    try:
        from openpyxl import load_workbook
        
        # Read current values
        wb = load_workbook(template_path)
        ws = wb.active
        
        def get_merged_value(sheet, cell_coord):
            for merged_range in sheet.merged_cells.ranges:
                if cell_coord in merged_range:
                    return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            return sheet[cell_coord].value

        val_c3 = get_merged_value(ws, 'C3')
        val_e1 = get_merged_value(ws, 'E1')
        wb.close()
        
        val_c3_str = str(val_c3).strip() if val_c3 is not None else ""
        val_e1_str = str(val_e1).strip() if val_e1 is not None else ""
        
        # ── Show auto-calculated values ──────────────────────────────────
        ctx.markdown("#### 📅 Valori Automatici (Anno Precedente)")
        ctx.caption(f"Anno di riferimento: **{prev_year}**")
        
        col1, col2 = ctx.columns(2)
        with col1:
            ctx.info(f"**C3 auto:** `{auto_c3}`")
            ctx.caption(f"Attuale: `{val_c3_str}`")
        with col2:
            ctx.info(f"**E1 auto:** `{auto_e1}`")
            ctx.caption(f"Attuale: `{val_e1_str}`")
        
        # Check if update is needed
        needs_update = (val_c3_str != auto_c3 or val_e1_str != auto_e1)
        
        if needs_update:
            ctx.warning("⚠️ I valori nel template non corrispondono all'anno corrente - 1.")
        else:
            ctx.success("✅ Il template è già aggiornato.")
        
        # ── Auto-update button ───────────────────────────────────────────
        if ctx.button("🔄 Aggiorna Automaticamente", type="primary", disabled=not needs_update):
            try:
                wb = load_workbook(template_path)
                ws = wb.active
                ws['C3'] = auto_c3
                ws['E1'] = auto_e1
                wb.save(template_path)
                ctx.success(f"✅ Template aggiornato: C3=`{auto_c3}`, E1=`{auto_e1}`")
                time.sleep(1)
                ctx.rerun()
            except Exception as e:
                ctx.error(f"❌ Errore: {e}")
        
        # ── Manual override (collapsed) ──────────────────────────────────
        with ctx.expander("✏️ Modifica Manuale (override)"):
            new_c3 = ctx.text_input("Cella C3 (Data Riferimento)", value=val_c3_str)
            new_e1 = ctx.text_input("Cella E1 (Periodo Variazioni)", value=val_e1_str)
            
            if ctx.button("💾 Salva Modifiche Manuali"):
                if new_c3 != val_c3_str or new_e1 != val_e1_str:
                    try:
                        wb = load_workbook(template_path)
                        ws = wb.active
                        ws['C3'] = new_c3
                        ws['E1'] = new_e1
                        wb.save(template_path)
                        ctx.success("✅ Template aggiornato!")
                        time.sleep(1)
                        ctx.rerun()
                    except Exception as e:
                        ctx.error(f"❌ Errore: {e}")
                else:
                    ctx.info("ℹ️ Nessuna modifica rilevata.")
                    
    except Exception as e:
        ctx.error(f"Errore nella lettura del template: {e}")
        
    return ""


def handle_clear_cache(params: Dict[str, Any]) -> str:
    """Displays a button to clear the geocoding cache."""
    ctx.warning(
        "⚠️ **PROMEMORIA:** Svuota la cache solo se hai modificato la logica API o se vuoi "
        "forzare una nuova geocodifica per tutti i comuni. La cache velocizza notevolmente "
        "l'elaborazione riutilizzando i risultati precedenti."
    )
    
    if ctx.button("🗑️ Svuota Cache API", type="secondary", use_container_width=True):
        if _CACHE_FILE.exists():
            try:
                _CACHE_FILE.unlink()
                ctx.success("✅ Cache geocoding svuotata con successo! Alla prossima esecuzione tutti i comuni verranno richiesti nuovamente all'API.")
            except Exception as e:
                ctx.error(f"❌ Errore durante la cancellazione della cache: {e}")
        else:
            ctx.info("ℹ️ La cache è già vuota (nessun file trovato).")
    
    # Show current cache status
    if _CACHE_FILE.exists():
        cache_size = _CACHE_FILE.stat().st_size
        ctx.caption(f"📊 Cache esistente: {cache_size:,} bytes")
    else:
        ctx.caption("📊 Nessuna cache presente")
    
    return ""


def manage_corrections(params: Dict[str, Any]) -> str:
    """Displays UI for managing custom name corrections."""
    ctx.info("📝 **Correzioni personalizzate** per nomi comuni problematici rilevati dall'API.")
    
    # Load current corrections
    corrections = _load_corrections()
    
    # Display current corrections in an editable way
    ctx.markdown("#### Correzioni Attive")
    
    # Convert dict to text for editing
    corrections_text = "\n".join([f"{k} → {v}" for k, v in corrections.items()])
    
    edited_text = ctx.text_area(
        "Formato: NOME_ERRATO → Nome Corretto (uno per riga)",
        value=corrections_text,
        height=200,
        help="Inserisci le correzioni nel formato: NOME_SBAGLIATO → Nome Corretto. Attenzione alle maiuscole/minuscole per il nome corretto!"
    )
    
    col1, col2 = ctx.columns(2)
    
    with col1:
        if ctx.button("💾 Salva Correzioni", use_container_width=True):
            # Parse text back to dict
            new_corrections = {}
            for line in edited_text.strip().split('\n'):
                if '→' in line:
                    parts = line.split('→')
                    if len(parts) == 2:
                        wrong = parts[0].strip()
                        correct = parts[1].strip()
                        new_corrections[wrong.upper()] = correct
           
            _save_corrections(new_corrections)
            ctx.success(f"✅ Salvate {len(new_corrections)} correzioni!")
            ctx.rerun()
    
    with col2:
        if ctx.button("🔄 Ripristina Predefinite", use_container_width=True):
            _save_corrections(_DEFAULT_CORRECTIONS.copy())
            ctx.success("✅ Correzioni ripristinate ai valori predefiniti!")
            ctx.rerun()
    
    ctx.caption(f"📊 Correzioni attive: **{len(corrections)}**")
    
    return ""


# ─────────────────────────────────────────────────────────────────────────────
#  LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

def load_lookup_data(path: Path) -> tuple:
    """Loads the lookup dictionary (Code -> Info) and a list of valid names."""
    try:
        # G=Name, L=ProvName, O=ProvSigla, U=CatastaleCode
        df = pd.read_excel(path, usecols="G,L,O,U", header=0)
        lookup = {}
        name_to_info = {}
        names = []
        for _, row in df.iterrows():
            name = str(row.iloc[0]).strip()
            prov_name = str(row.iloc[1]).strip()
            prov_sigla = str(row.iloc[2]).strip().upper()
            code = str(row.iloc[3]).strip().upper()
            
            if code and code != 'NAN':
                info = {
                    "name": name,
                    "prov_name": prov_name,
                    "prov_sigla": prov_sigla
                }
                lookup[code] = info
                if name and name != 'nan':
                    name_to_info[name.upper()] = info
                    names.append(name)
        
        # Remove duplicates from names list
        names = sorted(list(set(names)))
        
        return lookup, names, name_to_info
    except Exception as e:
        print(f"Error loading lookup file: {e}")
        return {}, [], {}



def _clean_str(s: str) -> str:
    """Strip accents, apostrophes, hyphens, extra spaces for fuzzy comparison."""
    # Decompose unicode, remove combining marks (accents)
    nfkd = unicodedata.normalize('NFKD', s)
    no_accents = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Remove apostrophes, hyphens, dots
    cleaned = no_accents.replace("'", "").replace("'", "").replace("-", " ").replace(".", "")
    # Collapse multiple spaces
    return ' '.join(cleaned.upper().split())


def normalize_city_name(raw_name: str, valid_names: List[str], 
                        corrections: Dict[str, str] = None) -> tuple:
    """Matches API result against official municipality list.
    1. Custom corrections (user-defined mappings)
    2. Exact match (case-insensitive)
    3. Fuzzy match (strips accents, apostrophes, hyphens)
    4. 'Comune di X' heuristic
    
    Returns:
        tuple: (normalized_name, found_in_excel)
        - normalized_name: the matched name from Excel or Title Case fallback
        - found_in_excel: True if matched in Excel, False otherwise
    """
    if not raw_name:
        return raw_name, False

    raw_upper = raw_name.upper()
    
    # 0. Check custom corrections first
    if corrections:
        for wrong_name, correct_name in corrections.items():
            if raw_upper == wrong_name.upper():
                # Found in corrections → verify it exists in Excel
                for vn in valid_names:
                    if vn.upper() == correct_name.upper():
                        return vn, True
                # Correction doesn't exist in Excel, use it anyway
                return correct_name, True

    # 1. Exact match (case-insensitive)
    for vn in valid_names:
        if vn.upper() == raw_upper:
            return vn, True

    # 2. Fuzzy match (accent/apostrophe/hyphen insensitive)
    raw_cleaned = _clean_str(raw_name)
    for vn in valid_names:
        if _clean_str(vn) == raw_cleaned:
            return vn, True

    # 3. Heuristic: "Comune di X"
    if "COMUNE DI " in raw_upper:
        stripped = raw_upper.replace("COMUNE DI ", "")
        for vn in valid_names:
            if vn.upper() == stripped or _clean_str(vn) == _clean_str(stripped):
                return vn, True

    # 4. No match found — return in Title Case (not UPPERCASE)
    return raw_name.title(), False


# ─────────────────────────────────────────────────────────────────────────────
#  PROCESS LINE (local lookup only — no API here)
# ─────────────────────────────────────────────────────────────────────────────

def process_line(line: str, file_type: str, lookup: Dict[str, Any],
                 whitelist: List[str], valid_names: List[str] = None,
                 name_to_info: Dict[str, Any] = None) -> Dict[str, Any]:
    """Processes a single line: filter by Sede, lookup Comune code.
    Does NOT call API — that's handled separately in batch mode.
    Risk 2 Update: Tries to resolve by name if code lookup fails."""
    clean_line = line.rstrip('\r\n')

    cfg = TRACCIATO.get(file_type)
    if not cfg:
        return {"line": line, "status": "ERROR_TYPE", "code": "", "name": "",
                "sede": "", "error": "Tipo file sconosciuto",
                "addr": "", "city": "", "cap": "", "prov_name": "", "prov_sigla": "", "cod_fun": ""}

    # Helper function to extract by config
    def _extract(field):
        fcfg = cfg.get(field)
        if not fcfg: return ""
        s = fcfg["start"] - 1
        l = fcfg["len"]
        if len(clean_line) >= s + l:
            return clean_line[s:s+l].strip()
        return ""

    # 1. Extract Sede
    sede_found = _extract("sede")
    if not sede_found:
         # Check if line was just too short for sede
         fcfg = cfg["sede"]
         if len(clean_line) < fcfg["start"] + fcfg["len"]:
            return {"line": line, "status": "ERROR_LEN", "code": "", "name": "",
                    "sede": "", "error": f"Riga troppo corta (pos {fcfg['start']})",
                    "addr": "", "city": "", "cap": "", "prov_name": "", "prov_sigla": "", "cod_fun": ""}
    
    # Risk 6: Regex Validation for Sede (numeric, 4 digits)
    if not re.match(r'^\d{4}$', sede_found):
         return {"line": line, "status": "ERROR_FORMAT", "code": "", "name": "",
                "sede": sede_found, "error": f"Sede '{sede_found}' non valida (attese 4 cifre)",
                "addr": "", "city": "", "cap": "", "prov_name": "", "prov_sigla": "", "cod_fun": ""}

    # 2. Filter Sede (Risk 1 handling is in caller, here we just mark status)
    if sede_found not in whitelist:
        return {"line": line, "status": "SCARTATO", "code": "", "name": "",
                "sede": sede_found, "error": "Sede Fuori Regione",
                "addr": "", "city": "", "cap": "", 
                "prov_name": "", "prov_sigla": "", "cod_fun": ""}

    # 3. Extract Fields
    code_found = _extract("comune").upper()
    t_addr = _extract("addr")
    t_city = _extract("city")
    t_cap = _extract("cap")
    cod_fun = _extract("fun") if cfg.get("fun") else ""

    # Risk 6: Regex Validation for Comune Code (1 letter + 3 digits typically, or just alphanumeric)
    # INPS codes are usually "Letter + 3 digits" (e.g. A123). Some might be different?
    # Let's stick to simple alphanumeric check to be safe but stricter than nothing.
    if code_found and not re.match(r'^[A-Z0-9]{4}$', code_found):
         # Soft error? Or just treat as empty? Let's treat as OK but it won't be found in lookup potentially.
         # Actually if it's garbage, better not to trust it.
         pass 

    # 4. Lookup Logic
    # 4a. Try by Code
    if code_found and code_found in lookup:
        info = lookup[code_found]
        return {
            "line": clean_line, "status": "OK",
            "code": code_found, "name": info["name"],
            "prov_name": info["prov_name"],
            "prov_sigla": info["prov_sigla"],
            "cod_fun": cod_fun,
            "sede": sede_found, "error": None, "tipo": file_type,
            "addr": t_addr, "city": t_city, "cap": t_cap
        }
    
    # 4b. Risk 2: Try by Name (Local Lookup) if code failed
    if t_city and valid_names and name_to_info:
        # Use existing normalization logic
        norm_name, found = normalize_city_name(t_city, valid_names)
        if found:
            info = name_to_info.get(norm_name.upper())
            if info:
                return {
                    "line": clean_line, "status": "OK_BY_NAME", # New status
                    "code": code_found, # Keep original (wrong/missing) code
                    "name": info["name"],
                    "prov_name": info["prov_name"],
                    "prov_sigla": info["prov_sigla"],
                    "cod_fun": cod_fun,
                    "sede": sede_found, 
                    "error": f"Codice '{code_found}' errato/mancante, recuperato per nome '{t_city}'",
                    "tipo": file_type,
                    "addr": t_addr, "city": t_city, "cap": t_cap
                }

    has_fields = bool(t_addr or t_city or t_cap)
    return {
        "line": clean_line,
        "status": "NEED_API" if has_fields else "NEED_API_EMPTY",
        "code": code_found, "name": "",
        "prov_name": "", "prov_sigla": "", "cod_fun": cod_fun,
        "sede": sede_found,
        "error": f"Codice '{code_found}' non trovato" if code_found else "Codice vuoto",
        "tipo": file_type,
        "addr": t_addr, "city": t_city, "cap": t_cap
    }



# ─────────────────────────────────────────────────────────────────────────────
#  PERSISTENT DASHBOARD (survives downloads and reruns)
# ─────────────────────────────────────────────────────────────────────────────

def _render_dashboard() -> None:
    """Display the summary dashboard from session_state. Persists across reruns."""
    data = ctx.session_state.get('prospetto_sicilia_dashboard')
    if not data:
        return

    total = data['total_processed']
    count_ok = data['count_ok']
    count_api = data['count_api']
    count_api_failed = data['count_api_failed']
    count_empty = data['count_empty']
    count_altri_total = count_api_failed + count_empty
    api_results = data.get('api_results', [])
    
    # New counters
    sindmens_counts = data.get('sindmens_counts', {"0": 0, "1": 0, "2": 0, "3": 0})
    total_rinnovi = data.get('total_rinnovi', 0)
    final_file = data.get('final_file')

    ctx.markdown("---")
    ctx.markdown("## 📊 Dashboard Riepilogativa (Sicilia)")

    # ── Summary Metrics ────────────────────────────────────────────────
    col1, col2 = ctx.columns(2)
    with col1:
        ctx.info(f"📋 **Totale Record Sicilia:** {total:,}")
        ctx.write(f"- 🔍 Da Codice Catastale: **{count_ok:,}**")
        ctx.write(f"- 🌐 Recuperati via API: **{count_api:,}**")
        ctx.write(f"- ❓ Assegnati a 'ALTRI': **{count_altri_total:,}**")
    
    with col2:
        ctx.success(f"📈 **Dettaglio Elaborazione Variazioni**")
        ctx.write(f"- 🔄 **RINNOVO** (SINDRINN): **{total_rinnovi:,}**")
        ctx.write(f"- 📍 **0 - CONCOMITANTI**: **{sindmens_counts.get('0', 0):,}**")
        ctx.write(f"- 📝 **2 - DELEGHE**: **{sindmens_counts.get('2', 0):,}**")
        ctx.write(f"- ❌ **1 - REVOCHE**: **{sindmens_counts.get('1', 0):,}**")
        ctx.write(f"- 🗑️ **3 - ELIMINATE**: **{sindmens_counts.get('3', 0):,}**")

    # ── Download persistence ──────────────────────────────────────────
    if final_file:
        ctx.markdown("### 📥 File Disponibile per il Download")
        with open(final_file, "rb") as f:
            ctx.download_button(
                label="📥 Scarica Prospetto Variazioni Compilato",
                data=f,
                file_name=os.path.basename(final_file),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    # ── Detail of ALTRI breakdown ────────────────────────────────────────
    if count_altri_total > 0:
        ctx.warning(
            f"⚠️ **{count_altri_total}** record assegnati a **'ALTRI'** — dettaglio:\n"
            f"- 📭 **{count_empty}** con campi indirizzo/comune/cap **vuoti** nel file "
            f"(impossibile chiamare API)\n"
            f"- ❌ **{count_api_failed}** inviati all'API ma **non risolti** "
            f"(nazione/comune non trovato)"
        )

    # ── API Results table (full width) ─────────────────────────────────
    ctx.markdown("### 🌐 Comuni Recuperati via API")
    ctx.caption("✅ = identico · ⚠️ = nome diverso (preferito file o API ha trovato comune diverso)")
    if api_results:
        # Define columns we want to display
        display_cols = [
            'match', 'n_records', 'comune_estratto', 'nome_excel', 
            'similarita', 'indirizzo', 'api_risultato', 'nome_normalizzato', 
            'fonte_usata', 'query_api'
        ]
        df_api = pd.DataFrame(api_results)[display_cols]
        df_api.columns = [
            '✅/⚠️', 'N. Record', '📄 Dal File: Comune',
            '📋 Excel: Nome Ufficiale', '📊 Similarità Nome',
            '📄 Dal File: Indirizzo', '🌐 API: Risultato', '✏️ → Output',
            '📌 Fonte Usata', '🔍 Query API'
        ]
        ctx.dataframe(df_api, use_container_width=True, hide_index=True, height=400)
    else:
        ctx.info("Nessun comune recuperato via API.")



# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def run(out_dir: Path, files: List[Any] = [], **kwargs) -> List[Path]:
    """
    Main entry point for the tool.
    params: Contains 'whitelist_sedi' string.
    """
    if not files:
        return []

    # Get whitelist from kwargs
    whitelist_str = kwargs.get('whitelist_sedi', '')
    if not whitelist_str:
        whitelist_str = ('2193, 2192, 0100, 8291, 7691, 5593, 4892, 0192, '
                         '2190, 1800, 0191, 5503, 8292, 2100, 5592, 2800, '
                         '1890, 2194, 7690, 0190, 8290, 2196, 8293, 4800, '
                         '4894, 5596, 6590, 2890, 7601, 5500, 5502, 5591, '
                         '2191, 4890, 5594, 6500, 4893, 4891, 0101, 7600, '
                         '5590, 8200, 6591')

    whitelist = [x.strip() for x in whitelist_str.split(',') if x.strip()]
    
    # Get validation flag
    enable_validation = kwargs.get('enable_validation', False)
    
    # Load corrections
    corrections = _load_corrections()
    
    # Load lookup
    base_path = Path(__file__).parent.parent / "Amministrazione" / "Supporto"

    # ── 1. Locate Lookup File ────────────────────────────────────────────
    base_path = Path(__file__).resolve().parent
    lookup_path = base_path / "Tabelle - Template" / "Elenco-comuni-italiani.xlsx"

    lookup_dict = {}
    valid_names = []
    lookup_error = None
    if not lookup_path.exists():
        lookup_error = f"File lookup non trovato: {lookup_path}"
        ctx.error(f"⚠️ {lookup_error}")
    else:
        lookup_dict, valid_names, name_to_info = load_lookup_data(lookup_path)
        ctx.info(f"📖 Caricati **{len(lookup_dict)}** codici comune dal file di riferimento.")

    # ── 2. Check API key ─────────────────────────────────────────────────
    has_api = bool(LOCATIONIQ_API_KEY)
    if not has_api:
        ctx.warning("⚠️ Chiave API LocationIQ non trovata. "
                    "I record senza codice saranno assegnati a 'ALTRI'.")

    # Init dashboard counters
    geocode_cache = _load_cache()
    count_api_resolved = 0
    count_api_failed = 0
    count_empty = 0
    api_log = []
    api_recovery_map = {}
    
    # Init accumulation lists
    output_files = []
    all_results = []
    sindrinn_lines = {}
    sindmens_accumulated_lines = []

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    scarti_stats = {} # Risk 1: track discarded sedi
    
    # ══════════════════════════════════════════════════════════════════════
    #  FASE 1: Parse all files (local lookup only)
    # ══════════════════════════════════════════════════════════════════════
    ctx.info("📝 **Fase 1:** Parsing file e lookup locale codici catastali...")

    for file_obj in files:
        filename = "unknown.txt"
        content_bytes = b""

        if hasattr(file_obj, "name"):
            filename = file_obj.name
        elif isinstance(file_obj, (str, Path)):
            filename = os.path.basename(str(file_obj))

        try:
            if hasattr(file_obj, "getvalue"):
                content_bytes = file_obj.getvalue()
            elif hasattr(file_obj, "read"):
                if hasattr(file_obj, "seek"):
                    file_obj.seek(0)
                content_bytes = file_obj.read()
                if isinstance(content_bytes, str):
                    content_bytes = content_bytes.encode('utf-8')
            elif isinstance(file_obj, (str, Path)) and os.path.exists(file_obj):
                with open(file_obj, "rb") as f:
                    content_bytes = f.read()
            else:
                content_bytes = bytes(file_obj)
        except Exception:
            continue

        if not content_bytes:
            continue

        ftype = None
        if "SINDRINN" in filename:
            ftype = "SINDRINN"
        elif "SINDMENS" in filename:
            ftype = "SINDMENS"

        if not ftype:
            continue

        try:
            # Risk 7: Encoding logic (UTF-8 first, then LATIN1)
            try:
                content_str = content_bytes.decode("utf-8")
            except UnicodeDecodeError:
                content_str = content_bytes.decode("latin1")
        except Exception:
             # Fallback extreme
             content_str = content_bytes.decode("utf-8", errors="replace")

        lines = content_str.splitlines()

        for idx, line in enumerate(lines):
            # Risk 2: Pass valid_names and name_to_info
            res = process_line(line, ftype, lookup_dict, whitelist, 
                               valid_names=valid_names, name_to_info=name_to_info)
            res['source_file'] = filename
            res['line_idx'] = idx
            
            # Risk 1: Accumulate scarti
            if res['status'] == 'SCARTATO':
                s_sede = res.get('sede', 'UNKNOWN')
                scarti_stats[s_sede] = scarti_stats.get(s_sede, 0) + 1
            
            all_results.append(res)

    # Count results by status
    ok_results = [r for r in all_results if r['status'] in ('OK', 'OK_BY_NAME')]
    need_api_results = [r for r in all_results if r['status'] == 'NEED_API']
    need_api_empty = [r for r in all_results if r['status'] == 'NEED_API_EMPTY']
    total_accepted = len(ok_results) + len(need_api_results) + len(need_api_empty)

    # Risk 1: Show warning for scarti
    if scarti_stats:
        tot_scarti = sum(scarti_stats.values())
        ctx.warning(f"⚠️ **ATTENZIONE:** {tot_scarti} record scartati per Sede non in Whitelist.")
        # Top 10 scarti
        df_scarti = pd.DataFrame(list(scarti_stats.items()), columns=['Codice Sede', 'Record Scartati'])
        df_scarti = df_scarti.sort_values('Record Scartati', ascending=False).head(10)
        ctx.dataframe(df_scarti, use_container_width=True, height=150)
        
        top_sede = df_scarti.iloc[0]['Codice Sede']
        ctx.info(f"💡 Se il codice **{top_sede}** è una sede valida, aggiungilo alla Whitelist nelle opzioni.")

    ctx.info(
        f"📊 **Fase 1 completata:** {total_accepted:,} record accettati (Sicilia)\n"
        f"- ✅ **{len(ok_results):,}** trovati da codice catastale\n"
        f"- 🌐 **{len(need_api_results):,}** da risolvere via API "
        f"(hanno campi indirizzo/comune/cap)\n"
        f"- 📭 **{len(need_api_empty):,}** con campi vuoti → assegnati direttamente a **ALTRI**"
    )

    # ══════════════════════════════════════════════════════════════════════
    #  FASE 2: Geocode missing records (dedup by Comune key)
    # ══════════════════════════════════════════════════════════════════════
    api_recovery_map = {}  # city_key -> resolved_name
    api_log = []           # log of API results for dashboard

    count_api_resolved = 0
    count_api_failed = 0

    if need_api_results and has_api:
        # Build unique city keys for deduplication
        city_keys = {}
        city_record_counts = {}
        for r in need_api_results:
            ckey = r['city'].upper() if r['city'] else ""
            if ckey not in city_keys:
                city_keys[ckey] = r
                city_record_counts[ckey] = 0
            city_record_counts[ckey] += 1

        unique_cities = list(city_keys.keys())
        n_unique = len(unique_cities)
        n_api_rows = len(need_api_results)

        ctx.info(
            f"🌐 **Fase 2:** Geocoding per **{n_api_rows:,}** record → "
            f"solo **{n_unique}** comuni unici da risolvere.\n\n"
            f"⏱️ Tempo stimato: **~{n_unique}** secondi "
            f"({n_api_rows - n_unique:,} duplicati saltati)"
        )

        progress_bar = ctx.progress(0, text="Geocoding comuni unici...")
        progress_text = ctx.empty()
        status_text = ctx.empty()
        start_time = time.time()

        for i, ckey in enumerate(unique_cities):
            sample = city_keys[ckey]
            t_addr = sample.get('addr', '')
            t_city = sample.get('city', '')
            t_cap = sample.get('cap', '')

            # Risk 3: OPTIMIZED API LOGIC
            # 1. Try local recovery FIRST (Risk 3 prevention)
            api_result = None
            
            # Check if t_city is valid (even if not found via code previously)
            if t_city:
                 norm, found = normalize_city_name(t_city, valid_names, corrections)
                 if found:
                     # IT IS VALID! No need to call API!
                     # Fake an API result to reuse downstream logic or just set it direct
                     api_result = (norm, "Recupero Locale (Nome Presente)")
            
            # 2. Call API only if local recovery failed
            if not api_result:
                # Decide if we force Sicily
                # If t_city looks like a non-sicilian city (e.g. Milano), maybe don't force?
                # For now let's keep force_sicilia=True by default unless we want to be smarter.
                # Risk 3 analysis suggested: "Rimuovere il suffisso... quando il campo comune del file contiene un nome non siciliano"
                # But detecting "non siciliano" is hard without a list of non-sicilian cities.
                # Let's trust LocationIQ with the user address if provided.
                
                api_result = _geocode_location(t_addr, t_city, t_cap,
                                            LOCATIONIQ_API_KEY, geocode_cache,
                                            valid_names=valid_names,
                                            force_sicilia=True) # Could make this smarter later

            n_records_for_key = city_record_counts.get(ckey, 1)

            if api_result:
                api_res, query_used = api_result
                
                # FILE-FIRST LOGIC: prefer city from file if present
                if t_city.strip():
                    # File has commune → use it
                    final_name, found_in_excel = normalize_city_name(t_city, valid_names, corrections)
                    fonte_usata = "📄 Dal file"
                else:
                    # File commune empty → use API result
                    final_name, found_in_excel = normalize_city_name(api_res, valid_names, corrections)
                    fonte_usata = "🌐 Da API"
                
                api_recovery_map[ckey] = final_name
                count_api_resolved += 1
                
                # Calculate similarity between file commune and Excel normalized name
                similarity_val = 0
                if t_city.strip():
                    similarity = difflib.SequenceMatcher(None, 
                                                        t_city.upper().strip(), 
                                                        final_name.upper().strip()).ratio()
                    similarity_pct = int(similarity * 100)
                    similarity_val = similarity_pct
                    if similarity_pct >= 90:
                        similarity_display = f"{similarity_pct}% ✅"
                    elif similarity_pct >= 70:
                        similarity_display = f"{similarity_pct}% ⚠️"
                    else:
                        similarity_display = f"{similarity_pct}% ❌"
                else:
                    similarity_display = "N/A"
                    similarity_val = 100 # Treat as good match for sorting purposes if N/A (no error in file name)
                
                #Excel match indicator
                excel_match = found_in_excel
                excel_nome = final_name if found_in_excel else "⚠️ Non in Excel"
                
                # Match indicator
                is_exact = (t_city.upper().strip() == final_name.upper().strip())
                api_log.append({
                    "match": "✅" if is_exact else "⚠️",
                    "n_records": n_records_for_key,
                    "comune_estratto": t_city,
                    "nome_excel": excel_nome,
                    "similarita": similarity_display,
                    "indirizzo": t_addr,
                    "api_risultato": api_res,
                    "nome_normalizzato": final_name,
                    "fonte_usata": fonte_usata,
                    "query_api": query_used,
                # Hidden keys for sorting
                    "sort_excel_bad": not excel_match,      # True if NOT in Excel (primary sort)
                    "sort_similarity": similarity_val       # Ascending similarity (secondary sort)
                })
            else:
                api_recovery_map[ckey] = "ALTRI"
                count_api_failed += 1
            
            # Timer
            elapsed = time.time() - start_time
            done = i + 1
            remaining = n_unique - done
            avg = elapsed / done if done > 0 else 1
            eta = remaining * avg
            el_m, el_s = divmod(int(elapsed), 60)
            et_m, et_s = divmod(int(eta), 60)

            # Update progress bar
            progress_text.text(f"Geocoding e normalizzazione: {done}/{n_unique} comuni...")
            progress_bar.progress(done / n_unique)
            status_text.markdown(
                f"⏱️ **Trascorso:** {el_m:02d}:{el_s:02d} · "
                f"**Rimanente:** ~{et_m:02d}:{et_s:02d} · "
                f"✅ {count_api_resolved} risolti · ❌ {count_api_failed} falliti"
            )

        # Clear progress bar
        progress_text.empty()
        progress_bar.empty()
        status_text.empty()
        
        # Sort api_log by error severity:
        # 1. Not in Excel (True first -> 0)
        # 2. Low similarity (Ascending)
        api_log.sort(key=lambda x: (0 if x['sort_excel_bad'] else 1, x['sort_similarity']))

        progress_bar.empty()
        status_text.empty()

        total_elapsed = time.time() - start_time
        tot_m, tot_s = divmod(int(total_elapsed), 60)

        # Count actual records affected
        records_resolved_via_api = sum(
            city_record_counts[ck] for ck in api_recovery_map
            if api_recovery_map[ck] != "ALTRI"
        )
        records_api_failed = sum(
            city_record_counts[ck] for ck in api_recovery_map
            if api_recovery_map[ck] == "ALTRI"
        )

        ctx.success(
            f"🌐 Geocoding completato in **{tot_m} min {tot_s} sec**:\n"
            f"- 🔍 **{n_unique}** comuni unici analizzati "
            f"(su {n_api_rows:,} record)\n"
            f"- ✅ **{count_api_resolved}** comuni risolti "
            f"(**{records_resolved_via_api:,}** record)\n"
            f"- ❌ **{count_api_failed}** comuni non risolti "
            f"(**{records_api_failed:,}** record → ALTRI)"
        )
    elif need_api_results and not has_api:
        for r in need_api_results:
            ckey = r['city'].upper() if r['city'] else ""
            api_recovery_map[ckey] = "ALTRI"
        ctx.warning(f"⚠️ Chiave API mancante: {len(need_api_results):,} record → ALTRI")
    else:
        ctx.success("✅ Nessun record da geocodificare via API.")

    # Save cache
    _save_cache(geocode_cache)

    # ══════════════════════════════════════════════════════════════════════
    #  FASE 3: Assemble output files
    # ══════════════════════════════════════════════════════════════════════
    ctx.info("📄 **Fase 3:** Assemblaggio file di output...")

    sindrinn_lines = {}
    report_rows = []
    final_count_ok = 0
    final_count_api = 0
    final_count_altri = 0

    for r in all_results:
        status = r['status']
        name = r.get('name', '')
        error = r.get('error', '')

        if status in ('OK', 'OK_BY_NAME'):
            final_count_ok += 1
        elif status == 'NEED_API':
            ckey = r['city'].upper() if r['city'] else ""
            resolved = api_recovery_map.get(ckey, "ALTRI")
            name = resolved
            if resolved != "ALTRI":
                # Get province info for resolved municipalitity
                inf = name_to_info.get(resolved.upper())
                
                # Risk 4: Fuzzy Province Lookup
                if not inf:
                    resolved_clean = _clean_str(resolved)
                    for k, v in name_to_info.items():
                        if _clean_str(k) == resolved_clean:
                            inf = v
                            break
                            
                if inf:
                    r["prov_name"] = inf["prov_name"]
                    r["prov_sigla"] = inf["prov_sigla"]
                else:
                    # Logs for debug (optional)
                    pass

                status = "RECUPERATO_API"
                error = f"API → {resolved}"
                final_count_api += 1
                # Risk 5: DO NOT increment count_api_resolved here (it tracks unique cities in Phase 2)
                # count_api_resolved += 1 
            else:
                status = "RECUPERATO_ALTRI"
                error = "API fallita → ALTRI"
                final_count_altri += 1
                # Risk 5: DO NOT increment count_api_failed here
                # count_api_failed += 1
        elif status == 'NEED_API_EMPTY':
            name = "ALTRI"
            status = "CAMPI_VUOTI"
            error = "Indirizzo/Comune/CAP vuoti → ALTRI"
            final_count_altri += 1
            count_empty += 1
        else:
            # SCARTATO, ERROR_TYPE, ERROR_LEN → skip output
            report_rows.append({
                "File": r.get('source_file', ''), "Tipo": r.get('tipo', ''),
                "Riga": r.get('line_idx', 0) + 1, "Status": status,
                "Sede": r.get('sede', ''), "Codice Comune": r.get('code', ''),
                "Comune": name, "Errore": error
            })
            continue

        # Build output line
        new_line = r['line'] + "    " + name + "\n"
        ftype = r.get('tipo', '')
        fname = r.get('source_file', '')

        if ftype == "SINDRINN":
            if fname not in sindrinn_lines:
                sindrinn_lines[fname] = []
            sindrinn_lines[fname].append(new_line)
        elif ftype == "SINDMENS":
            sindmens_accumulated_lines.append(new_line)

        report_rows.append({
            "File": fname, "Tipo": ftype,
            "Riga": r.get('line_idx', 0) + 1, "Status": status,
            "Sede": r.get('sede', ''), "Codice Comune": r.get('code', ''),
            "Comune": name, "Errore": error
        })

    # Write SINDRINN files
    for fname, out_lines in sindrinn_lines.items():
        out_name = f"PROCESSED_{fname}"
        out_path = out_dir / out_name
        with open(out_path, "w", encoding="latin1") as f:
            f.writelines(out_lines)
        output_files.append(out_path)

    # Write unified SINDMENS
    if sindmens_accumulated_lines:
        out_path = out_dir / "SINDMENS_UNIFICATO.txt"
        with open(out_path, "w", encoding="latin1") as f:
            f.writelines(sindmens_accumulated_lines)
        output_files.append(out_path)

    
    # ── GLOBAL AGGREGATION FOR DASHBOARD ─────────────────────────────────────
    global_rinnovi = 0
    global_sindmens = {"0": 0, "1": 0, "2": 0, "3": 0}
    
    for r in all_results:
        if r['status'] in ['SCARTATO', 'ERROR_TYPE', 'ERROR_LEN']:
            continue
        ftype = r.get('tipo', '')
        cfun = r.get('cod_fun', '')
        if ftype == "SINDRINN":
            global_rinnovi += 1
        elif ftype == "SINDMENS":
            if cfun in global_sindmens:
                global_sindmens[cfun] += 1

    # ── TEMPLATE POPULATION (Sicilia_Prospetto_Variazioni) ───────────────────
    ctx.info("📊 **Fase 4:** Popolamento template variazioni...")
    
    template_in = base_path / "Tabelle - Template" / "Template" / "Sicilia_Prospetto_Variazioni.xlsx"
    if not template_in.exists():
        ctx.error(f"❌ Template non trovato: {template_in}")
        # Update dashboard even if template fails
        ctx.session_state['prospetto_sicilia_dashboard'] = {
            'total_processed': len(all_results),
            'count_ok': final_count_ok,
            'count_api': final_count_api,
            'count_api_failed': count_api_failed,
            'count_empty': count_empty,
            'api_results': api_log,
            'total_rinnovi': global_rinnovi,
            'sindmens_counts': global_sindmens,
            'final_file': None
        }
        if report_rows:
            df_rep = pd.DataFrame(report_rows)
            rep_path = out_dir / "Report_Verifica_Prospetti.xlsx"
            df_rep.to_excel(rep_path, index=False)
            return [rep_path]
        return output_files

    # Categorize data for hierarchical sorting
    SICILIA_SIGLE = ['AG', 'CL', 'CT', 'EN', 'ME', 'PA', 'RG', 'SR', 'TP']
    SIGLA_TO_NAME = {
        'AG': 'AGRIGENTO', 'CL': 'CALTANISSETTA', 'CT': 'CATANIA', 
        'EN': 'ENNA', 'ME': 'MESSINA', 'PA': 'PALERMO', 
        'RG': 'RAGUSA', 'SR': 'SIRACUSA', 'TP': 'TRAPANI'
    }
    
    sicilia_groups = {sigla: {} for sigla in SICILIA_SIGLE}
    fuori_prov_group = {}
    estero_group = {}

    for r in all_results:
        # Skip discarded records
        if r['status'] in ['SCARTATO', 'ERROR_TYPE', 'ERROR_LEN']:
            continue
            
        status = r['status']
        name = r.get('name', 'ALTRI')
        
        # Resolve API names
        if status == 'NEED_API':
             ckey = r['city'].upper() if r['city'] else ""
             name = api_recovery_map.get(ckey, "ALTRI")
             
        # Normalize logic for orphans/failures
        # If name is generic, empty or failure placeholder, map to "Provincia non esistente"
        if not name or name.upper() in ["ALTRI", "FUORI PROVINCIA"]:
            name = "Provincia non esistente"

        psigla = r.get('prov_sigla', '')
        ftype = r.get('tipo', '')
        cfun = r.get('cod_fun', '')
        
        is_estero = (psigla == 'EE' or "ESTERO" in name.upper())
        
        # Target group
        if is_estero:
            target = estero_group
        elif psigla in SICILIA_SIGLE:
            target = sicilia_groups[psigla]
        else:
            # Everything else (Valid local non-Sicilian OR "Provincia non esistente") goes here
            target = fuori_prov_group
            
        if name not in target:
            target[name] = {"RINNOVO": 0, "0": 0, "2": 0, "1": 0, "3": 0}
        
        if ftype == "SINDRINN":
            target[name]["RINNOVO"] += 1
        elif ftype == "SINDMENS":
            if cfun in target[name]:
                target[name][cfun] += 1
                
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.formatting.rule import CellIsRule

        wb_tpl = load_workbook(template_in)
        ws_tpl = wb_tpl.active

        start_row = 4
        current_row = start_row

        # ── Styles ────────────────────────────────────────────────────────
        bold_font = Font(bold=True)
        normal_font = Font(bold=False)
        thin_side = Side(style='thin')
        medium_side = Side(style='medium')
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)
        subtotal_border = Border(left=thin_side, right=thin_side,
                                 top=medium_side, bottom=medium_side)


        # ── Helper: write formulas on a row ─────────────────────────────
        def _write_row_formulas(row):
            """Add Excel formulas for cols F, H, J, L, M, N, O, R.
            Division errors → 'N/D', zero values → empty cell."""
            r = row
            fmt2 = '0.00'

            # Division formulas: =IF(C=0,"N/D",IF(formula=0,"",formula))
            # F = E*100/C  (percentage number, e.g. 7.02)
            c = ws_tpl.cell(row=r, column=6,
                value=f'=IF(C{r}=0,"N/D",IF(E{r}*100/C{r}=0,"",E{r}*100/C{r}))')
            c.number_format = fmt2
            # H = G*100/C
            c = ws_tpl.cell(row=r, column=8,
                value=f'=IF(C{r}=0,"N/D",IF(G{r}*100/C{r}=0,"",G{r}*100/C{r}))')
            c.number_format = fmt2
            # J = I*100/C
            c = ws_tpl.cell(row=r, column=10,
                value=f'=IF(C{r}=0,"N/D",IF(I{r}*100/C{r}=0,"",I{r}*100/C{r}))')
            c.number_format = fmt2
            # L = K*100/C
            c = ws_tpl.cell(row=r, column=12,
                value=f'=IF(C{r}=0,"N/D",IF(K{r}*100/C{r}=0,"",K{r}*100/C{r}))')
            c.number_format = fmt2

            # M = E + G - I  (integer, no decimals)
            c = ws_tpl.cell(row=r, column=13,
                value=f'=IF(E{r}+G{r}-I{r}=0,"",E{r}+G{r}-I{r})')
            c.number_format = '0'
            # N = (E+G-I)*100/C
            c = ws_tpl.cell(row=r, column=14,
                value=f'=IF(C{r}=0,"N/D",IF((E{r}+G{r}-I{r})*100/C{r}=0,"",(E{r}+G{r}-I{r})*100/C{r}))')
            c.number_format = fmt2
            # O = G - I  (integer, no decimals)
            c = ws_tpl.cell(row=r, column=15,
                value=f'=IF(G{r}-I{r}=0,"",G{r}-I{r})')
            c.number_format = '0'
            # P = (G-I)*100/C
            c = ws_tpl.cell(row=r, column=16,
                value=f'=IF(C{r}=0,"N/D",IF((G{r}-I{r})*100/C{r}=0,"",(G{r}-I{r})*100/C{r}))')
            c.number_format = fmt2

        # ── Helper: apply border + font to cols B(2)..P(16) ───────────────
        center_align = Alignment(horizontal='center')

        def _style_row(row, border, font):
            for col in range(2, 17):  # B=2 to P=16 inclusive
                cell = ws_tpl.cell(row=row, column=col)
                cell.border = border
                cell.font = font
                if col >= 3:  # C onwards = numeric → center
                    cell.alignment = center_align

        # ── VALIDAZIONE DATI ──────────────────────────────────────────────────────
        def validate_data_quality(sicilia_groups, fuori_prov_group, estero_group, name_to_info, SICILIA_SIGLE, SIGLA_TO_NAME):
            """
            Esegue controlli di qualità sui dati:
            1. Duplicati/Simili
            2. Province Errate (via lookup name_to_info)
            """
            import difflib
            import unicodedata
            
            issues_map = {}
            
            def _clean(s):
                nfkd = unicodedata.normalize('NFKD', s)
                no_acc = ''.join(c for c in nfkd if not unicodedata.combining(c))
                return ' '.join(no_acc.replace("'","").replace("-"," ").replace(".","").upper().split())

            # Raccogli tutti i nomi con sezione
            all_names = []
            for sigla in SICILIA_SIGLE:
                for name in sicilia_groups.get(sigla, {}).keys():
                    all_names.append((name, sigla))
            for name in fuori_prov_group.keys():
                all_names.append((name, "FUORI"))
            for name in estero_group.keys():
                all_names.append((name, "ESTERO"))

            # 1. Check Duplicati
            checked = set()
            df_duplicati = []
            
            for i, (name1, sez1) in enumerate(all_names):
                clean1 = _clean(name1)
                for j, (name2, sez2) in enumerate(all_names):
                    if i >= j: continue
                    pair = tuple(sorted((name1, name2)))
                    if pair in checked: continue
                    checked.add(pair)

                    clean2 = _clean(name2)
                    
                    # Check similarità
                    ratio = difflib.SequenceMatcher(None, clean1, clean2).ratio()
                    is_contained = (clean1 in clean2 or clean2 in clean1) and min(len(clean1), len(clean2)) > 4
                    
                    issue_type = None
                    if clean1 == clean2:
                         issue_type = "Identico (accenti/format)"
                    elif ratio >= 0.85 or is_contained:
                         issue_type = "Simile/Contenuto"
                    
                    if issue_type:
                        msg = f"Possibile duplicato con: {name2} ({sez2})"
                        if name1 not in issues_map:
                            issues_map[name1] = {'color': 'FFD966', 'msg': msg} # Arancio
                        else:
                            issues_map[name1]['msg'] += f"; {msg}"
                            
                        if name2 not in issues_map:
                            issues_map[name2] = {'color': 'FFD966', 'msg': f"Possibile duplicato con: {name1} ({sez1})"}
                        
                        df_duplicati.append({'Comune 1': name1, 'Sez 1': sez1, 'Comune 2': name2, 'Sez 2': sez2, 'Tipo': issue_type})

            # 2. Check Province
            df_province = []
            for sigla in SICILIA_SIGLE:
                for name in sicilia_groups.get(sigla, {}).keys():
                    info = name_to_info.get(name.upper())
                    if not info:
                        # Fallback lookup con pulizia
                        c_name = _clean(name)
                        for k, v in name_to_info.items():
                            if _clean(k) == c_name:
                                info = v
                                break
                    
                    if not info:
                        if name not in issues_map:
                             issues_map[name] = {'color': 'FFE699', 'msg': "Comune non trovato nel database per verifica provincia"}
                        continue
                    
                    real_sigla = info.get('prov_sigla', '')
                    if real_sigla != sigla:
                         msg = f"Provincia errata: è {real_sigla} ({info.get('prov_name','')}), non {sigla}"
                         issues_map[name] = {'color': 'FF9999', 'msg': msg} # Rosso
                         df_province.append({'Comune': name, 'Assegnata': sigla, 'Corretta': real_sigla})

            # Visualizzazione in Streamlit
            if df_duplicati:
                 ctx.warning(f"⚠️ **Trovati {len(df_duplicati)} potenziali duplicati/simili**")
                 with ctx.expander("Dettagli Duplicati"):
                     ctx.dataframe(pd.DataFrame(df_duplicati))
            
            if df_province:
                 ctx.error(f"🔴 **Trovati {len(df_province)} comuni in provincia errata**")
                 with ctx.expander("Dettagli Errori Provincia"):
                     ctx.dataframe(pd.DataFrame(df_province))

            return issues_map

        # Esegui validazione se richiesto
        issues_map = {}
        if enable_validation:
            issues_map = validate_data_quality(sicilia_groups, fuori_prov_group, estero_group, name_to_info, SICILIA_SIGLE, SIGLA_TO_NAME)


        # ── Scrittura Template ──────────────────────────────────────────────────
        # Il workbook è già caricato in wb_tpl (da template_in)
        
        start_row = 15 # Ripristino riga corretta
        current_row = start_row

        # ── Helper: write a data row (municipality) ───────────────────────
        def write_data_row(cname, vals):
            nonlocal current_row
            r = current_row
            ws_tpl.cell(row=r, column=1, value="")
            
            cell_b = ws_tpl.cell(row=r, column=2, value=cname)
            
            # Applica highlights se ci sono problemi
            issue = issues_map.get(cname)
            if issue:
                color = issue.get('color', 'FFFF00')
                cell_b.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                # Nota in colonna Q (17)
                cell_q = ws_tpl.cell(row=r, column=17, value=issue.get('msg', ''))
                cell_q.font = Font(color="CC0000", italic=True, size=9)
                cell_q.alignment = Alignment(wrap_text=True)

            ws_tpl.cell(row=r, column=3, value=vals["RINNOVO"])
            ws_tpl.cell(row=r, column=5, value=vals["0"])
            ws_tpl.cell(row=r, column=7, value=vals["2"])
            ws_tpl.cell(row=r, column=9, value=vals["1"])
            ws_tpl.cell(row=r, column=11, value=vals["3"])
            
            _write_row_formulas(r)
            _style_row(r, thin_border, normal_font)
            current_row += 1

        # ── Helper: write a subtotal row with SUM formulas ────────────────
        def write_subtotal_row(label, group_start_row):
            """Write subtotal with SUM formulas referencing the range above."""
            nonlocal current_row
            r = current_row

            ws_tpl.cell(row=r, column=2, value=label)

            # SUM formulas for value columns: C(3), E(5), G(7), I(9), K(11)
            # M is now a formula (=E+G-I), so NOT included here
            for col_idx in [3, 5, 7, 9, 11]:
                col_let = ws_tpl.cell(row=r, column=col_idx).column_letter
                ws_tpl.cell(row=r, column=col_idx,
                            value=f'=SUM({col_let}{group_start_row}:{col_let}{r - 1})')

            # Percentage / calculation formulas
            _write_row_formulas(r)

            # Bold + thick border
            _style_row(r, subtotal_border, bold_font)
            current_row += 1

        # ── Track subtotal rows for grand total ──────────────────────────
        subtotal_rows = []

        # ══════════════════════════════════════════════════════════════════
        # 1. Sicilia Sections
        # ══════════════════════════════════════════════════════════════════
        for sigla in SICILIA_SIGLE:
            group = sicilia_groups[sigla]
            if not group:
                continue

            group_start = current_row

            for cname in sorted(group.keys()):
                write_data_row(cname, group[cname])

            write_subtotal_row(SIGLA_TO_NAME[sigla], group_start)
            subtotal_rows.append(current_row - 1)

        # ══════════════════════════════════════════════════════════════════
        # 2. Fuori Provincia o Assente
        # ══════════════════════════════════════════════════════════════════
        if fuori_prov_group:
            group_start = current_row
            for cname in sorted(fuori_prov_group.keys()):
                write_data_row(cname, fuori_prov_group[cname])

            write_subtotal_row("FUORI PROVINCIA O ASSENTE", group_start)
            subtotal_rows.append(current_row - 1)

        # ══════════════════════════════════════════════════════════════════
        # 3. Estero
        # ══════════════════════════════════════════════════════════════════
        if estero_group:
            group_start = current_row
            for cname in sorted(estero_group.keys()):
                write_data_row(cname, estero_group[cname])

            write_subtotal_row("ESTERO", group_start)
            subtotal_rows.append(current_row - 1)

        # ══════════════════════════════════════════════════════════════════
        # 4. Grand Total Row – SUM of subtotal rows only
        # ══════════════════════════════════════════════════════════════════
        # Skip a blank row before grand total
        current_row += 1
        r = current_row
        ws_tpl.cell(row=r, column=2, value="Sicilia")

        for col_idx in [3, 5, 7, 9, 11]:
            col_let = ws_tpl.cell(row=r, column=col_idx).column_letter
            # Build formula: =SUM(C46,C65,...) referencing only subtotal rows
            refs = ",".join([f"{col_let}{sr}" for sr in subtotal_rows])
            ws_tpl.cell(row=r, column=col_idx, value=f'=SUM({refs})')

        _write_row_formulas(r)
        _style_row(r, subtotal_border, bold_font)

        # ── Conditional formatting: bold "N/D" cells ──────────────────────
        last_data_row = current_row
        bold_nd_font = Font(bold=True)
        for col_letter in ['F', 'H', 'J', 'L', 'N', 'P']:
            cell_range = f'{col_letter}{start_row}:{col_letter}{last_data_row}'
            ws_tpl.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"N/D"'],
                           font=bold_nd_font)
            )

        # Save
        out_tpl_path = out_dir / "Sicilia_Prospetto_Variazioni_Compilato.xlsx"
        wb_tpl.save(out_tpl_path)
        
        # ── UPDATE PERSISTENT DASHBOARD ──────────────────────────────────────
        ctx.session_state['prospetto_sicilia_dashboard'] = {
            'total_processed': len(all_results),
            'count_ok': final_count_ok,
            'count_api': final_count_api,
            'count_api_failed': count_api_failed,
            'count_empty': count_empty,
            'api_results': api_log,
            'total_rinnovi': global_rinnovi,
            'sindmens_counts': global_sindmens,
            'final_file': str(out_tpl_path)
        }
        
        return [out_tpl_path]
        
    except Exception as e:
        import traceback
        full_tb = traceback.format_exc()
        ctx.error(f"❌ Errore durante il popolamento del template: {e}\n\n```\n{full_tb}\n```")
        # Update dashboard status even on error
        ctx.session_state['prospetto_sicilia_dashboard'] = {
            'total_processed': len(all_results),
            'count_ok': final_count_ok,
            'count_api': final_count_api,
            'count_api_failed': count_api_failed,
            'count_empty': count_empty,
            'api_results': api_log,
            'total_rinnovi': global_rinnovi,
            'sindmens_counts': global_sindmens,
            'final_file': None
        }
        if report_rows:
             rep_path = out_dir / "Report_Verifica_Prospetti.xlsx"
             pd.DataFrame(report_rows).to_excel(rep_path, index=False)
             return [rep_path]
        return output_files
