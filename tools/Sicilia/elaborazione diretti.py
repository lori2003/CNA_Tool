#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Tool per elaborazione anagrafiche da file Excel o SLK.
Filtra colonne specifiche, permette la selezione dei campi in output,
raggruppa per Sigla Provincia e calcola i totali parziali.
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Definizione delle colonne ammesse (come da richiesta)
ALLOWED_COLUMNS = [
    "Anno", "Cognome", "Nome", "Data di nascita", "quota", 
    "telefono", "cellulare", "e-mail", "via", "cap", 
    "comune", "sigla prov", "codice fiscale", "sede cna"
]

# Sinonimi per il riconoscimento automatico delle colonne
SYNONYMS = {
    "sigla prov": ["prov", "provincia", "pr", "sigla", "targa", "prov."],
    "comune": ["citta", "città", "paese", "luogo", "comune di nascita", "comune residenza"],
    "cap": ["c.a.p.", "zip", "code postal"],
    "e-mail": ["email", "mail", "indirizzo email", "e_mail"],
    "telefono": ["tel", "fisso", "telefono fisso", "tel."],
    "cellulare": ["cell", "mobile", "telefono cellulare", "cell."],
    "data di nascita": ["nato il", "data nascita", "nascita", "data_nascita", "d.nascita"],
    "codice fiscale": ["cf", "cod fisc", "cod.fisc.", "codice_fiscale", "codicefiscale"],
    "sede cna": ["sede", "ufficio", "struttura"],
    "quota": ["importo", "quota associativa", "valore"]
}

# Mappa per espandere le sigle nei totali
SICILY_PROVINCES = {
    "AG": "Agrigento", "CL": "Caltanissetta", "CT": "Catania", "EN": "Enna",
    "ME": "Messina", "PA": "Palermo", "RG": "Ragusa", "SR": "Siracusa", "TP": "Trapani"
}

TOOL = {'id': 'elaborazione_anagrafiche_slk',
 'name': 'Elaborazione Diretti (SLK/XLSX)',
 'description': (
    "#### 📌 1. FINALITÀ DEL TOOL\n"
    "Elabora liste anagrafiche pesanti (iscritti diretti) permettendo di filtrare i campi d'interesse, "
    "raggruppare i dati per provincia siciliana e generare prospetti riepilogativi formattati.\n\n"
    "#### 🚀 2. COME UTILIZZARLO\n"
    "1. **Input:** Carica il file (XLSX o SLK). Indica la riga d'intestazione se necessario.\n"
    "2. **Colonne:** Seleziona quali colonne vuoi mantenere nel file finale e, se vuoi, rinominale.\n"
    "3. **Esecuzione:** Il tool raggrupperà i dati, ordinandoli correttamente (es. mettendo i record mancanti alla fine).\n\n"
    "#### 🧠 3. LOGICA DI ELABORAZIONE (SPECIFICHE)\n"
    "* **Custom SLK Parser:** Implementa un motore di lettura SYLK personalizzato per gestire file non standard prodotti da vecchi software.\n"
    "* **Dynamic Parameters:** Legge il file riga per riga per mostrare le colonne reali nella dashboard in tempo reale.\n"
    "* **Geolocalizzazione:** Identifica e raggruppa gli iscritti basandosi sulla colonna 'sigla prov', espandendo le sigle (es. PA -> Palermo) nei totali.\n"
    "* **Layout Openpyxl:** Costruisce il file rigo per rigo, inserendo righi di anteprima provincia, totali parziali evidenziati in azzurro e il totale generale.\n\n"
    "#### 📂 4. RISULTATO FINALE\n"
    "Prospetto Anagrafico pulito, raggruppato e ordinato, con auto-ridimensionamento delle colonne e blocchi dei riquadri già impostati."
),
 'inputs': [{'key': 'input_file',
             'label': 'File Anagrafica (.xlsx, .xls, .slk)',
             'type': 'file_single',
             'required': True}],
 'params': [{'key': 'header_row',
             'label': 'Riga Intestazione (1 = prima riga)',
             'type': 'number',
             'default': 1,
             'min': 1,
             'step': 1,
             'help': 'Indica in quale riga si trovano i nomi delle colonne (es. Cognome, Nome...). Se il file ha righe '
                     "vuote o titoli all'inizio, aumenta questo numero."}]}

def get_standard_col_name(col: str) -> Optional[str]:
    """Restituisce il nome standard se la colonna corrisponde a un sinonimo."""
    c = str(col).strip().lower()
    # Match diretto
    for allowed in ALLOWED_COLUMNS:
        if c == allowed.lower():
            return allowed
    # Match sinonimi
    for standard, syns in SYNONYMS.items():
        if c in syns:
            return standard
    return None

def normalize_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalizza i nomi delle colonne del DataFrame per facilitare il matching:
    - usa la mappa dei sinonimi per rinominare le colonne trovate
    """
    new_cols = {}
    for col in df.columns:
        std = get_standard_col_name(col)
        if std:
            new_cols[col] = std
    
    # Rinomina le colonne trovate
    df = df.rename(columns=new_cols)
    return df

def read_sylk_custom(file_obj) -> pd.DataFrame:
    """Parser streaming per file SYLK (.slk) con gestione dello stato riga/colonna."""
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    
    data = {}
    curr_row = 1
    curr_col = 1
    
    for line_bytes in file_obj:
        line = line_bytes.decode('latin1', errors='replace').strip()
        if not line: continue
        
        parts = line.split(';')
        if parts[0] == 'C':
            val = None
            for p in parts[1:]:
                if not p: continue
                code = p[0]
                rest = p[1:]
                if code == 'Y': curr_row = int(rest)
                elif code == 'X': curr_col = int(rest)
                elif code == 'K':
                    val = rest
                    if val.startswith('"') and val.endswith('"'): val = val[1:-1]
            
            if val is not None:
                data[(curr_row, curr_col)] = val
        elif parts[0] == 'F':
            # Gestione avanzamento riga/colonna implicito se presente
            for p in parts[1:]:
                if p.startswith('Y'): curr_row = int(p[1:])
                elif p.startswith('X'): curr_col = int(p[1:])

    if not data: return pd.DataFrame()

    rows = sorted(list(set(k[0] for k in data.keys())))
    cols_idx = sorted(list(set(k[1] for k in data.keys())))
    
    matrix = []
    for r in rows:
        matrix.append([data.get((r, c), "") for c in cols_idx])

    return pd.DataFrame(matrix)

def load_data(file_obj, header_row: int = 1) -> pd.DataFrame:
    """Carica dati con rilevamento SYLK prioritario e fallback CSV."""
    filename = getattr(file_obj, "name", str(file_obj)).lower()
    header_idx = header_row - 1
    
    if hasattr(file_obj, "seek"): file_obj.seek(0)

    try:
        # 1. Controllo se è un SYLK vero (ID;P...)
        if hasattr(file_obj, "read"):
            start = file_obj.read(10).decode('latin1', errors='replace')
            file_obj.seek(0)
            if start.startswith("ID;"):
                df = read_sylk_custom(file_obj)
                if header_idx < len(df):
                    df.columns = df.iloc[header_idx]
                    return df.iloc[header_idx + 1:].reset_index(drop=True)
                return df

        # 2. Se è Excel standard
        if filename.endswith(('.xlsx', '.xls')):
            if hasattr(file_obj, "seek"): file_obj.seek(0)
            return pd.read_excel(file_obj, header=header_idx)
        
        # 3. Fallback CSV con rilevamento separatore
        if hasattr(file_obj, "seek"): file_obj.seek(0)
        return pd.read_csv(file_obj, sep=None, engine='python', header=header_idx, encoding='latin1', on_bad_lines='skip')
            
    except Exception as e:
        raise ValueError(f"Errore caricamento file: {e}")

def get_dynamic_params(inputs: dict, params: dict = None) -> List[dict]:
    """
    Legge il file caricato e restituisce le colonne reali come opzioni.
    """
    file_obj = inputs.get("input_file")
    if not file_obj:
        return []
    
    # Recupera la riga scelta dall'utente (default 1)
    h_row = 1
    if params and "header_row" in params:
        h_row = params["header_row"]

    # Non usiamo try-except qui, lasciamo che l'errore salga ad app.py
    # così l'utente vede il messaggio di errore (es. formato non valido)
    df = load_data(file_obj, header_row=h_row)
    df = normalize_headers(df)
    cols = list(df.columns)
    
    # Pre-seleziona quelle note
    default_sel = [c for c in cols if c in ALLOWED_COLUMNS]
    if not default_sel:
        default_sel = cols[:5] # Fallback

    dyn_params = [
        {
            'key': 'columns_to_keep',
            'label': f'Seleziona Colonne (Trovate {len(cols)})',
            'type': 'multiselect',
            'options': cols,
            'default': [],
            'help': 'Queste sono le colonne lette dal file.'
        }
    ]
    
    # Aggiunge dinamicamente i campi per rinominare le colonne selezionate
    selected_cols = params.get("columns_to_keep", [])
        
    for col in selected_cols:
        dyn_params.append({
            'key': f'rename_{col}',
            'label': f'Rinomina "{col}"',
            'type': 'text',
            'default': col
        })
        
    return dyn_params

def run(input_file: Path, columns_to_keep: List[str], header_row: int, out_dir: Path, **kwargs) -> List[Path]:
    """
    Funzione principale di esecuzione.
    """
    if not input_file:
        raise ValueError("Nessun file caricato.")
    
    if not columns_to_keep:
        raise ValueError("Devi selezionare almeno una colonna da visualizzare.")

    # 1. Caricamento Dati
    df = load_data(input_file, header_row=header_row)
    
    # 2. Normalizzazione Colonne
    df = normalize_headers(df)
    
    # 3. Filtro: tengo solo le colonne che esistono sia nel file che nella selezione utente
    # Nota: "sigla prov" serve per il raggruppamento, ci assicuriamo di averla
    available_cols = [c for c in columns_to_keep if c in df.columns]
    
    if not available_cols:
        # Se l'utente ha selezionato colonne che ora non ci sono (improbabile con dynamic params), errore.
        raise ValueError(f"Colonne richieste non trovate. Disponibili: {list(df.columns)}")
    
    # Assicuriamoci che 'sigla prov' sia presente per l'ordinamento, anche se l'utente non la vuole stampare
    # (La useremo per la logica, poi decideremo se stamparla)
    group_col = "sigla prov"
    has_group_col = group_col in df.columns
    
    cols_to_load = list(available_cols)
    if has_group_col and group_col not in cols_to_load:
        cols_to_load.append(group_col)
        
    df_filtered = df[cols_to_load].copy()
    
    # 4. Ordinamento e Gestione Mancanti
    if has_group_col:
        # Normalizza la sigla prov
        df_filtered[group_col] = df_filtered[group_col].fillna("").astype(str).str.upper().str.strip()
        # Gestione valori vuoti -> "Mancante"
        df_filtered.loc[df_filtered[group_col] == "", group_col] = "Mancante"
        
        # Logica di ordinamento: 0 per valori normali, 1 per "Mancante" (così va alla fine)
        df_filtered["_sort_key"] = df_filtered[group_col].apply(lambda x: 1 if x == "Mancante" else 0)
        df_filtered = df_filtered.sort_values(by=["_sort_key", group_col])
        df_filtered = df_filtered.drop(columns=["_sort_key"])
    
    # Riempie celle vuote o NaN in tutto il dataframe con "Mancante"
    df_filtered = df_filtered.fillna("Mancante")
    df_filtered = df_filtered.replace(r'^\s*$', 'Mancante', regex=True)
    
    # 5. Creazione Excel con Openpyxl (per gestire i totali e formattazione)
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospetto Riepilogativo"
    
    # Stili
    bold_font = Font(bold=True)
    total_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Mappa per rinominare le colonne
    rename_map = {}
    for col in columns_to_keep:
        # Cerca il parametro rename_{col} passato come argomento dinamico
        new_name = kwargs.get(f'rename_{col}', col)
        if new_name and str(new_name).strip():
            rename_map[col] = str(new_name).strip()
        else:
            rename_map[col] = col

    # Scrittura Intestazioni
    # Usiamo l'ordine scelto dall'utente
    headers = list(columns_to_keep)
    
    # Assicura che "sigla prov" sia sempre la prima colonna (se presente nel file)
    if "sigla prov" in headers:
        headers.remove("sigla prov")
        headers.insert(0, "sigla prov")
    elif has_group_col:
        headers.insert(0, "sigla prov")
    
    # Aggiunge colonna dedicata per i totali
    headers.append("T.Record")

    for col_idx, header in enumerate(headers, 1):
        display_name = rename_map.get(header, header)
        cell = ws.cell(row=1, column=col_idx, value=display_name)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Scrittura Dati e Totali
    current_row = 2
    current_prov = None
    prov_count = 0
    grand_total = 0
    
    # Indice della colonna su cui basare il raggruppamento (se presente nella selezione output)
    # Se l'utente non ha selezionato "sigla prov", i totali appariranno comunque "al cambio di gruppo"
    # ma visivamente potrebbe essere meno chiaro se la colonna manca.
    
    records = df_filtered.to_dict('records')
    
    for i, record in enumerate(records):
        this_prov = record.get(group_col, "N/D") if has_group_col else "Generale"
        
        # Se cambia la provincia (e non è la prima riga), stampa totale precedente
        if has_group_col and current_prov is not None and this_prov != current_prov:
            # Espansione nome provincia (es. PA -> Palermo)
            prov_name = SICILY_PROVINCES.get(current_prov, current_prov)
            
            # Riga Totale Provincia
            last_col_idx = len(headers)
            
            text_cell = ws.cell(row=current_row, column=1, value=prov_name)
            text_cell.font = bold_font
            text_cell.alignment = center_align

            if last_col_idx > 0:
                num_cell = ws.cell(row=current_row, column=last_col_idx, value=prov_count)
                num_cell.font = bold_font
                num_cell.alignment = center_align

            if last_col_idx > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col_idx - 1)
            
            # Stile riga totale
            for c in range(1, last_col_idx + 1):
                c_cell = ws.cell(row=current_row, column=c)
                c_cell.border = thin_border
                c_cell.fill = total_fill
                
            current_row += 1
            prov_count = 0 # Reset contatore
        
        current_prov = this_prov
        prov_count += 1
        grand_total += 1
        
        # Scrivi la riga dati
        for col_idx, col_name in enumerate(headers, 1):
            val = record.get(col_name, "")
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.alignment = left_align
            # Nessun bordo per le righe dati, per pulizia grafica
            
        current_row += 1
        
    # Stampa ultimo totale provincia
    if has_group_col and current_prov is not None:
        prov_name = SICILY_PROVINCES.get(current_prov, current_prov)
        last_col_idx = len(headers)

        text_cell = ws.cell(row=current_row, column=1, value=prov_name)
        text_cell.font = bold_font
        text_cell.alignment = center_align

        if last_col_idx > 0:
            num_cell = ws.cell(row=current_row, column=last_col_idx, value=prov_count)
            num_cell.font = bold_font
            num_cell.alignment = center_align

        if last_col_idx > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col_idx - 1)

        for c in range(1, last_col_idx + 1):
            c_cell = ws.cell(row=current_row, column=c)
            c_cell.border = thin_border
            c_cell.fill = total_fill
        current_row += 1

    # Riga Totale Generale
    current_row += 1
    last_col_idx = len(headers)

    text_cell = ws.cell(row=current_row, column=1, value="Totale Complessivo")
    text_cell.font = Font(bold=True, size=12)
    text_cell.alignment = center_align

    if last_col_idx > 0:
        num_cell = ws.cell(row=current_row, column=last_col_idx, value=grand_total)
        num_cell.font = Font(bold=True, size=12)
        num_cell.alignment = center_align

    if last_col_idx > 1:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=last_col_idx - 1)

    for c in range(1, last_col_idx + 1):
        c_cell = ws.cell(row=current_row, column=c)
        c_cell.border = thin_border
        c_cell.fill = total_fill
        
    # Auto-width colonne
    for i, col in enumerate(headers, 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 20

    # Salvataggio
    out_filename = "Prospetto_Anagrafica_Riepilogo.xlsx"
    out_path = out_dir / out_filename
    wb.save(out_path)
    
    return [out_path]
